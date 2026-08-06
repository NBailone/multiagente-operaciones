"""ControlMixin — Panel Cargar Datos / Control (T3+).

Legacy block extracted from ui_app.py. Mixed into the main App class; relies
on attributes/methods living on the composed instance (_log, _set_log_panel,
_cfg_obtener, _cfg_obtener_correo, _cfg_obtener_docs, _cfg_obtener_rutas,
_resolver_ruta, _abrir_excel_seguro, _guardar_excel_seguro, _mostrar_resumen,
_abrir_comparacion, _limpiar_log, _restaurar_log, _set_status, _icons,
_panel_frames, panel_container, tarea_activa, log_queue). Also imports
procesar_tickets locally within methods.
"""

import os
import re
import sys
import threading
import traceback
from datetime import datetime

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import xlrd

from constants import Palette, FONT_FAMILY


class ControlMixin:
    # ═══════════════════════════════════════════════════════════════════
    # PANEL: CARGA DE DATOS (T3+)
    # ═══════════════════════════════════════════════════════════════════
    def _panel_cargar_datos(self):
        if "cargar-datos" in self._panel_frames and hasattr(self, 'tree_coordinacion'):
            return
        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        # NOTA: _panel_frames se setea al FINAL del método para evitar
        # que una inicialización incompleta deje el panel huérfano.
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── T11: Verificar dependencias (Tesseract + Poppler) ───────────
        dependencias_ok = True
        errores_dep = []

        # ── Configurar rutas de binarios ──────────────────────────
        _app_base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        _res_base = getattr(sys, '_MEIPASS', _app_base)
        _POPPLER_BIN = os.path.join(_res_base, "poppler", "Library", "bin")
        _TESSERACT_EXE = os.path.join(_res_base, "engines", "tesseract", "tesseract.exe")
        if not os.path.isfile(_TESSERACT_EXE):
            _TESSERACT_EXE = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

        # Verificar Tesseract (solo archivo, sin ejecutarlo — evita ventana DOS)
        import pytesseract
        tesseract_sidecar = os.path.join(
            getattr(sys, '_MEIPASS', os.path.dirname(__file__)),
            "engines", "tesseract", "tesseract.exe")
        tesseract_paths = [
            tesseract_sidecar,
            _TESSERACT_EXE,
        ]
        if not any(os.path.isfile(p) for p in tesseract_paths):
            dependencias_ok = False
            errores_dep.append(
                "⚠ Tesseract-OCR no está instalado. "
                "Descargalo de https://github.com/UB-Mannheim/tesseract/wiki"
            )
        else:
            pytesseract.pytesseract.tesseract_cmd = next(p for p in tesseract_paths if os.path.isfile(p))

        # Verificar Poppler (pdf2image)
        poppler_binarios = [
            os.path.join(_POPPLER_BIN, "pdftoppm.exe"),
            os.path.join(_POPPLER_BIN, "pdfinfo.exe"),
        ]
        if not any(os.path.isfile(b) for b in poppler_binarios):
            dependencias_ok = False
            errores_dep.append(
                "⚠ Poppler no está instalado. "
                "Descargalo de https://github.com/oschwartz10612/poppler-windows/releases/"
            )

        if not dependencias_ok:
            error_frame = ctk.CTkFrame(
                frame, fg_color="#FFE0E0", corner_radius=8,
                border_width=1, border_color="#FFB0B0",
            )
            error_frame.pack(fill="x", pady=(0, 6))

            for msg in errores_dep:
                ctk.CTkLabel(
                    error_frame, text=msg,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                    text_color="red", wraplength=600, justify="left",
                ).pack(padx=12, pady=6, anchor="w")

            # ── Toolbar con botones deshabilitados ──────────────────────
            toolbar = ctk.CTkFrame(
                frame, fg_color=Palette.BG_CARD, corner_radius=8,
                border_width=1, border_color=Palette.BORDER, height=44
            )
            toolbar.pack(fill="x", pady=(0, 6))
            toolbar.pack_propagate(False)

            self.btn_controlar_tickets = ctk.CTkButton(
                toolbar,
                text="Controlar Tickets",
                image=self._icons["ticket"], compound="left",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.WHITE, corner_radius=6, height=34, width=170,
                state="disabled",
            )
            self.btn_controlar_tickets.pack(side="left", padx=4, pady=4)

            # Espacio para mantener layout
            ctk.CTkLabel(
                toolbar,
                text="Dependencias faltantes — resolver antes de continuar",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.ERROR,
            ).pack(side="left", padx=(8, 0))

            # ── Resultados vacío ──────────────────────────────────────
            scroll = ctk.CTkScrollableFrame(
                frame, fg_color=Palette.BG_CARD, corner_radius=8,
                border_width=1, border_color=Palette.BORDER
            )
            scroll.pack(fill="both", expand=True, pady=(0, 6))

            self._panel_frames["cargar-datos"] = frame
            return  # No seguir con el panel normal

        # ── Toolbar ──────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=44
        )
        toolbar.pack(fill="x", pady=(0, 6))
        toolbar.pack_propagate(False)

        # ── Grupo Control Tickets ──────────────────────────────────
        self.btn_controlar_tickets = ctk.CTkButton(
            toolbar,
            text="Controlar Tickets",
            image=self._icons["ticket"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            command=self._cargar_datos_seleccionar_pdfs,
        )
        self.btn_controlar_tickets.pack(side="left", padx=(10, 2), pady=5)

        self._ct_auto_switch = ctk.CTkSwitch(
            toolbar, text="Auto",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            variable=self._ct_auto_var,
            progress_color=Palette.ACCENT,
            button_color=Palette.TEXT_SECONDARY,
            button_hover_color=Palette.TEXT_PRIMARY,
            command=self._cargar_datos_switch_toggle,
        )
        self._ct_auto_switch.pack(side="left", padx=2, pady=5)
        # Restore persisted state
        if self.config.get("cargar_datos_auto", False):
            self._ct_auto_var.set(True)
            self.btn_controlar_tickets.configure(text="Controlar Tickets ⚡", image=self._icons["zap"])

        # ── Separador vertical ──────────────────────────────────────
        ctk.CTkFrame(
            toolbar, width=2, height=34, fg_color=Palette.BORDER
        ).pack(side="left", padx=6, pady=5)

        self.btn_controlar_coordinacion = ctk.CTkButton(
            toolbar,
            text="Controlar Coordinación",
            image=self._icons["clipboard-list"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            command=self._controlar_coordinacion,
        )
        self.btn_controlar_coordinacion.pack(side="left", padx=2, pady=5)

        # ── Separador vertical ──────────────────────────────────────
        ctk.CTkFrame(
            toolbar, width=2, height=34, fg_color=Palette.BORDER
        ).pack(side="left", padx=6, pady=5)

        self.btn_control_final = ctk.CTkButton(
            toolbar,
            text="Control Final",
            image=self._icons["zap"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.SECONDARY, hover_color=Palette.SECONDARY_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            state="normal",
            command=self._control_final_seleccionar,
        )
        self.btn_control_final.pack(side="left", padx=2, pady=5)

        self._cf_auto_switch = ctk.CTkSwitch(
            toolbar, text="Auto",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            variable=self._cf_auto_var,
            progress_color=Palette.ACCENT,
            button_color=Palette.TEXT_SECONDARY,
            button_hover_color=Palette.TEXT_PRIMARY,
            command=self._control_final_switch_toggle,
        )
        self._cf_auto_switch.pack(side="left", padx=(0, 6), pady=5)

        # Restore auto mode state from config
        if self.config.get("control_final_auto", False):
            self._cf_auto_var.set(True)
            self.btn_control_final.configure(text="Control Final ⚡", image=self._icons["zap"])

        self.progress_carga = ctk.CTkProgressBar(
            toolbar, width=140, height=8, corner_radius=4,
            mode="indeterminate",
            fg_color=Palette.BG_INPUT, progress_color=Palette.ACCENT,
        )
        # Oculto por defecto — se muestra al iniciar procesamiento

        self.btn_limpiar_carga = ctk.CTkButton(
            toolbar,
            text="Limpiar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color="transparent",
            hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            border_width=0,
            corner_radius=4,
            height=30,
            width=70,
            command=self._limpiar_carga,
        )
        self.btn_limpiar_carga.pack(side="right", padx=4, pady=4)

        # ── Tab bar compacta ────────────────────────────────
        tab_bar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_SIDEBAR, corner_radius=8, height=32,
            border_width=1, border_color=Palette.BORDER,
        )
        tab_bar.pack(fill="x", pady=(6, 2))
        tab_bar.pack_propagate(False)

        tab_font = ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold")
        _TAB_STEEL = "#546e7a"  # Distinct from toolbar blue/teal

        self._tab_tickets = ctk.CTkLabel(
            tab_bar, text="Control de Tickets", font=tab_font,
            fg_color=_TAB_STEEL, text_color=Palette.WHITE,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_tickets.pack(side="left", padx=4, pady=2)
        self._tab_tickets.bind("<Button-1>", lambda e: self._switch_tab("tickets"))

        self._tab_coord = ctk.CTkLabel(
            tab_bar, text="Coordinación", font=tab_font,
            fg_color="transparent", text_color=_TAB_STEEL,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_coord.pack(side="left", padx=2, pady=2)
        self._tab_coord.bind("<Button-1>", lambda e: self._switch_tab("coord"))

        self._tab_final = ctk.CTkLabel(
            tab_bar, text="Control Final", font=tab_font,
            fg_color="transparent", text_color=_TAB_STEEL,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_final.pack(side="left", padx=2, pady=2)
        self._tab_final.bind("<Button-1>", lambda e: self._switch_tab("final"))

        # ── Resultados (scroll frame + sections) ────────────────
        scroll = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER
        )
        scroll.pack(fill="both", expand=True, pady=(0, 6))

        # Treeview style oscuro
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "CargaDatos.Treeview",
            background=Palette.BG_TABLE,
            foreground=Palette.TEXT_PRIMARY,
            fieldbackground=Palette.BG_TABLE,
            borderwidth=0,
            font=(FONT_FAMILY, 10),
            rowheight=28,
        )
        style.map(
            "CargaDatos.Treeview",
            background=[("selected", Palette.ACCENT_DIM)],
            foreground=[("selected", Palette.WHITE)],
        )
        style.configure(
            "CargaDatos.Treeview.Heading",
            background=Palette.BG_SIDEBAR,
            foreground=Palette.TEXT_SECONDARY,
            font=(FONT_FAMILY, 10, "bold"),
            borderwidth=0,
            padding=(8, 6),
        )

        # ── Sección: Tickets ──────────────────────────────────────────
        self._section_tickets = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        columns = ("cliente", "camion", "semi", "conductor", "dni",
                   "neto", "tara", "contenedor", "permiso", "estado")
        headers = ("Cliente", "Camion", "Semi", "Conductor", "DNI",
                   "Neto", "Tara", "Contenedor", "Permiso", "Estado")
        anchos_default = (120, 80, 80, 100, 80, 70, 70, 100, 100, 100)
        _saved_cols = self.config.get("tree_carga_columns", {})

        self.tree_carga = ttk.Treeview(
            self._section_tickets, columns=columns, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(columns, headers, anchos_default):
            self.tree_carga.heading(col, text=hdr)
            w = _saved_cols.get(col, w_default)
            self.tree_carga.column(col, width=w, anchor="center", minwidth=40)

        # Scrollbar vertical
        scroll_y = ctk.CTkScrollbar(
            self._section_tickets, orientation="vertical",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        scroll_y.pack(side="right", fill="y", padx=(0, 2), pady=2)

        self.tree_carga.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self.tree_carga.yview)

        # Scrollbar horizontal
        self.scroll_carga_x = ctk.CTkScrollbar(
            self._section_tickets, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_carga.configure(xscrollcommand=self.scroll_carga_x.set)
        self.scroll_carga_x.configure(command=self.tree_carga.xview)
        self.scroll_carga_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_carga.pack(fill="both", expand=True, padx=2, pady=(0, 0))

        # Tags
        self.tree_carga.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_carga.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_carga.tag_configure("tag_no_match", background="#FFF3E0", foreground="#E65100")
        self.tree_carga.bind("<Double-1>", self._abrir_comparacion)

        # ── Sección: Coordinación ISO/FLEXI ──────────────────────────
        self._section_coord = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        cols_c = ("carpeta", "giro", "cliente", "destino",
                  "buque", "viaje", "booking", "pto_descarga", "pto_final",
                  "fecha_of_pe", "fecha_carga", "peso_flexi", "estado")
        hdrs_c = ("Carpeta", "Giro", "Cliente", "Destino",
                  "Buque", "Viaje", "Booking", "Pto Descarga", "Pto Final",
                  "Fec Of PE", "Fec Carga", "Peso Flexi", "Estado")
        anchos_c_default = (100, 70, 220, 120, 180, 70, 160, 120, 120, 110, 110, 80, 80)
        _saved_cols_c = self.config.get("tree_coordinacion_columns", {})

        self.tree_coordinacion = ttk.Treeview(
            self._section_coord, columns=cols_c, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(cols_c, hdrs_c, anchos_c_default):
            self.tree_coordinacion.heading(col, text=hdr)
            w = _saved_cols_c.get(col, w_default)
            self.tree_coordinacion.column(col, width=w, anchor="center", minwidth=40)

        self.tree_coordinacion.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_coordinacion.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_coordinacion.tag_configure("tag_error", background="#FFE0B0", foreground="#8B4500")

        self.tree_coordinacion.bind("<Double-1>", self._abrir_comparacion_coordinacion)

        # Scrollbar horizontal
        self.scroll_coord_x = ctk.CTkScrollbar(
            self._section_coord, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_coordinacion.configure(xscrollcommand=self.scroll_coord_x.set)
        self.scroll_coord_x.configure(command=self.tree_coordinacion.xview)
        self.scroll_coord_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_coordinacion.pack(fill="both", expand=True, padx=2, pady=(0, 8))

        # ── Sección: Control Final ──────────────────────────────────
        self._section_final = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        cols_f = ("cliente", "camion", "semi", "conductor", "dni",
                  "neto", "tara", "contenedor", "permiso", "estado", "salida_aduana")
        hdrs_f = ("Cliente", "Camion", "Semi", "Conductor", "DNI",
                  "Neto", "Tara", "Contenedor", "Permiso", "Estado", "Salida Aduana")
        anchos_f_default = (120, 80, 80, 100, 80, 70, 70, 100, 100, 100, 100)
        _saved_cols_f = self.config.get("tree_control_final_columns", {})

        self.tree_control_final = ttk.Treeview(
            self._section_final, columns=cols_f, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(cols_f, hdrs_f, anchos_f_default):
            self.tree_control_final.heading(col, text=hdr)
            w = _saved_cols_f.get(col, w_default)
            self.tree_control_final.column(col, width=w, anchor="center", minwidth=40)

        self.tree_control_final.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_control_final.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_control_final.tag_configure("tag_no_match", background="#FFF3E0", foreground="#E65100")
        self.tree_control_final.bind("<Double-1>", self._abrir_comparacion_final)

        # Scrollbar horizontal
        self.scroll_control_final_x = ctk.CTkScrollbar(
            self._section_final, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_control_final.configure(xscrollcommand=self.scroll_control_final_x.set)
        self.scroll_control_final_x.configure(command=self.tree_control_final.xview)
        self.scroll_control_final_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_control_final.pack(fill="both", expand=True, padx=2, pady=(0, 8))

        for section in (self._section_tickets, self._section_coord, self._section_final):
            section.pack(fill="both", expand=True, pady=(0, 6))
        self._section_coord.pack_forget()
        self._section_final.pack_forget()

        # Todo listo — registrar el frame como completo en _panel_frames
        self._panel_frames["cargar-datos"] = frame

    def _switch_tab(self, tab):
        """Switch visible section and highlight active tab."""
        _TAB_STEEL = "#546e7a"
        tabs = {
            "tickets": (self._tab_tickets, self._section_tickets),
            "coord":   (self._tab_coord,   self._section_coord),
            "final":   (self._tab_final,   self._section_final),
        }
        for name, (label, section) in tabs.items():
            if name == tab:
                label.configure(fg_color=_TAB_STEEL, text_color=Palette.WHITE)
                section.pack(fill="both", expand=True, pady=(0, 6))
            else:
                label.configure(fg_color="transparent", text_color=_TAB_STEEL)
                section.pack_forget()

    # ── Limpiar vista — planillas ──────────────────────────────────────
    def _limpiar_planillas(self):
        """Elimina filas y resetea estado del panel Planillas."""
        for row in self.tree_planillas.get_children():
            self.tree_planillas.delete(row)
        self.lbl_resumen_planillas.configure(text="Sin datos analizados")
        self.progress_planillas.set(0)

    # ── Limpiar vista — descargar ──────────────────────────────────────
    def _limpiar_mail(self):
        """Elimina filas y resetea estado del panel Descargar Mails."""
        for row in self._mail_tree.get_children():
            self._mail_tree.delete(row)
        self._mail_data.clear()
        self._mail_lbl_estado.configure(text="")
        self._mail_progress.set(0)

    # ── Limpiar vista — correos ────────────────────────────────────────
    def _limpiar_correos(self):
        """Elimina filas y resetea estado del panel Enviar Correos."""
        for row in self.tree_correos.get_children():
            self.tree_correos.delete(row)
        self.lbl_estado_correos.configure(text="Listo para despachar borradores")
        self.progress_correos.set(0)

    # ── Limpiar vista — cargar datos ───────────────────────────────────
    def _limpiar_carga(self):
        """Elimina todas las filas y resetea a vista tickets."""
        self.tree_carga.delete(*self.tree_carga.get_children())
        if hasattr(self, 'tree_control_final'):
            self.tree_control_final.delete(*self.tree_control_final.get_children())
            self.tree_control_final.heading("salida_aduana", text="Salida Aduana")
        if hasattr(self, 'tree_coordinacion'):
            self.tree_coordinacion.delete(*self.tree_coordinacion.get_children())
        # Resetear a vista tickets
        self._switch_tab("tickets")

    # ── Coordinación ISO/FLEXI ────────────────────────────────────────
    def _escanear_carpetas_coordinacion(self):
        """Escanea Desktop + 1 nivel buscando carpetas con CONTENEDORES Excel y PDF coordinación."""
        base = self._resolver_ruta("planillas_carga", "Desktop")
        carpetas = []

        def _tiene_coordinacion(nombre):
            """Detecta 'coordinacion' en el nombre sin importar acentos."""
            return "coordinaci" in nombre.lower()

        # 1 nivel de subcarpetas en Desktop
        for item in os.listdir(base):
            ruta_dir = os.path.join(base, item)
            if not os.path.isdir(ruta_dir):
                continue

            # Buscar CONTENEDORES Excel, PDF Coordinación, y Permiso dentro
            xls_path = None
            pdf_path = None
            permiso_path = None
            for f in os.listdir(ruta_dir):
                f_lower = f.lower()
                if "contenedor" in f_lower and (f_lower.endswith(".xlsx") or f_lower.endswith(".xls")):
                    xls_path = os.path.join(ruta_dir, f)
                elif f_lower.endswith(".pdf"):
                    if _tiene_coordinacion(f):
                        pdf_path = os.path.join(ruta_dir, f)
                    elif permiso_path is None:
                        permiso_path = os.path.join(ruta_dir, f)

            if xls_path:  # Si tiene Excel, lo consideramos
                carpetas.append({
                    "nombre": item,
                    "ruta": ruta_dir,
                    "pdf_path": pdf_path,
                    "xls_path": xls_path,
                    "permiso_path": permiso_path,
                    "pdf_ok": pdf_path is not None,
                    "xls_ok": True,
                    "permiso_ok": permiso_path is not None,
                })

        carpetas.sort(key=lambda c: c["nombre"])
        return carpetas

    def _controlar_coordinacion(self):
        """Escanea Desktop, muestra popup con checkboxes, procesa seleccionadas."""
        from procesar_tickets import (
            extraer_coordinacion, leer_choferes_coordinacion, comparar_coordinacion,
            extraer_fecha_permiso,
        )
        import unicodedata
        import tkinter as tk

        # 1. Escanear carpetas
        carpetas = self._escanear_carpetas_coordinacion()
        if not carpetas:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con archivos CONTENEDORES en el Desktop."
            )
            return

        # 2. Popup de selección con checkboxes
        seleccion = []

        top = ctk.CTkToplevel(self)
        top.title("Seleccionar envíos ISO/FLEXI para controlar")
        ancho, alto = 960, 520
        top.geometry(f"{ancho}x{alto}")
        top.transient(self)
        top.grab_set()
        top.update_idletasks()
        x = (top.winfo_screenwidth() - ancho) // 2
        y = (top.winfo_screenheight() - alto) // 2
        top.geometry(f"{ancho}x{alto}+{x}+{y}")

        # Encabezado
        ctk.CTkLabel(
            top,
            text="Carpetas con CONTENEDORES encontradas en el Desktop",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(12, 2))

        ctk.CTkLabel(
            top,
            text="Solo se pueden seleccionar carpetas con PDF + Excel presentes",
            font=ctk.CTkFont(family=FONT_FAMILY, size=10),
            text_color=Palette.TEXT_MUTED,
        ).pack(pady=(0, 8))

        # Scroll frame para la lista
        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        # Columnas fijas — una sola grilla compartida en el scroll
        cols_spec = [
            ("", 24, (6, 2)),           # checkbox
            ("Carpeta / Envío", 480, (4, 4)),
            ("PDF", 50, (0, 0)),
            ("Excel", 50, (0, 0)),
            ("Estado", 200, (4, 4)),
        ]

        # Configurar columnas UNA SOLA VEZ en el scroll
        for ci, (txt, w, pad) in enumerate(cols_spec):
            scroll.grid_columnconfigure(ci, minsize=w, weight=0)

        # Header de columnas — widgets directo en scroll, row 0
        for ci, (txt, w, pad) in enumerate(cols_spec):
            ctk.CTkLabel(
                scroll, text=txt, width=w,
                font=ctk.CTkFont(family=FONT_FAMILY, size=10, weight="bold"),
                text_color=Palette.TEXT_SECONDARY, anchor="w",
            ).grid(row=0, column=ci, padx=pad, pady=(8, 4), sticky="w")

        # Línea separadora — row 1
        sep = ctk.CTkFrame(scroll, height=1, fg_color=Palette.BORDER)
        sep.grid(row=1, column=0, columnspan=5, sticky="ew", padx=4, pady=2)

        vars_check = []

        for i, carp in enumerate(carpetas):
            row = i + 2
            puede = carp["pdf_ok"] and carp["xls_ok"]
            bg_fila = Palette.BG_CARD

            var = tk.BooleanVar(value=puede)
            vars_check.append(var)

            chk = ctk.CTkCheckBox(
                scroll, text="", variable=var, width=24,
                checkbox_width=20, checkbox_height=20,
                state="normal" if puede else "disabled",
                fg_color=Palette.ACCENT if puede else Palette.TEXT_MUTED,
            )
            chk.grid(row=row, column=0, padx=cols_spec[0][2], pady=3)
            if not puede:
                chk.configure(text=" " * 2)

            # Carpeta — tk.Label trunca al ancho de columna (no empuja ✅/❌)
            nombre_mostrar = carp["nombre"] if len(carp["nombre"]) <= 55 else carp["nombre"][:52] + "..."
            tk.Label(
                scroll, text=nombre_mostrar, width=58, anchor="w",
                font=(FONT_FAMILY, 10), fg=Palette.TEXT_PRIMARY,
                bg=bg_fila, bd=0, highlightthickness=0,
            ).grid(row=row, column=1, padx=cols_spec[1][2], sticky="w", pady=3)

            # PDF
            pdf_txt = "✅" if carp["pdf_ok"] else "❌"
            pdf_color = Palette.SUCCESS if carp["pdf_ok"] else Palette.ERROR
            ctk.CTkLabel(
                scroll, text=pdf_txt, width=50,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=pdf_color, anchor="center",
            ).grid(row=row, column=2, padx=cols_spec[2][2], pady=3)

            # Excel
            ctk.CTkLabel(
                scroll, text="✅", width=50,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.SUCCESS, anchor="center",
            ).grid(row=row, column=3, padx=cols_spec[3][2], pady=3)

            # Estado
            estado_txt = "✅ Listo para controlar" if puede else "❌ No se encontró PDF de Coordinación"
            ctk.CTkLabel(
                scroll, text=estado_txt, width=cols_spec[4][1],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                text_color=Palette.TEXT_MUTED, anchor="w",
            ).grid(row=row, column=4, padx=cols_spec[4][2], sticky="w", pady=3)

        # Botones
        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))

        def confirmar():
            nonlocal seleccion
            for i, (var, carp) in enumerate(zip(vars_check, carpetas)):
                if var.get() and carp["pdf_ok"] and carp["xls_ok"]:
                    seleccion.append(carp)
            top.destroy()

        def cancelar():
            top.destroy()

        ctk.CTkButton(
            btn_frame, text="Procesar seleccionados",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=190,
            command=confirmar,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="Cancelar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6, height=34, width=120,
            command=cancelar,
        ).pack(side="right", padx=4)

        top.wait_window()

        if not seleccion:
            return

        # 3. Procesar seleccionadas
        if not hasattr(self, 'tree_coordinacion'):
            self._panel_cargar_datos()
            if not hasattr(self, 'tree_coordinacion'):
                self.log("ERROR: Panel de carga no disponible. Reinicie la aplicación.")
                return
        self.tree_coordinacion.delete(*self.tree_coordinacion.get_children())
        self._coord_comparaciones = {}

        # Mostrar sección coordinación, ocultar otras
        self._switch_tab("coord")

        ok_count = 0
        diff_count = 0
        error_count = 0

        for carp in seleccion:
            try:
                datos_pdf = extraer_coordinacion(carp["pdf_path"])
                if "_error" in datos_pdf:
                    self._insertar_fila_coordinacion(
                        carp["nombre"], "⚠ Error en PDF: " + datos_pdf["_error"],
                        "tag_error", None
                    )
                    error_count += 1
                    continue

                datos_xls = leer_choferes_coordinacion(carp["xls_path"])
                if "_error" in datos_xls:
                    self._insertar_fila_coordinacion(
                        carp["nombre"], "⚠ " + datos_xls["_error"],
                        "tag_error", None
                    )
                    error_count += 1
                    continue

                # OCR del permiso de exportación para fecha de oficialización
                if carp.get("permiso_path"):
                    fecha_of = extraer_fecha_permiso(carp["permiso_path"])
                    datos_pdf["fecha_of_pe"] = fecha_of
                else:
                    datos_pdf["fecha_of_pe"] = ""

                comparaciones, estado = comparar_coordinacion(datos_pdf, datos_xls)

                tag = "tag_ok" if estado == "ok" else "tag_mismatch"
                estado_label = "✅ OK" if estado == "ok" else "❌ Diferencias"

                valores = (
                    datos_pdf.get("carpeta", ""),
                    datos_pdf.get("giro", ""),
                    datos_pdf.get("cliente", ""),
                    datos_pdf.get("destino", ""),
                    datos_pdf.get("buque", ""),
                    datos_pdf.get("viaje", ""),
                    datos_pdf.get("booking", ""),
                    datos_pdf.get("pto_descarga", ""),
                    datos_pdf.get("pto_final", ""),
                    comparaciones.get("fecha_of_pe", {}).get("pdf", ""),
                    datos_xls.get("fecha_carga", ""),
                    datos_pdf.get("peso_flexi", ""),
                    estado_label,
                )

                iid = self.tree_coordinacion.insert("", "end", values=valores, tags=(tag,))
                self._coord_comparaciones[iid] = comparaciones

                if estado == "ok":
                    ok_count += 1
                else:
                    diff_count += 1

            except Exception as e:
                self._insertar_fila_coordinacion(
                    carp["nombre"], "⚠ Error: " + str(e),
                    "tag_error", None
                )
                error_count += 1

        total = ok_count + diff_count + error_count

    def _insertar_fila_coordinacion(self, carpeta, estado_texto, tag, comparaciones):
        """Inserta una fila en el treeview de coordinación."""
        valores = (carpeta, "", "", "", "", "", "", "", "", "", "", "", estado_texto)
        iid = self.tree_coordinacion.insert("", "end", values=valores, tags=(tag,))
        if comparaciones is not None:
            self._coord_comparaciones[iid] = comparaciones

    def _abrir_comparacion_coordinacion(self, event):
        """Popup con tabla verde/rojo campo por campo — mismo estilo que el de tickets."""
        sel = self.tree_coordinacion.selection()
        if not sel or not hasattr(self, "_coord_comparaciones"):
            return

        comps = self._coord_comparaciones.get(sel[0])
        if not comps:
            return

        etiquetas = {
            "giro": "Puerto Salida",
            "carpeta": "Carpeta", "cliente": "Cliente",
            "destino": "Destino", "buque": "Buque", "viaje": "Viaje",
            "booking": "Booking", "pto_descarga": "Pto Descarga",
            "pto_final": "Pto Final",
            "fecha_of_pe": "Fecha Ofic.",
            "fecha_carga": "Fecha Carga",
            "peso_flexi": "Peso Flexi (kg)",
        }

        def _build_popup(level):
            fsizes = self._get_font_sizes(level)
            geom = self._get_popup_geometry(760, 480, level)
            ancho = int(geom.split("x")[0])

            dlg = ctk.CTkToplevel(self)
            dlg.title("Comparación — Coordinación vs Excel")
            dlg.transient(self)
            dlg.resizable(False, False)

            # ── Top bar: header + font level override ──────────────────
            top_bar = ctk.CTkFrame(dlg, fg_color=Palette.BG_SIDEBAR, corner_radius=6, height=36)
            top_bar.pack(fill="x", padx=12, pady=(12, 0))
            top_bar.pack_propagate(False)

            # Grid header: align with body columns
            top_bar.grid_columnconfigure(0, minsize=120, weight=0)
            top_bar.grid_columnconfigure(1, weight=1)
            top_bar.grid_columnconfigure(2, weight=1)

            for col_i, txt in enumerate(["Campo", "PDF (Coordinación)", "Excel (Choferes)"]):
                lbl = ctk.CTkLabel(
                    top_bar, text=txt, width=120 if col_i == 0 else 280,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                    text_color=Palette.TEXT_SECONDARY,
                )
                lbl.grid(row=0, column=col_i, sticky="ew", padx=4, pady=4)

            # Font level override selector
            override_menu = ctk.CTkOptionMenu(
                top_bar, values=["1", "2", "3"],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                width=50, height=26,
                command=lambda v: (dlg.destroy(), _build_popup(int(v))),
            )
            override_menu.set(str(level))
            override_menu.pack(side="right", padx=(0, 8))

            # ── Cuerpo: grid para columnas alineadas (igual que Control Final) ──
            body = ctk.CTkFrame(dlg, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=12, pady=8)
            body.grid_columnconfigure(0, minsize=120, weight=0)  # Campo — fijo
            body.grid_columnconfigure(1, weight=1)  # PDF — expand
            body.grid_columnconfigure(2, weight=1)  # Excel — expand

            for row_i, (campo, c) in enumerate(comps.items()):
                etq = etiquetas.get(campo, campo)
                v_pdf = c["pdf"] or "—"
                v_xls = c["excel"] or "—"
                ok = c["match"]

                bg = "#C8FFC8" if ok else "#FFC8C8"
                fg = "#006400" if ok else "#8B0000"

                ctk.CTkLabel(
                    body, text=etq, width=120, anchor="w",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"], weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).grid(row=row_i, column=0, sticky="w", padx=(4, 0), pady=2)

                for col_i, val in enumerate([v_pdf, v_xls], start=1):
                    lbl = ctk.CTkLabel(
                        body, text=val, anchor="center",
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                        fg_color=bg, text_color=fg, corner_radius=4,
                    )
                    lbl.grid(row=row_i, column=col_i, sticky="ew", padx=4, pady=2)

            # ── Leyenda al pie (centrada) ─────────────────────────────
            leyenda = ctk.CTkFrame(dlg, fg_color="transparent")
            leyenda.pack(fill="x", pady=(0, 12))
            inner = ctk.CTkFrame(leyenda, fg_color="transparent")
            inner.pack(anchor="center")
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#C8FFC8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Coincide",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#FFC8C8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Difiere",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            # Ajustar ventana al contenido real
            dlg.update_idletasks()
            alto = dlg.winfo_reqheight()
            x = (dlg.winfo_screenwidth() - ancho) // 2
            y = (dlg.winfo_screenheight() - alto) // 2
            dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

        _build_popup(self.config.get("font_level", 1))

    def _control_final_switch_toggle(self):
        """Toggle auto mode. Changes btn text hint and persists state."""
        if self._cf_auto_var.get():
            self.btn_control_final.configure(text="Control Final ⚡", image=self._icons["zap"])
        else:
            self.btn_control_final.configure(text="Control Final", image=self._icons["zap"])
        # Persist state
        self.config["control_final_auto"] = self._cf_auto_var.get()
        self._guardar_config()

    def _control_final_seleccionar(self):
        """Seleccionar PDFs (tickets + aduanas) + Excel para Control Final."""
        if self._cf_auto_var.get():
            self._control_final_auto_scan()
            return
        from tkinter import filedialog as tk_filedialog
        rutas = tk_filedialog.askopenfilenames(
            title="Seleccionar PDFs (tickets + salidas aduana) + Excel",
            filetypes=[("Archivos soportados", "*.pdf *.xls *.xlsx")],
        )
        if not rutas:
            return
        # Separar Excel de PDFs
        exceles = [r for r in rutas if r.lower().endswith(('.xls', '.xlsx'))]
        pdfs = [r for r in rutas if r.lower().endswith('.pdf')]
        if not exceles:
            self._log("ERROR: Debe seleccionar al menos un Excel de Contenedores")
            return
        if not pdfs:
            self._log("ERROR: Debe seleccionar al menos un PDF")
            return

        self._log(f"[Control Final] {len(pdfs)} PDFs, {len(exceles)} Excel(s)")

        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self.tarea_activa = True

        for btn_name in ('btn_control_final', 'btn_controlar_tickets', 'btn_controlar_coordinacion'):
            btn = getattr(self, btn_name, None)
            if btn and btn.winfo_exists():
                btn.configure(state="disabled")

        if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
            self.progress_carga.pack(side="right", padx=12)
            self.progress_carga.configure(mode="indeterminate")
            self.progress_carga.start()

        # Mostrar sección control final, ocultar otras
        self._switch_tab("final")

        # Limpiar tree viejo
        for item in self.tree_control_final.get_children():
            self.tree_control_final.delete(item)

        t = threading.Thread(
            target=self._control_final_worker,
            args=(list(pdfs), list(exceles)),
            daemon=True,
        )
        t.start()

    def _control_final_auto_scan(self):
        """Auto-scan Desktop for folders with CONTENEDORES Excel."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return

        self._scan_desktop_folders(callback=self._control_final_auto_scan_done)

    def _control_final_auto_scan_done(self, folders):
        """Callback after background scan completes for control final auto."""
        if not folders:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con CONTENEDORES en el Desktop"
            )
            return

        self._log(f"[Auto] {len(folders)} carpetas con CONTENEDORES encontradas")
        self._control_final_auto_popup(folders)

    def _control_final_auto_popup(self, folders):
        """Popup to select folders for auto-scan Control Final (planilla de carga style)."""
        top = ctk.CTkToplevel(self)
        top.title("Control Final Automático")
        top.geometry("600x400")
        top.configure(fg_color=Palette.BG_CARD)
        top.transient(self)
        top.lift()
        top.grab_set()
        top.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 600) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        top.geometry(f"600x400+{px}+{py}")

        ctk.CTkLabel(
            top, text="CONTROL FINAL AUTOMÁTICO",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        ctk.CTkLabel(
            top, text="Seleccioná las carpetas del Desktop a procesar:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_TABLE, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for carp in folders:
            var = ctk.BooleanVar(value=True)
            checks[carp["name"]] = var

        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        def _update_btn(*_args):
            any_selected = any(v.get() for v in checks.values())
            btn_procesar.configure(state="normal" if any_selected else "disabled")

        def _todas():
            for v in checks.values():
                v.set(True)

        def _ninguna():
            for v in checks.values():
                v.set(False)

        def _confirmar():
            selected = [c for c in folders if checks[c["name"]].get()]
            top.destroy()

            if not selected:
                return

            # Collect PDFs + Excel from selected folders
            # Only scan_* (API Vision tickets), get* (Salidas de Aduana),
            # and YYAR* (MIC/DTA terrestre, e.g. 26AR240914H)
            import re as _re
            _permiso_pat = _re.compile(r'^\d{2}069EC', _re.IGNORECASE)
            _mic_pat = _re.compile(r'^\d{2}AR', _re.IGNORECASE)
            all_pdfs = []
            all_excels = []
            for carp in selected:
                for f in os.scandir(carp["path"]):
                    if f.is_file() and f.name.lower().endswith('.pdf'):
                        if _permiso_pat.match(f.name):
                            continue  # skip permisos de exportación
                        name_lower = f.name.lower()
                        if not (name_lower.startswith("scan")
                                or name_lower.startswith("get")
                                or _mic_pat.match(f.name)):
                            continue  # only scan_*, get*, and YYAR* (MIC) PDFs
                        all_pdfs.append(f.path)
                if carp["excel_path"]:
                    all_excels.append(carp["excel_path"])

            if not all_pdfs:
                messagebox.showwarning(
                    "Sin PDFs",
                    "Las carpetas seleccionadas no contienen PDFs"
                )
                return
            if not all_excels:
                messagebox.showwarning(
                    "Sin Excel",
                    "Las carpetas seleccionadas no contienen Excel CONTENEDORES"
                )
                return

            self._log(f"[Auto] {len(all_pdfs)} PDFs, {len(all_excels)} Excel(s)")

            if self.tarea_activa:
                messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
                return
            self.tarea_activa = True

            for btn_name in ('btn_control_final', 'btn_controlar_tickets',
                             'btn_controlar_coordinacion'):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="disabled")

            if (hasattr(self, 'progress_carga')
                    and self.progress_carga.winfo_exists()):
                self.progress_carga.pack(side="right", padx=12)
                self.progress_carga.configure(mode="indeterminate")
                self.progress_carga.start()

            # Show control final section
            self._switch_tab("final")

            for item in self.tree_control_final.get_children():
                self.tree_control_final.delete(item)

            t = threading.Thread(
                target=self._control_final_worker,
                args=(list(all_pdfs), list(all_excels)),
                daemon=True,
            )
            t.start()

        # Todas / Ninguna (izquierda)
        ctk.CTkButton(
            btn_frame, text="Todas", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_todas,
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            btn_frame, text="Ninguna", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_ninguna,
        ).pack(side="left", padx=4)

        # Procesar (derecha)
        btn_procesar = ctk.CTkButton(
            btn_frame, text="Procesar", width=140, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=_confirmar,
        )
        btn_procesar.pack(side="right", padx=4)

        # Checkboxes
        for carp in folders:
            var = checks[carp["name"]]
            var.trace_add("write", _update_btn)
            ctk.CTkCheckBox(
                scroll, text=carp["name"], variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

    def _control_final_worker(self, pdf_paths, excel_paths):
        """Worker thread para Control Final."""
        import procesar_tickets
        from datetime import date
        import os

        try:
            result_count = 0
            # 1. Clasificar PDFs
            tickets_pdf = []   # scanned, need OCR
            aduanas_data = []  # PLT / Salida Aduana (texto)
            mic_data = []      # MIC/DTA (texto)

            for ruta in pdf_paths:
                import fitz
                doc = fitz.open(ruta)
                text = ""
                for p in doc:
                    text += p.get_text()
                doc.close()

                if not text.strip():
                    tickets_pdf.append(ruta)
                    continue

                if "MIC/DTA" in text.upper() or "MANIFIESTO INTERNACIONAL DE CARGA" in text.upper():
                    # MIC/DTA → terrestre
                    md = procesar_tickets.extraer_mic_dta(ruta)
                    if "_error" not in md:
                        md["_archivo"] = os.path.basename(ruta)
                        mic_data.append(md)
                elif "SALIDA DE ZONA PRIMARIA" in text.upper():
                    # PLT Aduana
                    data = procesar_tickets.extraer_salida_aduana(ruta, modo="flexi")
                    data["_archivo"] = os.path.basename(ruta)
                    aduanas_data.append(data)
                else:
                    tickets_pdf.append(ruta)

            self._log(f"[Control Final] {len(tickets_pdf)} tickets, "
                      f"{len(aduanas_data)} salidas aduana, {len(mic_data)} mic/dta")

            # 2. OCR tickets
            self._contenedores_cache = {}
            self._cargar_cache_contenedores(excel_paths)

            ocr_method = self.config.get("ocr_method", "api_vision")
            textos_por_pdf = {}
            api_datos_raw = {}

            # Sort by client folder name for grouped display (always, before any processing)
            tickets_pdf.sort(key=lambda r: os.path.basename(os.path.dirname(r)).split("_")[7].lower() if len(os.path.basename(os.path.dirname(r)).split("_")) > 7 else "")

            if ocr_method == "api_vision":
                api_vision_conf = self.config.get("api_vision", {})
                api_key_raw = api_vision_conf.get("api_key", "")
                api_key = self._decrypt_api_key(api_key_raw)
                model = api_vision_conf.get("model", procesar_tickets.MODELO_VISION_DEFAULT)
                parallel_enabled = api_vision_conf.get("parallel_enabled", False)

                if parallel_enabled and len(tickets_pdf) > 1:
                    # Parallel mode: distribute across enabled models
                    model_states = api_vision_conf.get("parallel_model_states", {})
                    enabled_models = [m for m, on in model_states.items() if on]
                    if not enabled_models:
                        enabled_models = [model]
                    # Ensure selected model is first
                    if model in enabled_models:
                        enabled_models.remove(model)
                    enabled_models.insert(0, model)
                    temperature = api_vision_conf.get("temperature", 0.1)
                    max_tokens = api_vision_conf.get("max_tokens", 4000)
                    timeout = api_vision_conf.get("timeout", 60)
                    result = procesar_tickets.api_vision_con_fallback(
                        tickets_pdf, api_key, enabled_models,
                        temperature=temperature, max_tokens=max_tokens,
                        timeout=timeout, log_callback=self._log,
                    )
                    api_datos_raw = result["datos"]
                    textos_por_pdf = result["textos"]
                else:
                    # Sequential fallback: try models one by one per PDF
                    all_models = api_vision_conf.get("custom_models", [])
                    if not all_models:
                        all_models = list(procesar_tickets.MODELOS_VISION)
                    # Ensure selected model is first
                    if model in all_models:
                        all_models.remove(model)
                    all_models.insert(0, model)

                    self._log(f"[Control Final] API Visión secuencial: {len(tickets_pdf)} PDF(s), {len(all_models)} modelo(s)")
                    for ruta in tickets_pdf:
                        stem = os.path.splitext(os.path.basename(ruta))[0]
                        self._log(f"  API Visión: {stem}")
                        success = False
                        for m in all_models:
                            datos = procesar_tickets.api_vision_extraer_datos(
                                ruta, api_key, model=m
                            )
                            if "error" in datos:
                                err = datos["error"]
                                if "is not a valid model ID" in err:
                                    err = "modelo no disponible"
                                elif "429" in err or "rate" in err.lower():
                                    err = "rate limit"
                                elif "timeout" in err.lower():
                                    err = "timeout"
                                self._log(f"  ERROR {m}: {err}")
                                continue
                            # Success
                            api_datos_raw[stem] = datos
                            self._log(f"  ✓ {stem} procesado con {m}")
                            success = True
                            break
                        if not success:
                            self._log(f"  ⚠ Todos los modelos fallaron para {stem}, usando PaddleOCR")
                            texto = procesar_tickets.pdf_a_texto(ruta, engine="paddleocr")
                            textos_por_pdf[stem] = texto
            else:
                textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(tickets_pdf, engine="paddleocr")

            # 3. Build ALL lookup structures (terrestre + flexi)
            self._log(f"[Control Final] Procesando tickets...")
            import re as _re

            def _norm_pat_match(s):
                """Normalizar patente: sacar espacios, uppercase."""
                if not s: return ""
                return _re.sub(r'\s+', '', s).upper()

            # ── Aduana terrestre lookups (MIC/DTA) ──
            aduana_por_patente = {}
            aduana_por_precinto_mic = {}
            aduana_por_dni = {}
            for md in mic_data:
                pat = _norm_pat_match(md.get("patente_camion", ""))
                if pat: aduana_por_patente[pat] = md
                if md.get("precinto"):
                    for p in md["precinto"].upper().split():
                        p = p.strip()
                        if p: aduana_por_precinto_mic[p] = md
                dni = _re.sub(r'\D', '', md.get("cuil", ""))
                if dni: aduana_por_dni[dni] = md

            # ── Aduana flexi lookups (Salida de Aduana) ──
            aduana_por_contenedor = {}
            aduana_por_precinto_aduana = {}
            for ad in aduanas_data:
                if ad.get("contenedor"):
                    aduana_por_contenedor[ad["contenedor"].upper()] = ad
                if ad.get("precinto"):
                    for p in ad["precinto"].upper().split():
                        p = p.strip()
                        if p: aduana_por_precinto_aduana[p] = ad

            # ── Excel lookups (ALL cache, list-based for shared trips) ──
            excel_por_patente = {}  # patente -> [(cont_data, cont_idx), ...]
            excel_por_precinto = {}
            excel_por_dni = {}
            excel_por_contenedor = {}
            for k, v in self._contenedores_cache.items():
                camiones = v.get("camiones", [])
                for ci, cam in enumerate(camiones):
                    pat = _norm_pat_match(cam.get("patente_camion", ""))
                    if pat:
                        excel_por_patente.setdefault(pat, []).append((v, ci))
                    prec = cam.get("precinto", "").upper().strip()
                    if prec:
                        # Normalizar: dividir por guión para crear key individual
                        for pp in prec.replace("-", " ").split():
                            if pp:
                                excel_por_precinto.setdefault(pp, []).append((v, ci))
                    dni = _re.sub(r'\D', '', str(cam.get("dni", "")))
                    if dni:
                        excel_por_dni.setdefault(dni, []).append((v, ci))
                    ctn = _re.sub(r'[\s\-]+', '', cam.get("contenedor", "").upper().strip())
                    if ctn:
                        excel_por_contenedor.setdefault(ctn, []).append((v, ci))

            # Merged precinto aduana lookup (try both MIC and Salida)
            aduana_por_precinto_all = {**aduana_por_precinto_mic, **aduana_por_precinto_aduana}

            # 4. Process each ticket — try terrestre first, then flexi
            for ruta in tickets_pdf:
                stem = os.path.splitext(os.path.basename(ruta))[0]
                carpeta = os.path.basename(os.path.dirname(ruta))
                cliente = carpeta.split("_")[7] if len(carpeta.split("_")) > 7 else carpeta
                try:
                    shared_excel_matches = []
                    modo_result = "flexi"  # default per-ticket

                    # Build ticket_data
                    if stem in api_datos_raw:
                        raw = api_datos_raw[stem]
                        ticket_data = {
                            "archivo":   stem,
                            "Patente":   raw.get("Patente", ""),
                            "Semirremolque": raw.get("Semirremolque", ""),
                            "Conductor": raw.get("Conductor", ""),
                            "DNI":       raw.get("DNI", ""),
                            "Neto":      raw.get("Neto", ""),
                            "Tara Contenedor": raw.get("Tara Contenedor", ""),
                            "Contenedor": raw.get("Contenedor", ""),
                            "Permiso":   raw.get("Permiso", ""),
                        }
                    else:
                        texto = textos_por_pdf.get(stem, "")
                        if not texto:
                            continue
                        extraido = procesar_tickets.extraer_datos(texto)
                        ticket_data = {
                            "archivo":   stem,
                            "Patente":   extraido.get("patente", ""),
                            "Semirremolque": extraido.get("semi", ""),
                            "Conductor": extraido.get("conductor", ""),
                            "DNI":       extraido.get("dni", ""),
                            "Neto":      extraido.get("neto", ""),
                            "Tara Contenedor": extraido.get("tara", ""),
                            "Contenedor": extraido.get("contenedor", ""),
                            "Permiso":   extraido.get("oferta", ""),
                        }

                    patente_ticket = _norm_pat_match(ticket_data.get("Patente", ""))
                    dni_ticket = _re.sub(r'\D', '', ticket_data.get("DNI", ""))
                    contenedor_ticket = _re.sub(r'[\s\-]+', '',
                                                ticket_data.get("Contenedor", "").upper().strip())

                    aduana = {}
                    cont_data = None
                    cont_idx = -1

                    # ── Step 1: Try terrestre (patente → MIC/DTA + patente → Excel) ──
                    if patente_ticket:
                        aduana = aduana_por_patente.get(patente_ticket, {})
                        all_excel = excel_por_patente.get(patente_ticket, [])
                        if all_excel:
                            cont_data, cont_idx = all_excel[0]
                            shared_excel_matches = all_excel[1:]
                        if aduana:
                            modo_result = "terrestre"

                        # Variants O/0 (Mercosur patentes)
                        if not cont_data and not aduana:
                            for i, ch in enumerate(patente_ticket):
                                variants = set()
                                if ch == 'O': variants.add(patente_ticket[:i] + '0' + patente_ticket[i+1:])
                                if ch == '0': variants.add(patente_ticket[:i] + 'O' + patente_ticket[i+1:])
                                if ch == 'I': variants.add(patente_ticket[:i] + '1' + patente_ticket[i+1:])
                                if ch == '1': variants.add(patente_ticket[:i] + 'I' + patente_ticket[i+1:])
                                if ch == 'S': variants.add(patente_ticket[:i] + '5' + patente_ticket[i+1:])
                                if ch == '5': variants.add(patente_ticket[:i] + 'S' + patente_ticket[i+1:])
                                for vr in variants:
                                    if vr in excel_por_patente:
                                        vm = excel_por_patente[vr]
                                        if not cont_data:
                                            cont_data, cont_idx = vm[0]
                                            shared_excel_matches = vm[1:] if len(vm) > 1 else []
                                            self._log(f"  ⚠ Patente OCR corregida: {patente_ticket}→{vr} (Excel)")
                                    if not aduana and vr in aduana_por_patente:
                                        aduana = aduana_por_patente[vr]
                                        self._log(f"  ⚠ Patente OCR corregida: {patente_ticket}→{vr} (Aduana)")
                                    if cont_data and aduana: break
                                if cont_data and aduana: break

                    # ── Step 2: If no terrestre match, try flexi (contenedor → Salida Aduana) ──
                    if not aduana and contenedor_ticket:
                        aduana = aduana_por_contenedor.get(contenedor_ticket, {})
                        if aduana:
                            modo_result = "flexi"
                            self._log(f"  ✓ {stem}: flexi match por contenedor='{contenedor_ticket}'")
                            # Match Excel by contenedor
                            all_excel = excel_por_contenedor.get(contenedor_ticket, [])
                            if all_excel:
                                cont_data, cont_idx = all_excel[0]
                                shared_excel_matches = all_excel[1:]

                    # ── Step 3: Fallback por precinto ──
                    if not cont_data or not aduana:
                        # 3a: Have Excel but need Aduana
                        if cont_data and cont_idx >= 0 and not aduana:
                            cam = cont_data["camiones"][cont_idx]
                            prec = cam.get("precinto", "").upper().strip()
                            if prec and prec in aduana_por_precinto_all:
                                aduana = aduana_por_precinto_all[prec]
                                modo_result = "terrestre" if prec in aduana_por_precinto_mic else "flexi"
                                self._log(f"  ⚠ Aduana x precinto Excel: {stem} — precinto={prec}")

                        # 3b: Have Aduana but need Excel
                        if aduana and (not cont_data or cont_idx < 0):
                            prec = aduana.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    self._log(f"  ⚠ Excel x precinto Aduana: {stem} — precinto={p}")
                                    break

                    # ── Step 3c: DNI fallback ──
                    if not cont_data and not aduana and dni_ticket:
                        if dni_ticket in excel_por_dni:
                            matches = excel_por_dni[dni_ticket]
                            cont_data, cont_idx = matches[0]
                            shared_excel_matches = matches[1:] if len(matches) > 1 else []
                            self._log(f"  ⚠ Excel x DNI: {stem} — dni={dni_ticket}")
                        if dni_ticket in aduana_por_dni:
                            aduana = aduana_por_dni[dni_ticket]
                            modo_result = "terrestre"
                            self._log(f"  ⚠ Aduana x DNI: {stem} — dni={dni_ticket}")

                    # ── Step 3d: Puente post-DNI ──
                    if not cont_data or not aduana:
                        if cont_data and cont_idx >= 0 and not aduana:
                            cam = cont_data["camiones"][cont_idx]
                            prec = cam.get("precinto", "").upper().strip()
                            if prec and prec in aduana_por_precinto_all:
                                aduana = aduana_por_precinto_all[prec]
                                self._log(f"  ⚠ Aduana x precinto post-DNI: {stem} — precinto={prec}")
                        if aduana and (not cont_data or cont_idx < 0):
                            prec = aduana.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    self._log(f"  ⚠ Excel x precinto post-DNI: {stem} — precinto={p}")
                                    break

                    # ── Step 3e: Brute force — scan all MIC by precinto ──
                    if not cont_data and not aduana:
                        for md in mic_data:
                            prec = md.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    aduana = md
                                    modo_result = "terrestre"
                                    self._log(f"  ⚠ Sin match x precinto (último): {stem} — precinto={p}")
                                    break
                            if cont_data: break

                    # ── Step 3f: Brute force — scan all Salida Aduana by precinto ──
                    if not cont_data and not aduana:
                        for ad in aduanas_data:
                            prec = ad.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    aduana = ad
                                    modo_result = "flexi"
                                    self._log(f"  ⚠ Sin match x precinto (flexi): {stem} — precinto={p}")
                                    break
                            if cont_data: break

                    result = self._build_fila_control_final(
                        ticket_data, cont_data, cont_idx, aduana, stem, cliente=cliente, modo=modo_result,
                        shared_excel_matches=shared_excel_matches
                    )
                    if result:
                        self.log_queue.put(("_CONTROL_FINAL_RESULT_", result))
                        result_count += 1

                except Exception as e:
                    self._log(f"  ✗ Error {stem}: {e}")
                    import traceback
                    self._log(traceback.format_exc())

        except Exception as e:
            self._log(f"[Control Final] Error general: {e}")
            import traceback
            self._log(traceback.format_exc())
        finally:
            try: rc = result_count
            except NameError: rc = 0
            self._log(f"[Control Final] Finalizado — {rc} ticket(s) procesado(s)")
            self.log_queue.put(("_TAREA_COMPLETA_", None))

    def _build_fila_control_final(self, ticket_data, cont_data, cont_idx, aduana, stem, cliente="",
                                   modo="flexi", shared_excel_matches=None):
        """Build row values and comparison data for Control Final.

        Args:
            modo: "flexi" (ISO/Flexi containers) or "terrestre" (bulk/granel).
            shared_excel_matches: list of (cont_data, cont_idx) from additional
                Excels matching the same truck (shared trips). The primary match
                is NOT in this list — it's passed via cont_data/cont_idx.

        Returns dict with keys:
          valores    → tuple for tree row (11 cols)
          tag        → tag_ok / tag_mismatch
          ticket     → dict {Patente, Semirremolque, ..., Permiso} (normalized ticket)
          contenedor → dict {Patente, Semirremolque, ..., Permiso} (normalized Excel)
          aduana     → dict {Patente Camión, ..., CUIL, Contenedor, ...}
          ok         → dict per field: True if ticket matches reference
          modo       → the mode used
          shared_excels → list of {peso_carga, archivo} for shared trips (empty if not shared)
          shared_neto_sum → sum of all Excel peso_carga (0 if not shared)
        """
        import procesar_tickets
        import re as _re

        # ── Normalization helpers ──────────────────────────────
        def _norm_pat(s):
            return str(s).replace(" ", "") if s else s

        def _norm_num(s):
            """Remove trailing .0 / .00 + thousands-separator dots from numeric values."""
            if not s:
                return s
            v = str(s).strip()
            v = _re.sub(r'\.0+$', '', v)      # "23440.0" → "23440"
            v = v.replace(".", "")             # "23.440" → "23440"
            return v

        def _fmt_neto(s):
            """Normalize number then format with dot as thousands separator: '27890' → '27.890'."""
            if not s:
                return s
            v = str(s).strip()
            # Quitar comas y puntos existentes
            v = v.replace(",", "")
            v = _re.sub(r'\.0+$', '', v)
            v = v.replace(".", "")
            if v.isdigit():
                return f"{int(v):,}".replace(",", ".")
            return v

        def _norm_ctn(s):
            """Remove spaces and dashes from contenedor numbers."""
            if not s:
                return s
            return _re.sub(r'[\s\-]+', '', str(s).strip())

        def _dni_solo(v):
            """Extraer solo DNI, limpiando CUIL/CUIT y no-dígitos."""
            if not v:
                return ""
            d = _re.sub(r'\D', '', v)
            if len(d) == 11 and d[:2] in ("20", "23", "27", "30"):
                return d[2:-1]
            return d

        # ── Ticket data (ORIGINAL, never overwrite) ──
        t_patente    = ticket_data.get("Patente", "")
        t_semi       = ticket_data.get("Semirremolque", "")
        t_conductor  = ticket_data.get("Conductor", "")
        t_dni        = ticket_data.get("DNI", "")
        t_neto       = ticket_data.get("Neto", "")
        t_tara       = ticket_data.get("Tara Contenedor", "")
        # Normalize ticket OCR values: strip thousand-separator dots
        # API sometimes returns "24.000" (with dots) instead of "24000"
        _tmp = str(t_neto).replace(".", "").replace(",", ".")
        if _tmp.replace(".", "", 1).isdigit():
            t_neto = _tmp
        _tmp = str(t_tara).replace(".", "").replace(",", ".")
        if _tmp.replace(".", "", 1).isdigit():
            t_tara = _tmp
        t_contenedor = ticket_data.get("Contenedor", "")
        t_permiso    = ticket_data.get("Permiso", "")
        t_dni        = _dni_solo(t_dni)

        # ── Aduana data ──
        a_plt            = aduana.get("plt", "")
        a_peso_bruto     = aduana.get("peso_bruto", "")
        a_cuil           = aduana.get("cuil", "")
        a_patente_camion = aduana.get("patente_camion", "")
        a_patente_semi   = aduana.get("patente_semi", "")
        a_conductor      = aduana.get("conductor", "")
        a_id_destinacion = aduana.get("id_destinacion", "")
        a_id_destinacion_1 = aduana.get("id_destinacion_1", "")
        a_id_destinacion_2 = aduana.get("id_destinacion_2", "")
        a_exportador     = aduana.get("exportador", "")
        a_contenedor     = aduana.get("contenedor", "")
        a_precinto       = aduana.get("precinto", "")

        a_dni = _dni_solo(a_cuil)

        # ── Excel / Contenedor data ──
        e_patente = ""
        e_semi = ""
        e_conductor = ""
        e_dni = ""
        e_neto = ""
        e_tara = ""
        e_contenedor = ""
        e_precinto = ""
        e_permiso = ""
        e_peso_flexi = "0"
        if cont_data and cont_idx >= 0:
            camiones = cont_data.get("camiones", [])
            if cont_idx < len(camiones):
                cam = camiones[cont_idx]
                e_patente    = cam.get("patente_camion", "")
                e_semi       = cam.get("patente_semi", "")
                e_conductor  = cam.get("conductor", "")
                e_dni        = cam.get("dni", "")
                e_neto       = str(cam.get("peso_carga", ""))
                e_tara       = str(cam.get("tara_cont", ""))
                e_contenedor = cam.get("contenedor", "")
                e_precinto   = cam.get("precinto", "")
                e_permiso    = cont_data.get("pe", "")
                e_peso_flexi = str(cam.get("peso_flexi", "0"))
        e_dni = _dni_solo(e_dni)

        # ── Shared Excel data ──
        shared_excels = []
        shared_neto_sum = 0
        if shared_excel_matches:
            for s_data, s_idx in shared_excel_matches:
                s_camiones = s_data.get("camiones", [])
                if s_idx < len(s_camiones):
                    s_cam = s_camiones[s_idx]
                    s_neto = s_cam.get("peso_carga", 0)
                    shared_excels.append({
                        "peso_carga": s_neto,
                        "archivo": os.path.basename(s_data.get("_archivo", "")),
                    })
                    shared_neto_sum += s_neto

        # Also count primary match
        if cont_data and cont_idx >= 0:
            camiones = cont_data.get("camiones", [])
            if cont_idx < len(camiones):
                primary_neto = camiones[cont_idx].get("peso_carga", 0)
                shared_excels.insert(0, {
                    "peso_carga": primary_neto,
                    "archivo": os.path.basename(cont_data.get("_archivo", "")),
                })
                shared_neto_sum += primary_neto

        is_shared = len(shared_excels) > 1

        # ── Normalize all values ──
        t_patente    = _norm_pat(t_patente)
        t_semi       = _norm_pat(t_semi)
        t_neto       = _norm_num(t_neto)
        t_contenedor = _norm_ctn(t_contenedor)

        e_patente    = _norm_pat(e_patente)
        e_semi       = _norm_pat(e_semi)
        e_neto       = _norm_num(e_neto)
        e_contenedor = _norm_ctn(e_contenedor)

        a_patente_camion = _norm_pat(a_patente_camion)
        a_patente_semi   = _norm_pat(a_patente_semi)
        a_peso_bruto     = _norm_num(a_peso_bruto)
        a_contenedor     = _norm_ctn(a_contenedor)

        # ── Calculate aduana neto (solo flexi: peso_bruto - flexi) ──
        if modo == "terrestre":
            a_neto_calc = a_peso_bruto
        else:
            a_neto_calc = a_peso_bruto
            if a_peso_bruto:
                try:
                    pb = float(a_peso_bruto.replace(",", ""))
                    pf_val = (cont_data.get("peso_flexi_global") if cont_data else None) or e_peso_flexi
                    pf = float(pf_val) if pf_val and pf_val != "0" else 0
                    a_neto_calc = str(int(pb - pf)) if pf > 0 else str(int(pb))
                except (ValueError, TypeError):
                    pass

        # ── Apply terrestre field overrides ──
        if modo == "terrestre":
            t_contenedor = "—"
            e_contenedor = "—"
            if not a_contenedor:
                a_contenedor = "—"

        # ── Format neto/tara with dot thousands separator ──
        t_neto       = _fmt_neto(t_neto)
        e_neto       = _fmt_neto(e_neto)
        a_neto_calc  = _fmt_neto(a_neto_calc)
        t_tara       = _fmt_neto(t_tara)
        e_tara       = _fmt_neto(e_tara)
        # No format a_peso_bruto display — a_neto_calc is used for comparison

        # ── Shared trips: override Excel Neto with TOTAL (all fractions) ──
        if is_shared and shared_neto_sum:
            e_neto = _fmt_neto(str(int(shared_neto_sum)))

        # ── Format ticket PE: si tiene "/", mantener primera parte + últimos 5 de la segunda ──
        if "/" in t_permiso:
            parts = t_permiso.split("/")
            if len(parts) >= 2 and len(parts[-1].strip()) >= 5:
                t_permiso = f"{parts[0].strip()}/{parts[-1].strip()[-5:]}"

        # ── For shared trips: combine PEs (Ticket + Excel + MIC) ──
        if is_shared:
            # Ticket: primary PE / last 5 of secondary PE (to match Excel format)
            # Only if ticket PE doesn't already have "/" (already formatted above)
            if t_permiso and "/" not in t_permiso and shared_excel_matches:
                s_data, s_idx = shared_excel_matches[0]
                s_pe = s_data.get("pe", "")
                if s_pe and len(s_pe) >= 5:
                    t_permiso = f"{t_permiso}/{s_pe[-5:]}"
            # Excel: primary PE / last 5 of secondary PE
            if e_permiso and shared_excel_matches:
                s_data, s_idx = shared_excel_matches[0]
                s_pe = s_data.get("pe", "")
                if s_pe and len(s_pe) >= 5:
                    e_permiso = f"{e_permiso}/{s_pe[-5:]}"
            # MIC: hoja 1 PE / last 5 of hoja 2 PE
            if modo == "terrestre" and a_id_destinacion_1 and a_id_destinacion_2:
                a_id_destinacion = f"{a_id_destinacion_1}/{a_id_destinacion_2[-5:]}"

        # ── Build display dicts (all normalized) ──
        ticket_d = {
            "Patente": t_patente, "Semirremolque": t_semi,
            "Conductor": t_conductor, "DNI": t_dni,
            "Neto (kg)": t_neto, "Tara (kg)": t_tara,
            "Contenedor": t_contenedor, "Permiso": t_permiso,
            "Precinto": "-",
        }
        cont_d = {
            "Patente": e_patente, "Semirremolque": e_semi,
            "Conductor": e_conductor, "DNI": e_dni,
            "Neto (kg)": e_neto, "Tara (kg)": e_tara,
            "Contenedor": e_contenedor, "Permiso": e_permiso,
            "Precinto": e_precinto,
        }
        dni_key = "DNI" if modo == "terrestre" else "CUIL"
        # MIC/DTA campo 42/44/46 — shared trip breakdown (terrestre only)
        a_campo_42 = aduana.get("campo_42", "")
        a_campo_44 = aduana.get("campo_44", "")
        a_campo_46 = aduana.get("campo_46", "")
        aduana_d = {
            "PLT": a_plt,
            "Peso Bruto Aduana": a_neto_calc,
            "Id Destinación": a_id_destinacion,
            "Exportador": a_exportador,
            "Patente Camión": a_patente_camion,
            "Patente Semi": a_patente_semi,
            "Conductor Aduana": a_conductor,
            dni_key: a_dni,
            "Contenedor": a_contenedor,
            "Precinto": a_precinto,
            "campo_42": a_campo_42,
            "campo_44": a_campo_44,
            "campo_46": a_campo_46,
        }

        # ── 3-way comparison ──
        ok_map = {}
        all_ok = True
        compare_fields = [
            "Patente", "Semirremolque", "Conductor",
            "DNI", "Neto (kg)", "Tara (kg)",
            "Contenedor", "Permiso", "Precinto",
        ]

        for campo in compare_fields:
            v_ticket = ticket_d.get(campo, "").strip().upper()
            v_cont   = cont_d.get(campo, "").strip().upper()

            # Map field → aduana variable
            aduana_val = ""
            if   campo == "Patente":      aduana_val = a_patente_camion.strip().upper()
            elif campo == "Semirremolque": aduana_val = a_patente_semi.strip().upper()
            elif campo == "Conductor":    aduana_val = a_conductor.strip().upper()
            elif campo == "DNI":          aduana_val = a_dni.strip().upper()
            elif campo == "Neto (kg)":
                # Shared terrestre: use campo_46 (total) instead of a_peso_bruto (hoja 1 only)
                if is_shared and modo == "terrestre" and a_campo_46:
                    aduana_val = _fmt_neto(a_campo_46).strip().upper()
                else:
                    aduana_val = a_neto_calc.strip().upper()
            elif campo == "Contenedor":   aduana_val = a_contenedor.strip().upper()
            elif campo == "Permiso":      aduana_val = a_id_destinacion.strip().upper()
            elif campo == "Precinto":     aduana_val = a_precinto.strip().upper()
            # Tara → no aduana equivalent, stays ""

            # Normalize conductor accents before compare
            if campo == "Conductor":
                import unicodedata
                _norm = lambda s: unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().strip()
                if v_ticket:   v_ticket   = _norm(v_ticket)
                if v_cont:     v_cont     = _norm(v_cont)
                if aduana_val: aduana_val = _norm(aduana_val)

                # Word-level matching: al menos 2 palabras de cada fuente en Aduana
                t_words = set(v_ticket.split()) if v_ticket else set()
                c_words = set(v_cont.split()) if v_cont else set()
                a_words = set(aduana_val.split()) if aduana_val else set()

                ticket_ok = len(t_words & a_words) >= 2 if len(t_words) >= 2 else t_words.issubset(a_words) if t_words else False
                excel_ok  = len(c_words & a_words) >= 2 if len(c_words) >= 2 else c_words.issubset(a_words) if c_words else False
                ok_map[campo] = ticket_ok and excel_ok
                if not ok_map[campo]:
                    all_ok = False
                continue  # skip the generic comparison below

            # Normalize DNI: strip non-digits for comparison
            if campo == "DNI":
                v_ticket   = _re.sub(r'\D', '', v_ticket)
                v_cont     = _re.sub(r'\D', '', v_cont)
                aduana_val = _re.sub(r'\D', '', aduana_val)

            # Collect non-empty values from all 3 sources
            vals = set()
            # Precinto: ticket siempre "-", no comparar
            if campo == "Precinto":
                if v_cont:     vals.add(v_cont)
                if aduana_val: vals.add(aduana_val)
            elif campo == "Tara (kg)":
                # Tara: solo Ticket vs Excel (Aduana no tiene tara)
                if v_ticket:   vals.add(v_ticket)
                if v_cont:     vals.add(v_cont)
            else:
                if v_ticket:   vals.add(v_ticket)
                if v_cont:     vals.add(v_cont)
                if aduana_val: vals.add(aduana_val)

            # 0 or 1 unique value → match; 2+ unique → mismatch
            if len(vals) <= 1:
                ok_map[campo] = True
            else:
                # Para Precinto: si hay multi-precinto, verificar solapamiento
                if campo == "Precinto" and len(vals) > 1:
                    cont_codes = set(v_cont.replace("-", " ").split()) if v_cont else set()
                    adu_codes = set(aduana_val.replace("-", " ").split()) if aduana_val else set()
                    if cont_codes & adu_codes:
                        ok_map[campo] = True
                    else:
                        ok_map[campo] = False
                        all_ok = False
                # Para Permiso en compartidos: al menos 2 de 3 fuentes coinciden
                elif campo == "Permiso" and is_shared:
                    sources = [v for v in [v_ticket, v_cont, aduana_val] if v]
                    if sources and max(sources.count(v) for v in sources) >= 2:
                        ok_map[campo] = True
                    else:
                        ok_map[campo] = False
                        all_ok = False
                else:
                    ok_map[campo] = False
                    all_ok = False

        estado = "ok" if all_ok else "mismatch"
        estado_label = "✅ OK" if all_ok else "❌ Diferencia"
        tag = "tag_ok" if all_ok else "tag_mismatch"

        # Tree row: all values normalized
        valores = (
            cliente,
            t_patente,
            t_semi,
            t_conductor,
            self._fmt_dni(t_dni),
            t_neto,
            t_tara,
            t_contenedor,
            t_permiso,
            estado_label,
            a_plt,
        )

        return {
            "valores": valores,
            "tag": tag,
            "ticket": ticket_d,
            "contenedor": cont_d,
            "aduana": aduana_d,
            "ok": ok_map,
            "modo": modo,
            "shared_excels": shared_excels,
            "shared_neto_sum": shared_neto_sum,
        }

    def _procesar_resultado_control_final(self, data):
        """Inserta fila en tree_control_final."""
        try:
            stem = data.get("valores", ())[0] if data.get("valores") else "?"
            iid = self.tree_control_final.insert("", "end", values=data["valores"], tags=(data["tag"],))
            # Actualizar header según modo (terrestre → MIC/DTA, flexi → Salida Aduana)
            if len(self.tree_control_final.get_children()) == 1:
                modo = data.get("modo", "flexi")
                self.tree_control_final.heading(
                    "salida_aduana",
                    text="MIC/DTA" if modo == "terrestre" else "Salida Aduana"
                )
            self._log(f"  + Fila insertada: {stem}")
            # Store comparison data
            if not hasattr(self, '_control_final_comparacion'):
                self._control_final_comparacion = {}
            self._control_final_comparacion[iid] = {
                "ticket": data["ticket"],
                "contenedor": data["contenedor"],
                "ok": data["ok"],
                "aduana": data["aduana"],
                "modo": data.get("modo", "flexi"),
                "shared_excels": data.get("shared_excels", []),
                "shared_neto_sum": data.get("shared_neto_sum", 0),
            }
        except Exception as e:
            self._log(f"Error insertando fila: {e}")

    def _abrir_comparacion_final(self, event):
        """Popup con tabla comparativa: Ticket vs Aduana vs Excel."""
        sel = self.tree_control_final.selection()
        if not sel or not hasattr(self, '_control_final_comparacion'):
            return

        datos = self._control_final_comparacion.get(sel[0])
        if not datos:
            return

        modo = datos.get("modo", "flexi")
        aduana_header = "MIC/DTA" if modo == "terrestre" else "Salida Aduana"

        def _build_popup(level):
            fsizes = self._get_font_sizes(level)
            geom = self._get_popup_geometry(700, 420, level)
            ancho = int(geom.split("x")[0])

            dlg = ctk.CTkToplevel(self)
            dlg.title("Comparación — Control Final")
            dlg.transient(self)
            dlg.grab_set()
            dlg.resizable(False, False)

            # ── Top bar: header + font level override ──────────────────
            top_bar = ctk.CTkFrame(dlg, fg_color=Palette.BG_SIDEBAR, corner_radius=6, height=36)
            top_bar.pack(fill="x", padx=12, pady=(12, 0))
            top_bar.pack_propagate(False)

            for col_i, txt in enumerate(["Campo", "Ticket (OCR)", aduana_header, "Contenedor (Excel)"]):
                lbl = ctk.CTkLabel(
                    top_bar, text=txt, width=100 if col_i == 0 else 160,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                    text_color=Palette.TEXT_SECONDARY,
                )
                kwargs = {"side": "left", "padx": 4}
                if col_i > 0:
                    kwargs["fill"] = "x"
                    kwargs["expand"] = True
                lbl.pack(**kwargs)

            # Font level override selector
            override_menu = ctk.CTkOptionMenu(
                top_bar, values=["1", "2", "3"],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                width=50, height=26,
                command=lambda v: (dlg.destroy(), _build_popup(int(v))),
            )
            override_menu.set(str(level))
            override_menu.pack(side="right", padx=(0, 8))

            # Cuerpo: grid para columnas alineadas
            body = ctk.CTkFrame(dlg, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=12, pady=8)
            body.grid_columnconfigure(0, minsize=100, weight=0)  # Campo — fijo
            body.grid_columnconfigure(1, weight=1)  # Ticket — expand
            body.grid_columnconfigure(2, weight=1)  # Aduana — expand
            body.grid_columnconfigure(3, weight=1)  # Contenedor — expand

            if modo == "terrestre":
                campos = [
                    ("Camion", "Patente", "Patente Camión"),
                    ("Semi", "Semirremolque", "Patente Semi"),
                    ("Conductor", "Conductor", "Conductor Aduana"),
                    ("DNI", "DNI", "DNI"),
                    ("Neto (kg)", "Neto (kg)", "Peso Bruto Aduana"),
                    ("Tara (kg)", "Tara (kg)", None),
                    ("Contenedor", "Contenedor", "Contenedor"),
                    ("Permiso", "Permiso", "Id Destinación"),
                    ("Precinto", "Precinto", "Precinto"),
                ]
            else:
                campos = [
                    ("Camion", "Patente", "Patente Camión"),
                    ("Semi", "Semirremolque", "Patente Semi"),
                    ("Conductor", "Conductor", "Conductor Aduana"),
                    ("DNI", "DNI", "CUIL"),
                    ("Neto (kg)", "Neto (kg)", "Peso Bruto Aduana"),
                    ("Tara (kg)", "Tara (kg)", None),
                    ("Contenedor", "Contenedor", "Contenedor"),
                    ("Permiso", "Permiso", "Id Destinación"),
                    ("Precinto", "Precinto", "Precinto"),
                ]

            # Campos con coloreado individual por celda (mayoría decide)
            campos_mayoria = {"Camion", "Semi", "DNI",
                              "Neto (kg)", "Contenedor", "Permiso"}

            shared_excels = datos.get("shared_excels", [])
            is_shared = len(shared_excels) > 1

            for row_i, (campo, key_ticket, key_aduana) in enumerate(campos):
                val_ticket = datos["ticket"].get(key_ticket, "")
                val_cont = datos["contenedor"].get(key_ticket, "")
                val_aduana = datos["aduana"].get(key_aduana, "—") if key_aduana else "—"
                # Format DNI with thousand separators: 36987654 -> 36.987.654
                if campo == "DNI":
                    val_ticket = self._fmt_dni(val_ticket)
                    val_cont = self._fmt_dni(val_cont)
                    val_aduana = self._fmt_dni(val_aduana)

                # Override Neto Excel column for shared trips
                if is_shared and campo == "Neto (kg)":
                    excel_neto_parts = " + ".join(
                        self._fmt_kg(ex["peso_carga"]) for ex in shared_excels
                    )
                    excel_neto_total = self._fmt_kg(datos["shared_neto_sum"])
                    val_cont = f"{excel_neto_parts} = {excel_neto_total}"

                # Override Neto MIC/DTA column for shared terrestre trips
                if is_shared and modo == "terrestre" and campo == "Neto (kg)":
                    c42 = datos["aduana"].get("campo_42", "")
                    c44 = datos["aduana"].get("campo_44", "")
                    c46 = datos["aduana"].get("campo_46", "")
                    if c42 and c44 and c46:
                        val_aduana = (f"{self._fmt_kg(c42)} + "
                                      f"{self._fmt_kg(c44)} = "
                                      f"{self._fmt_kg(c46)}")
                ok = datos["ok"].get(key_ticket, False)

                vals_list = [str(val_ticket), str(val_aduana), str(val_cont)]

                if campo in campos_mayoria:
                    colors = []
                    # For shared trips Neto: compare totals (after "="), not full strings
                    if is_shared and campo == "Neto (kg)":
                        totals = []
                        for v in vals_list:
                            if "=" in v:
                                totals.append(v.split("=")[-1].strip())
                            else:
                                totals.append(v.strip())
                        all_match = len(set(totals)) == 1
                        for _ in vals_list:
                            colors.append(("#C8FFC8", "#006400") if all_match else ("#FFC8C8", "#8B0000"))
                    else:
                        for v in vals_list:
                            count = sum(1 for x in vals_list if x == v)
                            if count >= 2:
                                colors.append(("#C8FFC8", "#006400"))
                            else:
                                colors.append(("#FFC8C8", "#8B0000"))
                else:
                    bg = "#C8FFC8" if ok else "#FFC8C8"
                    fg = "#006400" if ok else "#8B0000"
                    colors = [(bg, fg)] * 3

                # Campo label (column 0)
                ctk.CTkLabel(
                    body, text=campo, anchor="w",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"], weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).grid(row=row_i, column=0, padx=(4, 8), pady=2, sticky="w")

                # Data columns (1, 2, 3)
                for col_i, (val, (bg, fg)) in enumerate(zip(vals_list, colors)):
                    lbl = ctk.CTkLabel(
                        body, text=val, anchor="center",
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                        fg_color=bg, text_color=fg, corner_radius=4,
                    )
                    lbl.grid(row=row_i, column=col_i + 1, padx=4, pady=2, sticky="ew")

            # Legend
            leyenda = ctk.CTkFrame(dlg, fg_color="transparent")
            leyenda.pack(fill="x", pady=(0, 12))
            inner = ctk.CTkFrame(leyenda, fg_color="transparent")
            inner.pack(anchor="center")

            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#C8FFC8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Coincide",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#FFC8C8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Diferencia",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            # Ajustar ventana al contenido real
            dlg.update_idletasks()
            alto = dlg.winfo_reqheight()
            x = (dlg.winfo_screenwidth() - ancho) // 2
            y = (dlg.winfo_screenheight() - alto) // 2
            dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

        _build_popup(self.config.get("font_level", 1))

    def _cargar_datos_switch_toggle(self):
        """Toggle auto mode for Control Tickets. Changes btn text hint and persists state."""
        if self._ct_auto_var.get():
            self.btn_controlar_tickets.configure(text="Controlar Tickets ⚡", image=self._icons["zap"])
        else:
            self.btn_controlar_tickets.configure(text="Controlar Tickets", image=self._icons["ticket"])
        self.config["cargar_datos_auto"] = self._ct_auto_var.get()
        self._guardar_config()

    def _cargar_datos_auto_scan(self):
        """Auto-scan Desktop for folders with CONTENEDORES Excel (Control Tickets)."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self._scan_desktop_folders(
            pattern=None,
            callback=self._cargar_datos_auto_popup,
            button=self.btn_controlar_tickets,
        )

    def _cargar_datos_auto_popup(self, folders):
        """Popup to select folders for auto-scan Control Tickets (planilla de carga style)."""
        if not folders:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con CONTENEDORES en el Desktop"
            )
            return

        self._log(f"[Auto Tickets] {len(folders)} carpetas con CONTENEDORES encontradas")

        top = ctk.CTkToplevel(self)
        top.title("Control de Tickets Automático")
        top.geometry("600x400")
        top.configure(fg_color=Palette.BG_CARD)
        top.transient(self)
        top.lift()
        top.grab_set()
        top.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 600) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        top.geometry(f"600x400+{px}+{py}")

        ctk.CTkLabel(
            top, text="CONTROL DE TICKETS AUTOMÁTICO",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        ctk.CTkLabel(
            top, text="Seleccioná las carpetas del Desktop a procesar:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_TABLE, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for carp in folders:
            var = ctk.BooleanVar(value=True)
            checks[carp["name"]] = var

        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        def _update_btn(*_args):
            any_selected = any(v.get() for v in checks.values())
            btn_procesar.configure(state="normal" if any_selected else "disabled")

        def _todas():
            for v in checks.values():
                v.set(True)

        def _ninguna():
            for v in checks.values():
                v.set(False)

        def _confirmar():
            selected = [c for c in folders if checks[c["name"]].get()]
            top.destroy()

            if not selected:
                return

            # Build per-folder groups, merging COMPARTIDO folders together
            import re as _re
            _permiso_pat = _re.compile(r'^\d{2}069EC', _re.IGNORECASE)

            # Separate shared vs normal folders
            shared_groups = {}  # key -> [carp1, carp2, ...]
            normal_groups = []

            for carp in selected:
                name_upper = carp["name"].upper()
                if "COMPARTIDO" in name_upper:
                    # Extract a grouping key: the permiso code from folder name
                    # e.g. "19_06_2026_1_TERRESTRE_465R_560284_EWOS_F7_COMPARTIDO_CERRADO"
                    # Use the date prefix as grouping key (same trip, same date)
                    parts = carp["name"].split("_")
                    # Group by date+number prefix: "19_06_2026_1"
                    if len(parts) >= 3:
                        key = "_".join(parts[:3])  # "19_06_2026_1"
                    else:
                        key = carp["name"]
                    shared_groups.setdefault(key, []).append(carp)
                else:
                    normal_groups.append(carp)

            folder_groups = []
            total_pdfs = 0

            # Process shared folders: merge scans + combine Excels
            for key, shared_carp in shared_groups.items():
                all_pdfs = []
                all_excels = []
                for carp in shared_carp:
                    for f in os.scandir(carp["path"]):
                        if f.is_file() and f.name.lower().endswith('.pdf'):
                            if _permiso_pat.match(f.name):
                                continue
                            if not (f.name.lower().startswith("scan") or f.name.lower().startswith("escan")):
                                continue
                            all_pdfs.append(f.path)
                    if carp["excel_path"]:
                        all_excels.append(carp["excel_path"])

                if all_pdfs and all_excels:
                    # Pass ALL Excels for shared trips — worker sums Neto
                    folder_groups.append((all_pdfs, all_excels))
                    total_pdfs += len(all_pdfs)
                    names = [c["name"] for c in shared_carp]
                    self._log(
                        f"[Auto Tickets] Compartido detectado: {len(names)} carpetas, "
                        f"{len(all_excels)} Excels — Neto se sumará"
                    )

            # Process normal folders: each folder's scans paired with ITS OWN Excel
            for carp in normal_groups:
                pdfs = []
                for f in os.scandir(carp["path"]):
                    if f.is_file() and f.name.lower().endswith('.pdf'):
                        if _permiso_pat.match(f.name):
                            continue
                        if not (f.name.lower().startswith("scan") or f.name.lower().startswith("escan")):
                            continue
                        pdfs.append(f.path)
                if pdfs and carp["excel_path"]:
                    folder_groups.append((pdfs, [carp["excel_path"]]))
                    total_pdfs += len(pdfs)

            if not folder_groups:
                messagebox.showwarning(
                    "Sin datos",
                    "No se encontraron PDFs scan_ con Excel CONTENEDORES"
                )
                return

            self._log(
                f"[Auto Tickets] {len(folder_groups)} grupos, {total_pdfs} PDFs — "
                f"(compartidos: carpetas COMPARTIDO se procesan juntas)"
            )

            if self.tarea_activa:
                messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
                return
            self.tarea_activa = True

            for btn_name in ('btn_controlar_tickets',):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="disabled")

            if (hasattr(self, 'progress_carga')
                    and self.progress_carga.winfo_exists()):
                self.progress_carga.pack(side="right", padx=12)
                self.progress_carga.configure(mode="indeterminate")
                self.progress_carga.start()

            # Show tickets section
            self._switch_tab("tickets")

            # Store groups for serial processing
            self._ct_folder_groups = folder_groups
            self._ct_group_idx = 0

            # Start first group
            self._procesar_siguiente_grupo_ct()

        # Todas / Ninguna (izquierda)
        ctk.CTkButton(
            btn_frame, text="Todas", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_todas,
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            btn_frame, text="Ninguna", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_ninguna,
        ).pack(side="left", padx=4)

        # Procesar (derecha)
        btn_procesar = ctk.CTkButton(
            btn_frame, text="Procesar", width=140, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=_confirmar,
        )
        btn_procesar.pack(side="right", padx=4)

        # Checkboxes
        for carp in folders:
            var = checks[carp["name"]]
            var.trace_add("write", _update_btn)
            ctk.CTkCheckBox(
                scroll, text=carp["name"], variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

    def _procesar_siguiente_grupo_ct(self):
        """Process the next folder group (scan PDFs + Excel(s)) serially.
        Shared groups pass multiple Excels — worker sums Neto."""
        if self._ct_group_idx >= len(self._ct_folder_groups):
            # All groups done
            self._cargar_datos_done()
            return

        pdfs, excels = self._ct_folder_groups[self._ct_group_idx]
        self._ct_group_idx += 1
        remaining = len(self._ct_folder_groups) - self._ct_group_idx
        excel_note = f"{len(excels)} Excels" if len(excels) > 1 else "1 Excel"
        self._log(
            f"[Auto Tickets] Grupo {self._ct_group_idx}/{len(self._ct_folder_groups)}: "
            f"{len(pdfs)} PDFs, {excel_note} — {remaining} grupos restantes"
        )

        t = threading.Thread(
            target=self._cargar_datos_worker,
            args=("local", list(pdfs), list(excels)),
            daemon=True,
        )
        t.start()

    def _cargar_datos_seleccionar_pdfs(self):
        if self._ct_auto_var.get():
            self._cargar_datos_auto_scan()
            return
        from tkinter import filedialog as tk_filedialog
        rutas = tk_filedialog.askopenfilenames(
            title="Seleccionar PDFs y Excel CONTENEDORES",
            filetypes=[("Archivos soportados", "*.pdf *.xlsx *.xls")],
        )
        if not rutas:
            return
        rutas_pdf = [r for r in rutas if r.lower().endswith('.pdf')]
        rutas_excel = [r for r in rutas if r.lower().endswith(('.xlsx', '.xls'))]
        if not rutas_pdf:
            self._log("[Control Datos] ⚠ No seleccionaste ningún PDF.")
            return
        if not rutas_excel:
            self._log("[Control Datos] ⚠ No seleccionaste ningún archivo Excel CONTENEDORES.")
            return
        # Mostrar sección tickets, ocultar otras
        self._switch_tab("tickets")
        self._log(f"[Control Datos] Procesando {len(rutas_pdf)} PDFs contra {len(rutas_excel)} Excel(s)...")
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self.tarea_activa = True
        for btn_name in ('btn_controlar_tickets',):
            btn = getattr(self, btn_name, None)
            if btn and btn.winfo_exists():
                btn.configure(state="disabled")
        if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
            self.progress_carga.pack(side="right", padx=12)
            self.progress_carga.configure(mode="indeterminate")
            self.progress_carga.start()
        t = threading.Thread(
            target=self._cargar_datos_worker,
            args=("local", list(rutas_pdf), list(rutas_excel)),
            daemon=True,
        )
        t.start()

    def _cargar_datos_escribir(self):
        """T8: Escribe valores OCR validados en las celdas DATOS del CONTENEDOR."""
        ok_count = 0
        fail_count = 0
        items = list(self.tree_carga.get_children())

        if not items:
            self._log("[Control Datos] No hay filas para escribir.")
            return

        for item in items:
            vals = self.tree_carga.item(item, "values")
            if not vals or len(vals) < 9:
                continue
            estado = vals[-1]
            if estado != "✅ Ok":
                continue

            iid = item
            if iid not in self._cargar_datos_rutas:
                self._log(f"[Control Datos] ⚠ Sin ruta almacenada para {vals[0]}")
                fail_count += 1
                continue

            ruta, k_idx = self._cargar_datos_rutas[iid]
            neto_val = vals[5]   # columna neto
            tara_val = vals[6]   # columna tara
            nombre_archivo = vals[0]

            try:
                wb = openpyxl.load_workbook(ruta)
                if "DATOS" not in wb.sheetnames:
                    self._log(f"[Control Datos] ⚠ {nombre_archivo}: no tiene hoja DATOS")
                    wb.close()
                    fail_count += 1
                    continue
                ws = wb["DATOS"]

                # Calcular coordenadas según bloque k
                row_base = 20 + (k_idx // 2) * 13
                label_col = 1 + (k_idx % 2) * 12
                value_col = label_col + 6

                # PESO CARGA = row_base + 8, value_col
                ws.cell(row=row_base + 8, column=value_col, value=float(neto_val))
                # TARA CONT = row_base + 9, value_col
                ws.cell(row=row_base + 9, column=value_col, value=float(tara_val))

                wb.save(ruta)
                wb.close()
                ok_count += 1
                self._log(f"[Control Datos] ✓ {nombre_archivo}: Neto={neto_val}, Tara={tara_val}")

            except PermissionError:
                self._log(
                    f"[Control Datos] ✗ El archivo {os.path.basename(ruta)} "
                    "está bloqueado. Cerralo e intentá de nuevo."
                )
                fail_count += 1
            except Exception as e:
                self._log(f"[Control Datos] ✗ Error escribiendo {nombre_archivo}: {e}")
                fail_count += 1

        resumen = f"Escritura completada: {ok_count} ok, {fail_count} errores"
        self._log(f"[Control Datos] {resumen}")

    # ── T5: Matching CONTENEDOR ──────────────────────────────────────
    def _cargar_cache_contenedores(self, rutas_excel=None):
        """Escanea Desktop UNA VEZ y llena self._contenedores_cache abriendo
        cada workbook UNA SOLA VEZ (PE + DATOS + Choferes).
        Si se pasan rutas_excel explícitas, las usa en lugar de escanear."""
        if rutas_excel:
            self._contenedores_cache = {}
            for ruta in rutas_excel:
                data = self._leer_wb_completo(ruta)
                if data:
                    self._contenedores_cache[ruta] = data
            return
        self._contenedores_cache = {}
        base = self._resolver_ruta("planillas_carga", "Desktop")
        if not os.path.isdir(base):
            self._log("[CACHE] Desktop no encontrado")
            return

        archivos = []
        try:
            # Archivos sueltos en Desktop
            for f in os.listdir(base):
                ruta = os.path.join(base, f)
                if os.path.isfile(ruta) and 'CONTENEDORES' in f.upper() and \
                   (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')):
                    archivos.append(ruta)
            # 1 nivel de subcarpetas
            for item in os.listdir(base):
                ruta_item = os.path.join(base, item)
                if os.path.isdir(ruta_item):
                    for f in os.listdir(ruta_item):
                        if 'CONTENEDORES' in f.upper() and \
                           (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')):
                            archivos.append(os.path.join(ruta_item, f))
        except Exception:
            pass

        self._log(f"[CACHE] Escaneando {len(archivos)} archivos CONTENEDORES...")
        for ruta in archivos:
            data = self._leer_wb_completo(ruta)
            if data:
                self._contenedores_cache[ruta] = data
        n = len(self._contenedores_cache)
        self._log(f"[CACHE] Caché completado: {n} archivos")

    def _match_contenedor(self, permiso_ticket):
        """Busca archivos CONTENEDORES asociados al permiso del ticket.

        Usa self._contenedores_cache (cargado por _cargar_cache_contenedores).
        NO abre workbooks — lee PE desde caché.

        Soporta viajes compartidos: si el permiso tiene '/' (ej: '26069EC01000372Y/1000579A'),
        prueba cada parte por separado y combina resultados.

        Retorna lista de rutas de archivos que coinciden (ordenadas por
        estrategia 1 primero, luego estrategia 2), o lista vacía."""
        import re as _re

        if not permiso_ticket:
            self.log_queue.put((self._log_warning, "[MATCH] permiso_ticket vacío, abortando"))
            return []

        # ── Viajes compartidos: probar cada permiso por separado ──────
        partes = [p.strip() for p in str(permiso_ticket).split('/') if p.strip()]
        if len(partes) > 1:
            self.log_queue.put((self._log_warning,
                f"[MATCH] Viaje compartido: {len(partes)} permisos -> {[p for p in partes]}"))
            todas = []
            for parte in partes:
                sub = self._match_contenedor(parte)
                todas.extend(sub)
            # Deduplicar manteniendo orden
            seen = set()
            unicas = []
            for r in todas:
                if r not in seen:
                    seen.add(r)
                    unicas.append(r)
            if unicas:
                self.log_queue.put((self._log_warning,
                    f"[MATCH] ✅ Total candidatos (compartido): {len(unicas)}"))
            return unicas

        # ── Preparar patrones desde el permiso del ticket ────────────
        ticket_alfanum = _re.sub(r'[^a-zA-Z0-9]', '', str(permiso_ticket))
        self.log_queue.put((self._log_warning, f"[MATCH] Permiso limpio: '{ticket_alfanum}'"))

        if not ticket_alfanum:
            return []

        sufijo_ticket = ticket_alfanum[-5:].upper()
        self.log_queue.put((self._log_warning, f"[MATCH] Sufijo (últimos 5): '{sufijo_ticket}'"))

        codigo_corto = ""
        m = _re.search(r'0+([1-9A-Z][A-Z0-9]+)$', ticket_alfanum)
        if m:
            codigo_corto = m.group(1).upper()
            self.log_queue.put((self._log_warning, f"[MATCH] Código corto (tras ceros): '{codigo_corto}'"))

        # ── Buscar en caché en vez de abrir workbooks ──────────────
        if not self._contenedores_cache:
            self.log_queue.put((self._log_warning,
                f"[MATCH] Caché vacío, sin archivos CONTENEDORES"))
            return []

        self.log_queue.put((self._log_warning,
            f"[MATCH] Evaluando {len(self._contenedores_cache)} archivos desde caché..."))

        # ── Evaluar TODOS los archivos desde caché ─────────────────
        estrategia1 = []  # match exacto por sufijo PE
        estrategia2 = []  # código corto en nombre de carpeta

        for ruta, data in self._contenedores_cache.items():
            carpeta = os.path.basename(os.path.dirname(ruta))
            nombre = os.path.basename(ruta)
            pe_val = data.get("pe")
            try:
                if pe_val is not None:
                    pe_alfanum = _re.sub(r'[^a-zA-Z0-9]', '', str(pe_val))
                    pe_sufijo = pe_alfanum[-5:].upper() if len(pe_alfanum) >= 5 else pe_alfanum.upper()
                    if pe_sufijo == sufijo_ticket:
                        estrategia1.append(ruta)
                        self.log_queue.put((self._log_warning,
                            f"[MATCH]   ✅ E1: '{pe_sufijo}' == '{sufijo_ticket}' -> {ruta}"))
                        continue  # no evaluar E2 si ya es E1
                # Estrategia 2: código corto en carpeta
                if codigo_corto and codigo_corto in carpeta.upper():
                    estrategia2.append(ruta)
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   📁 E2: '{codigo_corto}' en carpeta -> {ruta}"))
            except Exception:
                pass

        # Combinar: E1 primero (match exacto), luego E2
        todas = estrategia1 + [r for r in estrategia2 if r not in estrategia1]

        if todas:
            self.log_queue.put((self._log_warning,
                f"[MATCH] ✅ Total candidatos: {len(todas)}"))
            return todas

        self.log_queue.put((self._log_warning,
            f"[MATCH] ❌ Sin candidatos para permiso '{permiso_ticket}' "
            f"(sufijo='{sufijo_ticket}')"))
        return []

    # ── Helper: leer PE de hoja Choferes ──────────────────────────
    def _leer_pe_choferes(self, ruta):
        """Busca etiqueta 'PE' en hoja Choferes y retorna el valor
        de la celda de la derecha. Retorna None si no encuentra."""
        # Cache check
        if ruta in getattr(self, '_contenedores_cache', {}):
            return self._contenedores_cache[ruta].get("pe")
        try:
            if ruta.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
                if 'Choferes' not in wb.sheetnames:
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   {os.path.basename(ruta)}: NO tiene hoja Choferes"))
                    wb.close()
                    return None
                ws = wb['Choferes']
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(r, c).value
                        if v is not None and str(v).strip().upper() == 'PE':
                            pe = ws.cell(r, c + 1).value
                            wb.close()
                            if pe:
                                return str(pe).strip()
                            return None
                wb.close()
                self.log_queue.put((self._log_warning,
                    f"[MATCH]   {os.path.basename(ruta)}: No se encontró etiqueta 'PE' en Choferes"))
                return None
            else:
                book = xlrd.open_workbook(ruta)
                if 'Choferes' not in book.sheet_names():
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   {os.path.basename(ruta)}: NO tiene hoja Choferes"))
                    return None
                ws = book.sheet_by_name('Choferes')
                for r in range(ws.nrows):
                    for c in range(ws.ncols):
                        v = ws.cell_value(r, c)
                        if v and str(v).strip().upper() == 'PE':
                            pe = ws.cell_value(r, c + 1)
                            if pe:
                                return str(pe).strip()
                            return None
                self.log_queue.put((self._log_warning,
                    f"[MATCH]   {os.path.basename(ruta)}: No se encontró etiqueta 'PE' en Choferes"))
                return None
        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[MATCH]   Error al leer {os.path.basename(ruta)}: {ex}"))
            return None

    def _leer_contenedores_desde_choferes(self, ruta_wb):
        """Lee la columna 'NUMERO DE CONTENEDORES' (col D) de hoja Choferes.

        Retorna lista de strings (container numbers vacíos o no),
        en el mismo orden que aparecen en la hoja (skip header row).
        """
        # Cache check
        if ruta_wb in getattr(self, '_contenedores_cache', {}):
            return self._contenedores_cache[ruta_wb].get("contenedores", [])
        import re as _re
        contenedores = []
        try:
            if ruta_wb.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)
                if 'Choferes' not in wb.sheetnames:
                    wb.close()
                    return contenedores
                ws = wb['Choferes']
                # Buscar header "NUMERO DE CONTENEDORES" en fila 1, col D
                col_idx = None
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(1, c).value
                    if v and 'NUMERO' in str(v).upper() and 'CONTENEDOR' in str(v).upper():
                        col_idx = c
                        break
                if col_idx is None:
                    col_idx = 4  # fallback a col D
                # Leer desde fila 2 hasta el final
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(r, col_idx).value
                    contenedores.append(str(val).strip() if val else "")
                wb.close()
            else:
                book = xlrd.open_workbook(ruta_wb)
                if 'Choferes' not in book.sheet_names():
                    return contenedores
                ws = book.sheet_by_name('Choferes')
                col_idx = None
                for c in range(ws.ncols):
                    v = ws.cell_value(0, c)
                    if v and 'NUMERO' in str(v).upper() and 'CONTENEDOR' in str(v).upper():
                        col_idx = c
                        break
                if col_idx is None:
                    col_idx = 3  # fallback a col D (0-indexed)
                for r in range(1, ws.nrows):
                    val = ws.cell_value(r, col_idx)
                    contenedores.append(str(val).strip() if val else "")

            n_reales = sum(1 for c in contenedores if c)
            self.log_queue.put((self._log_warning,
                f"[Choferes] {n_reales} contenedores encontrados en col {col_idx+1}"))
            return contenedores
        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[Choferes] Error leyendo contenedores: {ex}"))
            return contenedores

    # ── T5b: Lector combinado (una sola apertura) ──────────────────────
    def _leer_wb_completo(self, ruta_wb):
        """Abre el workbook UNA VEZ y extrae datos desde Choferes + peso/tara desde DATOS.

        Choferes es la fuente primaria: patente, semi, chofer, DNI, contenedor, precintos, PE.
        DATOS solo provee peso_carga y tara_cont.
        Patentes se limpian al leer (sin guiones/espacios).

        Retorna dict {pe, camiones[], contenedores[], precintos[], peso_flexi_global} o None.
        """
        import re as _re

        def _clean_pat(p):
            """Elimina guiones, espacios y puntuación de patentes."""
            if not p:
                return ""
            return _re.sub(r'[^A-Z0-9]', '', str(p).upper().strip())

        pe = None
        contenedores = []
        precintos = []
        camiones = []
        peso_flexi_global = None

        try:
            if ruta_wb.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)

                # ── Hoja Choferes: fuente primaria ──
                if 'Choferes' in wb.sheetnames:
                    ws = wb['Choferes']

                    # Detectar columnas por headers
                    h_cols = {}
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(1, c).value
                        if v:
                            h = str(v).upper()
                            if 'DOMINIO' in h and 'TRACTOR' in h:
                                h_cols['tractor'] = c
                            if 'DOMINIO' in h and 'SEMI' in h:
                                h_cols['semi'] = c
                            if 'NUMERO' in h and 'CONTENEDOR' in h:
                                h_cols['contenedor'] = c
                            if 'NOMBRE' in h and 'CHOFER' in h:
                                h_cols['chofer'] = c
                            if 'DNI' in h and 'CHOFER' in h:
                                h_cols['dni'] = c
                            if 'PRECINTO' in h and 'ADUANA' in h:
                                h_cols['precinto'] = c
                            if 'PRECINTO' in h and 'LINEA' in h:
                                h_cols['precinto_linea'] = c

                    # Buscar PE
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(r, c).value
                            if v is not None and str(v).strip().upper() == 'PE':
                                pe_v = ws.cell(r, c + 1).value
                                if pe_v:
                                    pe = str(pe_v).strip()
                                break
                        if pe is not None:
                            break

                    # Buscar PESO FLEXI
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(r, c).value
                            if v is not None and 'PESO FLEXI' in str(v).upper():
                                pfv = ws.cell(r, c + 1).value
                                if pfv is not None:
                                    try:
                                        peso_flexi_global = float(str(pfv).strip())
                                    except ValueError:
                                        pass
                                break
                        if peso_flexi_global is not None:
                            break

                    # Leer camiones desde filas de Choferes (empieza en fila 2)
                    for r in range(2, ws.max_row + 1):
                        tractor_val = ws.cell(r, h_cols.get('tractor', 1)).value if 'tractor' in h_cols else None
                        if not tractor_val:
                            break  # Fin de datos de camiones
                        camion = {"k": len(camiones), "peso_carga": 0, "tara_cont": 0}
                        if 'tractor' in h_cols:
                            v = ws.cell(r, h_cols['tractor']).value
                            camion["patente_camion"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_camion"] = ""
                        if 'semi' in h_cols:
                            v = ws.cell(r, h_cols['semi']).value
                            camion["patente_semi"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_semi"] = ""
                        if 'contenedor' in h_cols:
                            v = ws.cell(r, h_cols['contenedor']).value
                            camion["contenedor"] = str(v).strip() if v else ""
                        else:
                            camion["contenedor"] = ""
                        if 'chofer' in h_cols:
                            v = ws.cell(r, h_cols['chofer']).value
                            camion["conductor"] = str(v).strip() if v else ""
                        else:
                            camion["conductor"] = ""
                        if 'dni' in h_cols:
                            v = ws.cell(r, h_cols['dni']).value
                            if v is not None:
                                camion["dni"] = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
                            else:
                                camion["dni"] = ""
                        else:
                            camion["dni"] = ""
                        # Precinto: unir aduana + linea con guion
                        prec_parts = []
                        if 'precinto' in h_cols:
                            pv = ws.cell(r, h_cols['precinto']).value
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        if 'precinto_linea' in h_cols:
                            pv = ws.cell(r, h_cols['precinto_linea']).value
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        camion["precinto"] = "-".join(prec_parts) if prec_parts else ""
                        contenedores.append(camion["contenedor"])
                        precintos.append(camion["precinto"])
                        camiones.append(camion)

                # ── Hoja DATOS: SOLO peso_carga y tara_cont ──
                if 'DATOS' in wb.sheetnames and camiones:
                    ws_d = wb['DATOS']
                    max_row = ws_d.max_row or 0
                    max_col = ws_d.max_column or 0

                    def _celda(row, col):
                        if row > max_row or col > max_col:
                            return None
                        return ws_d.cell(row=row, column=col).value

                    def _val_num(raw_val):
                        if raw_val is None:
                            return 0
                        if isinstance(raw_val, (int, float)):
                            return float(raw_val)
                        return 0

                    # Buscar posiciones de CAMION / PATENTE CAMION en DATOS
                    camion_positions = []
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            raw = _celda(r, c)
                            if raw and isinstance(raw, str):
                                lbl = raw.strip().upper()
                                if lbl in ('CAMION', 'PATENTE CAMION'):
                                    camion_positions.append((r, c))

                    # Para cada CAMION, buscar PESO CARGA y TARA CONT en filas cercanas
                    peso_tara_list = []
                    for cam_row, cam_col in camion_positions:
                        peso = 0
                        tara = 0
                        for offset in range(1, 12):
                            check_row = cam_row + offset
                            if check_row > max_row:
                                break
                            raw = _celda(check_row, cam_col)
                            if raw and isinstance(raw, str):
                                lbl = raw.strip().upper()
                                if lbl in ('PESO CARGA', 'PESO CARGA (KG)'):
                                    peso = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 <= max_col else 0
                                elif lbl == 'TARA CONT':
                                    tara = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 <= max_col else 0
                        peso_tara_list.append((peso, tara))

                    # Asignar peso/tara a camiones por orden
                    for i, camion in enumerate(camiones):
                        if i < len(peso_tara_list):
                            camion["peso_carga"] = peso_tara_list[i][0]
                            camion["tara_cont"] = peso_tara_list[i][1]

                wb.close()

            else:
                # .xls con xlrd
                book = xlrd.open_workbook(ruta_wb)

                # ── Hoja Choferes: fuente primaria ──
                if 'Choferes' in book.sheet_names():
                    ws = book.sheet_by_name('Choferes')

                    # Detectar columnas por headers (row 0 en xlrd)
                    h_cols = {}
                    for c in range(ws.ncols):
                        v = ws.cell_value(0, c)
                        if v:
                            h = str(v).upper()
                            if 'DOMINIO' in h and 'TRACTOR' in h:
                                h_cols['tractor'] = c
                            if 'DOMINIO' in h and 'SEMI' in h:
                                h_cols['semi'] = c
                            if 'NUMERO' in h and 'CONTENEDOR' in h:
                                h_cols['contenedor'] = c
                            if 'NOMBRE' in h and 'CHOFER' in h:
                                h_cols['chofer'] = c
                            if 'DNI' in h and 'CHOFER' in h:
                                h_cols['dni'] = c
                            if 'PRECINTO' in h and 'ADUANA' in h:
                                h_cols['precinto'] = c
                            if 'PRECINTO' in h and 'LINEA' in h:
                                h_cols['precinto_linea'] = c

                    # Buscar PE
                    for r in range(ws.nrows):
                        for c in range(ws.ncols):
                            v = ws.cell_value(r, c)
                            if v and str(v).strip().upper() == 'PE':
                                pe_v = ws.cell_value(r, c + 1)
                                if pe_v:
                                    pe = str(pe_v).strip()
                                break
                        if pe is not None:
                            break

                    # Buscar PESO FLEXI
                    for r in range(ws.nrows):
                        for c in range(ws.ncols):
                            v = ws.cell_value(r, c)
                            if v and 'PESO FLEXI' in str(v).upper():
                                pfv = ws.cell_value(r, c + 1)
                                if pfv:
                                    try:
                                        peso_flexi_global = float(str(pfv).strip())
                                    except ValueError:
                                        pass
                                break
                        if peso_flexi_global is not None:
                            break

                    # Leer camiones desde filas de Choferes (empieza en row 1 en xlrd)
                    for r in range(1, ws.nrows):
                        tractor_val = ws.cell_value(r, h_cols.get('tractor', 0)) if 'tractor' in h_cols else None
                        if not tractor_val:
                            break  # Fin de datos de camiones
                        camion = {"k": len(camiones), "peso_carga": 0, "tara_cont": 0}
                        if 'tractor' in h_cols:
                            v = ws.cell_value(r, h_cols['tractor'])
                            camion["patente_camion"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_camion"] = ""
                        if 'semi' in h_cols:
                            v = ws.cell_value(r, h_cols['semi'])
                            camion["patente_semi"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_semi"] = ""
                        if 'contenedor' in h_cols:
                            v = ws.cell_value(r, h_cols['contenedor'])
                            camion["contenedor"] = str(v).strip() if v else ""
                        else:
                            camion["contenedor"] = ""
                        if 'chofer' in h_cols:
                            v = ws.cell_value(r, h_cols['chofer'])
                            camion["conductor"] = str(v).strip() if v else ""
                        else:
                            camion["conductor"] = ""
                        if 'dni' in h_cols:
                            v = ws.cell_value(r, h_cols['dni'])
                            if v is not None:
                                camion["dni"] = str(int(v)) if isinstance(v, float) and v == int(v) else str(v).strip()
                            else:
                                camion["dni"] = ""
                        else:
                            camion["dni"] = ""
                        # Precinto
                        prec_parts = []
                        if 'precinto' in h_cols:
                            pv = ws.cell_value(r, h_cols['precinto'])
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        if 'precinto_linea' in h_cols:
                            pv = ws.cell_value(r, h_cols['precinto_linea'])
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        camion["precinto"] = "-".join(prec_parts) if prec_parts else ""
                        contenedores.append(camion["contenedor"])
                        precintos.append(camion["precinto"])
                        camiones.append(camion)

                # ── Hoja DATOS: SOLO peso_carga y tara_cont ──
                if 'DATOS' in book.sheet_names() and camiones:
                    ws_d = book.sheet_by_name('DATOS')
                    max_row = ws_d.nrows
                    max_col = ws_d.ncols

                    def _celda(row, col):
                        if row > max_row or col > max_col:
                            return None
                        return ws_d.cell_value(row - 1, col - 1)

                    def _val_num(raw_val):
                        if raw_val is None:
                            return 0
                        if isinstance(raw_val, (int, float)):
                            return float(raw_val)
                        return 0

                    camion_positions = []
                    for r in range(1, max_row + 1):
                        for c in range(0, max_col):
                            raw = _celda(r, c)
                            if raw is not None:
                                lbl = str(raw).strip().upper()
                                if lbl in ('CAMION', 'PATENTE CAMION'):
                                    camion_positions.append((r, c))

                    peso_tara_list = []
                    for cam_row, cam_col in camion_positions:
                        peso = 0
                        tara = 0
                        for offset in range(1, 12):
                            check_row = cam_row + offset
                            if check_row > max_row:
                                break
                            raw = _celda(check_row, cam_col)
                            if raw is not None:
                                lbl = str(raw).strip().upper()
                                if lbl in ('PESO CARGA', 'PESO CARGA (KG)'):
                                    peso = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 < max_col else 0
                                elif lbl == 'TARA CONT':
                                    tara = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 < max_col else 0
                        peso_tara_list.append((peso, tara))

                    for i, camion in enumerate(camiones):
                        if i < len(peso_tara_list):
                            camion["peso_carga"] = peso_tara_list[i][0]
                            camion["tara_cont"] = peso_tara_list[i][1]

        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[CACHE] Error al leer {os.path.basename(ruta_wb)}: {ex}"))
            return None

        n_ctn = sum(1 for c in contenedores if c)
        self.log_queue.put((self._log_warning,
            f"[CACHE] Cacheado: {os.path.basename(ruta_wb)} — "
            f"PE={pe} camiones={len(camiones)} contenedores={n_ctn}"))
        return {"pe": pe, "camiones": camiones, "contenedores": contenedores,
                "peso_flexi_global": peso_flexi_global,
                "precintos": precintos}

    # ── T6: Lectura de DATOS y comparación ────────────────────────────
    def _leer_datos_contenedor(self, ruta_wb):
        """Lee la hoja DATOS del CONTENEDOR y devuelve dict con camiones.

        Retorna:
            {"camiones": [{patente_camion, patente_semi, conductor,
                           dni, peso_carga, tara_cont}, ...]}
            o None si no hay hoja DATOS.
        """
        # Cache check: si ya fue cargado por _leer_wb_completo, evitar otro open
        if ruta_wb in getattr(self, '_contenedores_cache', {}):
            camiones = self._contenedores_cache[ruta_wb].get("camiones", [])
            return {"camiones": camiones} if camiones else None

        if ruta_wb.lower().endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)
            except Exception as e:
                self.log_queue.put((self._log_warning, f"No se pudo abrir {os.path.basename(ruta_wb)}: {e}"))
                return None
            if 'DATOS' not in wb.sheetnames:
                wb.close()
                self.log_queue.put((self._log_warning, f"El archivo {os.path.basename(ruta_wb)} no tiene hoja DATOS"))
                return None
            ws = wb['DATOS']
            usar_xlrd = False
        else:
            # .xls con xlrd
            try:
                book = xlrd.open_workbook(ruta_wb)
            except Exception as e:
                self.log_queue.put((self._log_warning, f"No se pudo abrir {os.path.basename(ruta_wb)}: {e}"))
                return None
            if 'DATOS' not in book.sheet_names():
                self.log_queue.put((self._log_warning, f"El archivo {os.path.basename(ruta_wb)} no tiene hoja DATOS"))
                return None
            ws = book.sheet_by_name('DATOS')
            usar_xlrd = True

        # Obtener límites del sheet para evitar IndexError
        if usar_xlrd:
            max_row = ws.nrows
            max_col = ws.ncols
        else:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

        def _celda(row, col):
            """Lee valor de celda (1-indexed row/col)."""
            if row > max_row or col > max_col:
                return None
            if usar_xlrd:
                return ws.cell_value(row - 1, col - 1)
            else:
                return ws.cell(row=row, column=col).value

        def _val_num(raw_val):
            """Retorna float si raw_val es numérico, 0 en otro caso."""
            if raw_val is None:
                return 0
            if isinstance(raw_val, (int, float)):
                return float(raw_val)
            return 0

        camiones = []

        # ── Scan dinámico: busca todas las etiquetas PATENTE CAMION en DATOS ──
        # Soporta layout horizontal (C1..C3 en cols 1,13,25 misma fila)
        # y el layout antiguo de 2 por bloque de filas.
        bloques = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                raw = _celda(r, c)
                if raw is not None:
                    etiqueta = str(raw).strip().upper()
                    if 'PATENTE' in etiqueta and 'CAMION' in etiqueta:
                        if c + 6 <= max_col:
                            bloques.append((r, c))

        # Ordenar por fila, luego columna (izquierda→derecha, arriba→abajo)
        bloques.sort(key=lambda b: (b[0], b[1]))

        for k, (patente_row, label_col) in enumerate(bloques):
            row_base = patente_row - 1  # PATENTE CAMION está en row_base + 1
            value_col = label_col + 6

            patente_camion = str(_celda(patente_row, value_col) or "").strip()
            patente_semi   = str(_celda(patente_row + 1, value_col) or "").strip()
            conductor      = str(_celda(patente_row + 2, value_col) or "").strip()
            dni_raw        = _celda(patente_row + 3, value_col)
            if isinstance(dni_raw, (int, float)):
                dni_celda = str(int(dni_raw))
            else:
                dni_celda = str(dni_raw or "").strip()
            peso_carga_num = _val_num(_celda(patente_row + 7, value_col))
            tara_cont_num  = _val_num(_celda(patente_row + 8, value_col))

            camion = {
                "k": k,
                "patente_camion": patente_camion,
                "patente_semi": patente_semi,
                "conductor": conductor,
                "dni": dni_celda,
                "contenedor": "",  # se mergea desde Choferes después
                "peso_carga": peso_carga_num,
                "tara_cont": tara_cont_num,
            }
            camiones.append(camion)

        if not usar_xlrd:
            wb.close()

        # ── Merge container numbers from Choferes sheet (col D: NUMERO DE CONTENEDORES) ──
        try:
            contenedores = self._leer_contenedores_desde_choferes(ruta_wb)
            if contenedores and camiones:
                for i, camion in enumerate(camiones):
                    if i < len(contenedores) and contenedores[i]:
                        camion["contenedor"] = contenedores[i]
                        self.log_queue.put((self._log_warning,
                            f"[DATOS]   Contenedor[{i}]: '{contenedores[i]}' -> Patente: {camion['patente_camion']}"))
        except Exception as e:
            self.log_queue.put((self._log_warning, f"[DATOS] Error merge contenedores: {e}"))

        resultado = {"camiones": camiones} if camiones else None
        self.log_queue.put((self._log_warning,
            f"[DATOS] _leer_datos_contenedor: {len(camiones)} camiones encontrados "
            f"en {os.path.basename(ruta_wb)}"
            + (f" -> {[c['patente_camion'] for c in camiones]}" if camiones else "")))
        return resultado

    # ── T4: Worker OCR ────────────────────────────────────────────────
    def _cargar_datos_worker(self, fuente, rutas, rutas_excel=None):
        """Worker que corre en threading.Thread daemon.
        
        Procesa cada PDF: OCR → extracción → matching CONTENEDOR → lectura DATOS.
        Pushea resultados a log_queue como ("_OCR_RESULT_", data).
        Al finalizar pushea ("_OCR_DONE_", None).
        
        Args:
            rutas_excel: lista de rutas a archivos CONTENEDORES seleccionados
                         por el usuario. Si es None, escanea Desktop (viejo).
        """
        self._set_log_panel("cargar-datos")
        import procesar_tickets
        import re as _re
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] [Control Datos] Fuente: {fuente} — {len(rutas)} PDF(s)")

        total = len(rutas)
        # ── 0. Cargar caché de CONTENEDORES (una vez, zero opens repetidos) ──
        self._contenedores_cache = {}
        if rutas_excel:
            self._log(f"[Control Datos] Usando {len(rutas_excel)} Excel(s) seleccionados por el usuario")
            self._cargar_cache_contenedores(rutas_excel)
        else:
            self._cargar_cache_contenedores()  # escanea Desktop (compatibilidad)
        # ── 1. OCR: según método seleccionado ──
        ocr_method = self.config.get("ocr_method", "api_vision")
        textos_por_pdf = {}
        api_datos_raw = {}  # dict crudo de API Visión para mapeo directo

        if ocr_method == "api_vision":
            api_vision_conf = self.config.get("api_vision", {})
            api_key_raw = api_vision_conf.get("api_key", "")
            api_key = self._decrypt_api_key(api_key_raw)

            if not api_key:
                self.log_queue.put(f"[{timestamp}] [Control Datos] ⚠ No hay API Key configurada, usando OCR local")
                textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(rutas, engine="paddleocr")
            else:
                model = api_vision_conf.get("model", procesar_tickets.MODELO_VISION_DEFAULT)
                temperature = api_vision_conf.get("temperature", 0.1)
                max_tokens = api_vision_conf.get("max_tokens", 4000)
                timeout = api_vision_conf.get("timeout", 60)
                parallel_enabled = api_vision_conf.get("parallel_enabled", False)

                if parallel_enabled and total > 1:
                    # Parallel mode: distribute across enabled models
                    model_states = api_vision_conf.get("parallel_model_states", {})
                    enabled_models = [m for m, on in model_states.items() if on]
                    if not enabled_models:
                        enabled_models = [model]
                    # Ensure selected model is first
                    if model in enabled_models:
                        enabled_models.remove(model)
                    enabled_models.insert(0, model)

                    def _log_fn(msg):
                        self.log_queue.put(f"[{timestamp}] {msg}")

                    self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión paralelo: {total} PDF(s), {len(enabled_models)} modelo(s)...")
                    result = procesar_tickets.api_vision_con_fallback(
                        rutas, api_key, enabled_models,
                        temperature=temperature, max_tokens=max_tokens,
                        timeout=timeout, log_callback=_log_fn,
                    )
                    api_datos_raw = result["datos"]
                    textos_por_pdf = result["textos"]
                else:
                    # Sequential fallback: try models one by one per PDF
                    all_models = api_vision_conf.get("custom_models", [])
                    if not all_models:
                        all_models = list(procesar_tickets.MODELOS_VISION)
                    # Ensure selected model is first
                    if model in all_models:
                        all_models.remove(model)
                    all_models.insert(0, model)

                    self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión secuencial: {total} PDF(s), {len(all_models)} modelo(s) disponible(s)")
                    for i, ruta in enumerate(rutas):
                        stem = os.path.splitext(os.path.basename(ruta))[0]
                        carpeta = os.path.basename(os.path.dirname(ruta))
                        cliente = carpeta.split("_")[7] if len(carpeta.split("_")) > 7 else carpeta
                        self.log_queue.put(f"[{timestamp}] [Control Datos]   [{i+1}/{total}] {stem}")
                        success = False
                        for m in all_models:
                            try:
                                datos = procesar_tickets.api_vision_extraer_datos(
                                    ruta, api_key, model=m,
                                    temperature=temperature, max_tokens=max_tokens,
                                    timeout=timeout,
                                )
                                if "error" in datos:
                                    # Clean error message
                                    err = datos["error"]
                                    if "is not a valid model ID" in err:
                                        err = "modelo no disponible"
                                    elif "429" in err or "rate" in err.lower():
                                        err = "rate limit"
                                    elif "timeout" in err.lower():
                                        err = "timeout"
                                    self.log_queue.put(f"[{timestamp}] [Control Datos]   ERROR {m}: {err}")
                                    continue
                                # Success
                                api_datos_raw[stem] = datos
                                texto_simulado = "\n".join(f"{k}: {v}" for k, v in datos.items() if v)
                                textos_por_pdf[stem] = texto_simulado
                                self.log_queue.put(f"[{timestamp}] [Control Datos]   ✓ {stem} procesado con {m}")
                                success = True
                                break
                            except Exception as e:
                                self.log_queue.put(f"[{timestamp}] [Control Datos]   ERROR {m}: {e}")
                                continue
                        if not success:
                            self.log_queue.put(f"[{timestamp}] [Control Datos]   ⚠ Todos los modelos fallaron para {stem}, usando PaddleOCR")
                            textos_local = procesar_tickets.pdfs_a_texto_batch([ruta], engine="paddleocr")
                            textos_por_pdf[stem] = textos_local.get(stem, "")

                vision_ok = len(api_datos_raw)
                self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión completado: {vision_ok}/{total} PDFs OK")
        else:
            self.log_queue.put(f"[{timestamp}] [Control Datos] OCR batch: {total} PDF(s)...")
            textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(rutas, engine="paddleocr")
            self.log_queue.put(f"[{timestamp}] [Control Datos] OCR batch completado ({len(textos_por_pdf)} PDFs)")

        # ── 2. Parsear cada resultado + match CONTENEDOR ──
        def _normalizar_simple(s):
            return _re.sub(r'[^A-Z0-9]', '', str(s).upper().strip())

        def _procesar_texto(stem, texto):
            """Parsea el texto OCR y busca match CONTENEDOR.
            Si hay datos crudos de API Visión, mapea directamente."""
            
            if stem in api_datos_raw:
                # Mapeo directo desde JSON de API Visión
                raw = api_datos_raw[stem]
                patente  = raw.get("Patente", "")
                semi     = raw.get("Semirremolque", "")
                conductor = raw.get("Conductor", "")
                dni      = raw.get("DNI", "")
                permiso  = raw.get("Permiso", "")
                neto_str = raw.get("Neto", "")
                tara_str = raw.get("Tara Contenedor", "")
                contenedor_str = raw.get("Contenedor", "")
            else:
                # Parseo tradicional desde texto OCR
                datos = procesar_tickets.extraer_datos(texto)
                patente  = datos.get("Patente Camion", "")
                semi     = datos.get("Patente Acoplado", "")
                conductor_full = datos.get("Conductor", "")
                permiso  = datos.get("Merc./Permiso", "")
                neto_str = datos.get("Peso Neto (kg)", "")
                tara_str = datos.get("Tara Contenedor", "")
                contenedor_str = datos.get("Contenedor", "")
                dni = datos.get("DNI Conductor", "") or ""
                conductor = datos.get("Conductor", conductor_full) or conductor_full

            try:
                neto = float(neto_str.replace('.', '').replace(',', '.')) if neto_str else 0
            except (ValueError, AttributeError):
                neto = 0
            try:
                tara = float(tara_str.replace('.', '').replace(',', '.')) if tara_str else 0
            except (ValueError, AttributeError):
                tara = 0

            ticket_data = {
                "archivo": stem,
                "patente": patente, "semi": semi,
                "conductor": conductor, "dni": dni,
                "neto": neto, "tara": tara,
                "contenedor": contenedor_str,
                "permiso": permiso,
            }

            # Match CONTENEDOR — buscar por patente/semi/DNI en cache
            # Patentes en cache ya están limpias (sin guiones/espacios)
            ruta_match = None
            cont_data = None
            pe_val = None
            pn = _normalizar_simple(patente)
            sn = _normalizar_simple(semi)
            dn = _re.sub(r'\D', '', str(dni).strip())

            # ── Buscar por patente/semi/DNI en cache ──
            all_matches = []  # [(ruta, data, camion)]
            shared_excels = []
            shared_neto_sum = 0
            if self._contenedores_cache:
                for ruta, data in self._contenedores_cache.items():
                    if not data or not data.get("camiones"):
                        continue
                    for camion in data["camiones"]:
                        cp = camion.get("patente_camion", "")  # ya limpio en cache
                        cs = camion.get("patente_semi", "")    # ya limpio en cache
                        cdn = _re.sub(r'\D', '', str(camion.get("dni", "")).strip())
                        if (pn and cp and pn == cp) or \
                           (sn and cs and sn == cs) or \
                           (dn and cdn and dn == cdn):
                            all_matches.append((ruta, data, camion))
                            break

            if all_matches:
                ruta_match = all_matches[0][0]
                cont_data = all_matches[0][1]
                pe_val = data.get("pe")  # ya leído en cache

                if len(all_matches) > 1:
                    for ruta, data, camion in all_matches:
                        carpeta = os.path.basename(os.path.dirname(ruta))
                        shared_excels.append({
                            "ruta": ruta,
                            "nombre": carpeta,
                            "data": data,
                            "camion": camion,
                            "pe": data.get("pe"),
                        })
                        shared_neto_sum += camion.get("peso_carga", 0)
                    self.log_queue.put(
                        f"[MATCH] ✅ Fast compartido: patente '{patente}' "
                        f"-> {len(all_matches)} Excels, Neto total: {shared_neto_sum} kg")
                else:
                    self.log_queue.put(
                        f"[MATCH] ✅ Fast: patente '{patente}' -> {os.path.basename(ruta_match)}")

            return ticket_data, cont_data, ruta_match, pe_val, permiso, shared_excels, shared_neto_sum

        # ── 3. Procesar resultados y enviarlos a la UI ──
        # Sort by parent folder for grouped display
        rutas.sort(key=lambda r: os.path.basename(os.path.dirname(r)).split("_")[7].lower() if len(os.path.basename(os.path.dirname(r)).split("_")) > 7 else "")
        for hechos, ruta in enumerate(rutas, start=1):
            stem = os.path.splitext(os.path.basename(ruta))[0]
            has_data = stem in api_datos_raw or stem in textos_por_pdf

            if has_data:
                texto = textos_por_pdf.get(stem, "")
                ticket_data, cont_data, ruta_match, pe_val, permiso, \
                    shared_excels, shared_neto_sum = _procesar_texto(stem, texto)
                nombre = stem

                self.log_queue.put(
                    f"[{timestamp}] [{hechos}/{total}] ✓ {nombre} — permiso {permiso}")
                if shared_excels:
                    self.log_queue.put(
                        f"     ✓ COMPARTIDO: {len(shared_excels)} Excels, "
                        f"Neto total: {shared_neto_sum} kg")
                elif ruta_match:
                    self.log_queue.put(
                        f"     ✓ CONTENEDOR: {os.path.basename(ruta_match)}")
                else:
                    self.log_queue.put(
                        f"     ⚠ Sin CONTENEDOR match para permiso '{permiso}'")
                self.log_queue.put(("_OCR_RESULT_", {
                    "ticket": ticket_data,
                    "contenedor": cont_data,
                    "match": ruta_match,
                    "pe": pe_val,
                    "carpeta": os.path.basename(os.path.dirname(ruta)),
                    "cliente": os.path.basename(os.path.dirname(ruta)).split("_")[7] if len(os.path.basename(os.path.dirname(ruta)).split("_")) > 7 else "",
                    "shared_excels": shared_excels,
                    "shared_neto_sum": shared_neto_sum,
                }))
            else:
                nombre = os.path.basename(ruta)
                self.log_queue.put(
                    f"[{timestamp}] [{hechos}/{total}] ✗ {nombre}: OCR falló")

        # Finalizar
        self.log_queue.put(("_OCR_DONE_", None))
        self._contenedores_cache = {}  # liberar memoria

    def _cargar_datos_done(self):
        """Callback al recibir _OCR_DONE_. Desbloquea UI y habilita escritura.
        Si hay grupos pendientes (auto mode), procesa el siguiente."""
        # Chain: process next folder group if pending
        if (hasattr(self, '_ct_folder_groups')
                and self._ct_group_idx < len(self._ct_folder_groups)):
            self._procesar_siguiente_grupo_ct()
            return

        # No more groups — final cleanup
        self._ct_folder_groups = []
        self._ct_group_idx = 0
        self.tarea_activa = False
        try:
            for btn_name in ('btn_controlar_tickets',):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="normal")
            if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
                self.progress_carga.stop()
                self.progress_carga.pack_forget()
        except Exception:
            pass
        self._log("[Control Datos] OCR finalizado. Revisar resultados en la tabla.")

    def _finalizar_tarea(self):
        """Callback al recibir _TAREA_COMPLETA_. Rehabilita botones."""
        self.tarea_activa = False
        try:
            for btn_name in ('btn_control_final', 'btn_controlar_tickets', 'btn_controlar_coordinacion'):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="normal")
            if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
                self.progress_carga.stop()
                self.progress_carga.pack_forget()
        except Exception:
            pass
        self._log("[Control] Tarea completada.")

    def _on_ocr_method_change(self, choice):
        """Persiste el método OCR elegido."""
        val = "api_vision" if choice == "API Visión" else "local"
        self.config["ocr_method"] = val
        # Habilitar/deshabilitar selector de modelo según método
        es_api = val == "api_vision"
        self._modelo_vision_menu.configure(state="normal" if es_api else "disabled")

    def _on_modelo_vision_change(self, choice):
        """Persiste el modelo de API Visión elegido."""
        self.config.setdefault("api_vision", {})["model"] = choice
        # Sincronizar también el dropdown de Ajustes si existe
        if hasattr(self, '_ent_vision_model') and self._ent_vision_model.winfo_exists():
            self._ent_vision_model.set(choice)

    def _actualizar_modelos_vision(self):
        """Sincroniza el dropdown de modelo con la config actual."""
        import procesar_tickets
        modelos_guardados = self.config.get("api_vision", {}).get("custom_models", [])
        modelos = modelos_guardados if modelos_guardados else list(procesar_tickets.MODELOS_VISION)
        modelo_default = self.config.get("api_vision", {}).get("model", procesar_tickets.MODELO_VISION_DEFAULT)
        self._modelo_vision_menu.configure(values=modelos)
        if modelo_default in modelos:
            self._modelo_vision_menu.set(modelo_default)
        elif modelos:
            self._modelo_vision_menu.set(modelos[0])


