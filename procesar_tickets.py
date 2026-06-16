"""
procesar_tickets.py
-------------------
Extrae datos de Tickets de Pesaje (AGD) escaneados y los exporta a Excel.

Uso:
    # Un solo PDF:
    python procesar_tickets.py ticket.pdf

    # Varios PDFs (los acumula en una sola planilla):
    python procesar_tickets.py ticket1.pdf ticket2.pdf ticket3.pdf

    # Todos los PDFs de una carpeta:
    python procesar_tickets.py carpeta/

Dependencias:
    pip install pytesseract pdf2image openpyxl
    apt install tesseract-ocr tesseract-ocr-spa poppler-utils  (Linux)
    brew install tesseract poppler                              (Mac)
"""

import re
import subprocess
import sys
import tempfile
from pathlib import Path

import json
import os
import unicodedata
import cv2
import numpy as np
import pytesseract
from PIL import Image
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FILE = "tickets_pesaje.xlsx"

# --- Configuración de rutas de binarios ---
# --- Resolver base de la aplicación (exe o script) ---
def _app_dir():
    """Directorio del ejecutable (para sidecars como python39/)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _resource_dir():
    """Directorio de recursos embebidos (engines/, poppler/)."""
    if getattr(sys, 'frozen', False):
        return getattr(sys, '_MEIPASS', _app_dir())
    return os.path.dirname(os.path.abspath(__file__))

_app_base = _app_dir()
_res_base = _resource_dir()

# --- Resolver python39 (sidecar al lado del exe o bundleado en _internal/) ---
_python39_paths = [
    os.path.join(_app_base, "python39", "python.exe"),              # sidecar legacy al lado del exe
]
if getattr(sys, 'frozen', False):
    _python39_paths.extend([
        os.path.join(_res_base, "python39", "python.exe"),          # subdirectorio python39/ en _internal/
        os.path.join(_res_base, "python.exe"),                      # expandido directo en _internal/ (destino '.')
    ])
_python39_resolved = next((p for p in _python39_paths if os.path.isfile(p)), None)

# Debug: log de rutas al cargar el módulo
_debug_ok = True
_debug_msgs = []
for _name, _path in [
    ("python39/python.exe", _python39_resolved or _python39_paths[0]),
    ("engines/paddleocr/ocr_helper.py", os.path.join(_res_base, "engines", "paddleocr", "ocr_helper.py")),
    ("poppler/pdftoppm.exe", os.path.join(_res_base, "poppler", "Library", "bin", "pdftoppm.exe")),
]:
    if not os.path.isfile(_path):
        _debug_ok = False
        _debug_msgs.append(f"  [ERR] {_name}: NO ENCONTRADO en {_path}")
    else:
        _debug_msgs.append(f"  [OK]  {_name}")
if not _debug_ok:
    print(f"[OCR] ERROR DE RUTAS — app_base={_app_base} res_base={_res_base}")
    for m in _debug_msgs:
        print(m)

# --- Configuración de rutas de binarios ---
_poppler_sidecar = os.path.join(_res_base, "poppler", "Library", "bin")
POPPLER_PATH = _poppler_sidecar if os.path.isdir(_poppler_sidecar) else (os.environ.get("POPPLER_PATH") or "")
TESSERACT_CMD = os.environ.get("TESSERACT_CMD") or r"C:\Program Files\Tesseract-OCR\tesseract.exe"

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# --- Constantes de motores OCR ---
OCR_ENGINE_DEFAULT = "paddleocr"
TESSERACT_SIDECAR = os.path.join(_res_base, "engines", "tesseract")
PADDLE_SIDECAR = os.path.join(_res_base, "engines", "paddleocr")
PADDLE_PORTABLE_PYTHON = _python39_resolved or os.path.join(_app_base, "python39", "python.exe")
PADDLE_OCR_HELPER = os.path.join(PADDLE_SIDECAR, "ocr_helper.py")

# --- Sidecar detection for Tesseract ---
_tesseract_sidecar_exe = os.path.join(TESSERACT_SIDECAR, "tesseract.exe")
if os.path.isfile(_tesseract_sidecar_exe):
    pytesseract.pytesseract.tesseract_cmd = _tesseract_sidecar_exe

# --- PaddleOCR via subprocess con Python portable 3.9 ---
# Cache: almacena la ruta al python.exe portable, o None si no disponible.
_paddle_portable = None  # None = no verificado, False = no disponible, str = ruta

CAMPOS = [
    "Ticket Salida",
    "Fecha",
    "Turno",
    "Llegada",
    "Salida hora",
    "Emisor",
    "CUIT Emisor",
    "Remitente",
    "CUIT Remitente",
    "Transportista",
    "CUIT Transportista",
    "Destinatario",
    "CUIT Destinatario",
    "Clase",
    "N° Carta de Porte",
    "Producto",
    "Patente Camion",
    "Patente Acoplado",
    "Peso Bruto (kg)",
    "Peso Tara (kg)",
    "Peso Neto (kg)",
    "LOT",
    "ATA (Nombre-CUIT)",
    "Conductor",
    "DNI Conductor",
    "Merc./Permiso",
    "Pallet",
    "Contenedor",        # N° contenedor ISO/FLEXI (ej: MSMU 258531-2)
    "Tara Contenedor",   # Tara del contenedor (ej: 2.100)
]


# ---------------------------------------------------------------------------
# Corrección posicional de patentes
# ---------------------------------------------------------------------------

# ── Formatos de patentes Mercosur ──
# Cada tupla: (nombre, longitud, {posiciones_letra}, {posiciones_numero})
_FORMATOS_PATENTE = [
    # (nombre, longitud, {pos_letra}, {pos_numero}, prioridad)
    # Prioridad más baja = preferido en empate (ARG first por ser transporte arg)
    ('ARG_MERCOSUR', 7, {0,1,5,6}, {2,3,4},        0),   # LL NNN LL
    ('ARG_VIEJA',    6, {0,1,2},   {3,4,5},        1),   # LLL NNN
    ('PRY_VIEJO',    6, {0,1,2},   {3,4,5},        2),   # LLL NNN (igual ARG vieja)
    ('BRA_MERCOSUR', 7, {0,1,2,4}, {3,5,6},        3),   # LLL N L NN
    ('BRA_VIEJA',    7, {0,1,2},   {3,4,5,6},      4),   # LLL NNNN
    ('URY',          7, {0,1,2},   {3,4,5,6},      5),   # LLL NNNN (igual BRA vieja)
    ('PRY_MERCOSUR', 7, {0,1,2,3}, {4,5,6},        6),   # LLLL NNN
    ('CHL_NUEVO',    6, {0,1,2,3}, {4,5},          7),   # LLLL NN
    ('CHL_VIEJO',    6, {0,1},     {2,3,4,5},      8),   # LL NNNN
]

# Mapa: cuando OCR pone un DÍGITO donde debería ir una LETRA
_DIGITO_A_LETRA = {
    '0': 'O', '1': 'I', '2': 'Z', '3': 'E', '4': 'A',
    '5': 'S', '6': 'G', '7': 'T', '8': 'B',
}
# Mapa inverso: cuando OCR pone una LETRA donde debería ir un NÚMERO
_LETRA_A_DIGITO = {
    'A': '4', 'B': '8', 'E': '3', 'G': '6', 'I': '1',
    'L': '1', 'O': '0', 'S': '5', 'T': '7', 'Z': '2',
}
# Pares simétricos: 0↔O, 1↔I, 8↔B, 5↔S, 2↔Z, 3↔E, 4↔A, 6↔G, 7↔T


def _detectar_formato_patente(solo_alnum: str):
    """Detecta el formato de patente por patrón de letras/números en cada posición.
    
    En lugar de contar solo aciertos exactos (isalpha/isdigit), también cuenta
    caracteres *corregibles* vía los mapas OCR (ej: '6' donde esperamos letra
    porque 6→G es confusión típica de OCR).
    
    Retorna (nombre_formato, set_letras, set_números) o (None, None, None)
    si no reconoce el formato.
    """
    n = len(solo_alnum)
    best = (None, None, None)
    best_score = 0
    best_correctas = -1   # desempate 1: preferir menos correcciones
    best_prio = 999       # desempate 2: prioridad de formato (ARG > extranjero)

    for name, flen, letras_set, nums_set, prio in _FORMATOS_PATENTE:
        if flen != n:
            continue
        correctas = 0
        corregibles = 0
        for i, c in enumerate(solo_alnum):
            if i in letras_set:
                if c.isalpha():
                    correctas += 1
                elif c in _DIGITO_A_LETRA:   # dígito que puede ser letra
                    corregibles += 1
            elif i in nums_set:
                if c.isdigit():
                    correctas += 1
                elif c in _LETRA_A_DIGITO:   # letra que puede ser dígito
                    corregibles += 1
        score = correctas + corregibles
        mejora = (
            score > best_score
            or (score == best_score and correctas > best_correctas)
            or (score == best_score and correctas == best_correctas and prio < best_prio)
        )
        if mejora:
            best_score = score
            best = (name, letras_set, nums_set)
            best_correctas = correctas
            best_prio = prio

    # Tolerar 1 caracter que no sea ni correcto ni corregible
    if best_score >= n - 1:
        return best

    return (None, None, None)


def corregir_patente(valor: str) -> str:
    """Corrige patentes mal leídas por OCR usando la posición de cada carácter.
    
    Primero detecta el formato (ARG, BRA, CHL, PRY, URY) y aplica la corrección
    posicional SOLO para las posiciones que corresponden según ese formato.
    
    Si la patente no coincide con ningún formato conocido (extranjera no Mercosur),
    se devuelve limpia de caracteres no alfanuméricos pero sin corregir.
    """
    if not valor or not isinstance(valor, str):
        return valor

    valor = valor.upper().strip()
    solo_alnum = ''.join(c for c in valor if c.isalnum())
    if not solo_alnum:
        return valor

    fmt, letras_set, nums_set = _detectar_formato_patente(solo_alnum)
    if fmt is None:
        return solo_alnum  # formato desconocido → pasar limpio sin corregir

    chars = list(solo_alnum)
    for i, c in enumerate(chars):
        if i in letras_set and c.isdigit():
            chars[i] = _DIGITO_A_LETRA.get(c, c)
        elif i in nums_set and c.isalpha():
            chars[i] = _LETRA_A_DIGITO.get(c, c)

    return ''.join(chars)


# ---------------------------------------------------------------------------
# OCR
# ---------------------------------------------------------------------------

def _preprocess_tesseract(img):
    """Preprocesa imagen PIL para Tesseract OCR.
    
    Pipeline: gris → denoise → binarización OTSU → enderezar (deskew).
    Devuelve PIL Image en blanco y negro.
    """
    # PIL → numpy array (RGB)
    img_np = np.array(img)
    
    # 1. Escala de grises
    gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY) if len(img_np.shape) == 3 else img_np
    
    # 2. Denoise (reduce ruido de fondo, manchas)
    denoised = cv2.fastNlMeansDenoising(gray, h=12)
    
    # 3. Binarización OTSU (blanco y negro puro)
    _, binary = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # 4. Deskew (corrige inclinación del escaneo)
    coords = cv2.findNonZero(binary)
    if coords is not None:
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = 90 + angle
        if abs(angle) > 0.5:
            h, w = binary.shape
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            binary = cv2.warpAffine(binary, M, (w, h),
                                    flags=cv2.INTER_CUBIC,
                                    borderMode=cv2.BORDER_REPLICATE)
    
    return Image.fromarray(binary)


# Backward-compatible alias
_preprocess_image = _preprocess_tesseract


def _preprocess_paddle(img):
    """Preprocesa imagen PIL para PaddleOCR: mejora contraste + denoise.
    
    PaddleOCR espera imagen RGB (no binarizada). Aplica:
    1. Denoise más fuerte (h=10) para reducir ruido de escaneo
    2. CLAHE en canal Luminosidad para mejorar contraste local
    """
    if img.mode != "RGB":
        img = img.convert("RGB")
    img_np = np.array(img)

    # 1. Denoise (reduce ruido de fondo, manchas, textura de papel)
    denoised = cv2.fastNlMeansDenoisingColored(img_np, h=10, hColor=10)

    # 2. CLAHE: mejora contraste local sin alterar color (texto más legible)
    lab = cv2.cvtColor(denoised, cv2.COLOR_RGB2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=1.5, tileGridSize=(8, 8))
    l = clahe.apply(l)
    lab = cv2.merge([l, a, b])
    result = cv2.cvtColor(lab, cv2.COLOR_LAB2RGB)

    return Image.fromarray(result)


def _ocr_tesseract(img_proc):
    """Ejecuta OCR con Tesseract sobre imagen preprocesada."""
    # Try sidecar first, fallback to system installation
    sidecar = os.path.join(TESSERACT_SIDECAR, "tesseract.exe")
    if os.path.isfile(sidecar):
        pytesseract.pytesseract.tesseract_cmd = sidecar
    return pytesseract.image_to_string(img_proc, lang="spa")


def _detectar_paddle():
    """Detecta si PaddleOCR está disponible vía Python portable 3.9.

    Verifica que ./python39/python.exe exista y que pueda importar
    paddleocr correctamente (subprocess ping).

    Cachea el resultado para no repetir la verificación.

    Returns:
        str: Ruta al python.exe portable si está disponible.
        None: Si no está disponible.
    """
    global _paddle_portable
    if _paddle_portable is not None:  # Ya verificado
        return _paddle_portable if _paddle_portable else None

    portable_python = Path(PADDLE_PORTABLE_PYTHON).resolve()
    if not portable_python.is_file():
        _paddle_portable = False
        return None

    try:
        result = subprocess.run(
            [str(portable_python), "-c",
             "from paddleocr import PaddleOCR; print('ok')"],
            capture_output=True, timeout=30, text=True,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        if result.returncode == 0 and result.stdout.strip() == "ok":
            _paddle_portable = str(portable_python)
            return _paddle_portable
        else:
            print(f"[PaddleOCR] Import falló: {result.stderr.strip()}")
    except (subprocess.TimeoutExpired, OSError) as e:
        print(f"[PaddleOCR] Error detectando: {e}")

    _paddle_portable = False
    return None


def _ocr_paddle_batch(images: list) -> list:
    """Ejecuta OCR con PaddleOCR sobre múltiples imágenes en UNA sola llamada.

    Convierte todas las imágenes a PNGs temporales y las pasa al helper
    con --json. Si PaddleOCR no está disponible, fallback a Tesseract.

    Args:
        images: Lista de PIL Image (una por página del PDF).

    Returns:
        list[str]: Texto de cada página por separado, en orden.
    """
    python_path = _detectar_paddle()
    if not python_path:
        texts = []
        for img in images:
            proc = _preprocess_tesseract(img)
            texts.append(_ocr_tesseract(proc))
        return texts

    tmp_paths = []
    for img in images:
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        img.save(tmp.name, format="PNG")
        tmp.close()
        tmp_paths.append(tmp.name)

    try:
        result = subprocess.run(
            [python_path, PADDLE_OCR_HELPER] + tmp_paths
            + ["--json", "--min-conf", "0.15"],
            capture_output=True, timeout=180, text=True,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
        if result.returncode != 0:
            print(f"[PaddleOCR] Error subprocess: {result.stderr.strip()[:200]}")
            return [""] * len(images)

        data = json.loads(result.stdout)
        resultados = []
        for tmp_path in tmp_paths:
            key = os.path.basename(tmp_path)
            img_data = data.get(key, [])
            lines = []
            if isinstance(img_data, list):
                for item in img_data:
                    if isinstance(item, dict) and "text" in item:
                        lines.append(item["text"])
            resultados.append("\n".join(lines))
        return resultados

    except json.JSONDecodeError:
        return [""] * len(images)
    except subprocess.TimeoutExpired:
        print("[PaddleOCR] Timeout")
        return [""] * len(images)
    except Exception as e:
        print(f"[PaddleOCR] Error: {e}")
        return [""] * len(images)
    finally:
        for p in tmp_paths:
            try:
                os.unlink(p)
            except OSError:
                pass


def pdf_a_texto(ruta_pdf: str, engine: str = None, poppler_path: str = None) -> str:
    """Convierte el PDF a texto vía OCR, buscando solo la página con TICKET SALIDA.
    
    Args:
        ruta_pdf: Ruta al archivo PDF.
        engine: Motor OCR a usar ("tesseract" | "paddleocr").
        poppler_path: Ruta a los binarios de Poppler.

    Returns:
        str: Texto de la página que contiene "TICKET SALIDA", o "" si no se encuentra.
    """
    if poppler_path is None:
        poppler_path = POPPLER_PATH

    if engine is None:
        engine = OCR_ENGINE_DEFAULT

    dpi = 150 if engine == "paddleocr" else 300
    pages = convert_from_path(ruta_pdf, dpi=dpi, poppler_path=poppler_path)
    if not pages:
        return ""

    def _elegir_pagina_ticket(textos_por_pagina):
        """Devuelve el texto de la primera página que contiene TICKET SALIDA.
        
        Solo la página del ticket de pesaje tiene los datos completos
        (producto, pesos, patentes, conductor y DNI). El remito y otras
        páginas se descartan porque pueden tener datos parciales sin DNI.
        """
        for txt in textos_por_pagina:
            if re.search(r'TICKET\s+SALIDA', txt, re.IGNORECASE):
                return txt
        # Fallback: concatenar todo si no se encontró TICKET SALIDA
        return "\n\n".join(textos_por_pagina) if textos_por_pagina else ""

    if engine == "paddleocr":
        processed = [_preprocess_paddle(p) for p in pages]
        textos = _ocr_paddle_batch(processed)  # ahora devuelve list[str]
        texto = _elegir_pagina_ticket(textos)

        # Quality check: si devolvió muy poco texto, probar Tesseract
        if len(texto.strip()) < 50:
            fb_textos = []
            for p in pages:
                proc = _preprocess_tesseract(p)
                fb_textos.append(_ocr_tesseract(proc))
            fb = _elegir_pagina_ticket(fb_textos)
            if len(fb.strip()) > len(texto.strip()):
                print(f"[OCR] PaddleOCR poco texto, usando Tesseract")
                return fb
        return texto

    # Tesseract
    textos = []
    for p in pages:
        proc = _preprocess_tesseract(p)
        textos.append(_ocr_tesseract(proc))
    return _elegir_pagina_ticket(textos)


def pdfs_a_texto_batch(rutas_pdf: list, engine: str = None,
                       poppler_path: str = None) -> dict:
    """Procesa múltiples PDFs en un solo batch OCR (PaddleOCR).
    
    Convierte TODOS los PDFs a imágenes, las preprocesa, y las manda
    en UNA sola llamada al helper. PaddleOCR carga los modelos UNA vez
    para todos los PDFs.
    
    Args:
        rutas_pdf: Lista de rutas a archivos PDF.
        engine: Motor OCR ("tesseract" | "paddleocr").
        poppler_path: Ruta a poppler.
    
    Returns:
        dict: {nombre_archivo_sin_ext: texto_extraído}
              Los PDFs que fallan devuelven texto vacío.
    """
    if poppler_path is None:
        poppler_path = POPPLER_PATH

    # Detectar engine
    if engine is None:
        engine = OCR_ENGINE_DEFAULT

    dpi = 180 if engine == "paddleocr" else 300
    resultados = {}

    if engine == "paddleocr":
        # ── Batch PaddleOCR: convertir todos a PNGs, llamar helper UNA vez ──
        tmp_paths = []
        try:
            for ruta in rutas_pdf:
                try:
                    pages = convert_from_path(ruta, dpi=dpi, poppler_path=poppler_path)
                    for i, img in enumerate(pages):
                        processed = _preprocess_paddle(img)
                        tmp = tempfile.NamedTemporaryFile(
                            suffix=f"_{Path(ruta).stem}_p{i}.png", delete=False)
                        processed.save(tmp.name, format="PNG")
                        tmp.close()
                        tmp_paths.append(tmp.name)
                except Exception as e:
                    print(f"[Batch] Error convirtiendo {ruta}: {e}")

            if not tmp_paths:
                return {Path(p).stem: "" for p in rutas_pdf}

            python_path = _detectar_paddle()
            if not python_path:
                # Fallback a Tesseract individual
                for ruta in rutas_pdf:
                    resultados[Path(ruta).stem] = pdf_a_texto(
                        ruta, engine="tesseract", poppler_path=poppler_path)
                return resultados

            # Una sola llamada al helper con todas las imágenes
            result = subprocess.run(
                [python_path, PADDLE_OCR_HELPER] + tmp_paths
                + ["--json", "--min-conf", "0.15"],
                capture_output=True, timeout=300, text=True,
                creationflags=subprocess.CREATE_NO_WINDOW,
            )

            if result.returncode != 0:
                print(f"[Batch] Error subprocess: {result.stderr.strip()[:200]}")
                return {Path(p).stem: "" for p in rutas_pdf}

            # Parsear JSON: {"img_p0.png": [{"text":"...", "confidence":...}, ...]}
            data = json.loads(result.stdout)

            # Agrupar por PDF: separar texto por página, elegir la del ticket
            for ruta in rutas_pdf:
                stem = Path(ruta).stem
                paginas = []
                # Reconstruir per-page en orden
                for tmp_path in sorted(tmp_paths):
                    if f"_{stem}_p" in tmp_path:
                        img_data = data.get(os.path.basename(tmp_path), [])
                        lines = []
                        if isinstance(img_data, list):
                            for item in img_data:
                                if isinstance(item, dict) and "text" in item:
                                    lines.append(item["text"])
                        paginas.append("\n".join(lines))
                # Elegir la página con TICKET SALIDA (ignora remitos)
                texto = ""
                for txt in paginas:
                    if re.search(r'TICKET\s+SALIDA', txt, re.IGNORECASE):
                        texto = txt
                        break
                if not texto and paginas:
                    texto = "\n\n".join(paginas)
                resultados[stem] = texto

        except json.JSONDecodeError:
            print("[Batch] Error parseando JSON del helper")
            return {Path(p).stem: "" for p in rutas_pdf}
        except subprocess.TimeoutExpired:
            print("[Batch] Timeout")
            return {Path(p).stem: "" for p in rutas_pdf}
        except Exception as e:
            print(f"[Batch] Error: {e}")
            return {Path(p).stem: "" for p in rutas_pdf}
        finally:
            for p in tmp_paths:
                try:
                    os.unlink(p)
                except OSError:
                    pass

    else:
        # Tesseract: no hay batch real, procesar individualmente
        for ruta in rutas_pdf:
            resultados[Path(ruta).stem] = pdf_a_texto(
                ruta, engine="tesseract", poppler_path=poppler_path)

    return resultados


# ---------------------------------------------------------------------------
# Parseo
# ---------------------------------------------------------------------------

def extraer_datos(text: str) -> dict:
    lines = [l for l in text.split("\n") if l.strip()]

    def buscar(patron):
        m = re.search(patron, text, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    datos = {k: "" for k in CAMPOS}

    def _limpiar_dni(valor):
        """Corrige errores comunes de OCR en números de DNI.
        l→1, O→0, S→5, B→8, etc."""
        if not valor:
            return ""
        mapa = str.maketrans({
            'l': '1', 'I': '1', 'O': '0', 'o': '0',
            'S': '5', 's': '5', 'B': '8', 'b': '8',
            'g': '9', 'q': '9',
        })
        return valor.translate(mapa)

    def _limpiar_conductor(valor):
        """Saca la nacionalidad (ARGENTINO, BRASILERO, etc.) y separador,
        dejando solo el nombre. Tolera guión, =, coma como separador."""
        if not valor:
            return ""
        valor = re.sub(r'^(?:ARGENTINO|BRASILERO|CHILENO|URUGUAYO|PARAGUAYO|BOLIVIANO|PERUANO|ECUATORIANO)\s*[-=,]\s*', '', valor, flags=re.IGNORECASE).strip()
        # Corregir DNI dentro del texto
        valor = re.sub(r'DNI:\s*(\S+)',
                       lambda m: f'DNI: {_limpiar_dni(m.group(1))}',
                       valor)
        return valor

    def _extraer_conductor_robusto(text):
        """Extrae (conductor, dni) del texto OCR usando múltiples estrategias en cascada.
        
        E1 — Directo: captura 'NOMBRE APELLIDO-DNI:12345678' sin depender de labels.
             No importa el orden de líneas que devuelva PaddleOCR.
        E2 — Label Nacionalidad: 'Conductor y Nacionalidad ARGENTINO - NOMBRE'
             (regex actual mejorado, tolera orden invertido).
        E3 — Label simple: 'Nombre y Apellido del Conductor: NOMBRE'
        E4 — Label escueto: 'Conductor: NOMBRE'
        
        Returns: (conductor: str, dni: str)
        """
        def _es_nombre_valido(s):
            return len(s) >= 3 and any(c.isalpha() for c in s)

        def _extraer_dni_del_texto(t):
            """Busca DNI en cualquier parte del texto."""
            m = re.search(r'DNI[:\s]*(\d{7,8})', t, re.IGNORECASE)
            if m:
                return _limpiar_dni(m.group(1))
            # DNI suelto de 7-8 dígitos cerca del nombre
            m = re.search(r'\b(\d{7,8})\b', t)
            if m:
                return _limpiar_dni(m.group(1))
            return ""

        # ── E1: Directo NOMBRE-DNI:12345678 ──
        m1 = re.search(
            r'([A-ZÁÉÍÓÚÑÜ][A-ZÁÉÍÓÚÑÜ\s]{2,20})[\s-]*DNI[:\s]*(\d{7,8})',
            text, re.IGNORECASE
        )
        if m1:
            nombre = m1.group(1).strip().rstrip('-')
            dni = _limpiar_dni(m1.group(2))
            if _es_nombre_valido(nombre) and dni:
                return (nombre, dni)

        # ── E2: "Conductor y Nacionalidad ARGENTINO - NOMBRE" ──
        # Busca en la línea actual + posible continuación (OCR corta líneas)
        m2 = re.search(
            r"Conduc\w*r y Nacionalidad[:\s]+(.+(?:\n[^A-Z\n][^\n]*)?)",
            text, re.IGNORECASE
        )
        if m2:
            raw = _limpiar_conductor(m2.group(1).strip())
            if raw:
                # Separar nombre del DNI si viene pegado
                dni = _extraer_dni_del_texto(raw)
                nombre = re.sub(r'[\s-]+DNI[\s:]*\d{7,8}', '', raw, flags=re.IGNORECASE).strip()
                if _es_nombre_valido(nombre):
                    return (nombre, dni)
                # Fallback: el raw completo si no se pudo separar
                return (raw, dni)

        # ── E3: "Nombre y Apellido del Conductor: NOMBRE" ──
        m3 = re.search(r"Nombre\s+y\s+Apellido\s+del\s+Conductor[:\s]+(.+)", text, re.IGNORECASE)
        if m3:
            raw = m3.group(1).strip()
            dni = _extraer_dni_del_texto(raw)
            nombre = re.sub(r'[\s-]+DNI[\s:]*\d{7,8}', '', raw, flags=re.IGNORECASE).strip()
            if _es_nombre_valido(nombre):
                return (nombre, dni)

        # ── E4: "Conductor: NOMBRE" simple ──
        # Busca "Conductor:" y toma el nombre de la línea actual o siguiente
        m4 = re.search(r"Conductor[:\s]+(.+?)(?:\n|$)", text, re.IGNORECASE)
        if m4:
            raw = m4.group(1).strip()
            dni = _extraer_dni_del_texto(raw)
            nombre = re.sub(r'[\s-]+DNI[\s:]*\d{7,8}', '', raw, flags=re.IGNORECASE).strip()
            if _es_nombre_valido(nombre):
                return (nombre, dni)

        # Fallback total: devolver lo que se encontró como Conductor (vacío si nada)
        return ("", "")

    def _extraer_permiso(valor):
        """Del texto 'PRODUCTO - PERMISO' se queda solo con el permiso
        (último grupo alfanumérico)."""
        if not valor:
            return ""
        # Buscar el último bloque de letras+números (permiso)
        m = re.search(r'([A-Z0-9]{5,})$', valor, re.IGNORECASE)
        return m.group(1).upper() if m else valor.strip()

    # Campos simples con regex
    datos["Ticket Salida"]      = buscar(r"TICKET SALIDA[:\s]+([\w.]+)")
    datos["Turno"]               = (m.group(1) if (m := re.search(r"(?:^|Turno[:\s]+)(\d{5,8})\s*(?:LL|Llegada|Salida)", text, re.MULTILINE | re.IGNORECASE)) else "")
    # Llegada: busca después de "LLegada:" o fecha antes de "Salida:"
    llegada_m = re.search(r"(?:LL|Ll)egada[:\s]*\n*([\d/]+[\s]*[\d.]*)", text, re.IGNORECASE)
    if not llegada_m or not llegada_m.group(1).strip():
        # Fallback: fecha+hora justo antes de "Salida:" (sin el label Llegada)
        llegada_m = re.search(r"(\d{1,2}/\d{1,2}/\d{4}[\s.]*\d{2}\.\d{2})\s*\n\s*Salida", text, re.IGNORECASE)
    datos["Llegada"] = llegada_m.group(1).strip() if llegada_m and llegada_m.group(1).strip() else ""
    # Salida: espacio opcional entre fecha y hora (OCR junta "5/06/202610.58")
    # Salida hora: busca "Salida:" que no sea parte de "TICKET SALIDA:"
    salida_m = re.search(r"(?:^|\n)\s*Salida[:\s]+([\d/]+[\s]*[\d.]+)", text, re.IGNORECASE | re.MULTILINE)
    datos["Salida hora"] = salida_m.group(1).strip() if salida_m else ""
    datos["Transportista"]       = buscar(r"Transportista[:\s]+(.+)")
    # CUIT Transportista: tolera nombre entre "Transportista:" y "C.U.I.T.." 
    datos["CUIT Transportista"]  = buscar(r"Transportista[^\n]*\n[^\n]*\n\s*C[.\s]*[U0][.\s]*I[.\s]*T[.\s]*[:.\s]*(\S+)")
    datos["Clase"]               = buscar(r"Clase\s*:\s*(CARTA DE PORTE)")
    datos["N° Carta de Porte"]   = buscar(r"Numero[:\s]+([\d/]+)")
    # Producto: primero línea DESPUÉS del label, si es patente busca ANTES
    prod_m = re.search(r"Producto Transportado\s*\n(.+)", text, re.IGNORECASE)
    if prod_m:
        val = prod_m.group(1).strip()
        solo = ''.join(c for c in val if c.isalnum())
        if len(solo) in (6, 7) and solo.isalnum():
            # Parece patente, probar línea anterior
            prod_m = re.search(r"(.+)\n\s*Producto Transportado", text, re.IGNORECASE)
            val = prod_m.group(1).strip() if prod_m else val
        datos["Producto"] = val
    else:
        datos["Producto"] = buscar(r"Producto Transportado[:\s]+(.+)")
    # Patentes: extraer con tolerancia a formato y validar resultado
    def _valida_patente(v):
        if not v:
            return ""
        solo = ''.join(c for c in v if c.isalnum())
        if len(solo) in (6, 7) and solo.isalnum():
            return corregir_patente(solo)
        return ""

    def _extraer_patentes_contextual():
        """Busca TODAS las patentes en el texto en orden de aparición:
        primera → Camion, segunda → Acoplado.
        Acepta todos los formatos Mercosur:
        - ARG: AB123CD, ABC123
        - BRA: ABC1D23, ABC1234
        - CHL: HFBH63, PXDF80, AA1234
        - PRY: ABCD123, ABC123
        - URY: ABC1234
        """
        pats = re.findall(r'\b([A-Z]{2,4}\d{1,4}[A-Z]{0,2}\d{0,2})\b', text, re.IGNORECASE)
        validas = [p.upper() for p in pats if len(p) in (6, 7)]
        return validas[:2] if validas else ["", ""]

    pat_camion  = _valida_patente(buscar(r"patente del Camion[\.:\s]+(\S+)"))
    pat_acoplado = _valida_patente(buscar(r"paten[td]e del Acoplado[\.:\s]+\s*(\S+)"))
    # Si no encontró acoplado por etiqueta, buscar la línea después del camión
    # (el ticket tiene el acoplado entre "Patente del Camion" y "Patente del Acoplado")
    if pat_camion and not pat_acoplado:
        m_linea = re.search(
            r"patente del Camion[\.:\s]+(\S+)\s*\n\s*(\S+)",
            text, re.IGNORECASE
        )
        if m_linea:
            cam = _valida_patente(m_linea.group(1))
            acoplado = _valida_patente(m_linea.group(2))
            if cam and acoplado and cam != acoplado:
                pat_camion = cam
                pat_acoplado = acoplado

    # Si las etiquetas nos dan dos patentes distintas y válidas, perfecto
    if pat_camion and pat_acoplado and pat_camion != pat_acoplado:
        datos["Patente Camion"] = pat_camion
        datos["Patente Acoplado"] = pat_acoplado
    else:
        # Fallback contextual: orden de aparición en el texto
        ctx = _extraer_patentes_contextual()
        datos["Patente Camion"] = corregir_patente(ctx[0]) if ctx[0] else pat_camion
        datos["Patente Acoplado"] = corregir_patente(ctx[1]) if len(ctx) > 1 and ctx[1] else pat_acoplado
    datos["LOT"]                 = buscar(r"LOT[:\s]*(\d+)")
    # ATA: puede estar en la misma línea o en la siguiente
    ata_m = re.search(r"Nombre-CUIT ATA[:\s]+\n*(.+)", text, re.IGNORECASE)
    datos["ATA (Nombre-CUIT)"] = ata_m.group(1).strip() if ata_m else ""
    # Conductor: multi-estrategia robusta contra orden invertido de PaddleOCR
    conductor_nombre, conductor_dni = _extraer_conductor_robusto(text)
    datos["Conductor"] = conductor_nombre
    datos["DNI Conductor"] = conductor_dni
    datos["Merc./Permiso"]       = _extraer_permiso(buscar(r"Merc[,.]?\s*y\s*Permiso\s*Emb[ae]r[cgq]ue?[:\s]+(.+)"))
    datos["Pallet"]              = buscar(r"PALLET[:\s]+(\w+)")

    # Número de contenedor ISO/FLEXI: "Sigla Contenedor: MSMU 258531-2"
    datos["Contenedor"] = buscar(r"Sigla Contenedor[:\s]+([A-Z]+\s*\d+[\s-]+\d+)")
    # Tara Contenedor: para TODOS los tickets.
    # Orden: 1) valor cerca de Cert.Verif (más confiable), 2) inline tras label
    datos["Tara Contenedor"] = ""
    m_tc = re.search(r'(\d+[.,]\d{3})\s+Cert\.?Verif\.?INTI[\s-]*Balanza\s+E[gq]r', text, re.IGNORECASE)
    if not m_tc:
        m_tc = re.search(r'(\d+[.,]\d{3})\s+Cert\.?Verif\.?INTI[\s-]*Balanza\s+Ing', text, re.IGNORECASE)
    if not m_tc:
        m_tc = re.search(r'Tara Contenedor[:\s]+(\d+[.,]\d+)', text, re.IGNORECASE)
    if m_tc:
        datos["Tara Contenedor"] = m_tc.group(1)

    # Fecha: primer match dd/mm/yyyy en el texto
    fecha_m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    datos["Fecha"] = fecha_m.group(1) if fecha_m else ""

    # Emisor / Remitente / Destinatario:
    def _cuit_de_linea(linea):
        """Extrae CUIT de una línea tolerando variantes de OCR.
        
        Formatos soportados:
        - "C.U.I.T.: 3050287435/3"
        - "C.U.I.T .. 3050287435/3"
        - "C.0.I.T.:30.50287435.3"  (OCR lee 0 por U)
        - "3050287435/3"             (solo número)
        - "20-10980715-"             (con guiones)
        """
        if not linea:
            return ""
        # 1. Con etiqueta CUIT: flexible con dots, espacios, OCR 0/U
        m = re.search(r'C[.\s]*[U0][.\s]*I[.\s]*T[.\s]*[:.\s]*(\S+)', linea, re.IGNORECASE)
        if m:
            cuit = m.group(1).rstrip(".:,-")
            if cuit:
                return cuit
        # 2. Solo número: xx-xxxxxxxx/x o xxxxxxxxx/x
        m = re.search(r'(\d{2,3}[-.\s]?\d{6,10}[-/]\d[\d/-]*)', linea)
        if m:
            return m.group(1).strip().rstrip("-.:")
        return ""

    def _extraer_con_cuit(lines, label):
        """Busca línea con label 'Emisor/Remitente/Destinatario' y extrae nombre + CUIT.
        
        Soporta TRES formatos (en orden de prioridad):
        1. Inline:   "Emisor ...: NOMBRE" + CUIT en línea siguiente
        2. Nextline: "Emisor ." / "Emisor:" + NOMBRE en i+1 + CUIT en i+2..i+4
        3. Legacy:   NOMBRE en i-2, CUIT en i-1, label en i (formato anterior)
        """
        def _es_nombre_valido(s):
            """Un nombre válido tiene al menos 3 chars y contiene letras."""
            return s and len(s) >= 3 and any(c.isalpha() for c in s)

        for i, line in enumerate(lines):
            clean = line.strip().rstrip(":")
            if clean.upper() == label.upper() or clean.upper().startswith(label.upper()):
                # 1. Intento inline: "Emisor ...: NOMBRE" (con o sin colon)
                inline = re.search(rf"{re.escape(label)}[\s\.]*:?\s*(.+)", line, re.IGNORECASE)
                inline_val = inline.group(1).strip() if inline else ""

                if _es_nombre_valido(inline_val):
                    nombre = inline_val
                    cuit = ""
                    for j in range(i + 1, min(i + 4, len(lines))):
                        c = _cuit_de_linea(lines[j])
                        if c:
                            cuit = c
                            break
                elif i + 2 < len(lines):
                    # 2. Label sola → nombre en i+1, CUIT en i+2 o i+3
                    nombre = lines[i + 1].strip()
                    if not _es_nombre_valido(nombre):
                        # Si i+1 no es nombre válido, probar i+2
                        nombre = lines[i + 2].strip() if i + 2 < len(lines) else ""
                    cuit = ""
                    for j in range(i + 2, min(i + 5, len(lines))):
                        c = _cuit_de_linea(lines[j])
                        if c and c != nombre:
                            cuit = c
                            break
                elif i >= 2:
                    # 3. Formato legacy: nombre en i-2, CUIT en i-1
                    nombre = lines[i - 2]
                    cuit = _cuit_de_linea(lines[i - 1])
                else:
                    nombre, cuit = "", ""
                return nombre, cuit
        return "", ""

    datos["Emisor"],      datos["CUIT Emisor"]      = _extraer_con_cuit(lines, "Emisor")
    datos["Remitente"],   datos["CUIT Remitente"]   = _extraer_con_cuit(lines, "Remitente")
    datos["Destinatario"], datos["CUIT Destinatario"] = _extraer_con_cuit(lines, "Destinatario")

    # Pesos: intentar extracción con etiqueta primero (más confiable)
    # Formatos: "BRUTO: 44.900", "Peso total BRUTO: 44.560", "PSTARA: 16.980"
    def _valida_peso(v):
        """Valida que el valor sea un peso real (no un número de referencia)."""
        if not v:
            return ""
        v = v.strip().replace(" ", "")
        # Formatos válidos: 44.900, 16.98, 27.940 (decimal) o 44900, 27940 (entero 5+ dígitos)
        if re.match(r'^\d{2,3}[.,]\d{2,4}$', v):
            return v
        if re.match(r'^\d{5,6}$', v):
            return v
        return ""

    for etiqueta, campo in [("BRUTO", "Peso Bruto (kg)"),
                            ("TARA", "Peso Tara (kg)"),
                            ("NETO", "Peso Neto (kg)")]:
        datos[campo] = _valida_peso(buscar(r"(?:Peso (?:total )?)?(?:PS)?%s[:\s]+([\d.]+)" % etiqueta))

    def _peso_a_int(s):
        return int(s.replace(".", ""))

    def _int_a_peso(n):
        return f"{n:,}".replace(",", ".")

    # Si NETO no se encontró pero tenemos Bruto y Tara, calcularlo
    if not datos["Peso Neto (kg)"] and datos["Peso Bruto (kg)"] and datos["Peso Tara (kg)"]:
        try:
            neto = _peso_a_int(datos["Peso Bruto (kg)"]) - _peso_a_int(datos["Peso Tara (kg)"])
            datos["Peso Neto (kg)"] = _int_a_peso(neto)
        except ValueError:
            pass

    # Fallback solo para campos que siguen vacíos
    if not all([datos["Peso Bruto (kg)"], datos["Peso Tara (kg)"], datos["Peso Neto (kg)"]]):
        nums = re.findall(r"^(\d{2,3}\.\d{3})$", text, re.MULTILINE)
        if not datos["Peso Bruto (kg)"] and len(nums) >= 1:
            datos["Peso Bruto (kg)"] = nums[0]
        if not datos["Peso Tara (kg)"] and len(nums) >= 2:
            datos["Peso Tara (kg)"] = nums[1]
        if not datos["Peso Neto (kg)"] and len(nums) >= 3:
            datos["Peso Neto (kg)"] = nums[2]

    return datos


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _estilos():
    thin = Side(border_style="thin", color="BFBFBF")
    return {
        "border":      Border(left=thin, right=thin, top=thin, bottom=thin),
        "font_data":   Font(name="Arial", size=10),
        "font_header": Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "font_title":  Font(name="Arial", bold=True, size=12, color="FFFFFF"),
        "fill_title":  PatternFill("solid", start_color="1F4E79"),
        "fill_header": PatternFill("solid", start_color="2E75B6"),
        "fill_alt":    PatternFill("solid", start_color="DEEAF1"),
        "align_center": Alignment(horizontal="center", vertical="center"),
        "align_data":   Alignment(vertical="center", wrap_text=True),
    }


def crear_excel(lista_datos: list[dict], output_path: str):
    """
    Crea un Excel con una fila por ticket.
    lista_datos: lista de dicts, uno por PDF procesado.
    """
    s = _estilos()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Título
    ws.merge_cells(f"A1:{chr(64 + len(CAMPOS))}1")
    ws["A1"] = "TICKETS DE PESAJE — ACEITERA GENERAL DEHEZA S.A."
    ws["A1"].font = s["font_title"]
    ws["A1"].fill = s["fill_title"]
    ws["A1"].alignment = s["align_center"]
    ws.row_dimensions[1].height = 30

    # Encabezados
    for col, campo in enumerate(CAMPOS, start=1):
        cell = ws.cell(row=2, column=col, value=campo)
        cell.font = s["font_header"]
        cell.fill = s["fill_header"]
        cell.alignment = s["align_center"]
        cell.border = s["border"]
    ws.row_dimensions[2].height = 20

    # Filas de datos
    for fila_idx, datos in enumerate(lista_datos, start=3):
        fill = s["fill_alt"] if fila_idx % 2 == 0 else PatternFill()
        for col, campo in enumerate(CAMPOS, start=1):
            cell = ws.cell(row=fila_idx, column=col, value=datos.get(campo, ""))
            cell.font = s["font_data"]
            cell.alignment = s["align_data"]
            cell.border = s["border"]
            cell.fill = fill
        ws.row_dimensions[fila_idx].height = 18

    # Ancho de columnas
    anchos = {
        "Ticket Salida": 14, "Fecha": 12, "Turno": 10, "Llegada": 18,
        "Salida hora": 18, "Emisor": 28, "CUIT Emisor": 16,
        "Remitente": 28, "CUIT Remitente": 16, "Transportista": 30,
        "CUIT Transportista": 18, "Destinatario": 28, "CUIT Destinatario": 16,
        "Clase": 16, "N° Carta de Porte": 18, "Producto": 18,
        "Patente Camion": 14, "Patente Acoplado": 16,
        "Peso Bruto (kg)": 14, "Peso Tara (kg)": 13, "Peso Neto (kg)": 13,
        "LOT": 10, "ATA (Nombre-CUIT)": 22,         "Conductor": 40,
        "DNI Conductor": 16,
        "Merc./Permiso": 36, "Pallet": 8,
        "Contenedor": 20, "Tara Contenedor": 16,
    }
    for col, campo in enumerate(CAMPOS, start=1):
        ws.column_dimensions[chr(64 + col)].width = anchos.get(campo, 15)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Coordinación ISO/FLEXI
# ---------------------------------------------------------------------------

# Equivalencias de puertos en distinto idioma (singapore ↔ singapur, etc.)
EQUIV_PUERTOS = {
    "singapore": "singapur",
    "bangkok": "bangkok",
    "philadelphia": "filadelfia",
    "rotterdam": "roterdam",
    "new york": "nueva york",
    "los angeles": "los angeles",
    "thailand": "tailandia",
}

def _normalizar_ciudad(txt):
    """Normaliza nombre de ciudad: lower, sin acentos, sin espacios extra."""
    if not txt:
        return ""
    txt = str(txt).lower().strip()
    txt = unicodedata.normalize("NFKD", txt)
    txt = txt.encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()

def _comparar_ciudades(a, b):
    """Compara dos nombres de ciudad usando normalización + tabla EQUIV_PUERTOS."""
    if not a and not b:
        return True
    if not a or not b:
        return False
    na = EQUIV_PUERTOS.get(_normalizar_ciudad(a), _normalizar_ciudad(a))
    nb = EQUIV_PUERTOS.get(_normalizar_ciudad(b), _normalizar_ciudad(b))
    return na == nb


def _comparar_texto(a, b):
    """Compara dos textos: upper, sin puntuación, espacios normalizados,
       NFKD para acentos, y limpia caracteres invisibles (NBSP, etc.)."""
    if not a and not b:
        return True
    if not a or not b:
        return False

    def norm(t):
        t = t.upper().strip()
        t = unicodedata.normalize("NFKD", t)   # descompone acentos
        t = re.sub(r"[^\w\s]", "", t)          # saca puntuación
        t = t.replace("\u00A0", " ")           # NBSP → espacio normal
        t = re.sub(r"\s+", " ", t)             # colapsa espacios
        return t.strip()

    na = norm(a)
    nb = norm(b)

    # Debug: si normalizan igual pero el strip original difiere → carácter invisible
    if na == nb and a.strip().upper() != b.strip().upper():
        print(f"[DEBUG _comparar_texto] a={repr(a)}, b={repr(b)}")

    return na == nb

def extraer_coordinacion(pdf_path):
    """
    Extrae campos de un PDF de Coordinación digital (NO escaneado).
    Retorna dict con: contenedor, giro, carpeta, cliente, destino,
                      buque, viaje, booking, pto_descarga, pto_final.
    """
    import fitz
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    # Si no se extrajo texto → el PDF es escaneado (imagen)
    if not text.strip():
        return {"_error": "El PDF no tiene texto seleccionable (parece escaneado)"}

    def buscar(patron, flags=re.IGNORECASE | re.DOTALL):
        m = re.search(patron, text, flags)
        return m.group(1).strip() if m else ""

    # Buque y Viaje: la línea "Vapor\nCOSCO SHIPPING DANUBE 047E"
    buque_completo = buscar(r"Vapor\s*\n\s*(.+?)(?:\n|$)")
    buque = ""
    viaje = ""
    if buque_completo:
        partes = buque_completo.strip().rsplit(None, 1)
        buque = partes[0] if len(partes) > 1 else buque_completo
        viaje = partes[1] if len(partes) > 1 else ""

    # Puerto Final lo sacamos de Observaciones (ej: "Destino: Bangkok")
    texto_obs = buscar(r"Observaciones\s*\n\s*(.+?)(?:$)")
    pto_final = ""
    if texto_obs:
        m_obs = re.search(r"Destino[:\s]+(.+?)$", texto_obs, re.IGNORECASE | re.MULTILINE)
        if m_obs:
            pto_final = m_obs.group(1).strip()

    # Contenedor: buscar por patrón de código (4 letras + 7 dígitos)
    m_contenedor = re.search(r"\b([A-Z]{4}\d{7})\b", text)
    contenedor = m_contenedor.group(1) if m_contenedor else ""

    # Puerto de descarga: extraer solo el nombre del puerto
    descarga = buscar(r"Pto\.?\s*de\s*transbordo[:\s]*\n?\s*([A-Za-z]+?)(?=Proveedor|Tank|\s|$)")

    # Cliente: puede no tener valor (siguiente línea es otra etiqueta, no un nombre)
    _cliente_raw = buscar(r"Cliente\s*\n\s*(.+?)(?:\n|$)")
    _etq_conocidas = {"destino", "vapor", "bandera", "puerto", "giro", "carpeta",
                      "pedido", "sector", "calidad", "exportaciones", "chabas",
                      "nro.", "surveyor", "agencia", "booking"}
    if _cliente_raw.strip().lower() in _etq_conocidas:
        cliente = ""
    else:
        cliente = _cliente_raw

    resultado = {
        "contenedor":   contenedor,
        "giro":         buscar(r"Giro[:\s]*\n?\s*(\S+)"),
        "carpeta":      buscar(r"Carpeta[:\s]*\n?\s*(\d+)"),
        "cliente":      cliente,
        "destino":      buscar(r"Destino\s*\n\s*(.+?)(?:\n|$)"),
        "buque":        buque,
        "viaje":        viaje,
        "booking":      buscar(r"Booking[:\s]*\n?\s*(\S+)"),
        "pto_descarga": descarga,
        "pto_final":    pto_final,
        "peso_flexi":   buscar(r"Peso del Flexi[:\s]*(\d+)"),
    }
    resultado["_archivo"] = os.path.basename(pdf_path)
    return resultado


def extraer_fecha_permiso(pdf_path):
    """
    Extrae la fecha de oficialización de un permiso de exportación escaneado
    (ej: 26069EC01000530K.pdf). OCR Tesseract en la primera página.
    Retorna 'YYYY-MM-DD' o '' si no se encuentra.
    """
    import fitz
    from PIL import Image

    doc = fitz.open(pdf_path)
    if len(doc) == 0:
        doc.close()
        return ""

    page = doc[0]
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()

    texto = _ocr_tesseract(img)

    # Buscar "Oficialización" + fecha dd/mm/aaaa
    m = re.search(r'Oficializaci[oó]n[^0-9]*(\d{2})/(\d{2})/(\d{4})',
                  texto, re.IGNORECASE)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

    # Fallback: buscar "OFICIALIZADO" + fecha
    m2 = re.search(r'OFICIALIZADO[^0-9]*(\d{2})/(\d{2})/(\d{4})', texto)
    if m2:
        return f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"

    return ""


def _es_xlrd(ws):
    """Detecta si ws es una hoja xlrd (vs openpyxl)."""
    return hasattr(ws, 'nrows')


def _xlrd_datemode(ws):
    """Retorna el datemode del workbook xlrd."""
    try:
        return ws.book.datemode
    except Exception:
        return 0


def _iter_celdas_con_idx(ws):
    """Itera sobre (row_idx, col_idx, valor) de una hoja (xlrd o openpyxl)."""
    if _es_xlrd(ws):
        import xlrd
        datemode = _xlrd_datemode(ws)
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                if cell.ctype == xlrd.XL_CELL_DATE:
                    date_tuple = xlrd.xldate_as_tuple(cell.value, datemode)
                    val = f"{date_tuple[0]:04d}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    v = cell.value
                    val = str(int(v)) if v == int(v) else str(v)
                elif cell.ctype == xlrd.XL_CELL_TEXT:
                    val = cell.value
                else:
                    continue
                yield row_idx, col_idx, val.strip()
    else:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    yield cell.row - 1, cell.column - 1, str(cell.value).strip()


def _obtener_valor(ws, row_idx, col_idx):
    """Obtiene valor de celda por índices 0-based (xlrd u openpyxl)."""
    if _es_xlrd(ws):
        if row_idx < ws.nrows and col_idx < ws.ncols:
            import xlrd
            cell = ws.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate_as_tuple(cell.value, _xlrd_datemode(ws))
                return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d}"
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                return str(int(v)) if v == int(v) else str(v)
            if cell.ctype == xlrd.XL_CELL_TEXT:
                return cell.value.strip()
            return ""
        return ""
    else:
        c = ws.cell(row=row_idx + 1, column=col_idx + 1)
        return str(c.value).strip() if c.value is not None else ""


def _buscar_valor_choferes(ws, etiqueta):
    """Busca etiqueta en cualquier celda (label-value) y devuelve valor a la derecha."""
    etq_up = re.sub(r"\s+", " ", etiqueta.upper().strip())
    for row_idx, col_idx, val in _iter_celdas_con_idx(ws):
        limpio = re.sub(r"\s+", " ", val.upper())
        if etq_up in limpio:
            return _obtener_valor(ws, row_idx, col_idx + 1)
    return ""


def _buscar_valor_tabular(ws, etiqueta):
    """Busca header en cualquier fila y devuelve valor en la fila siguiente (misma columna)."""
    etq_up = re.sub(r"\s+", " ", etiqueta.upper().strip())
    for row_idx, col_idx, val in _iter_celdas_con_idx(ws):
        limpio = re.sub(r"\s+", " ", val.upper())
        if etq_up in limpio:
            return _obtener_valor(ws, row_idx + 1, col_idx)
    return ""


def leer_choferes_coordinacion(excel_path):
    """
    Lee hoja Choferes de CONTENEDORES Excel y devuelve dict con campos
    clave para comparar contra Coordinación.
    """
    is_xls = excel_path.lower().endswith('.xls')
    ws = None
    if is_xls:
        import xlrd
        wb = xlrd.open_workbook(excel_path)
        for name in wb.sheet_names():
            if "chofer" in name.lower():
                ws = wb.sheet_by_name(name)
                break
    else:
        wb = load_workbook(excel_path, data_only=True)
        for name in wb.sheetnames:
            if "chofer" in name.lower():
                ws = wb[name]
                break
        wb.close()
    if ws is None:
        return {"_error": f"No se encontró hoja 'Choferes' en {os.path.basename(excel_path)}"}

    def _normalizar_fecha(val):
        """Convierte fecha Excel (YYYY-MM-DD HH:MM:SS) a YYYY-MM-DD."""
        if not val:
            return ""
        s = str(val).strip()
        # Ya está en formato ISO
        m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
        if m:
            return m.group(1)
        # DD/MM/YYYY
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
        return s  # no se pudo parsear, devolver raw

    return {
        "contenedor":   _buscar_valor_tabular(ws, "NUMERO DEL CONTENEDOR"),
        "giro":         _buscar_valor_choferes(ws, "PUERTO SALIDA"),
        "carpeta":      _buscar_valor_choferes(ws, "CARPETA"),
        "cliente":      _buscar_valor_choferes(ws, "DESTINATARIO"),
        "destino":      _buscar_valor_choferes(ws, "DESTINO"),
        "buque":        _buscar_valor_choferes(ws, "BUQUE"),
        "viaje":        _buscar_valor_choferes(ws, "VIAJE"),
        "booking":      _buscar_valor_choferes(ws, "BOOKING"),
        "pto_descarga": _buscar_valor_choferes(ws, "PUERTO DESCARGA"),
        "pto_final":    _buscar_valor_choferes(ws, "PUERTO FINAL"),
        "fecha_of_pe":  _normalizar_fecha(
                            _buscar_valor_choferes(ws, "FECHA OF PE")),
        "fecha_carga":  _normalizar_fecha(
                            _buscar_valor_choferes(ws, "FECHA CARGA")),
        "peso_flexi":   _buscar_valor_choferes(ws, "PESO FLEXI"),
    }


def comparar_coordinacion(datos_pdf, datos_excel):
    """
    Compara los datos extraídos del PDF contra los del Excel.
    Retorna: (comparaciones: dict, estado: str)
    - comparaciones: {campo: {"pdf": ..., "excel": ..., "match": bool}}
    - estado: "ok" / "mismatch" / "error"
    """
    if "_error" in datos_pdf:
        return {}, "error"
    if "_error" in datos_excel:
        return {}, "error"

    from datetime import date

    campos = [
        "giro", "carpeta", "cliente", "destino",
        "buque", "viaje", "booking", "pto_descarga", "pto_final",
        "fecha_of_pe", "fecha_carga", "peso_flexi",
    ]
    # Ciudades/puertos usan _comparar_ciudades con tabla EQUIV_PUERTOS; texto general usa _comparar_texto
    ciudades = {"giro", "pto_descarga", "pto_final", "destino"}
    texto_general = {"cliente", "buque", "carpeta"}
    fechas = {"fecha_of_pe", "fecha_carga"}

    comparaciones = {}
    errores = 0
    for campo in campos:
        v_pdf = (datos_pdf.get(campo) or "").strip()
        v_xls = (datos_excel.get(campo) or "").strip()

        if campo == "fecha_carga":
            # fecha_carga: PDF side es la fecha de hoy
            v_pdf = date.today().strftime("%Y-%m-%d")
        elif campo in fechas:
            # fecha_of_pe: PDF viene de extraer_fecha_permiso()
            pass  # v_pdf ya está seteado

        if campo in ciudades:
            match = _comparar_ciudades(v_pdf, v_xls)
        elif campo in texto_general:
            match = _comparar_texto(v_pdf, v_xls)
        elif campo in fechas:
            # Normalizar fechas a YYYY-MM-DD antes de comparar
            from datetime import datetime as _dt
            v_pdf_norm, v_xls_norm = v_pdf, v_xls
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y/%m/%d"):
                try:
                    v_pdf_norm = _dt.strptime(v_pdf, fmt).strftime("%Y-%m-%d")
                except (ValueError, AttributeError):
                    pass
                try:
                    v_xls_norm = _dt.strptime(v_xls, fmt).strftime("%Y-%m-%d")
                except (ValueError, AttributeError):
                    pass
            match = v_pdf_norm.upper() == v_xls_norm.upper()
        else:
            # contenedor, viaje, booking, fechas — comparación exacta
            match = v_pdf.upper() == v_xls.upper()

        # Si es campo obligatorio (todos menos contenedor) y PDF no trajo valor → no match
        if not v_pdf and campo != "contenedor":
            match = False

        # Peso Flexi: ambos vacíos o PDF vacío + Excel "0" = ISO, no marcar diferencia
        if campo == "peso_flexi":
            if not v_pdf and not v_xls:
                match = True
            elif not v_pdf and v_xls == "0":
                match = True
            else:
                # Comparación numérica exacta
                match = v_pdf.upper() == v_xls.upper()

        comparaciones[campo] = {
            "pdf": v_pdf,
            "excel": v_xls,
            "match": match,
        }
        if not match:
            errores += 1

    if errores == 0:
        estado = "ok"
    elif not any(c["match"] for c in comparaciones.values()):
        estado = "error"
    else:
        estado = "mismatch"

    return comparaciones, estado


# ---------------------------------------------------------------------------
# API Visión (OpenRouter / modelos vision-compatibles)
# ---------------------------------------------------------------------------

OPENROUTER_BASE = "https://openrouter.ai/api/v1"

MODELOS_VISION = [
    "google/gemini-2.0-flash-exp:free",
    "meta-llama/llama-3.2-11b-vision:free",
    "google/gemini-2.5-flash",
    "qwen/qwen3.6-flash",
]

MODELO_VISION_DEFAULT = MODELOS_VISION[0]


def cuil_a_dni(valor: str) -> str:
    """Convierte CUIL (11 dígitos, empieza 20/27) a DNI sacando prefijo y dígito verificador."""
    val = valor.strip()
    if len(val) == 11 and val.isdigit() and val[:2] in ("20", "27"):
        return val[2:-1]
    return val


def extraer_salida_aduana(pdf_path: str, modo: str = "flexi") -> dict:
    """
    Extrae datos de un PDF de Salida de Zona Primaria Aduanera (getjobid*.pdf o PLT*).
    Estos PDFs contienen texto (no son escaneados).

    Args:
        pdf_path: Ruta al PDF.
        modo: "flexi" (ISO/Flexi con CUIL) o "terrestre" (carga a granel con DNI).

    Retorna dict con:
        plt, contenedor, patente_camion, patente_semi, conductor,
        cuil/dni, peso_bruto, id_destinacion, exportador
    """
    import fitz
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    if not text.strip():
        return {"_error": "El PDF de aduana no tiene texto"}

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    data = {}

    # PLT: PLT seguido de dígitos
    m = re.search(r"(PLT\d+)", text)
    data["plt"] = m.group(1) if m else ""

    # Número de contenedor: 4 letras + 7 dígitos (solo flexi)
    if modo == "flexi":
        m = re.search(r"\b([A-Z]{4}\d{7})\b", text)
        data["contenedor"] = m.group(1) if m else ""
    else:
        data["contenedor"] = ""

    # Peso bruto:
    #   flexi: ej 23,525.000
    #   terrestre: ej 27,890  o  27890  (después del Doc. Transporte/Manifiesto)
    m = re.search(r"([\d,]+\.\d{3})", text)  # formato flexi
    if not m:
        # terrestre: número después de referencia tipo 069AR00902026
        m = re.search(r"\b(069[A-Z0-9]{8,})\s*\n\s*(\d{4,6})", text)
        if m:
            data["peso_bruto"] = m.group(2)
        else:
            # fallback general: último bloque numérico grande antes de nombre de empresa
            m = re.search(r"\b(\d{4,6})\b\s*\n\s*(?:\d[\d,]*\s*\n\s*)?[A-ZÁÉÍÓÚÑÜ\s]{4,}(?:SA|SRL|LTDA)\b", text)
            data["peso_bruto"] = m.group(1) if m else ""
    else:
        data["peso_bruto"] = m.group(1).replace(",", "") if m else ""

    if modo == "terrestre":
        # DNI: buscar label "DNI" seguido del número
        m = re.search(r"\bDNI\s*\n?\s*(\d{7,8})", text)
        data["cuil"] = m.group(1) if m else ""
        # Conductor: línea antes de "DNI"
        for i, line in enumerate(lines):
            if line.strip() == "DNI" and i > 0:
                data["conductor"] = lines[i - 1].strip()
                break
        else:
            data["conductor"] = ""
        # Exportador: línea antes del CUIT
        m = re.search(r"PLT\d+\s*\n\s*(.+?)\s*\n\s*\d{9,11}", text)
        data["exportador"] = m.group(1).strip() if m else ""
    else:
        # CUIL: buscar "CUIL" seguido del número (misma línea o siguiente)
        m = re.search(r"CUIL\s*(\d{11})", text)
        if not m:
            # Fallback: cualquier número de 11 dígitos
            m = re.search(r"\b(\d{11})\b", text)
        data["cuil"] = m.group(1) if m else ""
        # Exportador: línea antes del CUIT (9-11 dígitos)
        m = re.search(r"PLT\d+\s*\n\s*(.+?)\s*\n\s*\d{9,11}", text)
        data["exportador"] = m.group(1).strip() if m else ""

    # Precinto(s): extraer código(s) después de "Salida Reversada/Anulada"
    m = re.search(r"Salida Reversada/Anulada\s*\n(.+?)(?=\n\s*$|\nPEMA|\n(?:A GRANEL|Tipo Embalaje|\d{3,4}\s*$))", text, re.DOTALL)
    if m:
        block = m.group(1).strip()
        parts = re.split(r'[\s\n\-]+', block)
        codes = [p.upper() for p in parts if p.strip() and len(p.strip()) >= 6 and not p.strip().isdigit()]
        data["precinto"] = " ".join(sorted(set(codes)))
    else:
        data["precinto"] = ""

    # Id Destinación: código alfanumérico largo después de EXPORTACION A CONSUMO
    m = re.search(r"EXPORTACION A CONSUMO\s*\n?\s*(\S+)", text)
    data["id_destinacion"] = m.group(1) if m else ""

    # Patentes argentinas: formato nuevo (AA111AA) o viejo (AAA111)
    pat_all = re.findall(r"\b(?:[A-Z]{2}\d{3}[A-Z]{2}|[A-Z]{3}\d{3})\b", text)
    # Filtrar falsos positivos como "PLT075"
    patentes = [p for p in pat_all if not p.startswith("PLT")]
    if len(patentes) >= 2:
        data["patente_camion"] = patentes[0]
        data["patente_semi"] = patentes[1]
    elif len(patentes) == 1:
        data["patente_camion"] = patentes[0]
        data["patente_semi"] = ""
    else:
        data["patente_camion"] = ""
        data["patente_semi"] = ""

    # Conductor (fallback para flexi si no se encontró antes)
    if modo == "flexi" and not data.get("conductor"):
        for i, line in enumerate(lines):
            if line.strip() == "CUIL" and i > 0:
                data["conductor"] = lines[i - 1].strip()
                break
        else:
            data["conductor"] = ""

    return data


def extraer_mic_dta(pdf_path: str) -> dict:
    """
    Extrae datos de un PDF de MIC/DTA (Manifiesto Internacional de Carga).
    Extrae por campos numerados (1-41) del formulario MIC/DTA.

    Retorna dict con las mismas claves que espera _build_fila_control_final:
        plt, patente_camion, patente_semi, peso_bruto, precinto,
        id_destinacion, exportador, conductor, cuil, contenedor
    """
    import fitz
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()

    if not text.strip():
        return {"_error": "El PDF de MIC/DTA no tiene texto"}

    # Separar sección MIC (campos 1-41) de la sección CRT que sigue después
    mic_end = text.find("CRT\nCarta de Porte")
    mic_text = text[:mic_end] if mic_end > 0 else text

    lines = mic_text.split("\n")
    field_pat = re.compile(r"^(\d{1,2})(?:\s|$)")
    blocks = {}
    current_num = None
    current_lines = []

    for line in lines:
        m = field_pat.match(line.strip())
        if m:
            num = int(m.group(1))
            if 1 <= num <= 41:
                if current_num is not None:
                    blocks[current_num] = "\n".join(current_lines)
                current_num = num
                current_lines = [line.strip()]
                continue
        if current_num is not None:
            current_lines.append(line.rstrip())

    if current_num is not None:
        blocks[current_num] = "\n".join(current_lines)

    def _f(n):
        return blocks.get(n, "")

    data = {}

    # Campo 4 — N° MIC/DTA
    txt = _f(4)
    m = re.search(r"(\d{2}[A-Z]{2}\d{5,6}[A-Z]?)", txt)
    data["plt"] = m.group(1) if m else ""

    # Campo 11 — Patente camión (arg/chilena/extranjera)
    txt = _f(11)
    # Buscar la línea después de "Placa de Camion" / "Placa do caminhao"
    m = re.search(r"Placa\s*(?:de\s*)?Camion[^\n]*\n[^\n]*\n\s*([A-Z0-9]{4,8})\b", txt, re.IGNORECASE)
    if not m:
        # Fallback: cualquier token de 4-8 alfanumérico que no sea label
        m = re.search(r"\b([A-Z0-9]{4,8})\b", txt)
    data["patente_camion"] = m.group(1) if m else ""

    # Campo 15 — Patente semi (campo 22 no suele tener el valor)
    txt = _f(15)
    labels_15 = {"X", "SEMIREMOLQUE", "REMOLQUE", "SEMI", "SEMI-REBOQUE", "REBOQUE", "PLACA", "PLACA:"}
    pat_all = re.findall(r"\b([A-Z0-9]{4,8})\b", txt.upper())
    patentes = [p for p in pat_all if p not in labels_15 and not p.isdigit()]
    data["patente_semi"] = patentes[0] if patentes else ""

    # Campo 31 — Cantidad de bultos / Peso bruto
    txt = _f(31)
    m = re.search(r"(?:Cantidad de bultos|Quantidade de volumes)[^\n]*\n\s*(\d[\d.,]*)", txt)
    if not m:
        m = re.search(r"(\d[\d.,]*)", txt)
    data["peso_bruto"] = m.group(1).replace(",", "").replace(".", "") if m else ""

    # Campo 37 — Precinto
    txt = _f(37)
    m = re.search(r"([A-Z0-9]{6,})", txt)
    data["precinto"] = m.group(1).strip() if m else ""

    # Campo 36 — Documentos Anexos / Destinación
    txt = _f(36)
    m = re.search(r"Destinacion[:\s]*(\S+)", txt)
    data["id_destinacion"] = m.group(1) if m else ""

    # Campo 33 — Exportador / Remitente
    txt = _f(33)
    m = re.search(r"(?:Remitente|Remetente)\s*\n\s*/\s*\w+[^\n]*\n\s*([^\n]+)", txt)
    if not m:
        m = re.search(r"Remitente\s*\n[^/][^\n]*", txt)
    data["exportador"] = m.group(1).strip() if m else ""

    # Campo 40 — Conductor + DNI/RUT (Argentina DNI o Chile CI/RUT)
    txt = _f(40)
    m = re.search(r"CONDUCTOR 1:\s*(.+?)\s+DOC:\s*(?:DNI|CI)\s*([\d-]+)", txt)
    if m:
        data["conductor"] = m.group(1).strip()
        data["cuil"] = m.group(2).strip()
    else:
        data["conductor"] = ""
        data["cuil"] = ""

    # Contenedor: no aplica en MIC terrestre
    data["contenedor"] = ""

    return data


def api_vision_extraer_datos(ruta_pdf: str, api_key: str,
                             model: str = MODELO_VISION_DEFAULT,
                             temperature: float = 0.1,
                             max_tokens: int = 8192,
                             timeout: int = 60,
                             api_base: str = OPENROUTER_BASE) -> dict:
    """Convierte PDF a imagen, la envía a una API visión (OpenRouter)
    y devuelve datos estructurados del ticket.

    Args:
        ruta_pdf: Ruta al archivo PDF.
        api_key: API key de OpenRouter (o del provider elegido).
        model: Modelo con visión a usar.
        temperature: Temperatura del modelo (0-1).
        max_tokens: Máximo de tokens en la respuesta.
        timeout: Timeout de la petición HTTP en segundos.
        api_base: Base URL de la API.

    Returns:
        dict: Datos extraídos con claves del ticket, o {"error": mensaje} si falla.
    """
    try:
        import requests
    except ImportError:
        return {"error": "requests no instalado. Ejecutá: pip install requests"}

    import base64
    import json
    import os

    # Leer PDF directamente como bytes
    filename = os.path.basename(ruta_pdf)
    try:
        with open(ruta_pdf, "rb") as f:
            pdf_bytes = f.read()
    except Exception as e:
        return {"error": f"Error leyendo PDF: {e}"}

    b64 = base64.b64encode(pdf_bytes).decode("utf-8")

    # System prompt con los nombres exactos del ticket
    system_prompt = """Extraé los datos del ticket de balanza y devolvé SOLO un JSON con estos campos exactos:

{
  "Patente": "",
  "Semirremolque": "",
  "Conductor": "",
  "DNI": "",
  "Neto": "",
  "Tara Contenedor": "",
  "Contenedor": "",
  "Permiso": ""
}

Instrucciones por campo:
- "Patente": buscar la etiqueta "Patente" o "Camión". Las patentes pueden ser:
  - ARGENTINA nuevo (Mercosur): AB123CD (2 letras + 3 números + 2 letras)
  - ARGENTINA viejo: ABC123 (3 letras + 3 números)
  - BRASILERA nuevo: ABC1D23 (3 letras + 1 número + 1 letra + 2 números)
  - BRASILERA viejo: ABC1234 (3 letras + 4 números)
  - CHILENA: 4 letras + 2 números (ej: HFBH63) o 2 letras + 4 números
  - PARAGUAYA: 4 letras + 3 números o 3 letras + 3 números
  - URUGUAYA: 3 letras + 4 números
  - Si ves una combinación de letras y números de 5 a 7 caracteres, probablemente es la patente.
  - IMPORTANTE: no confundas la letra O (oh) con el número 0 (cero). Si una patente tiene números donde deberían ir letras (ej: "AF1910A" → es "AF191OA"), usá la interpretación correcta según el formato.
- "Semirremolque": buscar "Acoplado", "Semirremolque" o "Semi"
- "Conductor": buscar "Conductor" o "Chofer"
- "DNI": puede ser DNI argentino (solo números), CI chileno (ej: 24581338-4 con guión), o RUT. Tomar el número completo incluyendo el guión si aparece.
- "Neto": solo el número, buscar "Neto" o "Peso Neto"
- "Tara Contenedor": solo el número, buscar "Tara" o "Tara Contenedor"
- "Contenedor": código de 4 letras + 7 números, buscar "Contenedor" o "ISO" o "CNTR"
- "Permiso": buscar "Merc. y Permiso Embarque" o "Permiso Embarque". El permiso es un código largo que empieza con "26" (ej: 26069EC01000398W). Si aparece "LECITINA DE SOJA - 26069EC01000398W", tomar SOLO el código después del guión.

Si un campo no se ve, dejalo como string vacío. No inventes datos."""

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    # Headers específicos de OpenRouter
    if "openrouter" in api_base.lower():
        headers["HTTP-Referer"] = "https://github.com/usuario/Multiagente"
        headers["X-Title"] = "Multiagente Control de Datos"

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": [
                {"type": "file", "file": {
                    "filename": filename,
                    "file_data": f"data:application/pdf;base64,{b64}"
                }},
                {"type": "text", "text": "Extraé los datos del ticket en formato JSON."}
            ]},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }

    try:
        resp = requests.post(
            f"{api_base}/chat/completions",
            headers=headers,
            json=payload,
            timeout=timeout,
        )
        resp.raise_for_status()
        result = resp.json()
        content = result["choices"][0]["message"]["content"].strip()
        # Limpiar posibles bloques markdown ```json ... ```
        if content.startswith("```"):
            import re as _re
            content = _re.sub(r'^```(?:json)?\s*', '', content)
            content = _re.sub(r'\s*```$', '', content)
        return json.loads(content)
    except requests.exceptions.HTTPError as e:
        body = e.response.text[:500] if e.response is not None else ""
        return {"error": f"HTTP {e.response.status_code}: {body}"}
    except requests.exceptions.Timeout:
        return {"error": "timeout"}
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def resolver_pdfs(args: list[str]) -> list[Path]:
    """Recibe paths de archivos o carpetas y devuelve lista de PDFs."""
    pdfs = []
    for arg in args:
        p = Path(arg)
        if p.is_dir():
            pdfs.extend(sorted(p.glob("*.pdf")))
        elif p.suffix.lower() == ".pdf":
            pdfs.append(p)
        else:
            print(f"  ⚠ Ignorado (no es PDF ni carpeta): {arg}")
    return pdfs


def main():
    args = sys.argv[1:]
    if not args:
        print("Uso: python procesar_tickets.py archivo.pdf [archivo2.pdf ...] [carpeta/]")
        sys.exit(1)

    pdfs = resolver_pdfs(args)
    if not pdfs:
        print("No se encontraron PDFs.")
        sys.exit(1)

    lista_datos = []
    for pdf in pdfs:
        print(f"📄 Procesando: {pdf.name}")
        try:
            texto = pdf_a_texto(str(pdf))
            datos = extraer_datos(texto)
            datos["_archivo"] = pdf.name  # referencia interna, no va al Excel
            lista_datos.append(datos)
            vacios = [k for k, v in datos.items() if not v and not k.startswith("_")]
            if vacios:
                print(f"   ⚠ Campos vacíos: {', '.join(vacios)}")
            else:
                print(f"   ✓ Todos los campos extraídos")
        except Exception as e:
            print(f"   ✗ Error: {e}")

    if lista_datos:
        crear_excel(lista_datos, OUTPUT_FILE)
        print(f"\n✅ Excel generado: {OUTPUT_FILE}  ({len(lista_datos)} ticket/s)")
    else:
        print("\n✗ No se pudo procesar ningún PDF.")


if __name__ == "__main__":
    main()
