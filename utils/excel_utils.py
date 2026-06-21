"""
Utilidades de Excel: detección de duplicados, celdas mergeadas, reintentos.
Extraído de ui_app.py — Paso 3 de la refactorización modular.
"""
import threading
import tkinter.messagebox as messagebox
from datetime import datetime

# circular import evitado: se importa bajo demanda dentro de las funciones


def preguntar_reintentar(nombre_archivo, parent=None):
    """Muestra diálogo preguntando si reintentar. Retorna True si reintentar, False si cancelar.
    Funciona tanto desde hilo principal como desde hilo de fondo."""
    resultado = threading.Event()
    respuesta = {"reintentar": False}

    def _mostrar():
        ok = messagebox.askretrycancel(
            "Archivo en uso",
            f"El archivo:\n\n    {nombre_archivo}\n\n"
            f"está abierto en Excel.\n\n"
            f"Cerrá el archivo en Excel y presioná 'Reintentar' para continuar,\n"
            f"o 'Cancelar' para abortar la operación.",
            parent=parent,
        )
        respuesta["reintentar"] = ok
        resultado.set()

    if threading.current_thread() is threading.main_thread():
        _mostrar()
    elif parent is not None and hasattr(parent, 'after'):
        parent.after(0, _mostrar)
        resultado.wait(timeout=300)
    else:
        _mostrar()

    return respuesta["reintentar"]


def celda_es_mergeada(ws, row, col):
    """True si la celda está dentro de un merged range como no-anchor (no se puede escribir)."""
    from openpyxl.cell.cell import MergedCell
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def primera_fila_libre(ws, inicio=3):
    """Busca la primera fila completamente vacía y sin celdas mergeadas."""
    fila = inicio
    while True:
        libre = True
        for col in range(1, 7):
            cell = ws.cell(row=fila, column=col)
            if cell.value is not None:
                libre = False
                break
            if celda_es_mergeada(ws, fila, col):
                libre = False
                break
        if libre:
            return fila
        fila += 1


def ya_existe_en_hoja(ws, valores_fila, excluir_columnas=None):
    """Verifica si ya existe una fila con TODOS los mismos valores en las columnas indicadas.
    Resuelve celdas mergeadas para comparar correctamente.
    excluir_columnas: set de números de columna a no comparar (ej: {9} para B/L)."""
    from .pendrive import formatear_fecha_excel

    if excluir_columnas is None:
        excluir_columnas = set()

    def _normalizar(val):
        if val is None:
            return ""
        if isinstance(val, datetime):
            return formatear_fecha_excel(val)
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        if isinstance(val, (int, float)):
            return str(val)
        s = str(val).strip()
        if s.startswith("="):
            return s
        return s

    # Pre-mapear celdas mergeadas: {(row, col): valor_real}
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        anchor_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = anchor_val

    def _valor_real(row, col):
        key = (row, col)
        if key in merge_map:
            return merge_map[key]
        return ws.cell(row=row, column=col).value

    for row in range(3, ws.max_row + 1):
        coincide = True
        for col, val_esperado in valores_fila.items():
            if col in excluir_columnas:
                continue
            val_celda = _valor_real(row, col)
            if _normalizar(val_celda) != _normalizar(val_esperado):
                coincide = False
                break
        if coincide:
            return True
    return False
