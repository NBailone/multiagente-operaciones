"""PlanillasMixin — Panel de Planillas / SOBRES.

Legacy block extracted from ui_app.py. Methods are mixed into the main
App class; they rely on attributes/methods living on the composed instance
(_log, _set_log_panel, _cfg_obtener, _cfg_obtener_rutas, _resolver_ruta,
_abrir_excel_seguro, _guardar_excel_seguro, _mostrar_resumen, _scan_desktop_folders,
_clasificar_tipo_transporte, _limpiar_planillas, _set_status, _icons,
_panel_frames, panel_container, tarea_activa, datos_planillas, log_queue).
"""

import os
import re
from copy import copy
import string
import threading
import time
from datetime import datetime
import zipfile

import customtkinter as ctk
from tkinter import ttk, messagebox
import openpyxl
import xlrd

from constants import Palette, FONT_FAMILY
from utils import (
    buscar_archivo_en_pendrive,
    buscar_bl_por_carpeta_xls,
    buscar_bl_por_carpeta_xlsx,
    formatear_fecha_excel,
    preguntar_reintentar,
    primera_fila_libre,
    ya_existe_en_hoja,
)


class PlanillasMixin:
    # PANEL 2: PLANILLAS / SOBRES
    # ═══════════════════════════════════════════════════════════════════
    def _panel_planillas(self):
        if "planillas" in self._panel_frames:
            return
        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        self._panel_frames["planillas"] = frame
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── Toolbar ──────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        toolbar.pack(fill="x", pady=(0, 6))

        # Botones principales a la izquierda
        btns_left = ctk.CTkFrame(toolbar, fg_color="transparent")
        btns_left.pack(side="left", fill="x", expand=True, padx=4, pady=4)

        _btn_width = 140

        self.btn_ejecutar_planillas = ctk.CTkButton(
            btns_left,
            text="Completar Planillas",
            image=self._icons["play"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT,
            hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE,
            corner_radius=6, height=34, width=_btn_width,
            command=self._popup_completar_planillas,
        )

        self.btn_agregar_guarda = ctk.CTkButton(
            btns_left,
            text="Agregar Guarda",
            image=self._icons["shield"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.SECONDARY,
            hover_color=Palette.SECONDARY_HOVER,
            text_color=Palette.WHITE,
            corner_radius=6, height=34, width=_btn_width,
            command=self._popup_agregar_guarda,
        )

        self.btn_editar_excels = ctk.CTkButton(
            btns_left, text="Editar Excels",
            image=self._icons["pencil"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34, width=_btn_width,
            command=self._popup_editar_excels,
        )

        self.btn_planilla_carga = ctk.CTkButton(
            btns_left, text="Planilla de Carga",
            image=self._icons["clipboard-list"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34, width=_btn_width,
            command=self._popup_planilla_carga,
        )

        # Progress + Limpiar a la derecha
        right_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        right_frame.pack(side="right", padx=(0, 8), pady=4)
        right_frame.pack_propagate(False)
        right_frame.configure(width=220, height=34)

        self.progress_planillas = ctk.CTkProgressBar(
            right_frame, width=120, height=8, corner_radius=4,
            fg_color=Palette.BG_INPUT, progress_color=Palette.ACCENT,
        )
        self.progress_planillas.pack(side="left", padx=(0, 8), pady=13)
        self.progress_planillas.set(0)

        self.btn_limpiar_planillas = ctk.CTkButton(
            right_frame,
            text="Limpiar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color="transparent",
            hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            corner_radius=4, height=30, width=70,
            command=self._limpiar_planillas,
        )
        self.btn_limpiar_planillas.pack(side="left", pady=2)

        _toolbar_btns = [
            self.btn_ejecutar_planillas,
            self.btn_agregar_guarda,
            self.btn_editar_excels,
            self.btn_planilla_carga,
        ]
        for b in _toolbar_btns:
            b.grid(row=0, column=0)

        def _reflow_toolbar_btns(event=None):
            fw = btns_left.winfo_width()
            if fw < 50:
                return
            pitch = _btn_width + 16  # 140 + 16 = 156
            cols = max(1, fw // pitch)
            for i, b in enumerate(_toolbar_btns):
                b.grid(row=i // cols, column=i % cols, padx=4, pady=3, sticky="w")

        btns_left.bind("<Configure>", _reflow_toolbar_btns)
        btns_left.after_idle(_reflow_toolbar_btns)

        # ── Tabla de resultados ─────────────────────────────────────
        self._crear_tabla_planillas(frame)

        # ── Resumen ──────────────────────────────────────────────────
        resumen_frame = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=48
        )
        resumen_frame.pack(fill="x", pady=(6, 0))
        resumen_frame.pack_propagate(False)

        self.lbl_resumen_planillas = ctk.CTkLabel(
            resumen_frame,
            text="Sin datos analizados",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            text_color=Palette.TEXT_MUTED,
        )
        self.lbl_resumen_planillas.pack(side="left", padx=16, pady=12)

    def _crear_tabla_planillas(self, parent):
        """Tabla estilo Treeview dentro de un frame oscuro."""
        tabla_frame = ctk.CTkFrame(
            parent, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER
        )
        tabla_frame.pack(fill="both", expand=True)

        # Treeview con estilo oscuro
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Contenedores.Treeview",
            background=Palette.BG_TABLE,
            foreground=Palette.TEXT_PRIMARY,
            fieldbackground=Palette.BG_TABLE,
            borderwidth=0,
            font=(FONT_FAMILY, 10),
            rowheight=28,
        )
        style.configure(
            "Contenedores.Treeview.Heading",
            background=Palette.BG_SIDEBAR,
            foreground=Palette.TEXT_SECONDARY,
            font=(FONT_FAMILY, 10, "bold"),
            borderwidth=0,
            padding=(8, 6),
        )
        style.map(
            "Contenedores.Treeview",
            background=[("selected", Palette.ACCENT_DIM)],
            foreground=[("selected", Palette.WHITE)],
        )

        columnas = (
            "fecha", "cant", "pe", "carpeta", "terminal",
            "transporte", "destinatario", "bl", "fraccion", "servicio"
        )
        headers = (
            "Fecha Carga", "Cant.", "P.E.", "Carpeta", "Terminal",
            "Transporte", "Destinatario", "B/L", "Fracción", "Servicio (ATA)"
        )
        anchos = (90, 50, 65, 65, 70, 120, 110, 70, 80, 100)

        self.tree_planillas = ttk.Treeview(
            tabla_frame, columns=columnas, show="headings",
            style="Contenedores.Treeview", selectmode="browse",
        )
        for col, hdr, w in zip(columnas, headers, anchos):
            self.tree_planillas.heading(col, text=hdr)
            self.tree_planillas.column(col, width=w, anchor="center", minwidth=w)

        # Scrollbar horizontal (arriba)
        scroll_planillas_x = ctk.CTkScrollbar(
            tabla_frame, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_planillas.configure(xscrollcommand=scroll_planillas_x.set)
        scroll_planillas_x.configure(command=self.tree_planillas.xview)
        scroll_planillas_x.pack(fill="x", padx=2, pady=(2, 0))

        # Scrollbar vertical
        scroll_y = ctk.CTkScrollbar(
            tabla_frame, orientation="vertical",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        scroll_y.pack(side="right", fill="y", padx=(0, 2), pady=2)

        self.tree_planillas.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self.tree_planillas.yview)

        self.tree_planillas.pack(fill="both", expand=True, padx=2, pady=(0, 0))

        # Tags de color por empresa
        self.tree_planillas.tag_configure("vitapro", background="#003340", foreground="#87ceeb")
        self.tree_planillas.tag_configure("ewos", background="#003300", foreground="#00ff00")
        self.tree_planillas.tag_configure("dicoal", background="#4d2e00", foreground="#ffb74d")
        self.tree_planillas.tag_configure("nutreco", background="#4d4100", foreground="#ffee58")
        self.tree_planillas.tag_configure("cargill", background="#0d2a59", foreground="#66b2ff")
        self.tree_planillas.tag_configure("biomar", background="#1a4d2e", foreground="#69f0ae")
    # ═══════════════════════════════════════════════════════════════════
    # AGREGAR GUARDA
    # ═══════════════════════════════════════════════════════════════════
    def _popup_agregar_guarda(self):
        """Popup para elegir un guarda y las carpetas donde aplicarlo."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución. Espere a que finalice.")
            return

        self._scan_desktop_folders(callback=self._popup_agregar_guarda_done)

    def _popup_agregar_guarda_done(self, results):
        """Callback: muestra popup de agregar guarda con las carpetas encontradas."""
        guardas = self._cfg_obtener("valores", "guardas", ["Gonzalez", "Rodriguez", "Martinez", "Perez"])
        carpetas_encontradas = [(r["name"], r["path"]) for r in results]

        if not carpetas_encontradas:
            messagebox.showinfo("Sin planillas", "No se encontraron archivos Contenedores en el escritorio.")
            return

        # ── Popup ──────────────────────────────────────────────────
        popup = ctk.CTkToplevel(self)
        popup.title("Agregar Guarda")
        popup.geometry("550x520")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.grab_set()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 550) // 2
        py = self.winfo_y() + (self.winfo_height() - 520) // 2
        popup.geometry(f"550x520+{px}+{py}")

        ctk.CTkLabel(
            popup, text="Seleccioná un Guarda",
            font=ctk.CTkFont(family=FONT_FAMILY, size=15, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        # Dropdown de guardas
        guarda_var = ctk.StringVar(value=guardas[0] if guardas else "")
        ctk.CTkOptionMenu(
            popup, variable=guarda_var, values=guardas,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
            button_hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6,
            width=240, height=36,
        ).pack(pady=(0, 16))

        ctk.CTkLabel(
            popup, text="Carpetas donde se aplicará el guarda:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(popup, fg_color=Palette.BG_TABLE, corner_radius=8,
                                         border_width=1, border_color=Palette.BORDER)
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for nombre, ruta_carp in carpetas_encontradas:
            var = ctk.BooleanVar(value=True)
            checks[nombre] = var
            ctk.CTkCheckBox(
                scroll, text=nombre, variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

        # Botones
        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        ctk.CTkButton(
            btn_frame, text="Aplicar Guarda", width=180, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=lambda: self._aplicar_guarda(
                popup, guarda_var.get(),
                [c[1] for c in carpetas_encontradas if checks[c[0]].get()]
            ),
        ).pack(side="right", padx=4)
        ctk.CTkButton(
            btn_frame, text="Cancelar", width=100, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=popup.destroy,
        ).pack(side="right", padx=4)

    def _popup_editar_excels(self):
        """Popup con botones para abrir/editar archivos Excel."""
        popup = ctk.CTkToplevel(self)
        popup.title("Editar Excels")
        popup.geometry("360x360")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.grab_set()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 360) // 2
        py = self.winfo_y() + (self.winfo_height() - 360) // 2
        popup.geometry(f"360x360+{px}+{py}")

        ctk.CTkLabel(popup, text="EDITAR EXCELS",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
                     text_color=Palette.TEXT_PRIMARY).pack(pady=(20, 12))

        excels = [
            ("SOBRES_2026.xlsx", self._cfg_obtener_rutas("sobres", os.path.join("TRABAJO", "01_PLANILLAS"))),
            ("COBRO_2026.xlsx", self._cfg_obtener_rutas("cobro", os.path.join("TRABAJO", "01_PLANILLAS"))),
            ("PC.xlsx", self._cfg_obtener_rutas("pc", os.path.join("TRABAJO", "01_PLANILLAS"))),
            ("FECHA MIC Y SELLOS.xlsx", self._cfg_obtener_rutas("mic_sellos", os.path.join("TRABAJO", "01_PLANILLAS"))),
            ("FECHA CRT Y ORIGINAL.xlsx", self._cfg_obtener_rutas("crt_original", os.path.join("TRABAJO", "01_PLANILLAS"))),
            (self._cfg_obtener_rutas("carga_terrestre_nombre", "CARGA TERRESTRE.xlsx"),
             self._cfg_obtener_rutas("carga_terrestre_carpeta", os.path.join("TRABAJO", "01_PLANILLAS"))),
        ]

        def _abrir(nombre, carpeta):
            ruta = buscar_archivo_en_pendrive(nombre, carpeta)
            if ruta:
                os.startfile(ruta)
                popup.destroy()
            else:
                messagebox.showwarning("No encontrado", f"No se encontró:\n{nombre}\n\nEn: {carpeta}")

        for nombre, carpeta in excels:
            ctk.CTkButton(popup, text=f"📝 {nombre}",
                          font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                          fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
                          text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34, width=280,
                          command=lambda n=nombre, c=carpeta: _abrir(n, c)).pack(pady=3)

        ctk.CTkButton(popup, text="Cerrar", width=100, height=32,
                      font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                      fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
                      text_color=Palette.TEXT_PRIMARY, corner_radius=6,
                      command=popup.destroy).pack(pady=(8, 10))

    # ═══════════════════════════════════════════════════════════════════
    # POPUP: PLANILLA DE CARGA
    # ═══════════════════════════════════════════════════════════════════

    def _popup_planilla_carga(self):
        """Popup para extraer la hoja 'Planilla de Carga' de archivos Excel."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución. Espere a que finalice.")
            return

        escritorio = self._resolver_ruta("planillas_carga", "Desktop")
        patron = re.compile(r"^\d{2}_\d{2}_\d{4}_\d+_.+$")

        carpetas_encontradas = []
        try:
            for item in sorted(os.listdir(escritorio)):
                ruta = os.path.join(escritorio, item)
                if os.path.isdir(ruta) and patron.match(item):
                    carpetas_encontradas.append((item, ruta))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el escritorio:\n{e}")
            return

        if not carpetas_encontradas:
            messagebox.showinfo("Sin carpetas",
                                "No se encontraron carpetas con formato\n"
                                "DD_MM_AAAA_NNN_Nombre en el escritorio.")
            return

        popup = ctk.CTkToplevel(self)
        popup.title("Planilla de Carga")
        popup.geometry("600x400")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.grab_set()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 600) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        popup.geometry(f"600x400+{px}+{py}")

        ctk.CTkLabel(
            popup, text="PLANILLA DE CARGA",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        ctk.CTkLabel(
            popup, text="Seleccioná las carpetas de las que extraer la hoja:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(
            popup, fg_color=Palette.BG_TABLE, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for nombre, ruta_carp in carpetas_encontradas:
            var = ctk.BooleanVar(value=True)
            checks[nombre] = var

        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        seleccionadas_var = []

        def _obtener_seleccionadas():
            seleccionadas_var.clear()
            for nombre, ruta in carpetas_encontradas:
                if checks[nombre].get():
                    seleccionadas_var.append(ruta)

        def _update_generar_btn(*_args):
            any_selected = any(v.get() for v in checks.values())
            btn_generar.configure(state="normal" if any_selected else "disabled")

        def _seleccionar_todas():
            for v in checks.values():
                v.set(True)

        def _borrar_seleccion():
            for v in checks.values():
                v.set(False)

        # BotonesSeleccionar todas / Ninguna (izquierda)
        ctk.CTkButton(
            btn_frame, text="Todas", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_seleccionar_todas,
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            btn_frame, text="Ninguna", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_borrar_seleccion,
        ).pack(side="left", padx=4)

        btn_generar = ctk.CTkButton(
            btn_frame, text="Generar", width=140, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            state="disabled",
            command=lambda: (_obtener_seleccionadas(),
                             self._generar_planilla_carga_thread(popup, seleccionadas_var)),
        )
        btn_generar.pack(side="right", padx=4)

        for nombre, var in checks.items():
            var.trace_add("write", _update_generar_btn)
            ctk.CTkCheckBox(
                scroll, text=nombre, variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

        _update_generar_btn()

        ctk.CTkButton(
            btn_frame, text="Cancelar", width=100, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=popup.destroy,
        ).pack(side="right", padx=4)

    def _extraer_planilla_carga(self, folder_path):
        """Extrae la hoja 'planilla de carga' del primer .xlsx en la carpeta.

        Returns:
            (success, message)
        """
        try:
            archivos = os.listdir(folder_path)
        except Exception as e:
            return False, f"No se pudo leer la carpeta: {e}"

        # Buscar primer archivo Excel con preferencia a *CONTENEDORES*
        xlsx_preferido = None
        xlsx_cualquiera = None
        xls_preferido = None
        xls_cualquiera = None
        for archivo in archivos:
            # Ignorar archivos temporales de Excel (~$xxx.xlsx) — se crean cuando el archivo está abierto
            if archivo.startswith("~$"):
                continue
            up = archivo.upper()
            ruta = os.path.join(folder_path, archivo)
            if up.endswith(".XLSX"):
                if "CONTENEDORES" in up:
                    xlsx_preferido = ruta
                if xlsx_cualquiera is None:
                    xlsx_cualquiera = ruta
            elif up.endswith(".XLS"):
                if "CONTENEDORES" in up:
                    xls_preferido = ruta
                if xls_cualquiera is None:
                    xls_cualquiera = ruta

        source_path = xlsx_preferido or xls_preferido or xlsx_cualquiera or xls_cualquiera
        if not source_path:
            return False, "No se encontró archivo .xlsx/.xls en la carpeta"

        is_xls = source_path.lower().endswith(".xls") and not source_path.lower().endswith(".xlsx")

        # Cargar workbook y buscar hoja
        if is_xls:
            # xlrd 2.0+ no soporta estilos. Usamos win32com para convertir .xls → .xlsx
            # temporal y luego procesar como .xlsx (preserva formato, fórmulas, protección).
            import win32com.client as _wincom
            import pythoncom
            pythoncom.CoInitialize()
            excel = None
            wb_com = None
            temp_xlsx = None
            try:
                excel = _wincom.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                # Workbook-level password "123" may be set on .xls
                try:
                    wb_com = excel.Workbooks.Open(source_path, ReadOnly=True, Password="123")
                except Exception:
                    wb_com = excel.Workbooks.Open(source_path, ReadOnly=True)
                # Save as .xlsx (FileFormat 51 = xlOpenXMLWorkbook)
                temp_xlsx = source_path + ".__temp__.xlsx"
                if os.path.exists(temp_xlsx):
                    try:
                        os.unlink(temp_xlsx)
                    except OSError:
                        pass
                wb_com.SaveAs(temp_xlsx, FileFormat=51)
                wb_com.Close(SaveChanges=False)
                wb_com = None
                excel.Quit()
                excel = None
            except Exception as e:
                return False, f"Error al convertir .xls a .xlsx: {e}"
            finally:
                try:
                    if wb_com:
                        wb_com.Close(SaveChanges=False)
                except Exception:
                    pass
                try:
                    if excel:
                        excel.Quit()
                except Exception:
                    pass
                pythoncom.CoUninitialize()

            if not temp_xlsx or not os.path.exists(temp_xlsx):
                return False, "No se pudo convertir el archivo .xls a .xlsx"

            # Ahora procesar el .xlsx temporal con openpyxl (data_only para valores cacheados)
            try:
                wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
            except Exception as e:
                try:
                    os.unlink(temp_xlsx)
                except OSError:
                    pass
                return False, f"Error al abrir .xls convertido: {e}"

            ws_origen_name = None
            for sn in wb.sheetnames:
                if sn.strip().lower() == "planilla de carga":
                    ws_origen_name = sn
                    break

            if not ws_origen_name:
                wb.close()
                try:
                    os.unlink(temp_xlsx)
                except OSError:
                    pass
                return False, "No se encontró la hoja 'Planilla de Carga' en .xls"

            ws_origen = wb[ws_origen_name]

            # Resolver XML path del .xlsx temporal para lectura de anchos de columna
            target_sheet_xml_path = None
            try:
                import zipfile as _zf
                import xml.etree.ElementTree as _ET
                _NS_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                _NS_SHEET = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                with _zf.ZipFile(temp_xlsx, 'r') as _z:
                    _wb_xml = _ET.fromstring(_z.read('xl/workbook.xml'))
                    _sheets = _wb_xml.find(f'{{{_NS_SHEET}}}sheets')
                    _rid = None
                    for _s in _sheets:
                        if (_s.get('name') or '').strip().lower() == 'planilla de carga':
                            _rid = _s.get(f'{{{_NS_REL}}}id')
                            break
                    if _rid:
                        _rels = _ET.fromstring(_z.read('xl/_rels/workbook.xml.rels'))
                        for _r in _rels:
                            if _r.get('Id') == _rid:
                                target_sheet_xml_path = 'xl/' + _r.get('Target')
                                break
            except Exception:
                target_sheet_xml_path = None

            # Registrar cleanup del temp al final del procesamiento
            # (se elimina tras guardar el archivo de salida)
            self._temp_xlsx_to_clean = temp_xlsx

        else:
            # --- .xlsx path: load with data_only=True to get cached values ---
            # data_only=True returns cached formula results instead of formula strings.
            # This avoids #REF errors when the referenced sheets (DATOS, etc.) are removed.
            try:
                wb = openpyxl.load_workbook(source_path, data_only=True)
            except Exception as e:
                return False, f"Error al abrir .xlsx: {e}"

            ws_origen_name = None
            for sn in wb.sheetnames:
                if sn.strip().lower() == "planilla de carga":
                    ws_origen_name = sn
                    break

            if not ws_origen_name:
                wb.close()
                return False, "No se encontró la hoja 'Planilla de Carga' en .xlsx"

            ws_origen = wb[ws_origen_name]

            # Resolve the sheet's XML path inside the .xlsx ZIP (needed below for column widths)
            target_sheet_xml_path = None
            try:
                import zipfile as _zf
                import xml.etree.ElementTree as _ET
                _NS_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                _NS_SHEET = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                with _zf.ZipFile(source_path, 'r') as _z:
                    _wb_xml = _ET.fromstring(_z.read('xl/workbook.xml'))
                    _sheets = _wb_xml.find(f'{{{_NS_SHEET}}}sheets')
                    _rid = None
                    for _s in _sheets:
                        if (_s.get('name') or '').strip().lower() == 'planilla de carga':
                            _rid = _s.get(f'{{{_NS_REL}}}id')
                            break
                    if _rid:
                        _rels = _ET.fromstring(_z.read('xl/_rels/workbook.xml.rels'))
                        for _r in _rels:
                            if _r.get('Id') == _rid:
                                target_sheet_xml_path = 'xl/' + _r.get('Target')
                                break
            except Exception:
                target_sheet_xml_path = None
            # Falls through to shared cell-by-cell copy code below

        # ── xls path: copy cell by cell into new workbook ─────────────
        # Crear nuevo workbook
        wb_nuevo = openpyxl.Workbook()
        ws_destino = wb_nuevo.active
        ws_destino.title = ws_origen_name if ws_origen_name else "Planilla de Carga"

        # Copiar celdas: valores y estilos
        for row in ws_origen.iter_rows():
            for cell in row:
                nueva_celda = ws_destino.cell(row=cell.row, column=cell.column)
                nueva_celda.value = cell.value
                if cell.has_style:
                    nueva_celda.font = copy(cell.font)
                    nueva_celda.border = copy(cell.border)
                    nueva_celda.fill = copy(cell.fill)
                    nueva_celda.number_format = cell.number_format
                    nueva_celda.protection = copy(cell.protection)
                    nueva_celda.alignment = copy(cell.alignment)

        # Copiar celdas combinadas
        for merge in ws_origen.merged_cells.ranges:
            ws_destino.merge_cells(str(merge))

        # Copiar anchos de columna
        # openpyxl no crea ColumnDimension para columnas agrupadas en un <col min="2" max="3">
        # (solo asigna el ancho a la primera letra del rango). Leemos el XML original
        # para obtener los anchos reales por rango y aplicarlos a cada columna individual.
        col_widths = {}  # col_letter -> width
        # Para .xls usamos el .xlsx temporal convertido; para .xlsx usamos el source directo
        _xml_source = temp_xlsx if is_xls else source_path
        if _xml_source and _xml_source.lower().endswith(".xlsx") and target_sheet_xml_path:
            try:
                import zipfile as _zf
                import xml.etree.ElementTree as _ET
                _NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                with _zf.ZipFile(_xml_source, 'r') as _z:
                    _sheet_xml = _ET.fromstring(_z.read(target_sheet_xml_path))
                for _col in _sheet_xml.findall(f'{{{_NS}}}cols/{{{_NS}}}col'):
                    _min = int(_col.get('min', '1'))
                    _max = int(_col.get('max', '1'))
                    _w = _col.get('width')
                    if _w:
                        for _ci in range(_min, _max + 1):
                            col_widths[openpyxl.utils.get_column_letter(_ci)] = float(_w)
            except Exception:
                pass  # fallback below
        # Fallback / complement: use openpyxl's column_dimensions for anything not covered
        for col_idx in range(1, 27):  # A-Z
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_letter in col_widths:
                ws_destino.column_dimensions[col_letter].width = col_widths[col_letter]
            else:
                src_dim = ws_origen.column_dimensions.get(col_letter)
                if src_dim and src_dim.width:
                    ws_destino.column_dimensions[col_letter].width = src_dim.width

        # Auto-fit: expand columns whose content es más ancho que el ancho seteado
        for col_idx in range(1, 27):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 0
            for row in ws_destino.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_len:
                            max_len = cell_len
            if max_len > 0:
                fit_width = max_len * 1.3 + 3
                current = ws_destino.column_dimensions[col_letter].width
                if current is None or fit_width > current:
                    ws_destino.column_dimensions[col_letter].width = fit_width

        # Copiar alturas de fila
        _max = ws_origen.max_row or 50
        for row_idx in range(1, _max + 1):
            src_dim = ws_origen.row_dimensions.get(row_idx)
            if src_dim and src_dim.height:
                ws_destino.row_dimensions[row_idx].height = src_dim.height

        # Copiar protección de hoja
        if ws_origen.protection.sheet:
            ws_destino.protection.sheet = True
            ws_destino.protection.enable()
            # openpyxl no preserva el hash del password al cargar; setear "123" explícitamente
            ws_destino.protection.password = self._cfg_obtener("valores", "clave_pdf", "123")

        # Generar nombre de archivo de salida
        partes = os.path.basename(folder_path).split("_")
        sufijo = "_".join(partes[5:]) if len(partes) > 5 else ""
        nombre_salida = f"PLANILLA DE CARGA_{sufijo}.xlsx" if sufijo else "PLANILLA DE CARGA.xlsx"
        ruta_salida = os.path.join(folder_path, nombre_salida)

        try:
            wb_nuevo.save(ruta_salida)
        except PermissionError:
            if preguntar_reintentar(nombre_salida, parent=self):
                try:
                    wb_nuevo.save(ruta_salida)
                except Exception as e:
                    wb.close()
                    wb_nuevo.close()
                    self._limpiar_temp_xlsx()
                    return False, f"No se pudo guardar: {e}"
            else:
                wb.close()
                wb_nuevo.close()
                self._limpiar_temp_xlsx()
                return False, "Operación cancelada (archivo en uso)"
        except Exception as e:
            wb.close()
            wb_nuevo.close()
            self._limpiar_temp_xlsx()
            return False, f"Error al guardar: {e}"

        wb.close()
        wb_nuevo.close()
        self._limpiar_temp_xlsx()
        return True, nombre_salida

    def _limpiar_temp_xlsx(self):
        """Elimina el archivo .xlsx temporal generado al convertir .xls."""
        temp = getattr(self, "_temp_xlsx_to_clean", None)
        if temp:
            try:
                if os.path.exists(temp):
                    os.unlink(temp)
            except OSError:
                pass
            self._temp_xlsx_to_clean = None

    def _generar_planilla_carga_thread(self, popup, carpetas_seleccionadas):
        """Procesa la extracción de planillas de carga en hilo de fondo."""
        if not carpetas_seleccionadas:
            messagebox.showwarning("Incompleto", "Seleccioná al menos una carpeta.")
            return

        popup.destroy()

        if self.tarea_activa:
            return
        self.tarea_activa = True
        self._cancelar_tarea.clear()
        self.btn_planilla_carga.configure(text="⏳  Procesando...", state="disabled")
        self.btn_ejecutar_planillas.configure(state="disabled")
        self.progress_planillas.set(0)
        self.progress_planillas.configure(mode="indeterminate")
        self.progress_planillas.start()
        self._limpiar_log()
        self._log("📋 Generando Planillas de Carga")

        def worker():
            self._set_log_panel("planillas")
            total = len(carpetas_seleccionadas)
            errores = []
            for i, ruta_carp in enumerate(carpetas_seleccionadas, 1):
                if self._cancelar_tarea.is_set():
                    self.log_queue.put("[...] ⚠ Tarea cancelada.")
                    break
                nombre_carp = os.path.basename(ruta_carp)
                self.log_queue.put(f"[...] ({i}/{total}) 📁 {nombre_carp}")

                ok, msg = self._extraer_planilla_carga(ruta_carp)
                if ok:
                    self.log_queue.put(f"[...]     ✓ {msg}")
                else:
                    self.log_queue.put(f"[...]     ⚠ {msg}")
                    errores.append(f"{nombre_carp}: {msg}")

            resumen = f"[...] ✓ Planillas de Carga completadas: {total - len(errores)}/{total}"
            if errores:
                resumen += f"\n[...] ⚠ Errores ({len(errores)}):"
                for err in errores:
                    resumen += f"\n[...]     • {err}"
            self.log_queue.put(resumen)

            def _mostrar_resumen():
                if errores:
                    messagebox.showwarning(
                        "Planilla de Carga",
                        f"Procesadas: {total - len(errores)}/{total}\n\n"
                        f"Errores:\n" + "\n".join(f"• {e}" for e in errores),
                    )
                else:
                    messagebox.showinfo(
                        "Planilla de Carga",
                        f"✅ {total} carpeta(s) procesada(s) correctamente.",
                    )

            self.after(0, _mostrar_resumen)
            self.after(0, self._planilla_carga_done)

        t = threading.Thread(target=worker, daemon=True)
        t.start()

    def _planilla_carga_done(self):
        self.tarea_activa = False
        try:
            self.btn_planilla_carga.configure(
                text="Planilla de Carga", state="normal",
            )
            self.btn_ejecutar_planillas.configure(state="normal")
            self.progress_planillas.stop()
        except (AttributeError, Exception):
            pass

    def _aplicar_guarda(self, popup, guarda_elegido, carpetas_seleccionadas):
        """Escribe el guarda elegido en la hoja Choferes de cada planilla de carga."""
        if not guarda_elegido or not carpetas_seleccionadas:
            messagebox.showwarning("Incompleto", "Elegí un guarda y al menos una carpeta.")
            return
        popup.destroy()

        if self.tarea_activa:
            return
        self.tarea_activa = True
        self._cancelar_tarea.clear()
        self.btn_agregar_guarda.configure(text="⏳  Procesando...", state="disabled")
        self.btn_ejecutar_planillas.configure(state="disabled")
        self.progress_planillas.set(0)
        self.progress_planillas.configure(mode="indeterminate")
        self.progress_planillas.start()
        self._limpiar_log()
        self._log(f"🛡 Agregando Guarda: {guarda_elegido}")

        def worker():
            self._set_log_panel("planillas")
            for ruta_carp in carpetas_seleccionadas:
                if self._cancelar_tarea.is_set():
                    self.log_queue.put("[...] ⚠ Tarea cancelada.")
                    break
                nombre_carp = os.path.basename(ruta_carp)
                self.log_queue.put(f"[...]   📁 {nombre_carp}")

                ruta_contenedores = None
                try:
                    for archivo in os.listdir(ruta_carp):
                        up = archivo.upper()
                        if "CONTENEDORES" in up and (up.endswith(".XLSX") or up.endswith(".XLS")):
                            ruta_contenedores = os.path.join(ruta_carp, archivo)
                            break
                except Exception as e:
                    self.log_queue.put(f"[...]     ⚠ No se pudo leer: {e}")
                    continue

                if not ruta_contenedores:
                    self.log_queue.put(f"[...]     ⚠ Sin archivo Contenedores")
                    continue

                archivo_nombre = os.path.basename(ruta_contenedores)
                self.log_queue.put(f"[...]     📄 {archivo_nombre}")

                try:
                    if not self._escribir_guarda_en_archivo(ruta_contenedores, guarda_elegido, self.log_queue.put):
                        self.log_queue.put(f"[...]     ⚠ 'Guarda' no hallada en col G")
                    else:
                        self.log_queue.put(f"[...]     ✓ '{guarda_elegido}' escrito correctamente")
                except Exception as e:
                    self.log_queue.put(f"[...]     ⚠ Error: {e}")

            self.log_queue.put(f"[...] ✓ Guarda '{guarda_elegido}' completado.")
            self.after(0, self._guarda_done)

        t = threading.Thread(target=worker, daemon=True)
        t.start()

    def _guarda_done(self):
        self.tarea_activa = False
        try:
            self.btn_agregar_guarda.configure(
                text="Agregar Guarda", state="normal",
                fg_color=Palette.SECONDARY)
            self.btn_ejecutar_planillas.configure(state="normal")
            self.progress_planillas.stop()
        except (AttributeError, Exception):
            pass

    # ═══════════════════════════════════════════════════════════════════
    # POPUP: SELECCIONAR PLANILLAS A COMPLETAR
    # ═══════════════════════════════════════════════════════════════════
    def _popup_completar_planillas(self):
        """Popup para elegir qué planillas completar (SOBRES, COBRO, PC)."""
        if self.tarea_activa:
            messagebox.showwarning(
                "Agente ocupado",
                "Hay una tarea en ejecución. Espere a que finalice."
            )
            return

        popup = ctk.CTkToplevel(self)
        popup.title("Completar Planillas")
        popup.geometry("360x300")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.resizable(False, False)

        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 360) // 2
        py = self.winfo_y() + (self.winfo_height() - 300) // 2
        popup.geometry(f"360x300+{px}+{py}")

        ctk.CTkLabel(
            popup, text="Seleccionar Planillas",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(20, 16))

        var_sobres = ctk.BooleanVar(value=True)
        var_cobro  = ctk.BooleanVar(value=True)
        var_pc     = ctk.BooleanVar(value=True)
        vars_tipos = {"sobres": var_sobres, "cobro": var_cobro, "pc": var_pc}

        def _actualizar_boton():
            alguna = any(v.get() for v in vars_tipos.values())
            btn_completar.configure(
                state="normal" if alguna else "disabled",
                fg_color=Palette.ACCENT if alguna else Palette.ACCENT_DIM,
            )

        def _seleccionar_todo():
            for v in vars_tipos.values():
                v.set(True)
            _actualizar_boton()

        def _ninguno():
            for v in vars_tipos.values():
                v.set(False)
            _actualizar_boton()

        # ── Checkboxes ────────────────────────────────────────────
        frame_cbs = ctk.CTkFrame(popup, fg_color="transparent")
        frame_cbs.pack(pady=(0, 8))

        for label, key in [("SOBRES", "sobres"), ("COBRO", "cobro"), ("PC (Precintos/Cables)", "pc")]:
            cb = ctk.CTkCheckBox(
                frame_cbs, text=label,
                font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                variable=vars_tipos[key], command=_actualizar_boton,
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            cb.pack(anchor="w", pady=4, padx=20)

        # ── Botones Seleccionar Todo / Ninguno ────────────────────
        frame_sel = ctk.CTkFrame(popup, fg_color="transparent")
        frame_sel.pack(pady=(4, 12))

        ctk.CTkButton(
            frame_sel, text="Seleccionar Todo",
            image=self._icons["check"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.SECONDARY, hover_color=Palette.SECONDARY_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=32, width=150,
            command=_seleccionar_todo,
        ).pack(side="left", padx=6)

        ctk.CTkButton(
            frame_sel, text="Ninguno",
            image=self._icons["x"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6, height=32, width=110,
            command=_ninguno,
        ).pack(side="left", padx=6)

        # ── Botón Completar ───────────────────────────────────────
        btn_completar = ctk.CTkButton(
            popup, text="Completar",
            image=self._icons["play"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=38,
            command=lambda: self._confirmar_completar(popup, vars_tipos),
        )
        btn_completar.pack(fill="x", padx=30, pady=(6, 16))

    def _confirmar_completar(self, popup, vars_tipos):
        """Cierra el popup y lanza la ejecución con los tipos seleccionados."""
        tipos = [nombre for nombre, var in vars_tipos.items() if var.get()]
        if not tipos:
            return
        popup.destroy()
        self._ejecutar_agente_planillas(tipos)

    # ═══════════════════════════════════════════════════════════════════
    # AGENTE 2: EJECUCIÓN EN HILO
    # ═══════════════════════════════════════════════════════════════════
    def _ejecutar_agente_planillas(self, tipos=None):
        if self.tarea_activa:
            messagebox.showwarning(
                "Agente ocupado",
                "Hay una tarea en ejecución. Espere a que finalice."
            )
            return
        self.tarea_activa = True
        self._tipos_planillas = tipos
        self._cancelar_tarea.clear()
        self.btn_ejecutar_planillas.configure(
            text="⏳  Analizando...", state="disabled",
            fg_color=Palette.ACCENT_DIM
        )
        self.progress_planillas.set(0)
        self.progress_planillas.configure(mode="indeterminate")
        self.progress_planillas.start()
        self._set_status("Analizando planillas del Escritorio...")

        # Limpiar pantalla antes de empezar
        self._limpiar_log()
        for row in self.tree_planillas.get_children():
            self.tree_planillas.delete(row)
        self.lbl_resumen_planillas.configure(text="Sin datos analizados")

        # Mostrar en log qué planillas se van a completar
        nombres = {"sobres": "SOBRES", "cobro": "COBRO", "pc": "PC"}
        seleccionadas = [nombres[t] for t in (tipos or ["sobres", "cobro", "pc"])]
        self.log_queue.put(f"[...] Planillas seleccionadas: {', '.join(seleccionadas)}")

        t = threading.Thread(target=self._agente_planillas_worker, daemon=True)
        t.start()

    def _agente_planillas_worker(self):
        """Ejecuta el agente de planillas en hilo de fondo."""
        self._set_log_panel("planillas")
        try:
            self._planillas_core(self._tipos_planillas)
        except Exception as e:
            self.after(0, lambda: self._planillas_error(str(e)))
        finally:
            self.after(0, self._planillas_done)

    def _buscar_buque_en_chofer(self, ruta_excel):
        """Busca 'BUQUE' en la hoja CHOFERES del Excel y devuelve el valor de la derecha.

        Soporta .xlsx (openpyxl) y .xls (xlrd). Case insensitive.
        Devuelve string vacío si no encuentra.
        """
        try:
            if str(ruta_excel).lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(ruta_excel, data_only=True)
                try:
                    pestana = next((s for s in wb.sheetnames if "CHOFER" in s.upper()), None)
                    if not pestana:
                        return ""
                    ws = wb[pestana]
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            val = ws.cell(row=r, column=c).value
                            if val and isinstance(val, str) and "BUQUE" in val.strip().upper():
                                nxt = ws.cell(row=r, column=c + 1).value
                                if nxt:
                                    return str(nxt).strip()
                finally:
                    wb.close()
            else:
                import xlrd
                book = xlrd.open_workbook(ruta_excel)
                pestana = next((s for s in book.sheet_names() if "CHOFER" in s.upper()), None)
                if not pestana:
                    return ""
                sheet = book.sheet_by_name(pestana)
                for r in range(sheet.nrows):
                    for c in range(sheet.ncols):
                        val = sheet.cell_value(r, c)
                        if val and isinstance(val, str) and "BUQUE" in val.strip().upper():
                            if c + 1 < sheet.ncols:
                                nxt = sheet.cell_value(r, c + 1)
                                if nxt:
                                    return str(nxt).strip()
        except Exception as e:
            self._log(f"  ⚠ Error buscando BUQUE en {os.path.basename(ruta_excel)}: {e}")
        return ""

    def _planillas_core(self, tipos=None):
        """Lógica real del agente de planillas (modificada para GUI)."""
        escritorio = self._resolver_ruta("planillas_carga", "Desktop")
        self._log(f"Escaneando: {escritorio}")

        # Tracking de resultados para el popup final
        resultados = {
            "sobres": {"ok": False, "detalle": "No procesado"},
            "cobro":  {"ok": False, "detalle": "No procesado"},
            "pc":     {"ok": False, "detalle": "No procesado"},
        }

        if not os.path.exists(escritorio):
            self._log("ERROR: No se pudo localizar la ruta del Escritorio.")
            self.after(0, lambda: self._mostrar_resumen(resultados))
            return

        # Buscar archivos en el pendrive (rutas configurables por tipo)
        ruta_sobres = buscar_archivo_en_pendrive(
            "SOBRES_2026.xlsx", self._cfg_obtener_rutas("sobres", os.path.join("TRABAJO", "01_PLANILLAS")))

        wb_sobres = None
        ws_sobres = None
        if ruta_sobres:
            try:
                wb_sobres = self._abrir_excel_seguro(ruta_sobres)
                meses_es = {
                    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
                }
                mes_actual = meses_es[datetime.now().month]
                for sheet in wb_sobres.sheetnames:
                    if mes_actual in sheet.upper():
                        ws_sobres = wb_sobres[sheet]
                        break
                if not ws_sobres:
                    self._log(f"ERROR: No se encontró hoja del mes '{mes_actual}' en SOBRES_2026.xlsx.")
                    self.after(0, lambda m=mes_actual: messagebox.showerror(
                        "Hoja del mes no encontrada",
                        f"No se encontró la hoja \"{m}\" en SOBRES_2026.xlsx.\n\n"
                        f"Verifique que el archivo contenga una hoja con el nombre del mes actual.\n\n"
                        f"Operación cancelada."
                    ))
                    wb_sobres.close()
                    return
                self._log(f"SOBRES_2026.xlsx detectada → {ruta_sobres} · hoja '{ws_sobres.title}'")
            except Exception as e:
                self._log(f"Error al abrir SOBRES_2026.xlsx: {e}")
        else:
            self._log("ERROR: No se encontró SOBRES_2026.xlsx.")
            pendrive_path = os.path.join("PENDRIVE:\\", "TRABAJO", "01_PLANILLAS", "SOBRES_2026.xlsx")
            self.after(0, lambda: messagebox.showwarning(
                "Planilla SOBRES no encontrada",
                f"No se encontró SOBRES_2026.xlsx.\n\n"
                f"Ruta esperada en el pendrive:\n"
                f"  {pendrive_path}\n\n"
                f"Conecte el pendrive y vuelva a ejecutar el análisis."
            ))
            # FRENAR: no seguir escaneando el Escritorio sin SOBRES
            return

        carpetas_encontradas = 0
        datos_extraidos = []

        for item in sorted(os.listdir(escritorio)):
            time.sleep(0.01) # Ceder el GIL
            ruta_carpeta = os.path.join(escritorio, item)
            if not os.path.isdir(ruta_carpeta):
                continue
            if item.startswith(".") or item.upper() in ("RECYCLED", "RECYCLER"):
                continue

            try:
                archivos_en_carpeta = os.listdir(ruta_carpeta)
            except Exception:
                continue

            # Filtrar excels de CONTENEDORES
            excels = []
            for archivo in archivos_en_carpeta:
                up = archivo.upper().strip()
                if archivo.startswith("~$"):
                    continue
                if (up.endswith(".XLSX") or up.endswith(".XLS")) and "CONTENEDORES" in up:
                    excels.append(archivo)

            if not excels:
                continue

            carpetas_encontradas += 1
            self._log(f"Carpeta localizada: \"{item}\"")

            match_frac = re.search(r"(F(?:RACCION)?\s*\d+)", item, re.IGNORECASE)
            frac_carpeta = match_frac.group(1).upper() if match_frac else "No detectada"

            match_pe_folder = re.search(r"(\d{3,4}[A-Z])", item, re.IGNORECASE)
            pe_carpeta = match_pe_folder.group(1).upper() if match_pe_folder else ""

            for archivo in excels:
                time.sleep(0.01)  # Ceder el GIL para evitar que la UI se congele
                ruta_excel = os.path.join(ruta_carpeta, archivo)
                self._log(f"  Leyendo: {archivo}")

                match_cant = re.match(r"^(\d+)", archivo.strip())
                cant_final = int(match_cant.group(1)) if match_cant else 1

                try:
                    if archivo.lower().endswith(".xlsx"):
                        datos = self._leer_xlsx_moderno(ruta_excel)
                    else:
                        datos = self._leer_xls_antiguo(ruta_excel)
                    f_celda, pe_celda, carp_celda, dest_celda, bl_celda, trans_final, primera_patente, precintos, guarda, puerto_salida, peso_flexi = datos

                    pe_origen = pe_celda if pe_celda else pe_carpeta
                    pe_recortado = pe_origen[-5:].lstrip("0") if pe_origen else "No detectado"
                    if not trans_final:
                        trans_final = "No detectado"

                    # Determinar tipo de transporte: primero desde Excel, fallback desde nombre de carpeta
                    tipo_transporte = self._clasificar_tipo_transporte(puerto_salida, peso_flexi)
                    if tipo_transporte == "TERRESTRE":
                        # Fallback: el nombre de carpeta ya tiene el tipo correcto (índice 4)
                        partes_carpeta = item.split("_")
                        if len(partes_carpeta) > 4 and partes_carpeta[4] in ("ISO", "FLEXI"):
                            tipo_transporte = partes_carpeta[4]
                    if tipo_transporte in ("ISO", "FLEXI"):
                        # Marítimo: buscar BUQUE en CHOFERES
                        bl_detectado = self._buscar_buque_en_chofer(ruta_excel)
                        bl_final = bl_detectado if bl_detectado else "-"
                    else:
                        bl_final = bl_celda if bl_celda else "-"

                    # Convertir carpeta a número si es posible (para evitar alerta de texto en Excel)
                    try:
                        carp_celda_num = int(carp_celda)
                    except (ValueError, TypeError):
                        carp_celda_num = carp_celda

                    registro = {
                        "f_celda": f_celda,
                        "cant_final": cant_final,
                        "pe_recortado": pe_recortado,
                        "pe_completo": pe_origen if pe_origen else "No detectado",
                        "carp_celda": carp_celda_num,
                        "tipo_transporte": "TERRES" if tipo_transporte == "TERRESTRE" else tipo_transporte,
                        "terminal": (puerto_salida[:9] if puerto_salida else "CHILE") if tipo_transporte in ("ISO", "FLEXI") else "CHILE",
                        "trans_final": trans_final,
                        "dest_celda": dest_celda,
                        "bl_final": bl_final,
                        "frac_carpeta": "-" if tipo_transporte in ("ISO", "FLEXI") else frac_carpeta,
                        "servicio": "-" if tipo_transporte in ("ISO", "FLEXI") else cant_final * int(self._cfg_obtener("valores", "ata_tares", 65000)),
                        "precintos": precintos,
                        "guarda": guarda if guarda else "No detectado",
                        "source_folder": item,
                        "source_file": archivo,
                    }
                    if "COMPARTIDO" in item.upper():
                        self._log(f"  🔀 Carpeta COMPARTIDO: {item}")
                    self.after(0, lambda r=registro: self._agregar_fila_planillas(r))

                    if ws_sobres is not None:
                        datos_extraidos.append(registro)

                except Exception as err:
                    self._log(f"  ERROR procesando {archivo}: {err}")

        # Guardar en SOBRES
        if datos_extraidos and ws_sobres is not None and (not tipos or "sobres" in tipos):
            self._log("Escribiendo datos en SOBRES_2026.xlsx...")
            try:

                def sort_date(item):
                    try:
                        p = item["f_celda"].split("/")
                        return datetime(int(p[2]), int(p[1]), int(p[0]))
                    except Exception:
                        return datetime.max

                datos_extraidos.sort(key=sort_date)

                def abreviar_dest(nombre):
                    n = (nombre or "").upper()
                    if "VITAPRO" in n: return "VITAPRO"
                    if "EWOS" in n: return "EWOS"
                    if "NUTRECO" in n: return "NUTRECO"
                    if "DICOAL" in n: return "DICOAL"
                    if "CARGILL" in n: return "CARGILL"
                    if "BIOMAR" in n: return "BIOMAR"
                    return nombre

                COLORES_EMPRESA = {
                    "VITAPRO": "7EB8E0",
                    "EWOS": "8CC56E", "DICOAL": "F4B183",
                    "NUTRECO": "FFEB9C", "CARGILL": "BDD7EE", "BIOMAR": "E2EFDA",
                }
                from openpyxl.styles import PatternFill, Border, Side, Alignment as Aln, Font

                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Aln(horizontal="center", vertical="center", wrap_text=True)
                font_normal = Font(name="Calibri", size=11)
                font_mediano = Font(name="Calibri", size=9)   # columnas H, I
                font_chico = Font(name="Calibri", size=8)     # columna G

                # Agrupar pares comparte: detectar por carpeta (ambas con "COMPARTIDO")
                def _pares_comparte(datos):
                    usados = set()
                    items = []
                    for i, d1 in enumerate(datos):
                        if i in usados: continue
                        f1 = d1.get("source_folder", "")
                        if "COMPARTIDO" not in f1.upper():
                            items.append(d1); usados.add(i); continue
                        pareja = None
                        for j, d2 in enumerate(datos):
                            if j == i or j in usados: continue
                            f2 = d2.get("source_folder", "")
                            if ("COMPARTIDO" in f2.upper()
                                    and d1.get("dest_celda") == d2.get("dest_celda")
                                    and d1.get("f_celda") == d2.get("f_celda")):
                                pareja = j; break
                        if pareja is not None:
                            items.append((d1, datos[pareja]))
                            usados.add(i); usados.add(pareja)
                        else:
                            items.append(d1); usados.add(i)
                    return items

                def _escribir_celda(ws, row, col, val, font, fill=None):
                    cell = ws.cell(row=row, column=col)
                    cell.value = val
                    cell.number_format = 'General'
                    cell.border = border
                    cell.alignment = center
                    cell.font = font
                    if fill:
                        cell.fill = fill

                def _escribir_fecha(ws, row, f_celda_str):
                    nueva_fila = row
                    while True:
                        time.sleep(0.005)
                        v1 = ws.cell(row=nueva_fila, column=1).value
                        v5 = ws.cell(row=nueva_fila, column=5).value
                        if not v1 and not v5:
                            break
                        nueva_fila += 1
                    if nueva_fila != row:
                        self._log(f"  → Fila ajustada de {row} a {nueva_fila}")
                    r_prev = nueva_fila - 1
                    val_prev = None
                    while r_prev >= 3:
                        val_prev = ws.cell(row=r_prev, column=1).value
                        if val_prev is not None:
                            break
                        r_prev -= 1
                    if val_prev and formatear_fecha_excel(val_prev) == f_celda_str:
                        for m_range in list(ws.merged_cells.ranges):
                            if m_range.min_col == 1 and m_range.max_col == 1:
                                if (m_range.min_row <= nueva_fila <= m_range.max_row
                                        or m_range.min_row <= r_prev <= m_range.max_row):
                                    ws.merged_cells.ranges.remove(m_range)
                        ws.merge_cells(start_row=r_prev, start_column=1, end_row=nueva_fila, end_column=1)
                        fecha_merge = ws.cell(row=r_prev, column=1)
                        fecha_merge.number_format = 'General'
                        fecha_merge.alignment = Aln(horizontal="center", vertical="center")
                    else:
                        ws.cell(row=nueva_fila, column=1).value = f_celda_str
                    c1 = ws.cell(row=nueva_fila, column=1)
                    c1.number_format = 'General'
                    c1.border = border
                    c1.alignment = Aln(horizontal="center", vertical="center", wrap_text=True)
                    c1.font = font_normal
                    return nueva_fila

                items = _pares_comparte(datos_extraidos)
                for item in items:
                    es_par = isinstance(item, tuple) and len(item) == 2
                    if es_par:
                        d1, d2 = item
                        self._log(f"  🔀 Par COMPARTE: {d1.get('pe_recortado')} + {d2.get('pe_recortado')}")
                    else:
                        d1 = item
                        d2 = None

                    f_celda = d1["f_celda"]
                    dest_abrev = abreviar_dest(d1["dest_celda"])
                    color_hex = COLORES_EMPRESA.get(dest_abrev, "FFFFFF")
                    fill_color = PatternFill(fill_type="solid", fgColor=color_hex)

                    tipo = d1.get("tipo_transporte", "TERRESTRE")
                    es_maritimo = tipo in ("ISO", "FLEXI")
                    valores_check_1 = {
                        1: f_celda, 2: d1["cant_final"], 3: tipo, 4: d1["pe_recortado"],
                        5: d1["carp_celda"], 6: d1.get("terminal", "CHILE"), 7: d1["trans_final"], 8: dest_abrev,
                        9: d1["bl_final"], 10: "-" if es_maritimo else d1["frac_carpeta"],
                        11: "-" if es_maritimo else d1["cant_final"] * int(self._cfg_obtener("valores", "ata_tares", 65000)),
                    }
                    if ya_existe_en_hoja(ws_sobres, valores_check_1, excluir_columnas={9}):
                        self._log(f"  Omitido (ya existe): Carpeta {d1.get('carp_celda','?')} | P.E. {d1.get('pe_recortado','?')} | {f_celda}")
                        continue

                    r1 = _escribir_fecha(ws_sobres, 3, f_celda)
                    for col in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11):
                        val = valores_check_1[col]
                        font = font_chico if col == 7 else (font_mediano if col in (6, 8, 9) else font_normal)
                        fl = fill_color if col == 8 else None
                        _escribir_celda(ws_sobres, r1, col, val, font, fl)

                    if es_par:
                        dest2 = abreviar_dest(d2["dest_celda"])
                        color2 = COLORES_EMPRESA.get(dest2, "FFFFFF")
                        fill2 = PatternFill(fill_type="solid", fgColor=color2)
                        tipo2 = d2.get("tipo_transporte", "TERRESTRE")
                        es_maritimo2 = tipo2 in ("ISO", "FLEXI")
                        valores_check_2 = {
                            4: d2["pe_recortado"], 5: d2["carp_celda"],
                            9: d2["bl_final"], 10: "-" if es_maritimo2 else d2["frac_carpeta"],
                        }
                        r2 = r1 + 1
                        # Fecha (col 1) merge con r1
                        for m_range in list(ws_sobres.merged_cells.ranges):
                            if m_range.min_col == 1 and m_range.max_col == 1:
                                if m_range.min_row <= r2 <= m_range.max_row or m_range.min_row <= r1 <= m_range.max_row:
                                    ws_sobres.merged_cells.ranges.remove(m_range)
                        ws_sobres.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
                        ws_sobres.cell(row=r1, column=1).alignment = Aln(horizontal="center", vertical="center")

                        # Columnas individuales para fila 2
                        for col, val in valores_check_2.items():
                            font = font_chico if col == 7 else (font_mediano if col in (6, 8, 9) else font_normal)
                            fl = fill2 if col == 8 else None
                            _escribir_celda(ws_sobres, r2, col, val, font, fl)
                        # Cols compartidas (merge) para fila 2: B,C,F,G,H,K
                        for col in (2, 3, 6, 7, 8, 11):
                            val = valores_check_1[col]
                            font = font_chico if col == 7 else (font_mediano if col in (6, 8) else font_normal)
                            _escribir_celda(ws_sobres, r2, col, val, font)
                        # Merge vertical para B(2), C(3), F(6), G(7), H(8), K(11)
                        for col in (2, 3, 6, 7, 8, 11):
                            ws_sobres.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                            ws_sobres.cell(row=r1, column=col).alignment = center

                        self._log(f"Dato copiado en SOBRES_2026 (Filas {r1}-{r2}) → Par COMPARTE")
                    else:
                        self._log(f"Dato copiado en SOBRES_2026 (Fila {r1})")

                # Auto-ajustar alto de filas para que el texto entre
                for row_idx in range(1, ws_sobres.max_row + 1):
                    ws_sobres.row_dimensions[row_idx].height = 20

                self._guardar_excel_seguro(wb_sobres, ruta_sobres)

                # Post-procesado: completar B/L faltantes buscando por Carpeta
                bls_completados = self._completar_bl_por_carpeta(wb_sobres, ws_sobres)
                if bls_completados > 0:
                    self._guardar_excel_seguro(wb_sobres, ruta_sobres)
                    self._log(f"  → {bls_completados} B/L completados por coincidencia de Carpeta.")

                self._log("COMPLETADO: SOBRES_2026.xlsx guardada correctamente.")
                resultados["sobres"] = {"ok": True, "detalle": f"{len(datos_extraidos)} registros cargados"}
            except Exception as e:
                self._log(f"Error al guardar SOBRES_2026.xlsx: {e}")
                resultados["sobres"] = {"ok": False, "detalle": f"Error: {str(e)[:60]}"}

        if carpetas_encontradas == 0:
            self._log("No se detectaron carpetas con archivos de CONTENEDORES.")

        self.datos_planillas = datos_extraidos
        self._log(f"Escaneo finalizado. {len(datos_extraidos)} registros extraídos.")

        # ═══════════════════════════════════════════════════════════════
        # COMPLETAR PLANILLA DE COBRO
        # ═══════════════════════════════════════════════════════════════
        if datos_extraidos and (not tipos or "cobro" in tipos):
            resultados["cobro"] = self._completar_cobro(datos_extraidos)

        # ═══════════════════════════════════════════════════════════════
        # COMPLETAR PLANILLA PC (PRECINTOS/CABLES)
        # ═══════════════════════════════════════════════════════════════
        if datos_extraidos and (not tipos or "pc" in tipos):
            resultados["pc"] = self._completar_pc(datos_extraidos)

        # Guardar resultados para el popup (se muestra en el hilo principal)
        self._resultados_pendientes = resultados

    def _completar_cobro(self, datos_extraidos):
        """Llena la planilla COBRO_2026.xlsx con los datos extraídos."""
        ruta_cobro = buscar_archivo_en_pendrive(
            "COBRO_2026.xlsx", self._cfg_obtener_rutas("cobro", os.path.join("TRABAJO", "01_PLANILLAS")))

        if not ruta_cobro:
            self._log("ERROR: No se encontró COBRO_2026.xlsx.")
            pendrive_path = os.path.join("PENDRIVE:\\", "TRABAJO", "01_PLANILLAS", "COBRO_2026.xlsx")
            self.after(0, lambda: messagebox.showwarning(
                "Planilla COBRO no encontrada",
                f"No se encontró COBRO_2026.xlsx.\n\n"
                f"Ruta esperada en el pendrive:\n"
                f"  {pendrive_path}\n\n"
                f"Conecte el pendrive y vuelva a ejecutar el análisis."
            ))
            return {"ok": False, "detalle": "Archivo no encontrado"}

        wb_cobro = None
        ws_cobro = None
        try:
            wb_cobro = self._abrir_excel_seguro(ruta_cobro)
            meses_es = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
            }
            mes_actual = meses_es[datetime.now().month]
            for sheet in wb_cobro.sheetnames:
                if mes_actual in sheet.upper():
                    ws_cobro = wb_cobro[sheet]
                    break
            if not ws_cobro:
                self._log(f"ERROR: No se encontró hoja del mes '{mes_actual}' en COBRO_2026.xlsx.")
                self.after(0, lambda m=mes_actual: messagebox.showerror(
                    "Hoja del mes no encontrada",
                    f"No se encontró la hoja \"{m}\" en COBRO_2026.xlsx.\n\n"
                    f"Verifique que el archivo contenga una hoja con el nombre del mes actual.\n\n"
                    f"Operación cancelada."
                ))
                wb_cobro.close()
                return {"ok": False, "detalle": f"Hoja '{mes_actual}' no encontrada"}
            self._log(f"COBRO_2026.xlsx detectada → hoja '{ws_cobro.title}'")
        except Exception as e:
            self._log(f"ERROR al abrir COBRO_2026.xlsx: {e}")
            self.after(0, lambda: messagebox.showerror(
                "Error al abrir COBRO",
                f"No se pudo abrir COBRO_2026.xlsx:\n{e}\n\n"
                f"Verifique que el archivo no esté abierto por otro programa."
            ))
            if wb_cobro:
                try:
                    wb_cobro.close()
                except Exception:
                    pass
            return {"ok": False, "detalle": f"Error al abrir: {str(e)[:40]}"}

        try:
            from openpyxl.styles import Border, Side, Alignment as Aln
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Aln(horizontal="center", vertical="center")

            # Ordenar por fecha
            def sort_date(item):
                try:
                    p = item["f_celda"].split("/")
                    return datetime(int(p[2]), int(p[1]), int(p[0]))
                except Exception:
                    return datetime.max

            datos_ordenados = sorted(datos_extraidos, key=sort_date)

            # Pre-escanear fechas que YA existen en COBRO
            fechas_existentes = set()
            for row in range(3, ws_cobro.max_row + 1):
                val_fecha = ws_cobro.cell(row=row, column=1).value
                if val_fecha:
                    fechas_existentes.add(formatear_fecha_excel(val_fecha))

            fecha_anterior = None

            # Agrupar pares comparte por carpeta (ambas con "COMPARTIDO")
            def _pares_cobro(datos):
                usados = set()
                items = []
                for i, d1 in enumerate(datos):
                    if i in usados: continue
                    f1 = d1.get("source_folder", "")
                    if "COMPARTIDO" not in f1.upper():
                        items.append(d1); usados.add(i); continue
                    pareja = None
                    for j, d2 in enumerate(datos):
                        if j == i or j in usados: continue
                        f2 = d2.get("source_folder", "")
                        if ("COMPARTIDO" in f2.upper()
                                and d1.get("dest_celda") == d2.get("dest_celda")
                                and d1.get("f_celda") == d2.get("f_celda")):
                            pareja = j; break
                    if pareja is not None:
                        items.append((d1, datos[pareja]))
                        usados.add(i); usados.add(pareja)
                    else:
                        items.append(d1); usados.add(i)
                return items

            items = _pares_cobro(datos_ordenados)
            for item in items:
                es_par = isinstance(item, tuple) and len(item) == 2
                if es_par:
                    d1, d2 = item
                    self._log(f"  🔀 Par COMPARTE COBRO: {d1.get('pe_recortado')} + {d2.get('pe_recortado')}")
                else:
                    d1 = item
                    d2 = None

                f_celda = d1["f_celda"]
                if f_celda != fecha_anterior:
                    fecha_anterior = f_celda
                    es_primero_del_dia = f_celda not in fechas_existentes
                else:
                    es_primero_del_dia = False

                precio_base = int(self._cfg_obtener("valores", "precio_carpeta", 49000))
                precio_carpeta = precio_base if es_primero_del_dia else precio_base // 2
                tipo_cobro = d1.get("tipo_transporte", "TERRESTRE")
                es_maritimo_cobro = tipo_cobro in ("ISO", "FLEXI")
                servicio_ata = "-" if es_maritimo_cobro else d1["cant_final"] * int(self._cfg_obtener("valores", "ata_tares", 65000))

                valores_fila = {
                    1: f_celda, 2: d1["cant_final"], 3: tipo_cobro, 4: d1["pe_recortado"],
                    5: d1["carp_celda"], 6: "-" if es_maritimo_cobro else d1["frac_carpeta"],
                    7: servicio_ata, 8: precio_carpeta,
                }
                valores_check = {c: v for c, v in valores_fila.items() if c != 8}
                if ya_existe_en_hoja(ws_cobro, valores_check):
                    self._log(f"  Omitido (ya existe en COBRO): Carpeta {d1.get('carp_celda','?')} | {f_celda}")
                    continue

                fechas_existentes.add(f_celda)
                r1 = primera_fila_libre(ws_cobro)

                for col, val in valores_fila.items():
                    cell = ws_cobro.cell(row=r1, column=col)
                    cell.value = val
                    cell.border = border
                    cell.alignment = center

                if es_par:
                    # Segunda fila con datos del otro Contenedores
                    precio_carpeta_2 = precio_base // 2
                    tipo2_cobro = d2.get("tipo_transporte", "TERRESTRE")
                    es_maritimo2_cobro = tipo2_cobro in ("ISO", "FLEXI")
                    servicio_ata_2 = "-" if es_maritimo2_cobro else d2["cant_final"] * int(self._cfg_obtener("valores", "ata_tares", 65000))
                    r2 = r1 + 1

                    vals2 = {1: f_celda, 2: d2["cant_final"], 3: tipo2_cobro, 4: d2["pe_recortado"],
                             5: d2["carp_celda"], 6: "-" if es_maritimo2_cobro else d2["frac_carpeta"],
                             7: servicio_ata_2, 8: precio_carpeta_2}
                    for col, val in vals2.items():
                        cell = ws_cobro.cell(row=r2, column=col)
                        cell.value = val
                        cell.border = border
                        cell.alignment = center

                    # Merge fecha (col 1)
                    ws_cobro.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
                    ws_cobro.cell(row=r1, column=1).alignment = center
                    # Merge Cantidad (col 2), Tipo (col 3) y Servicio ATA (col 7)
                    for col in (2, 3, 7):
                        ws_cobro.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                        ws_cobro.cell(row=r1, column=col).alignment = center

                    self._log(f"Dato copiado en COBRO_2026 (Filas {r1}-{r2}) → Par COMPARTE")
                else:
                    # Merge de fecha normal
                    if not es_primero_del_dia:
                        r_prev = r1 - 1
                        val_prev = None
                        while r_prev >= 3:
                            val_prev = ws_cobro.cell(row=r_prev, column=1).value
                            if val_prev is not None:
                                break
                            r_prev -= 1
                        if val_prev and formatear_fecha_excel(val_prev) == f_celda:
                            for m_range in list(ws_cobro.merged_cells.ranges):
                                if m_range.min_col == 1 and m_range.max_col == 1:
                                    if (m_range.min_row <= r1 <= m_range.max_row
                                            or m_range.min_row <= r_prev <= m_range.max_row):
                                        ws_cobro.merged_cells.ranges.remove(m_range)
                            ws_cobro.merge_cells(start_row=r_prev, start_column=1, end_row=r1, end_column=1)
                            ws_cobro.cell(row=r_prev, column=1).alignment = center

                    self._log(f"Dato copiado en COBRO_2026 (Fila {r1}) → Carpetas: ${precio_carpeta:,}")

            self._guardar_excel_seguro(wb_cobro, ruta_cobro)
            self._log("COMPLETADO: COBRO_2026.xlsx guardada correctamente.")
            return {"ok": True, "detalle": f"{len(datos_extraidos)} registros cargados"}
        except Exception as e:
            self._log(f"ERROR al guardar COBRO_2026.xlsx: {e}")
            self.after(0, lambda: messagebox.showerror(
                "Error al guardar COBRO",
                f"No se pudo guardar COBRO_2026.xlsx:\n{e}\n\n"
                f"Verifique que el archivo no esté abierto por otro programa."
            ))
            return {"ok": False, "detalle": f"Error: {str(e)[:60]}"}
        finally:
            if wb_cobro:
                try:
                    wb_cobro.close()
                except Exception:
                    pass

    # ── Precintos disponibles ──────────────────────────────────────────
    def _contar_precintos_disponibles(self):
        """Cuenta precintos SIN ASIGNAR en PC.xlsx (col 1 llena, col 2 vacía).
        Retorna el número o None si no encuentra la planilla."""
        ruta_pc = buscar_archivo_en_pendrive(
            "PC.xlsx", self._cfg_obtener_rutas("pc", os.path.join("TRABAJO", "01_PLANILLAS")))
        if not ruta_pc:
            ruta_pc = buscar_archivo_en_pendrive(
                "PC_2026.xlsx", self._cfg_obtener_rutas("pc", os.path.join("TRABAJO", "01_PLANILLAS")))
        if not ruta_pc:
            return None
        try:
            wb = openpyxl.load_workbook(ruta_pc)
            try:
                ws = wb["SIN ASIGNACION"]
            except KeyError:
                ws = wb.active
            disponibles = 0
            for fila in range(2, ws.max_row + 1):
                nro = ws.cell(row=fila, column=1).value
                if nro and not ws.cell(row=fila, column=2).value:
                    disponibles += 1
            wb.close()
            return disponibles
        except Exception:
            return None

    def _actualizar_titulo_precintos(self):
        """Actualiza lbl_titulo_panel con el conteo de precintos disponibles."""
        disponibles = self._contar_precintos_disponibles()
        if disponibles is None:
            self.lbl_titulo_panel.configure(
                text="Completar Planillas — No se encontró planilla PC")
        else:
            self.lbl_titulo_panel.configure(
                text=f"Completar Planillas — Precintos disponibles: {disponibles}")

    def _completar_pc(self, datos_extraidos):
        """Llena la planilla PC (Precintos/Cables) en la hoja SIN ASIGNACION.
        Busca cada precinto de abajo hacia arriba y completa: Fecha, P.E., Guarda, Carpeta."""
        ruta_pc = buscar_archivo_en_pendrive(
            "PC.xlsx", self._cfg_obtener_rutas("pc", os.path.join("TRABAJO", "01_PLANILLAS")))
        if not ruta_pc:
            ruta_pc = buscar_archivo_en_pendrive(
                "PC_2026.xlsx", self._cfg_obtener_rutas("pc", os.path.join("TRABAJO", "01_PLANILLAS")))
        if not ruta_pc:
            self._log("ERROR: No se encontró PC.xlsx.")
            pendrive_path = os.path.join("PENDRIVE:\\", "TRABAJO", "01_PLANILLAS", "PC.xlsx")
            self.after(0, lambda: messagebox.showwarning(
                "Planilla PC no encontrada",
                f"No se encontró PC.xlsx.\n\n"
                f"Ruta esperada en el pendrive:\n"
                f"  {pendrive_path}\n\n"
                f"Conecte el pendrive y vuelva a ejecutar el análisis."
            ))
            return {"ok": False, "detalle": "Archivo no encontrado"}

        wb_pc = None
        ws_pc = None
        try:
            wb_pc = self._abrir_excel_seguro(ruta_pc)
            # Buscar hoja SIN ASIGNACION
            for sheet in wb_pc.sheetnames:
                if "SIN ASIGNACION" in sheet.upper() or "SINASIGNACION" in sheet.upper():
                    ws_pc = wb_pc[sheet]
                    break
            if not ws_pc:
                # Intentar la primera hoja como fallback
                ws_pc = wb_pc.active
                self._log(f"PC.xlsx: hoja 'SIN ASIGNACION' no encontrada, usando '{ws_pc.title}'")
            else:
                self._log(f"PC.xlsx detectada → hoja '{ws_pc.title}'")
        except Exception as e:
            self._log(f"ERROR al abrir PC.xlsx: {e}")
            self.after(0, lambda: messagebox.showerror(
                "Error al abrir PC",
                f"No se pudo abrir PC.xlsx:\n{e}"
            ))
            if wb_pc:
                try:
                    wb_pc.close()
                except Exception:
                    pass
            return {"ok": False, "detalle": f"Error al abrir: {str(e)[:40]}"}

        try:
            from openpyxl.styles import PatternFill, Border, Side, Alignment as Aln

            # Color de resalte para precintos encontrados (dorado claro)
            fill_encontrado = PatternFill(fill_type="solid", fgColor="FFD966")
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Aln(horizontal="center", vertical="center")

            # Agrupar pares comparte por carpeta (misma lógica que SOBRES/COBRO)
            def _pares_pc(datos):
                usados = set()
                items = []
                for i, d1 in enumerate(datos):
                    if i in usados: continue
                    f1 = d1.get("source_folder", "")
                    if "COMPARTIDO" not in f1.upper():
                        items.append(d1); usados.add(i); continue
                    pareja = None
                    for j, d2 in enumerate(datos):
                        if j == i or j in usados: continue
                        f2 = d2.get("source_folder", "")
                        if ("COMPARTIDO" in f2.upper()
                                and d1.get("dest_celda") == d2.get("dest_celda")
                                and d1.get("f_celda") == d2.get("f_celda")):
                            pareja = j; break
                    if pareja is not None:
                        items.append((d1, datos[pareja]))
                        usados.add(i); usados.add(pareja)
                    else:
                        items.append(d1); usados.add(i)
                return items

            # Recolectar precintos, manejando pares comparte
            todos_precintos = []
            items_pc = _pares_pc(datos_extraidos)
            for item in items_pc:
                es_par = isinstance(item, tuple) and len(item) == 2
                if es_par:
                    d1, d2 = item
                    # d1 = comparte, d2 = cierra (o viceversa). Usar precintos del que NO tiene CERRADO? No, del que cierra.
                    # Identificar cuál es cierra: el que tiene "CERRADO" en source_folder
                    d_cierra = d1 if "CERRADO" in d1.get("source_folder", "").upper() else d2
                    d_comparte = d1 if d_cierra is d2 else d2

                    pe_cierra = d_cierra.get("pe_completo", d_cierra.get("pe_recortado", ""))
                    pe_comparte = d_comparte.get("pe_completo", d_comparte.get("pe_recortado", ""))
                    pe_comparte_5 = pe_comparte.strip()[-5:] if pe_comparte else ""
                    pe_combinado = f"{pe_cierra}/{pe_comparte_5}" if pe_cierra and pe_comparte_5 else pe_cierra
                    carp_combinado = f"{d_cierra.get('carp_celda','')}/{d_comparte.get('carp_celda','')}"

                    # Crear un dato combinado con los precintos del cierra
                    dato_combinado = dict(d_cierra)
                    dato_combinado["pe_completo"] = pe_combinado
                    dato_combinado["carp_celda"] = carp_combinado
                    dato_combinado["_es_par_comparte"] = True

                    for p in d_cierra.get("precintos", []):
                        if p:
                            todos_precintos.append((p, dato_combinado))
                    self._log(f"  PC Par COMPARTE: PE={pe_combinado} | Carpeta={carp_combinado}")
                else:
                    for p in item.get("precintos", []):
                        if p:
                            todos_precintos.append((p, item))

            if not todos_precintos:
                self._log("PC: No se encontraron precintos en los archivos de carga.")
                return {"ok": False, "detalle": "Sin precintos en archivos"}

            encontrados = 0

            for precinto, dato in todos_precintos:
                # Buscar de abajo hacia arriba (pocos sin asignar al final)
                encontrado = False
                for row in range(ws_pc.max_row, 0, -1):
                    val_celda = ws_pc.cell(row=row, column=1).value
                    if val_celda and str(val_celda).strip() == precinto:
                        # Completar datos: Fecha, P.E., Guarda, Carpeta (columnas 2-5)
                        ws_pc.cell(row=row, column=2).value = dato["f_celda"]
                        ws_pc.cell(row=row, column=3).value = dato.get("pe_completo", dato["pe_recortado"])
                        ws_pc.cell(row=row, column=4).value = dato.get("guarda", "")
                        ws_pc.cell(row=row, column=5).value = dato["carp_celda"]

                        # Aplicar formato y color a toda la fila
                        for c in range(1, 6):
                            cell = ws_pc.cell(row=row, column=c)
                            cell.fill = fill_encontrado
                            cell.border = border
                            cell.alignment = center

                        encontrados += 1
                        encontrado = True
                        self._log(f"  PC: Precinto {precinto} → Fila {row} | P.E. {dato.get('pe_completo', dato['pe_recortado'])} | Carpeta {dato['carp_celda']} | Guarda {dato.get('guarda','')}")
                        break

                # Si no se encontró, simplemente seguimos sin loguear (ya fue asignado antes)

            # Auto-ajustar alto de filas
            for row_idx in range(1, ws_pc.max_row + 1):
                ws_pc.row_dimensions[row_idx].height = 20
            self._guardar_excel_seguro(wb_pc, ruta_pc)
            self._log(f"COMPLETADO: PC.xlsx — {encontrados} precintos asignados.")
            return {"ok": True, "detalle": f"{encontrados} precintos asignados"}
        except Exception as e:
            self._log(f"ERROR al guardar PC.xlsx: {e}")
            self.after(0, lambda: messagebox.showerror(
                "Error al guardar PC",
                f"No se pudo guardar PC.xlsx:\n{e}\n\n"
                f"Verifique que el archivo no esté abierto por otro programa."
            ))
            return {"ok": False, "detalle": f"Error: {str(e)[:60]}"}
        finally:
            if wb_pc:
                try:
                    wb_pc.close()
                except Exception:
                    pass

    def _leer_xls_antiguo(self, ruta_excel):
        book = xlrd.open_workbook(ruta_excel)
        sheet_name = next(
            (s for s in book.sheet_names() if "CHOFER" in s.upper()), None
        )
        if not sheet_name:
            raise ValueError(f"Falta pestaña 'CHOFER'. Hojas: {book.sheet_names()}")
        sheet = book.sheet_by_name(sheet_name)

        f_celda = pe_celda = carp_celda = dest_celda = bl_celda = ""
        transporte_encontrado = primera_patente = guarda = ""
        puerto_salida = peso_flexi = ""
        col_patente = None

        for r in range(sheet.nrows):
            time.sleep(0.01) # Ceder GIL
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val and isinstance(val, str):
                    val_up = val.strip().upper()
                    if "FECHA CARGA" in val_up or "FECHA DE CARGA" in val_up or val_up == "FECHA":
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt and not f_celda:
                            if sheet.cell_type(r, c + 1) == xlrd.XL_CELL_DATE:
                                nxt = xlrd.xldate_as_datetime(nxt, book.datemode)
                            f_celda = formatear_fecha_excel(nxt)
                    elif val_up in ("P.E", "P.E.", "PE", "PERMISO"):
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt:
                            pe_celda = str(nxt).strip()
                    elif "CARPETA" in val_up or val_up == "CARP.":
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt and not carp_celda:
                            carp_celda = str(int(nxt)) if isinstance(nxt, float) else str(nxt).strip()
                    elif "DESTINATARIO" in val_up or "CLIENTE" in val_up:
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt and not dest_celda:
                            dest_celda = str(nxt).strip().upper()
                    elif val_up in ("BL", "B/L", "CONOCIMIENTO"):
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt and not bl_celda:
                            bl_celda = str(nxt).strip()
                    elif "PUERTO SALIDA" in val_up or "P. SALIDA" in val_up:
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt and not puerto_salida:
                            puerto_salida = str(nxt).strip()
                    elif "PESO FLEXI" in val_up:
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt != "" and nxt is not None and not peso_flexi:
                            # xlrd: float → int si es entero, sino string
                            if isinstance(nxt, float):
                                peso_flexi = str(int(nxt)) if nxt.is_integer() else str(nxt)
                            else:
                                peso_flexi = str(nxt).strip()
                    if "PATENTE" in val_up:
                        col_patente = c
                    # La columna de precintos no se busca más por etiqueta, se toma directo de col 1
                    if "TRANSPORT" in val_up or "EMPRESA" in val_up:
                        # Puede haber hasta 4 transportes en grilla 2x2 a la derecha
                        transportes = []
                        for dr in (0, 1):
                            for dc in (1, 2):
                                if r + dr < sheet.nrows and c + dc < sheet.ncols:
                                    val_t = sheet.cell_value(r + dr, c + dc)
                                    if val_t:
                                        transportes.append(str(val_t).strip())
                        # Eliminar duplicados preservando orden
                        unicos = list(dict.fromkeys(transportes))
                        transporte_encontrado = " - ".join(unicos) if unicos else ""
                    if "GUARDA" in val_up or "CHOFER" in val_up:
                        if c + 1 < sheet.ncols:
                            val_g = sheet.cell_value(r, c + 1)
                            if val_g:
                                guarda = str(val_g).strip()
                    if val_up == "GUARDA" and not guarda:
                        nxt = sheet.cell_value(r, c + 1) if c + 1 < sheet.ncols else None
                        if nxt:
                            guarda = str(nxt).strip()

        if col_patente is not None:
            for row_p in range(sheet.nrows):
                val_p = sheet.cell_value(row_p, col_patente)
                if val_p and not any(
                    k in str(val_p).upper() for k in ("PATENTE", "TOTAL", "CHOFER")
                ):
                    primera_patente = str(val_p).strip()
                    break

        # Precintos: SIEMPRE columna 1 (A), desde fila 2 (xlrd: row>=1)
        precintos = []
        for row_pr in range(1, sheet.nrows):
            val_pr = sheet.cell_value(row_pr, 0)  # columna 0 = A
            if val_pr:
                val_str = str(val_pr).strip()
                palabras = val_str.split()
                es_nombre_largo = len(palabras) >= 3 or len(val_str) > 30
                if len(val_str) >= 2 and val_str != "-" and not es_nombre_largo:
                    # Blacklist de palabras que son etiquetas/headers, no precintos reales
                    labels_prohibidas = (
                        "PRECINTO", "SELLO", "TOTAL", "CHOFER", "GUARDA", "ADUANA",
                        "PUERTO", "CARPETA", "BOOKING", "TRANSPORTE", "PESO", "FLEXI",
                        "FECHA", "DESTINATARIO", "TERMINAL", "PATENTE", "PERMISO",
                        "EMPRESA", "SALIDA", "CARGA", "EMBARQUE", "CONTENEDOR",
                    )
                    if not any(k in val_str.upper() for k in labels_prohibidas):
                        # Separar precintos múltiples unidos por guion (ej: "12345-67890")
                        for parte in val_str.split("-"):
                            parte = parte.strip()
                            if parte:
                                precintos.append(parte)
        # Si no se encontró B/L, buscar por número de carpeta en la misma hoja
        if not bl_celda and carp_celda:
            bl_celda = buscar_bl_por_carpeta_xls(sheet, carp_celda)
        return f_celda, pe_celda, carp_celda, dest_celda, bl_celda, transporte_encontrado, primera_patente, precintos, guarda, puerto_salida, peso_flexi

    def _completar_bl_por_carpeta(self, wb, ws_actual):
        """Recorre las filas de la hoja SOBRES. Si B/L (col 9) es '-', busca
           en filas de arriba (misma hoja) u hoja anterior otra fila con la
           misma Carpeta (col 5) que tenga B/L válido, y lo copia.
           Omite filas con Fracción F1 (col 10), que no llevan B/L."""
        def normalizar(val):
            if val is None:
                return ""
            if isinstance(val, float):
                if val == int(val):
                    return str(int(val))
                return str(val).strip()
            s = str(val).strip().lstrip("'")
            return s

        completados = 0
        nombre_actual = ws_actual.title
        idx_actual = wb.sheetnames.index(nombre_actual)

        def recolectar_datos(ws):
            filas = []  # [(r, carpeta, bl, fraccion)]
            for r in range(3, ws.max_row + 1):
                carp = ws.cell(row=r, column=5).value
                bl = ws.cell(row=r, column=9).value
                if carp is not None:
                    frac = normalizar(ws.cell(row=r, column=10).value)
                    filas.append((r, normalizar(carp), normalizar(bl), frac))
            return filas

        filas_actual = recolectar_datos(ws_actual)

        filas_anterior = []
        if idx_actual > 0:
            ws_ant = wb[wb.sheetnames[idx_actual - 1]]
            filas_anterior = recolectar_datos(ws_ant)

        for r, carpeta, bl, fraccion in filas_actual:
            if bl and bl not in ("-", "—"):
                continue
            if not carpeta or carpeta == "0":
                continue
            # Fracción F1 no lleva B/L
            if fraccion.upper() == "F1":
                continue

            bl_encontrado = ""
            for r2, carp2, bl2, _ in filas_actual:
                if r2 >= r:
                    continue
                if carp2 == carpeta and bl2 and bl2 not in ("-", "—"):
                    bl_encontrado = bl2
                    break

            if not bl_encontrado and filas_anterior:
                for r2, carp2, bl2, _ in filas_anterior:
                    if carp2 == carpeta and bl2 and bl2 not in ("-", "—"):
                        bl_encontrado = bl2
                        break

            if bl_encontrado:
                from openpyxl.styles import Border, Side, Alignment as Aln, Font
                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Aln(horizontal="center", vertical="center", wrap_text=True)
                font_bl = Font(name="Calibri", size=9)
                cell_bl = ws_actual.cell(row=r, column=9)
                cell_bl.value = bl_encontrado
                cell_bl.number_format = 'General'
                cell_bl.border = border
                cell_bl.alignment = center
                cell_bl.font = font_bl
                completados += 1

        if completados > 0:
            self._log(f"  → {completados} B/L completados por coincidencia de Carpeta.")
        return completados

    def _leer_xlsx_moderno(self, ruta_excel):
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        pestana_chofer = next(
            (s for s in wb.sheetnames if "CHOFER" in s.upper()), None
        )
        if not pestana_chofer:
            wb.close()
            raise ValueError(f"Falta pestaña 'CHOFER'. Hojas: {wb.sheetnames}")
        ws = wb[pestana_chofer]

        f_celda = pe_celda = carp_celda = dest_celda = bl_celda = ""
        transporte_encontrado = primera_patente = guarda = ""
        puerto_salida = peso_flexi = ""
        col_patente = None

        for r in range(1, ws.max_row + 1):
            time.sleep(0.01) # Ceder GIL
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    val_up = val.strip().upper()
                    if "FECHA CARGA" in val_up or "FECHA DE CARGA" in val_up or val_up == "FECHA":
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt and not f_celda:
                            f_celda = formatear_fecha_excel(nxt)
                    elif val_up in ("P.E", "P.E.", "PE", "PERMISO"):
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt:
                            pe_celda = str(nxt).strip()
                    elif "CARPETA" in val_up or val_up == "CARP.":
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt and not carp_celda:
                            carp_celda = str(nxt).strip()
                    elif "DESTINATARIO" in val_up or "CLIENTE" in val_up:
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt and not dest_celda:
                            dest_celda = str(nxt).strip().upper()
                    elif val_up in ("BL", "B/L", "CONOCIMIENTO"):
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt and not bl_celda:
                            bl_celda = str(nxt).strip()
                    elif "PUERTO SALIDA" in val_up or "P. SALIDA" in val_up:
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt and not puerto_salida:
                            puerto_salida = str(nxt).strip()
                    elif "PESO FLEXI" in val_up:
                        nxt = ws.cell(row=r, column=c + 1).value
                        if nxt != "" and nxt is not None and not peso_flexi:
                            # openpyxl: float → int si es entero, sino string
                            if isinstance(nxt, float):
                                peso_flexi = str(int(nxt)) if nxt.is_integer() else str(nxt)
                            else:
                                peso_flexi = str(nxt).strip()
                    if "PATENTE" in val_up:
                        col_patente = c
                    # La columna de precintos no se busca más por etiqueta, se toma directo de col 1
                    if "TRANSPORT" in val_up or "EMPRESA" in val_up:
                        # Puede haber hasta 4 transportes en grilla 2x2 a la derecha
                        transportes = []
                        for dr in (0, 1):
                            for dc in (1, 2):
                                val_t = ws.cell(row=r + dr, column=c + dc).value
                                if val_t:
                                    transportes.append(str(val_t).strip())
                        # Eliminar duplicados preservando orden
                        unicos = list(dict.fromkeys(transportes))
                        transporte_encontrado = " - ".join(unicos) if unicos else ""
                    if "GUARDA" in val_up or "CHOFER" in val_up:
                        val_g = ws.cell(row=r, column=c + 1).value
                        if val_g:
                            guarda = str(val_g).strip()

        if col_patente is not None:
            for row_p in range(1, ws.max_row + 1):
                val_p = ws.cell(row=row_p, column=col_patente).value
                if val_p and not any(
                    k in str(val_p).upper() for k in ("PATENTE", "TOTAL", "CHOFER")
                ):
                    primera_patente = str(val_p).strip()
                    break

        # Precintos: SIEMPRE columna 1 (A), desde fila 2 (openpyxl: row>=2)
        precintos = []
        for row_pr in range(2, ws.max_row + 1):
            val_pr = ws.cell(row=row_pr, column=1).value  # columna 1 = A
            if val_pr:
                val_str = str(val_pr).strip()
                palabras = val_str.split()
                es_nombre_largo = len(palabras) >= 3 or len(val_str) > 30
                if len(val_str) >= 2 and val_str != "-" and not es_nombre_largo:
                    # Blacklist de palabras que son etiquetas/headers, no precintos reales
                    labels_prohibidas = (
                        "PRECINTO", "SELLO", "TOTAL", "CHOFER", "GUARDA", "ADUANA",
                        "PUERTO", "CARPETA", "BOOKING", "TRANSPORTE", "PESO", "FLEXI",
                        "FECHA", "DESTINATARIO", "TERMINAL", "PATENTE", "PERMISO",
                        "EMPRESA", "SALIDA", "CARGA", "EMBARQUE", "CONTENEDOR",
                    )
                    if not any(k in val_str.upper() for k in labels_prohibidas):
                        # Separar precintos múltiples unidos por guion (ej: "12345-67890")
                        for parte in val_str.split("-"):
                            parte = parte.strip()
                            if parte:
                                precintos.append(parte)
        # Si no se encontró B/L, buscar por número de carpeta en la hoja actual y la anterior
        if not bl_celda and carp_celda:
            bl_celda = buscar_bl_por_carpeta_xlsx(wb, pestana_chofer, carp_celda)

        wb.close()
        return f_celda, pe_celda, carp_celda, dest_celda, bl_celda, transporte_encontrado, primera_patente, precintos, guarda, puerto_salida, peso_flexi

    def _agregar_fila_planillas(self, registro):
        """Agrega una fila a la tabla de planillas (thread-safe, en hilo principal)."""
        dest = (registro.get("dest_celda") or "").upper()

        if "VITAPRO" in dest:
            tag = "vitapro"
        elif "EWOS" in dest:
            tag = "ewos"
        elif "DICOAL" in dest:
            tag = "dicoal"
        elif "NUTRECO" in dest:
            tag = "nutreco"
        elif "CARGILL" in dest:
            tag = "cargill"
        elif "BIOMAR" in dest:
            tag = "biomar"
        else:
            tag = ""

        valores = (
            registro["f_celda"],
            registro["cant_final"],
            registro["pe_recortado"],
            registro["carp_celda"],
            registro["terminal"],
            registro["trans_final"],
            registro["dest_celda"],
            registro["bl_final"],
            registro["frac_carpeta"],
            "-" if registro["servicio"] == "-" else f'${registro["servicio"]:,}',
        )
        self.tree_planillas.insert("", "end", values=valores, tags=(tag,) if tag else ())

        # Actualizar resumen
        count = len(self.tree_planillas.get_children())
        total_cant = sum(
            int(self.tree_planillas.item(child, "values")[1])
            for child in self.tree_planillas.get_children()
        )
        self.lbl_resumen_planillas.configure(
            text=f"{count} registros  ·  {total_cant} contenedores totales"
        )

    def _planillas_error(self, error_msg):
        self._log(f"ERROR CRÍTICO: {error_msg}")
        messagebox.showerror(
            "Error en Agente de Contenedores",
            f"Ocurrió un error durante el análisis:\n\n{error_msg}",
        )

    def _planillas_done(self):
        self.tarea_activa = False
        try:
            self.btn_ejecutar_planillas.configure(
                text="Completar Planillas",
                state="normal",
                fg_color=Palette.ACCENT,
            )
            self.progress_planillas.stop()
        except (AttributeError, Exception):
            pass
        self._set_status("Análisis de planillas finalizado")

        # Actualizar conteo de precintos disponibles
        self._actualizar_titulo_precintos()

        # Mostrar popup de resumen si hay resultados pendientes
        if hasattr(self, "_resultados_pendientes") and self._resultados_pendientes:
            try:
                self._mostrar_resumen(self._resultados_pendientes)
            except Exception:
                pass
            self._resultados_pendientes = None


