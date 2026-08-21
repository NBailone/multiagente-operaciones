"""ImpresionMixin — Panel de Impresión Documental.

Legacy block extracted from ui_app.py. Methods are mixed into the main
App class; they rely on attributes/methods living on the composed instance
(_log, _set_log_panel, _cfg_obtener, _cfg_obtener_docs, _cfg_obtener_rutas,
_resolver_ruta, _abrir_excel_seguro, _guardar_config, _icons, _panel_frames,
_excel_com_ok, tarea_activa, panel_container).
"""

import os
import re
import sys
import time
import threading
import subprocess
import traceback
from datetime import datetime

import customtkinter as ctk
from tkinter import messagebox
import openpyxl
import xlrd
import win32com.client
import pythoncom

from constants import Palette, FONT_FAMILY
from utils import buscar_archivo_en_pendrive


class ImpresionMixin:
    def _panel_impresion(self):
        if "impresion" in self._panel_frames:
            return
        self._imp_carpetas_vars = {}
        self._imp_select_all_var = ctk.BooleanVar(value=True)

        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        self._panel_frames["impresion"] = frame
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── Toolbar: solo Refrescar ──────────────────────────────────
        toolbar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=44
        )
        toolbar.pack(fill="x", pady=(0, 6))
        toolbar.pack_propagate(False)

        self._imp_btn_refresh = ctk.CTkButton(
            toolbar,
            text="Refrescar",
            image=self._icons["refresh-cw"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34,
            command=self._imp_escanear_carpetas,
        )
        self._imp_btn_refresh.pack(side="left", padx=4, pady=4)

        self._imp_btn_dorsos = ctk.CTkButton(
            toolbar,
            text="Dorsos",
            image=self._icons["file-text"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34,
            command=self._imp_popup_dorsos,
        )
        self._imp_btn_dorsos.pack(side="left", padx=4, pady=4)

        self._imp_btn_sellos_mic = ctk.CTkButton(
            toolbar, text="Sellos MIC",
            image=self._icons["bookmark"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34,
            command=self._imp_popup_sellos_mic,
        )
        self._imp_btn_sellos_mic.pack(side="left", padx=4, pady=4)

        self._imp_btn_sellos_crt = ctk.CTkButton(
            toolbar, text="Sellos CRT",
            image=self._icons["bookmark"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6, height=34,
            command=self._imp_popup_sellos_crt,
        )
        self._imp_btn_sellos_crt.pack(side="left", padx=4, pady=4)

        self._imp_lbl_estado = ctk.CTkLabel(
            toolbar,
            text="", font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
        )
        self._imp_lbl_estado.pack(side="right", padx=12)

        # ── Split izquierda-derecha ──────────────────────────────────
        split_frame = ctk.CTkFrame(frame, fg_color="transparent")
        split_frame.pack(fill="both", expand=True)

        # Columna izquierda: lista de carpetas
        left_frame = ctk.CTkFrame(
            split_frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER
        )
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 4))

        ctk.CTkLabel(
            left_frame, text="CARPETAS DE CARGA",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(anchor="w", padx=14, pady=(10, 2))

        self._imp_scroll_carpetas = ctk.CTkScrollableFrame(
            left_frame, fg_color="transparent"
        )
        self._imp_scroll_carpetas.pack(fill="both", expand=True, padx=4, pady=(0, 6))

        # Columna derecha: opciones de impresión
        right_frame = ctk.CTkFrame(
            split_frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, width=310
        )
        right_frame.pack(side="right", fill="y", padx=(4, 0))
        right_frame.pack_propagate(False)

        ctk.CTkLabel(
            right_frame, text="OPCIONES DE IMPRESIÓN",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(anchor="w", padx=12, pady=(8, 6))

        opciones = [
            ("sobre", "Sobre (Planilla de Carga)", 1),
            ("permiso", "Permiso de Exportación", 2),
            ("hoja_ruta", "Hoja de Ruta", 2),
            ("servicio_ata", "Servicio ATA / Recibo ATA", 0),
        ]
        def _verificar_boton_imprimir(*args):
            """Activa el botón IMPRIMIR solo si hay carpetas y opciones seleccionadas."""
            hay_carpetas = any(v.get() for v in self._imp_carpetas_vars.values())
            hay_opciones = any(v.get() for v in self._imp_opciones_vars.values())
            if hay_carpetas and hay_opciones:
                self._imp_btn_imprimir.configure(state="normal", fg_color=Palette.ACCENT)
            else:
                self._imp_btn_imprimir.configure(state="disabled", fg_color=Palette.ACCENT_DIM)

        self._imp_verificar_btn = _verificar_boton_imprimir

        self._imp_opciones_vars = {}
        self._imp_opciones_copias = {}
        for key, label, default_copias in opciones:
            row = ctk.CTkFrame(right_frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=1)
            var = ctk.BooleanVar(value=True)
            var.trace_add("write", _verificar_boton_imprimir)
            cb = ctk.CTkCheckBox(
                row, text=label,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                variable=var,
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            cb.pack(side="left")
            if key != "servicio_ata":
                entry = ctk.CTkEntry(
                    row, width=45, height=26,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                    fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                    text_color=Palette.TEXT_PRIMARY, corner_radius=4,
                )
                entry.insert(0, str(default_copias))
                entry.pack(side="right", padx=(0, 4))
                ctk.CTkLabel(
                    row, text="copias",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                    text_color=Palette.TEXT_MUTED,
                ).pack(side="right")
                self._imp_opciones_copias[key] = entry
            else:
                self._imp_opciones_copias[key] = None
            self._imp_opciones_vars[key] = var

        # Botón IMPRIMIR (arranca activo porque hay carpetas y opciones por defecto)
        self._imp_btn_imprimir = ctk.CTkButton(
            right_frame,
            text="IMPRIMIR",
            image=self._icons["printer"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=36,
            command=self._imp_ejecutar_desde_panel,
        )
        self._imp_btn_imprimir.pack(fill="x", padx=12, pady=(10, 8))

    def _scan_desktop_folders(self, pattern=None, callback=None, button=None):
        """Scan Desktop for folders containing CONTENEDORES Excel files.
        
        Thread-safe: launches a background daemon thread, results delivered
        via self.after(0, callback, results) on the main thread.
        
        Args:
            pattern: regex to filter folder names (None = all folders)
            callback: function(results) called on main thread with list of dicts
            button: tkinter button to disable during scan (optional)
        
        Returns list of dicts: [{"name", "path", "excel_path", "pdf_count"}]
        """
        if button:
            button.configure(state="disabled")

        def _worker():
            try:
                desktop = self._resolver_ruta("planillas_carga", "Desktop")
                results = []
                if os.path.isdir(desktop):
                    for entry in os.scandir(desktop):
                        if not entry.is_dir() or entry.name.startswith('.'):
                            continue
                        if entry.name.upper() in ("RECYCLED", "RECYCLER"):
                            continue
                        if pattern and not re.search(pattern, entry.name, re.IGNORECASE):
                            continue
                        excel_path = None
                        pdf_count = 0
                        try:
                            for f in os.scandir(entry.path):
                                if f.is_file():
                                    fname_upper = f.name.upper()
                                    if ("CONTENEDORES" in fname_upper
                                            and fname_upper.endswith(('.XLS', '.XLSX'))):
                                        if excel_path is None:
                                            excel_path = f.path
                                    if f.name.lower().endswith('.pdf'):
                                        pdf_count += 1
                        except OSError:
                            continue
                        if excel_path:
                            results.append({
                                "name": entry.name,
                                "path": entry.path,
                                "excel_path": excel_path,
                                "pdf_count": pdf_count,
                            })
                if callback:
                    self.after(0, callback, results)
            except Exception as e:
                self._log(f"ERROR al escanear Desktop: {e}")
                if callback:
                    self.after(0, callback, [])
            finally:
                if button:
                    self.after(0, lambda: button.configure(state="normal"))

        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    def _imp_escanear_carpetas(self):
        """Escanea el Escritorio y lista las carpetas de carga en la columna izquierda."""
        for w in self._imp_scroll_carpetas.winfo_children():
            w.destroy()
        self._imp_carpetas_vars.clear()
        self._imp_select_all_var.set(True)

        # Mostrar indicador de búsqueda en la toolbar
        self._imp_lbl_estado.configure(text="🔍 Buscando carpetas...")
        self.update_idletasks()

        self._scan_desktop_folders(
            callback=self._imp_poblar_carpetas,
            button=self._imp_btn_refresh,
        )

    def _imp_poblar_carpetas(self, results):
        """Callback: puebla el scroll frame con las carpetas encontradas."""
        carpetas = []
        for folder in results:
            match_frac = re.search(r"(F(?:RACCION)?\s*\d+)", folder["name"], re.IGNORECASE)
            frac = match_frac.group(1).upper() if match_frac else ""
            carpetas.append((folder["path"], folder["name"], frac))

        if not carpetas:
            ctk.CTkLabel(
                self._imp_scroll_carpetas,
                text="No se encontraron carpetas\nde carga en el Escritorio",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.TEXT_MUTED, justify="center",
            ).pack(pady=40)
            self._imp_lbl_estado.configure(text="Sin carpetas")
            return

        self._imp_lbl_estado.configure(text=f"{len(carpetas)} carpetas")

        # Select All
        cb_all = ctk.CTkCheckBox(
            self._imp_scroll_carpetas,
            text="SELECCIONAR TODAS",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
            variable=self._imp_select_all_var,
            command=self._imp_toggle_all,
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
            text_color=Palette.TEXT_PRIMARY,
        )
        cb_all.pack(anchor="w", padx=12, pady=(8, 4))

        ctk.CTkFrame(self._imp_scroll_carpetas, fg_color=Palette.DIVIDER, height=1).pack(fill="x", padx=8, pady=(0, 4))

        for ruta, nombre, frac in carpetas:
            var = ctk.BooleanVar(value=True)
            if hasattr(self, '_imp_verificar_btn'):
                var.trace_add("write", self._imp_verificar_btn)
            label = nombre
            if frac:
                label += f"  [{frac}]"
            cb = ctk.CTkCheckBox(
                self._imp_scroll_carpetas, text=label,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                variable=var,
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            cb.pack(anchor="w", padx=16, pady=1)
            self._imp_carpetas_vars[ruta] = var

        # Verificar estado del botón IMPRIMIR
        if hasattr(self, '_imp_verificar_btn'):
            self._imp_verificar_btn()

    def _imp_toggle_all(self):
        estado = self._imp_select_all_var.get()
        for var in self._imp_carpetas_vars.values():
            var.set(estado)

    def _imp_diagnosticar(self):
        """Analiza las carpetas seleccionadas y muestra los archivos encontrados SIN imprimir."""
        seleccionadas = [ruta for ruta, var in self._imp_carpetas_vars.items() if var.get()]
        if not seleccionadas:
            messagebox.showinfo("Sin selección", "Seleccione al menos una carpeta para diagnosticar.")
            return
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return

        self.tarea_activa = True
        self._imp_btn_imprimir.configure(state="disabled")
        self._imp_btn_refresh.configure(state="disabled")
        self._limpiar_log()

        opciones = {k: v.get() for k, v in self._imp_opciones_vars.items()}
        self._log("══════════ DIAGNÓSTICO DE IMPRESIÓN ══════════")
        self._log(f"Carpetas seleccionadas: {len(seleccionadas)}")
        self._log(f"Opciones activas: {', '.join(k for k, v in opciones.items() if v)}")
        self._log("")

        t = threading.Thread(target=self._imp_diag_worker, args=(seleccionadas, opciones), daemon=True)
        t.start()

    def _imp_diag_worker(self, seleccionadas, opciones):
        """Worker de diagnóstico: lista archivos encontrados sin imprimir."""
        self._set_log_panel("impresion")
        anio = datetime.now().strftime("%y")
        prefijo_permiso = f"{anio}069EC"

        for ruta in seleccionadas:
            nombre = os.path.basename(ruta)
            self._log(f"📂 {nombre}")
            self._log(f"   Ruta: {ruta}")
            try:
                archivos = sorted(os.listdir(ruta))
            except Exception as e:
                self._log(f"   ❌ Error al leer carpeta: {e}")
                continue

            # Clasificar archivos
            sobres = [a for a in archivos if "CONTENEDORES" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
            permisos = [a for a in archivos if a.upper().startswith(prefijo_permiso) and a.upper().endswith(".PDF")]
            # Buscar Hoja de Ruta: flexible, que contenga "HOJA" y "RUTA" (no requiere "DE")
            hojas_ruta = [a for a in archivos if "HOJA" in a.upper() and "RUTA" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
            # Fallback: buscar hoja interna 'Hoja de Ruta' dentro de los Excel
            hojas_ruta_internas = []
            if not hojas_ruta:
                hojas_ruta_internas = self._imp_buscar_hoja_ruta_interna(ruta, archivos)

            # Buscar también con patrón más amplio si no se encontró
            if not permisos:
                # Buscar cualquier PDF que empiece con el año
                permisos = [a for a in archivos if a.upper().endswith(".PDF") and a.upper()[:2] == anio]

            self._log(f"   ┌─ SOBRE (Excel CONTENEDORES):")
            if sobres:
                for a in sobres:
                    self._log(f"   │  ✓ {a}")
            else:
                self._log(f"   │  ❌ No encontrado (busca: *CONTENEDORES*.xlsx)")

            self._log(f"   ├─ PERMISO EXPORTACIÓN (PDF {prefijo_permiso}*):")
            if permisos:
                for a in permisos:
                    self._log(f"   │  ✓ {a}")
            else:
                self._log(f"   │  ❌ No encontrado (busca: {prefijo_permiso}*.PDF)")
                # Mostrar todos los PDF para ayudar a diagnosticar
                todos_pdf = [a for a in archivos if a.upper().endswith(".PDF")]
                if todos_pdf:
                    self._log(f"   │  ⚠ PDFs en la carpeta: {todos_pdf}")

            self._log(f"   ├─ HOJA DE RUTA (nombre u hoja interna):")
            if hojas_ruta:
                for a in hojas_ruta:
                    self._log(f"   │  ✓ {a}")
            elif hojas_ruta_internas:
                for archivo, hoja in hojas_ruta_internas:
                    self._log(f"   │  ✓ {archivo} → hoja '{hoja}'")
            else:
                self._log(f"   │  ❌ No encontrado (busca archivo con 'HOJA'+'RUTA', o hoja interna 'Hoja de Ruta')")

            # Detectar hojas Recibo ATA: nombres exactos "Recibo Ata", "Recibo Ata 2", ... "Recibo Ata 8"
            hojas_ata = []
            todas_hojas = []
            nombres_ata = [f"RECIBO ATA"] + [f"RECIBO ATA {i}" for i in range(2, 9)]
            if sobres:
                for sobre in sobres:
                    ruta_excel = os.path.join(ruta, sobre)
                    try:
                        if sobre.lower().endswith(".xlsx"):
                            wb = openpyxl.load_workbook(ruta_excel, read_only=True)
                            todas_hojas = list(wb.sheetnames)
                            for sn in wb.sheetnames:
                                if sn.upper() in nombres_ata:
                                    hojas_ata.append((sobre, sn))
                            wb.close()
                        elif sobre.lower().endswith(".xls"):
                            book = xlrd.open_workbook(ruta_excel)
                            todas_hojas = list(book.sheet_names())
                            for sn in book.sheet_names():
                                if sn.upper() in nombres_ata:
                                    hojas_ata.append((sobre, sn))
                    except Exception as e:
                        self._log(f"   │  ⚠ Error leyendo hojas de {sobre}: {e}")

            self._log(f"   ├─ RECIBOS ATA (hojas dentro del Excel, {len(hojas_ata)} choferes):")
            if hojas_ata:
                for archivo, hoja in hojas_ata:
                    self._log(f"   │  ✓ {archivo} → Hoja: \"{hoja}\"")
            else:
                self._log(f"   │  ❌ No encontrado. Hojas disponibles: {todas_hojas}")

            # Mostrar TODOS los archivos reales de la carpeta
            self._log(f"   └─ TODOS los archivos ({len(archivos)}):")
            for a in archivos:
                self._log(f"      · {a}")
            self._log("")

        self._log("══════════ DIAGNÓSTICO COMPLETADO ══════════")
        self.after(0, self._imp_diag_done)

    def _imp_diag_done(self):
        self.tarea_activa = False
        try:
            self._imp_btn_imprimir.configure(state="normal")
            self._imp_btn_refresh.configure(state="normal")
        except (AttributeError, Exception):
            pass

    def _imp_popup_dorsos(self):
        """Popup para imprimir dorsos con cantidad de hojas editable."""
        popup = ctk.CTkToplevel(self)
        popup.title("Impresión de Dorsos")
        popup.geometry("400x330")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()

        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 330) // 2
        popup.geometry(f"400x330+{px}+{py}")

        ctk.CTkLabel(
            popup, text="IMPRESIÓN DE DORSOS",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(20, 16))

        var_mic = ctk.BooleanVar(value=False)
        var_crt = ctk.BooleanVar(value=False)
        var_pe = ctk.BooleanVar(value=False)

        def _actualizar_boton():
            if var_mic.get() or var_crt.get() or var_pe.get():
                btn_imprimir.configure(state="normal", fg_color=Palette.ACCENT)
            else:
                btn_imprimir.configure(state="disabled", fg_color=Palette.ACCENT_DIM)

        def _crear_fila(label, var, default_cant):
            frame = ctk.CTkFrame(popup, fg_color="transparent")
            frame.pack(fill="x", padx=30, pady=4)
            cb = ctk.CTkCheckBox(
                frame, text=label,
                font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                variable=var, command=_actualizar_boton,
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            cb.pack(side="left")
            entry = ctk.CTkEntry(
                frame, width=60, height=28,
                font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                text_color=Palette.TEXT_PRIMARY, corner_radius=4,
            )
            entry.insert(0, str(default_cant))
            entry.pack(side="right")
            ctk.CTkLabel(
                frame, text="hojas",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="right", padx=(0, 6))
            return entry

        def_mic = self._cfg_obtener_docs("dorso_mic", 15)
        def_crt = self._cfg_obtener_docs("dorso_crt", 4)
        def_pe  = self._cfg_obtener_docs("dorso_pe", 2)
        entry_mic = _crear_fila("Dorso MIC", var_mic, def_mic)
        entry_crt = _crear_fila("Dorso CRT", var_crt, def_crt)
        entry_pe = _crear_fila("Dorso PE", var_pe, def_pe)

        def _imprimir_dorsos():
            base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            archivos = {
                "MIC": os.path.join(base, "DORSO MIC.pdf"),
                "CRT": os.path.join(base, "DORSO CRT.pdf"),
                "PE":  os.path.join(base, "dorso PE.pdf"),
            }
            # Leer estado de tkinter en hilo principal antes de destruir el popup
            tipos = []
            for tipo, var, entry in [("MIC", var_mic, entry_mic),
                                      ("CRT", var_crt, entry_crt),
                                      ("PE", var_pe, entry_pe)]:
                if var.get():
                    try:
                        n = int(entry.get())
                    except ValueError:
                        n = (self._cfg_obtener_docs("dorso_mic", 15) if tipo == "MIC" else
                             self._cfg_obtener_docs("dorso_crt", 4) if tipo == "CRT" else
                             self._cfg_obtener_docs("dorso_pe", 2))
                    tipos.append((tipo, n, archivos[tipo]))
            popup.destroy()

            def _esperar_cola_vacia():
                # Espera 3s para que el visor PDF envíe las últimas copias al spooler
                time.sleep(3.0)
                try:
                    import win32print
                    printer = win32print.GetDefaultPrinter()
                    deadline = time.time() + 120
                    while time.time() < deadline:
                        h = win32print.OpenPrinter(printer)
                        try:
                            jobs = win32print.EnumJobs(h, 0, 100, 1)
                        finally:
                            win32print.ClosePrinter(h)
                        if not jobs:
                            return
                        time.sleep(1.5)
                except Exception:
                    time.sleep(10.0)

            def _worker():
                self._set_log_panel("impresion")
                total = 0
                sumatra = self._imp_sumatra_exe()
                for idx, (tipo, n, ruta) in enumerate(tipos):
                    if not os.path.exists(ruta):
                        self._log(f"⚠ Dorso {tipo}: archivo no encontrado ({ruta})")
                        continue
                    if n > 1 and sumatra:
                        # Un solo trabajo con N copias (rápido)
                        subprocess.run(
                            [sumatra, "-print-to-default", "-print-settings", f"{n}x",
                             "-exit-when-done", ruta],
                            creationflags=subprocess.CREATE_NO_WINDOW, timeout=300,
                        )
                    else:
                        for copia in range(n):
                            os.startfile(ruta, "print")
                            time.sleep(0.5)
                    total += n
                    self._log(f"Dorso {tipo}: {n} copias enviadas → {os.path.basename(ruta)}")
                    # Si hay más tipos después, esperar que la cola se vacíe antes de continuar
                    if idx < len(tipos) - 1:
                        self._log(f"  ↳ Esperando cola de impresión antes de siguiente dorso...")
                        _esperar_cola_vacia()
                self._log(f"Total dorsos: {total} hojas")

            threading.Thread(target=_worker, daemon=True).start()

        btn_imprimir = ctk.CTkButton(
            popup, text="IMPRIMIR DORSOS", image=self._icons["printer"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=40,
            command=_imprimir_dorsos,
        )
        btn_imprimir.pack(fill="x", padx=30, pady=(20, 12))

    def _imp_popup_sellos_mic(self):
        """Popup para imprimir Sellos MIC: guardas = nombres de hojas del Excel."""
        # Usar la misma lógica de búsqueda que CRT
        base = self._cfg_obtener_rutas("mic_sellos", os.path.join("TRABAJO", "01_PLANILLAS"))
        ruta_mic = buscar_archivo_en_pendrive("FECHA MIC Y SELLOS.xlsx", base)
        if not ruta_mic:
            ruta_mic = buscar_archivo_en_pendrive("FECHA MIC Y SELLOS.xls", base)

        guardas = []
        if ruta_mic:
            try:
                wb = self._abrir_excel_seguro(ruta_mic)
                guardas = list(wb.sheetnames)
                wb.close()
            except Exception:
                pass

        if not guardas:
            messagebox.showerror("Archivo no encontrado",
                f"No se encontró FECHA MIC Y SELLOS.\n\nRuta configurada: {base}\n"
                f"Se buscó en D-Z y en la carpeta del programa.")
            return

        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        ultimo_guarda_dia = self._cfg_obtener("super_auto", f"mic_guarda_{fecha_hoy}", "")

        popup = ctk.CTkToplevel(self)
        popup.title("Sellos MIC")
        popup.geometry("400x270")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.grab_set()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 270) // 2
        popup.geometry(f"400x270+{px}+{py}")

        ctk.CTkLabel(popup, text="SELLOS MIC",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
                     text_color=Palette.TEXT_PRIMARY).pack(pady=(20, 16))

        ctk.CTkLabel(popup, text="Seleccioná el guarda:",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                     text_color=Palette.TEXT_SECONDARY).pack(anchor="w", padx=30, pady=(0, 4))
        guarda_var = ctk.StringVar(value=ultimo_guarda_dia if ultimo_guarda_dia in guardas else (guardas[0] if guardas else ""))
        ctk.CTkOptionMenu(popup, variable=guarda_var, values=guardas,
                          font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                          fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                          button_hover_color=Palette.ACCENT_HOVER,
                          text_color=Palette.TEXT_PRIMARY, corner_radius=6,
                          width=240, height=34).pack(pady=(0, 12))

        ctk.CTkLabel(popup, text="Cantidad de copias:",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                     text_color=Palette.TEXT_SECONDARY).pack(anchor="w", padx=30, pady=(0, 4))
        copia_entry = ctk.CTkEntry(popup, width=80, height=30,
                                   font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                                   fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                                   text_color=Palette.TEXT_PRIMARY, corner_radius=4)
        copia_entry.insert(0, "1")
        copia_entry.pack(pady=(0, 12))

        def _imprimir():
            guarda = guarda_var.get()
            try: copias = int(copia_entry.get().strip())
            except ValueError: copias = 1
            self.config.setdefault("super_auto", {})[f"mic_guarda_{fecha_hoy}"] = guarda
            self._guardar_config()
            popup.destroy()
            self._imp_sellos_mic_worker(guarda, copias, ruta_mic)

        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(pady=(8, 0))
        ctk.CTkButton(btn_frame, text="Imprimir", width=120, height=34,
                      image=self._icons["printer"], compound="left",
                      font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                      fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                      text_color=Palette.WHITE, corner_radius=6,
                      command=_imprimir).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Cancelar", width=100, height=34,
                      font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                      fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
                      text_color=Palette.TEXT_SECONDARY, corner_radius=6,
                      command=popup.destroy).pack(side="left", padx=6)

    def _imp_sellos_mic_worker(self, guarda, copias, ruta_mic=None):
        """Busca e imprime la hoja del guarda en FECHA MIC Y SELLOS."""
        if not ruta_mic:
            base = self._cfg_obtener_rutas("mic_sellos", os.path.join("TRABAJO", "01_PLANILLAS"))
            ruta_mic = buscar_archivo_en_pendrive("FECHA MIC Y SELLOS.xlsx", base)
            if not ruta_mic:
                ruta_mic = buscar_archivo_en_pendrive("FECHA MIC Y SELLOS.xls", base)
        if not ruta_mic:
            self._log(f"⚠ Sellos MIC: no se encontró FECHA MIC Y SELLOS")
            return

        def _worker():
            self._set_log_panel("impresion")
            try:
                wb = self._abrir_excel_seguro(ruta_mic)
                hoja_encontrada = None
                for sn in wb.sheetnames:
                    if guarda.upper() in sn.upper():
                        hoja_encontrada = sn; break
                if not hoja_encontrada:
                    self.log_queue.put(f"[...] ⚠ Sellos MIC: guarda '{guarda}' no hallado en ninguna hoja")
                    wb.close(); return
                wb.close()
                impresora = self._detectar_impresoras()[0] if self._detectar_impresoras() else "Default"
                self.log_queue.put(f"[...] 🖨 Sellos MIC: {guarda} → hoja '{hoja_encontrada}' ({copias} copias)")
                self._imp_enviar(ruta_mic, impresora, f"  Sellos MIC - {guarda}", hojas=[hoja_encontrada], copias=copias)
            except Exception as e:
                self.log_queue.put(f"[...] ⚠ Error Sellos MIC: {e}")
        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    def _imp_popup_sellos_crt(self):
        """Popup para imprimir Sellos CRT: cantidad de copias."""
        popup = ctk.CTkToplevel(self)
        popup.title("Sellos CRT")
        popup.geometry("360x180")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.grab_set()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 360) // 2
        py = self.winfo_y() + (self.winfo_height() - 180) // 2
        popup.geometry(f"360x180+{px}+{py}")

        ctk.CTkLabel(popup, text="SELLOS CRT",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
                     text_color=Palette.TEXT_PRIMARY).pack(pady=(20, 16))

        ctk.CTkLabel(popup, text="Cantidad de copias:",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                     text_color=Palette.TEXT_SECONDARY).pack(pady=(0, 4))
        copia_entry = ctk.CTkEntry(popup, width=80, height=30,
                                   font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                                   fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                                   text_color=Palette.TEXT_PRIMARY, corner_radius=4)
        copia_entry.insert(0, "1")
        copia_entry.pack(pady=(0, 12))

        def _imprimir():
            try: copias = int(copia_entry.get().strip())
            except ValueError: copias = 1
            popup.destroy()
            self._imp_sellos_crt_worker(copias)

        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(pady=(4, 0))
        ctk.CTkButton(btn_frame, text="Imprimir", width=120, height=34,
                      image=self._icons["printer"], compound="left",
                      font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                      fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                      text_color=Palette.WHITE, corner_radius=6,
                      command=_imprimir).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Cancelar", width=100, height=34,
                      font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                      fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
                      text_color=Palette.TEXT_SECONDARY, corner_radius=6,
                      command=popup.destroy).pack(side="left", padx=6)

    def _imp_sellos_crt_worker(self, copias):
        """Imprime hoja 'Hoja1' de FECHA CRT Y ORIGINAL."""
        base = self._cfg_obtener_rutas("crt_original", os.path.join("TRABAJO", "01_PLANILLAS"))
        ruta_crt = buscar_archivo_en_pendrive("FECHA CRT Y ORIGINAL.xlsx", base)
        if not ruta_crt:
            ruta_crt = buscar_archivo_en_pendrive("FECHA CRT Y ORIGINAL.xls", base)
        if not ruta_crt:
            self._log(f"⚠ Sellos CRT: no se encontró FECHA CRT Y ORIGINAL en {base} (D-Z)")
            return

        def _worker():
            self._set_log_panel("impresion")
            try:
                wb = self._abrir_excel_seguro(ruta_crt)
                hoja = wb.sheetnames[0]
                wb.close()
                impresora = self._detectar_impresoras()[0] if self._detectar_impresoras() else "Default"
                self.log_queue.put(f"[...] 🖨 Sellos CRT: hoja '{hoja}' ({copias} copias)")
                self._imp_enviar(ruta_crt, impresora, f"  Sellos CRT", hojas=[hoja], copias=copias)
            except Exception as e:
                self.log_queue.put(f"[...] ⚠ Error Sellos CRT: {e}")
        t = threading.Thread(target=_worker, daemon=True)
        t.start()

    def _imp_ejecutar_desde_panel(self):
        seleccionadas = [ruta for ruta, var in self._imp_carpetas_vars.items() if var.get()]
        if not seleccionadas:
            messagebox.showinfo("Sin selección", "Seleccione al menos una carpeta para imprimir.")
            return
        self._imp_ejecutar(seleccionadas)

    def _detectar_impresoras(self):
        """Devuelve lista de impresoras disponibles con nombre completo (Name on Port:)."""
        try:
            import subprocess
            result = subprocess.run(
                ["powershell", "-Command",
                 "Get-Printer | ForEach-Object { $_.Name + ' on ' + $_.PortName + ':' }"],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0 and result.stdout.strip():
                printers = [p.strip() for p in result.stdout.strip().split("\n") if p.strip()]
                # Poner Microsoft Print to PDF primero si existe
                pdf = [p for p in printers if "PDF" in p.upper()]
                otros = [p for p in printers if "PDF" not in p.upper()]
                return pdf + otros
        except Exception:
            pass
        return ["Microsoft Print to PDF (simulación)"]

    def _detectar_tipo_carpeta(self, nombre_carpeta: str) -> str:
        """Return transport type from folder name: TERRESTRE, ISO, or FLEXI.

        Folder format: DD_MM_YYYY_CANT_TIPO_PE_CARPETA_DEST_[SUFFIX]
        The type is at index 4 (0-indexed). Old-format folders default to TERRESTRE.
        """
        partes = nombre_carpeta.split("_")
        if len(partes) < 5:
            return "TERRESTRE"
        tipo = partes[4]
        if tipo not in ("TERRESTRE", "ISO", "FLEXI"):
            return "TERRESTRE"
        return tipo

    def _imp_ejecutar(self, seleccionadas):
        """Ejecuta la cola de impresión."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return

        self.tarea_activa = True
        self._imp_btn_imprimir.configure(text="⏳  Imprimiendo...", state="disabled")
        self._imp_btn_refresh.configure(state="disabled")
        self._limpiar_log()

        opciones = {k: v.get() for k, v in self._imp_opciones_vars.items()}
        # Leer copias de los entries (usar defaults si el entry no existe)
        copias = {}
        for key, ent in self._imp_opciones_copias.items():
            if ent is not None:
                try:
                    copias[key] = int(ent.get().strip())
                except ValueError:
                    copias[key] = 2
            else:
                copias[key] = 0
        impresora = self._detectar_impresoras()[0] if self._detectar_impresoras() else "Default"
        self._log(f"IMPRESORA: {impresora}")
        self._log(f"Carpetas a procesar: {len(seleccionadas)}")
        self._log(f"Opciones: {', '.join(k for k, v in opciones.items() if v)}")
        self._log("─" * 40)

        t = threading.Thread(
            target=self._imp_worker,
            args=(seleccionadas, opciones, impresora, copias),
            daemon=True,
        )
        t.start()

    def _imp_worker(self, seleccionadas, opciones, impresora, copias=None):
        """Hilo de fondo: procesa cada carpeta y lanza las impresiones."""
        self._set_log_panel("impresion")
        if copias is None:
            copias = {}
        anio = datetime.now().strftime("%y")
        # El permiso de exportación empieza con el año + 069EC (ej: 26069EC...)
        prefijo_permiso = f"{anio}069EC"
        total_ok = 0

        try:
            for ruta in seleccionadas:
                nombre = os.path.basename(ruta)
                self._log(f"Procesando: {nombre}")
                self._log(f"  Carpeta: {ruta}")
                archivos = sorted(os.listdir(ruta))

                # Listar todo lo que hay en la carpeta
                self._log(f"  Archivos en la carpeta ({len(archivos)}):")
                for a in archivos:
                    self._log(f"    · {a}")

                # Detectar archivos por tipo
                sobres = [a for a in archivos if "CONTENEDORES" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
                permisos = [a for a in archivos if a.upper().startswith(prefijo_permiso) and a.upper().endswith(".PDF")]
                # Buscar Hoja de Ruta: flexible, que contenga "HOJA" y "RUTA" (no requiere "DE")
                hojas_ruta = [a for a in archivos if "HOJA" in a.upper() and "RUTA" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
                # Fallback: si el archivo se llama distinto (ej: "ATA xxx.xls"),
                # buscar dentro de los Excel una hoja interna que sea la Hoja de Ruta
                hojas_ruta_internas = []
                if not hojas_ruta:
                    hojas_ruta_internas = self._imp_buscar_hoja_ruta_interna(ruta, archivos)
                    if hojas_ruta_internas:
                        self._log(f"  Hoja de Ruta hallada por nombre de hoja interna: {hojas_ruta_internas}")

                # Detectar hojas Recibo ATA: nombres exactos
                nombres_ata = ["RECIBO ATA"] + [f"RECIBO ATA {i}" for i in range(2, 9)]
                hojas_ata = []
                if sobres:
                    for sobre in sobres:
                        ruta_excel = os.path.join(ruta, sobre)
                        try:
                            if sobre.lower().endswith(".xlsx"):
                                wb = openpyxl.load_workbook(ruta_excel, read_only=True)
                                for sn in wb.sheetnames:
                                    if sn.upper() in nombres_ata:
                                        hojas_ata.append((sobre, sn))
                                wb.close()
                            elif sobre.lower().endswith(".xls"):
                                book = xlrd.open_workbook(ruta_excel)
                                for sn in book.sheet_names():
                                    if sn.upper() in nombres_ata:
                                        hojas_ata.append((sobre, sn))
                        except Exception:
                            pass
                cant_choferes = len(hojas_ata) if hojas_ata else 1

                self._log(f"  Detectado: Sobre={sobres}, Permiso={permisos}, HojaRuta={hojas_ruta}, RecibosATA(hojas)={hojas_ata}, Choferes={cant_choferes}")

                # 1. Sobre: imprimir SOLO la hoja "SOBRE"
                if opciones.get("sobre", True):
                    for a in sobres:
                        hojas_sobre = self._imp_hojas_sobre(os.path.join(ruta, a))
                        self._imp_enviar(os.path.join(ruta, a), impresora, f"Sobre: {a}", hojas=hojas_sobre)
                        total_ok += 1

                # 2. Permiso de Exportación (PDF)
                if opciones.get("permiso"):
                    if permisos:
                        n_copias_permiso = copias.get("permiso", 2) or self._cfg_obtener_docs("permiso_exportacion", 2)
                        for a in permisos:
                            # Un solo trabajo con N copias (SumatraPDF) en vez de un envío por copia
                            self._imp_enviar(os.path.join(ruta, a), impresora, f"Permiso Exp. ({n_copias_permiso} copias): {a}", copias=n_copias_permiso)
                            total_ok += 1
                    else:
                        self._log(f"  ⚠ No se encontró Permiso de Exportación (busca: {prefijo_permiso}*.PDF)")

                # 3. Hoja de Ruta (Excel)
                if opciones.get("hoja_ruta"):
                    n_copias_hr = copias.get("hoja_ruta", 2) or self._cfg_obtener_docs("hoja_ruta", 2)
                    if hojas_ruta:
                        for a in hojas_ruta:
                            self._imp_enviar(os.path.join(ruta, a), impresora, f"Hoja Ruta ({n_copias_hr} copias): {a}", copias=n_copias_hr)
                            total_ok += 1
                    elif hojas_ruta_internas:
                        for archivo, hoja in hojas_ruta_internas:
                            self._imp_enviar(os.path.join(ruta, archivo), impresora, f"Hoja Ruta ({n_copias_hr} copias): {archivo} → {hoja}", hojas=[hoja], copias=n_copias_hr)
                            total_ok += 1
                    else:
                        self._log(f"  ⚠ No se encontró Hoja de Ruta (busca archivo con 'HOJA'+'RUTA' en el nombre, o una hoja interna 'Hoja de Ruta')")

                # 4. Servicio ATA / Recibo ATA: imprimir SOLO las hojas exactas
                if opciones.get("servicio_ata"):
                    tipo = self._detectar_tipo_carpeta(nombre)
                    if tipo in ("ISO", "FLEXI"):
                        self._log(f"  ⏭ Saltando Recibo ATA: carpeta marítima ({tipo})")
                    else:
                        if hojas_ata:
                            for archivo, hoja in hojas_ata:
                                self._imp_enviar(os.path.join(ruta, archivo), impresora, f"Recibo ATA: {archivo} → {hoja}", hojas=[hoja])
                                total_ok += 1
                        else:
                            self._log(f"  ⚠ No se encontraron hojas Recibo Ata en el Excel")

            self._log(f"COMPLETADO: Impresión finalizada — {total_ok} documentos enviados.")
        except Exception as e:
            self._log(f"ERROR en impresión: {e}")
            traceback.print_exc()
        finally:
            self.after(0, self._imp_done)

    def _imp_hojas_sobre(self, ruta_excel):
        """Busca la hoja 'SOBRE'. Si no existe, devuelve la primera hoja."""
        try:
            if ruta_excel.lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(ruta_excel, read_only=True)
                nombres = list(wb.sheetnames)
                for sn in nombres:
                    if sn.upper() == "SOBRE":
                        wb.close()
                        return [sn]
                wb.close()
                return [nombres[0]] if nombres else None
            elif ruta_excel.lower().endswith(".xls"):
                book = xlrd.open_workbook(ruta_excel)
                nombres = list(book.sheet_names())
                for sn in nombres:
                    if sn.upper() == "SOBRE":
                        return [sn]
                return [nombres[0]] if nombres else None
        except Exception:
            pass
        return None  # error, imprimir todo

    def _imp_buscar_hoja_ruta_interna(self, ruta, archivos):
        """Fallback Hoja de Ruta: busca en los Excel de la carpeta una hoja
        interna cuyo nombre contenga 'HOJA' y 'RUTA'.

        Returns:
            list[tuple[str, str]]: pares (archivo, nombre_de_hoja).
        """
        resultados = []
        for a in archivos:
            if not (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS")):
                continue
            ruta_excel = os.path.join(ruta, a)
            try:
                if a.lower().endswith(".xlsx"):
                    wb = openpyxl.load_workbook(ruta_excel, read_only=True)
                    nombres = wb.sheetnames
                    wb.close()
                else:
                    nombres = xlrd.open_workbook(ruta_excel).sheet_names()
                for sn in nombres:
                    su = sn.upper()
                    if "HOJA" in su and "RUTA" in su:
                        resultados.append((a, sn))
            except Exception:
                continue
        return resultados

    def _imp_sumatra_exe(self):
        """Ruta al SumatraPDF portable (engines/sumatra/) o None si no está.

        SumatraPDF permite enviar N copias como UN solo trabajo de impresión
        a nivel del driver (-print-settings "Nx"), igual que el diálogo manual.
        """
        candidatos = []
        if getattr(sys, "frozen", False):
            candidatos.append(os.path.join(getattr(sys, "_MEIPASS", ""), "engines", "sumatra", "SumatraPDF.exe"))
            candidatos.append(os.path.join(os.path.dirname(sys.executable), "engines", "sumatra", "SumatraPDF.exe"))
        raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        candidatos.append(os.path.join(raiz, "engines", "sumatra", "SumatraPDF.exe"))
        for c in candidatos:
            if c and os.path.isfile(c):
                return c
        return None

    def _imp_enviar(self, ruta_archivo, impresora, descripcion, hojas=None, copias=1):
        """Envía un archivo a la impresora seleccionada. hojas=lista, copias=N."""
        nombre_archivo = os.path.basename(ruta_archivo)
        self._log(f"  {descripcion}")
        self._log(f"     Archivo: {nombre_archivo}")
        if hojas:
            self._log(f"     Hojas: {hojas}")
        try:
            if "simulación" in impresora.lower():
                self._log(f"     → SIMULADO (sin impresora)")
                return True

            ext = os.path.splitext(ruta_archivo)[1].upper()
            es_excel = ext in (".XLSX", ".XLS")

            # PDFs: un solo trabajo con N copias vía SumatraPDF (rápido).
            if ext == ".PDF" and copias > 1:
                sumatra = self._imp_sumatra_exe()
                if sumatra:
                    self._log(f"     → Impresora predeterminada ({copias} copias en un solo trabajo)")
                    subprocess.run(
                        [sumatra, "-print-to-default", "-print-settings", f"{copias}x",
                         "-exit-when-done", ruta_archivo],
                        creationflags=subprocess.CREATE_NO_WINDOW, timeout=300,
                    )
                    return True
                self._log(f"     → SumatraPDF no disponible: enviando copias una por una")

            self._log(f"     → Impresora predeterminada")
            if es_excel and self._excel_com_ok:
                result = self._imp_excel_com(ruta_archivo, impresora, hojas, copias)
            else:
                if es_excel and not self._excel_com_ok:
                    self._log(f"     → Modo visible (COM no disponible, se abrirá Excel)")
                for i in range(copias):
                    os.startfile(ruta_archivo, "print")
                    if copias > 1:
                        self._log(f"     → Copia {i+1}/{copias}")
                    time.sleep(2.0)  # os.startfile es asíncrono, pausa entre copias
                result = True
            return result
        except Exception as e:
            self._log(f"     → ERROR: {e}")
            return False

    def _imp_excel_com(self, ruta, impresora, hojas=None, copias=1):
        """Imprime Excel con Python win32com (sin abrir ventana)."""
        try:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            # Imprimir siempre en la impresora predeterminada del sistema:
            # no se asigna ActivePrinter (falla con formatos de puerto como PORTPROMPT).
            try:
                import win32print
                self._log(f"     → Impresora predeterminada: {win32print.GetDefaultPrinter()}")
            except Exception:
                pass

            wb = excel.Workbooks.Open(ruta)

            if hojas:
                for nombre_hoja in hojas:
                    try:
                        ws = wb.Worksheets(nombre_hoja)
                        # PrintOut(From, To, Copies, Preview, ActivePrinter, PrintToFile, Collate)
                        ws.PrintOut(1, 9999, copias)
                        self._log(f"     → Hoja '{nombre_hoja}' enviada ({copias} copias)")
                    except Exception as e:
                        self._log(f"     → Error hoja '{nombre_hoja}': {e}")
            else:
                wb.PrintOut(1, 9999, copias)
                self._log(f"     → OK ({copias} copias)")

            wb.Close(SaveChanges=False)
            excel.Quit()
            return True
        except Exception as e:
            self._log(f"     → Error COM: {e}")
            return False
        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _imp_done(self):
        self.tarea_activa = False
        try:
            self._imp_btn_imprimir.configure(
                text="IMPRIMIR", state="normal",
                image=self._icons["printer"], compound="left",
                fg_color=Palette.ACCENT,
            )
            self._imp_btn_refresh.configure(text="Refrescar", image=self._icons["refresh-cw"], compound="left", state="normal", command=self._imp_escanear_carpetas)
        except (AttributeError, Exception):
            pass

