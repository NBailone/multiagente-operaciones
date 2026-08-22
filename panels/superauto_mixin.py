"""SuperAutoMixin.

Legacy block extracted from ui_app.py. Mixed into the main App class; relies
on attributes/methods living on the composed instance (_log, _set_log_panel,
_cfg_obtener, _cfg_obtener_rutas, _resolver_ruta, _abrir_excel_seguro,
_guardar_excel_seguro, _detectar_impresoras, _detectar_tipo_carpeta,
_limpiar_log, _set_status, _icons, _panel_frames, tarea_activa). Also uses
ImpresionMixin._imp_* and PlanillasMixin._planillas_core via composed App.
"""

import os
import re
import time
import threading
import traceback
import shutil
from datetime import datetime

import openpyxl
import xlrd

import customtkinter as ctk

from constants import Palette, FONT_FAMILY


class SuperAutoMixin:
    # ═══════════════════════════════════════════════════════════════════
    # SÚPER AUTO
    # ═══════════════════════════════════════════════════════════════════
    def _super_toggle(self):
        """Activa/desactiva el modo Súper Automatización."""
        if self._super_switch.get():
            guardas = self._cfg_obtener("valores", "guardas", ["Gonzalez", "Rodriguez", "Martinez", "Perez"])
            popup = ctk.CTkToplevel(self)
            popup.title("Súper Auto — Elegir Guarda")
            popup.geometry("380x180")
            popup.configure(fg_color=Palette.BG_CARD)
            popup.transient(self)
            popup.lift()
            popup.grab_set()
            popup.update_idletasks()
            px = self.winfo_x() + (self.winfo_width() - 380) // 2
            py = self.winfo_y() + (self.winfo_height() - 180) // 2
            popup.geometry(f"380x180+{px}+{py}")

            ctk.CTkLabel(popup, text="Seleccioná el guarda para Súper Auto:",
                         font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                         text_color=Palette.TEXT_PRIMARY).pack(pady=(20, 10))
            guarda_var = ctk.StringVar(value=guardas[0] if guardas else "")
            ctk.CTkOptionMenu(popup, variable=guarda_var, values=guardas,
                              font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                              fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                              button_hover_color=Palette.ACCENT_HOVER,
                              text_color=Palette.TEXT_PRIMARY, corner_radius=6,
                              width=220, height=34).pack(pady=(0, 12))

            def _confirmar():
                self._super_guarda = guarda_var.get()
                self._super_auto = True
                self._super_lbl_guarda.configure(text=f"Guarda: {self._super_guarda}")
                popup.destroy()

            def _cancelar():
                self._super_switch.deselect()
                self._super_auto = False
                self._super_guarda = ""
                self._super_lbl_guarda.configure(text="")
                popup.destroy()

            ctk.CTkButton(popup, text="Activar", width=100, height=32,
                          font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                          fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                          text_color=Palette.WHITE, corner_radius=6,
                          command=_confirmar).pack(side="left", padx=(100, 8))
            ctk.CTkButton(popup, text="Cancelar", width=100, height=32,
                          font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                          fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
                          text_color=Palette.TEXT_SECONDARY, corner_radius=6,
                          command=_cancelar).pack(side="left")
        else:
            self._super_auto = False
            self._super_guarda = ""
            self._super_lbl_guarda.configure(text="")

    def _super_ejecutar_cadena(self, resultados):
        """Ejecuta la cadena: imprimir → guarda → planillas."""
        carpetas = []
        for _, _, nombre_carpeta in resultados:
            escritorio = self._resolver_ruta("descarga_mails", "Desktop")
            ruta = os.path.join(escritorio, nombre_carpeta)
            if os.path.isdir(ruta):
                carpetas.append(ruta)

        if not carpetas:
            self._log("⚡ Súper Auto: sin carpetas para procesar.")
            return

        self._log(f"⚡ Súper Auto: procesando {len(carpetas)} carpetas...")
        self._ultimas_carpetas = carpetas

        def _cadena():
            cfg = self._cfg_obtener("super_auto", "pasos", {})
            res_imp = self._super_imprimir(carpetas, cfg) if cfg.get("sobre", True) or cfg.get("permiso", True) or cfg.get("hoja_ruta", True) or cfg.get("recibo_ata", True) else 0
            res_gua = self._super_aplicar_guarda(carpetas) if cfg.get("aplicar_guarda", True) else 0
            res_pla = self._super_completar_planillas() if cfg.get("completar_planillas", True) else False
            self.after(0, lambda: self._super_popup_final(res_imp, res_gua, res_pla))

        t = threading.Thread(target=_cadena, daemon=True)
        t.start()

    def _super_imprimir(self, carpetas, cfg=None):
        """Imprime usando la misma lógica que el panel de Impresión Documental."""
        if cfg is None:
            cfg = {}
        hacer_sobre    = cfg.get("sobre", True)
        hacer_permiso  = cfg.get("permiso", True)
        hacer_hoja_ruta = cfg.get("hoja_ruta", True)
        hacer_recibo   = cfg.get("recibo_ata", True)
        anio = datetime.now().strftime("%y")
        prefijo_permiso = f"{anio}069EC"
        impresora = self._imp_impresora_default()
        total_ok = 0

        for ruta in carpetas:
            nombre = os.path.basename(ruta)
            self.log_queue.put(f"[...] 📁 {nombre}")
            try:
                archivos = sorted(os.listdir(ruta))
            except Exception:
                continue

            # Detectar archivos (misma lógica que _imp_worker)
            sobres = [a for a in archivos if "CONTENEDORES" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
            permisos = [a for a in archivos if a.upper().startswith(prefijo_permiso) and a.upper().endswith(".PDF")]
            hojas_ruta = [a for a in archivos if "HOJA" in a.upper() and "RUTA" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]

            # 1. SOBRE: imprimir hoja "SOBRE" de Contenedores (1 copia)
            if hacer_sobre and sobres:
                for sobre in sobres:
                    ruta_excel = os.path.join(ruta, sobre)
                    hojas = self._imp_hojas_sobre(ruta_excel)
                    if hojas:
                        self.log_queue.put(f"[...]   📄 Sobre: {sobre} → {hojas}")
                        ok = self._imp_enviar(ruta_excel, impresora, f"  Sobre ({sobre[:40]})", hojas=hojas, copias=1)
                        if ok:
                            total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin archivo Contenedores")

            # 2. Permisos (2 copias)
            if hacer_permiso and permisos:
                for a in permisos:
                    self.log_queue.put(f"[...]   📄 Permiso: {a} (2 copias)")
                    ok = self._imp_enviar(os.path.join(ruta, a), impresora, f"  Permiso ({a[:40]})", copias=2)
                    if ok:
                        total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Permiso (busca: {prefijo_permiso}*.PDF)")

            # 3. Hojas de Ruta (2 copias)
            if hacer_hoja_ruta and hojas_ruta:
                for a in hojas_ruta:
                    self.log_queue.put(f"[...]   📄 Hoja Ruta: {a} (2 copias)")
                    ok = self._imp_enviar(os.path.join(ruta, a), impresora, f"  Hoja Ruta ({a[:40]})", copias=2)
                    if ok:
                        total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Hoja de Ruta")

            # 4. Servicio ATA / Recibo ATA: hojas del Excel Contenedores
            if hacer_recibo and sobres:
                tipo = self._detectar_tipo_carpeta(nombre)
                if tipo in ("ISO", "FLEXI"):
                    self.log_queue.put(f"[...]   ⏭ Saltando Recibo ATA: carpeta marítima ({tipo})")
                else:
                    nombres_ata = ["RECIBO ATA"] + [f"RECIBO ATA {i}" for i in range(2, 9)]
                    hojas_ata = []
                    for sobre in sobres:
                        ruta_excel = os.path.join(ruta, sobre)
                        try:
                            if sobre.lower().endswith(".xlsx"):
                                wb_tmp = self._abrir_excel_seguro(ruta_excel)
                                for sn in wb_tmp.sheetnames:
                                    if sn.upper() in [n.upper() for n in nombres_ata]:
                                        hojas_ata.append((sobre, sn))
                                wb_tmp.close()
                            else:
                                book = xlrd.open_workbook(ruta_excel)
                                for sn in book.sheet_names():
                                    if sn.upper() in [n.upper() for n in nombres_ata]:
                                        hojas_ata.append((sobre, sn))
                        except Exception:
                            pass
                    if hojas_ata:
                        for archivo, hoja in hojas_ata:
                            self.log_queue.put(f"[...]   📄 Recibo ATA: {archivo} → {hoja}")
                            ok = self._imp_enviar(os.path.join(ruta, archivo), impresora, f"  Recibo ATA ({archivo[:30]})", hojas=[hoja])
                            if ok:
                                total_ok += 1
                    else:
                        self.log_queue.put(f"[...]   ⚠ Sin hoja Recibo ATA en Contenedores")
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Contenedores para Recibo ATA")

        self.log_queue.put(f"[...] ✓ Impresión: {total_ok} documentos enviados de {len(carpetas)} carpetas")
        return total_ok

    def _super_aplicar_guarda(self, carpetas):
        """Escribe el guarda en los archivos Contenedores de las carpetas."""
        aplicados = 0
        for carpeta in carpetas:
            try:
                archivos = os.listdir(carpeta)
            except Exception:
                continue
            for archivo in archivos:
                up = archivo.upper()
                if "CONTENEDORES" not in up or not (up.endswith(".XLSX") or up.endswith(".XLS")):
                    continue
                ruta = os.path.join(carpeta, archivo)
                es_xls = ruta.lower().endswith(".xls") and not ruta.lower().endswith(".xlsx")
                # Reintentos automáticos: el archivo puede estar bloqueado brevemente
                # si COM lo usó para imprimir justo antes.
                max_reintentos = 6
                ultimo_error = None
                for _reintento in range(max_reintentos):
                    try:
                        _abrir_ok = True
                        if es_xls:
                            import xlrd as xlrd_local
                            xlrd_local.open_workbook(ruta, formatting_info=True).release_resources()
                        else:
                            openpyxl.load_workbook(ruta, read_only=True).close()
                        break
                    except PermissionError as _pe:
                        _abrir_ok = False
                        ultimo_error = _pe
                        if _reintento < max_reintentos - 1:
                            self.log_queue.put(f"[...]   ⏳ {archivo[:40]} en uso, esperando... ({_reintento+1}/{max_reintentos-1})")
                            time.sleep(1.5)
                else:
                    self.log_queue.put(f"[...]   ⚠ Error guarda en {archivo}: {ultimo_error}")
                    continue
                try:
                    if not self._escribir_guarda_en_archivo(ruta, self._super_guarda, self.log_queue.put):
                        self.log_queue.put(f"[...]   ⚠ 'Guarda' no hallada en {archivo[:50]}")
                    else:
                        aplicados += 1
                except Exception as e:
                    self.log_queue.put(f"[...]   ⚠ Error guarda en {archivo}: {e}")
                    import traceback
                    self.log_queue.put(f"[...]     {traceback.format_exc()}")
        self.log_queue.put(f"[...] ✓ Guarda '{self._super_guarda}': {aplicados} planillas actualizadas")
        return aplicados

    def _escribir_guarda_en_archivo(
        self, ruta: str, guarda_nombre: str, log_func=None,
        excel_app=None, clave: str = "123",
    ) -> bool:
        """Write guarda_nombre to column H in the CHOFER sheet of an Excel file.

        Si se pasa `excel_app` (sesión compartida de Excel), ambos formatos
        (.xls/.xlsx) se procesan por COM sin relanzar Excel por archivo —
        mucho más rápido en lotes. La apertura usa `clave` por si el workbook
        tiene password (config valores.clave_pdf).
        Sin sesión, comportamiento legacy: sesión propia para .xls y openpyxl
        para .xlsx.
        Returns True if written, False if no 'GUARDA' found or no CHOFER sheet.
        """
        es_xls = ruta.lower().endswith(".xls") and not ruta.lower().endswith(".xlsx")
        escrito = False

        if excel_app is not None:
            # Camino rápido por lote: reusar la sesión de Excel existente
            try:
                try:
                    wb = excel_app.Workbooks.Open(ruta, Password=clave)
                except Exception:
                    wb = excel_app.Workbooks.Open(ruta)
                ws_chofer = None
                for s in wb.Sheets:
                    if "CHOFER" in str(s.Name).upper():
                        ws_chofer = s
                        break
                if not ws_chofer:
                    wb.Close(SaveChanges=False)
                    if log_func:
                        log_func("[...]   Hoja 'Choferes' no hallada")
                    return False
                protegida = bool(ws_chofer.ProtectContents)
                if protegida:
                    ws_chofer.Unprotect(Password=clave)
                for row in range(2, 16):
                    val = ws_chofer.Cells(row, 7).Value  # col 7 = G
                    if val is not None and "GUARDA" in str(val).strip().upper():
                        ws_chofer.Cells(row, 8).Value = guarda_nombre  # col 8 = H
                        escrito = True
                        break
                if protegida:
                    ws_chofer.Protect(Password=clave)
                wb.Save()
                wb.Close(SaveChanges=False)
            except Exception as e:
                if log_func:
                    log_func(f"[...]     ⚠ Error: {e}")
                return False
            return escrito

        if es_xls:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    wb = excel.Workbooks.Open(ruta)
                    ws_chofer = None
                    for s in wb.Sheets:
                        if "CHOFER" in s.Name.upper():
                            ws_chofer = s
                            break
                    if not ws_chofer:
                        if log_func:
                            log_func("[...]   Hoja 'Choferes' no hallada")
                        return False
                    for row in range(2, 16):
                        val = ws_chofer.Cells(row, 7).Value  # col 7 = G
                        if val is not None and "GUARDA" in str(val).strip().upper():
                            ws_chofer.Cells(row, 8).Value = guarda_nombre  # col 8 = H
                            escrito = True
                            break
                    wb.Save()
                    wb.Close()
                finally:
                    excel.Quit()
            finally:
                pythoncom.CoUninitialize()
        else:
            wb = self._abrir_excel_seguro(ruta)
            try:
                ws_chofer = None
                for s in wb.sheetnames:
                    if "CHOFER" in s.upper():
                        ws_chofer = wb[s]
                        break
                if not ws_chofer:
                    if log_func:
                        log_func("[...]   Hoja 'Choferes' no hallada")
                    return False
                for row in range(2, 16):
                    cell = ws_chofer.cell(row=row, column=7)  # col 7 = G (1-indexed)
                    val = cell.value
                    if val is None:
                        for mr in ws_chofer.merged_cells.ranges:
                            if (mr.min_col <= 7 <= mr.max_col and mr.min_row <= row <= mr.max_row):
                                val = ws_chofer.cell(row=mr.min_row, column=mr.min_col).value
                                break
                    if val is not None and "GUARDA" in str(val).strip().upper():
                        ws_chofer.cell(row=row, column=8).value = guarda_nombre  # col 8 = H
                        escrito = True
                        break
                self._guardar_excel_seguro(wb, ruta)
            finally:
                wb.close()

        return escrito

    def _super_completar_planillas(self):
        """Ejecuta Completar Planillas con los datos extraídos."""
        try:
            self._planillas_core()
            self.log_queue.put(f"[...] ✓ Contenedores completadas (SOBRES, COBRO, PC)")
            return True
        except Exception as e:
            self.log_queue.put(f"[...] ⚠ Error completando planillas: {e}")
            return False

    def _super_popup_final(self, impresos, guardas, planillas_ok):
        """Popup resumen del Súper Auto."""
        popup = ctk.CTkToplevel(self)
        popup.title("Súper Auto Completado")
        popup.geometry("400x300")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 300) // 2
        popup.geometry(f"400x300+{px}+{py}")

        ctk.CTkLabel(popup, text="⚡ SÚPER AUTO COMPLETADO",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=18, weight="bold"),
                     text_color=Palette.TEXT_PRIMARY).pack(pady=(24, 20))

        ctk.CTkLabel(popup, text=f"🖨  {impresos} archivos impresos",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS).pack(pady=3)
        txt_guarda = f"🛡  Guarda '{self._super_guarda}' aplicado a {guardas} planillas" if guardas > 0 else f"⚠  Guarda '{self._super_guarda}' — no se encontró etiqueta GUARDA"
        ctk.CTkLabel(popup, text=txt_guarda,
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS if guardas > 0 else Palette.WARNING).pack(pady=3)
        ctk.CTkLabel(popup, text=f"{'✓' if planillas_ok else '✗'} Planillas completadas (SOBRES, COBRO, PC)",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS if planillas_ok else Palette.ERROR).pack(pady=3)

        ctk.CTkButton(popup, text="Cerrar", width=120, height=34,
                      font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                      fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                      text_color=Palette.WHITE, corner_radius=6,
                      command=popup.destroy).pack(pady=(20, 16))


