"""
Sistema de Automatización de Operaciones — Interfaz Profesional
================================================================
Aplicación de escritorio para la gestión de agentes operativos:
  1. Agente de Impresión Documental
  2. Agente de Procesamiento de Contenedores
  3. Agente de Despacho de Correos

Construido con CustomTkinter sobre Tkinter.
"""

import sys
import os
import re
import threading
import queue
import time
import traceback
import json
import subprocess
import string
import imaplib
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import email
import shutil
import tempfile
import dotenv
from copy import copy
import zipfile
import xml.etree.ElementTree as ET
from PIL import Image

# ── Auto-install UI dependencies ──────────────────────────────────────────
def _instalar_deps_ui():
    deps = [
        ("customtkinter", "customtkinter"),
        ("openpyxl", "openpyxl"),
        ("xlrd", "xlrd"),
        ("win32com", "pywin32"),
    ]
    for mod, pip_name in deps:
        try:
            __import__(mod)
        except ImportError:
            print(f"[!] Instalando {pip_name}...")
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", pip_name],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                # pywin32 necesita post-instalación para registrar las DLLs COM
                if pip_name == "pywin32":
                    try:
                        subprocess.check_call(
                            [sys.executable, "-m", "pywin32_postinstall", "-install"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                        )
                    except Exception:
                        pass  # no es fatal si falla
                print(f"[OK] {pip_name} instalado.")
            except Exception as e:
                print(f"[X] No se pudo instalar {pip_name}: {e}")

if not getattr(sys, 'frozen', False):
    _instalar_deps_ui()

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import xlrd

from constants import Palette, FONT_FAMILY, FONT_MONO, FONT_LEVEL_SCALES, FONT_BASE_SIZES
from constants import IMAP_SERVER, PUERTO_IMAP
from constants import DESTINATARIOS_GRUPAL, DESTINATARIOS_INDIVIDUAL

from utils import (buscar_archivo_en_pendrive, formatear_fecha_excel,
                    adjuntar_archivo, preguntar_reintentar,
                    celda_es_mergeada, primera_fila_libre, ya_existe_en_hoja,
                    buscar_bl_por_carpeta_xlsx, buscar_bl_por_carpeta_xls)

from panels import ImpresionMixin, PlanillasMixin, CorreosMixin, DescargaMixin

# ── .env loading for secret key ─────────────────────────────────────────
_frozen = getattr(sys, 'frozen', False)
_base_dir = os.path.dirname(sys.executable) if _frozen else os.path.dirname(__file__)
_dotenv_path = os.path.join(_base_dir, ".env")

dotenv.load_dotenv(_dotenv_path)

def _ensure_secret_key(dotenv_path: str) -> None:
    key = os.environ.get("MULTIAGENTE_SECRET_KEY")
    if key:
        return
    import secrets
    key = secrets.token_hex(32)
    try:
        with open(dotenv_path, "w", encoding="utf-8") as f:
            f.write(f"MULTIAGENTE_SECRET_KEY={key}\n")
    except OSError:
        print(f"[WARN] Could not write {dotenv_path} — key in memory only for this session.")
    os.environ["MULTIAGENTE_SECRET_KEY"] = key

_ensure_secret_key(_dotenv_path)

# ── Redirector de stdout al widget de log ───────────────────────────────
class OutputRedirector:
    """Captura stdout/stderr y los redirige al widget de log con colores."""
    def __init__(self, log_widget, app):
        self.log_widget = log_widget
        self.app = app
        self.buffer = ""

    def write(self, text):
        self.buffer += text
        if "\n" in self.buffer:
            lines = self.buffer.split("\n")
            self.buffer = lines.pop()
            for line in lines:
                if line.strip():
                    self.app.emit_log(line.strip())
        # flush on \n line endings

    def flush(self):
        if self.buffer.strip():
            self.app.emit_log(self.buffer.strip())
            self.buffer = ""

# ── Ventana Principal ───────────────────────────────────────────────────
class App(ctk.CTk, ImpresionMixin, PlanillasMixin, CorreosMixin, DescargaMixin):
    def __init__(self):
        super().__init__()

        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(__file__)
        self.config_file = os.path.join(base_dir, "ui_config.json")
        self.config = {}
        self._master_pw_cache = ""
        self._cargar_config()

        # ── Configuración de ventana (antes de mostrar) ───────────────
        self.title("Sistema de Automatización de Operaciones")
        self.minsize(900, 550)
        self.configure(fg_color=Palette.BG_MAIN)

        # Restaurar geometría guardada o usar 80% de la pantalla
        if "window_geo" in self.config:
            self.geometry(self.config["window_geo"])
        else:
            self._centrar_ventana()

        # Forzar centrado si la ventana quedó fuera de pantalla (ej. monitor más chico)
        self.update_idletasks()
        self.after(100, self._asegurar_visible)

        # Icono (si existe)
        icon_dir = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
        icon_path = os.path.join(icon_dir, "icono.ico")
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except Exception:
                pass

        # ── Contraseña maestra al iniciar ─────────────────────────────
        # Se ejecuta vía after() al final de __init__ para evitar conflictos
        pw = self._cfg_obtener("seguridad", "password", "")
        self._pw_inicio_valida = not pw  # si no hay pass, ya está validado

        # ── Cola de logs thread-safe ──────────────────────────────────
        self.log_queue = queue.Queue()
        self._poll_log_queue()

        # ── Estados ───────────────────────────────────────────────────
        self.tarea_activa = False
        self._cancelar_tarea = threading.Event()  # flag para cancelar tareas en curso
        self._log_ctx = threading.local()  # per-thread log panel context
        self.datos_planillas = []  # Resultados del agente 2
        self._mail_data = {}       # {item_id: {mid, subject, date, checked, downloaded}}
        self.panel_actual = ""      # Panel activo actual
        self._panel_frames = {}    # nombre -> CTkFrame, persistente entre cambios de panel
        self._resultados_pendientes = None  # Resultados para popup de resumen
        self._excel_com_ok = True
        self._super_auto = False          # Toggle de súper automatización
        self._super_guarda = ""           # Guarda elegido para súper auto
        self._ultimas_carpetas = []       # Carpetas descargadas en la última tanda
        self._backup_resultados = None    # Resultados del último backup
        self.logs_por_panel = {    # Historial de consola por agente
            "impresion": [],
            "planillas": [],
            "descargar": [],
            "correos": [],
            "backup": [],
            "cargar-datos": [],
        }
        # T7/T8: almacén de rutas CONTENEDOR + k por fila TreeView
        self._cargar_datos_rutas = {}
        self._cargar_datos_comparacion = {}  # iid → dict para popup comparación
        self._cargar_datos_idx = 0
        self._cf_auto_var = ctk.BooleanVar(value=False)  # Control Final auto mode
        self._ct_auto_var = ctk.BooleanVar(value=False)  # Control Tickets auto mode

        # ── Construir UI ──────────────────────────────────────────────
        self._crear_sidebar()
        self._crear_area_principal()

        # Seleccionar panel inicial
        self._cambiar_panel("descargar")

        # Bind de teclas
        self.bind("<Escape>", lambda e: self._confirmar_salida())
        self.protocol("WM_DELETE_WINDOW", self._confirmar_salida)
        # Destroy popups on minimize to prevent grab deadlock
        self._open_popups = set()
        self.bind("<Unmap>", self._on_main_unmap)

        # Verificar contraseña maestra al iniciar, luego diagnóstico
        self.after(300, self._verificar_inicio)

    def _centrar_ventana(self):
        """Centra la ventana en la pantalla actual al 80% del tamaño."""
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = min(int(sw * 0.8), 1280)
        h = min(int(sh * 0.8), 800)
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _asegurar_visible(self):
        """Si la ventana está fuera del área visible, la re-centra."""
        try:
            x = self.winfo_x()
            y = self.winfo_y()
            w = self.winfo_width()
            h = self.winfo_height()
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            ajustar = False
            if x + w < 100:  # muy a la izquierda o invisible
                x = max(0, (sw - w) // 2)
                ajustar = True
            if y + h < 50:  # muy arriba
                y = max(0, (sh - h) // 2)
                ajustar = True
            if x > sw - 50:  # muy a la derecha
                x = max(0, (sw - w) // 2)
                ajustar = True
            if ajustar:
                self.geometry(f"+{x}+{y}")
        except Exception:
            pass

    def _register_popup(self, popup):
        """Track a Toplevel popup so it's destroyed on minimize."""
        self._open_popups.add(popup)
        popup.bind("<Destroy>", lambda e, p=popup: self._open_popups.discard(p))

    def _on_main_unmap(self, event):
        """When main window is minimized, destroy all open popups to release grabs."""
        for popup in list(self._open_popups):
            try:
                popup.grab_release()
            except Exception:
                pass
            try:
                popup.destroy()
            except Exception:
                pass
        self._open_popups.clear()
        # Also destroy any untracked Toplevel children (safety net)
        for w in self.winfo_children():
            if isinstance(w, ctk.CTkToplevel):
                try:
                    w.grab_release()
                except Exception:
                    pass
                try:
                    w.destroy()
                except Exception:
                    pass

    def _verificar_inicio(self):
        if not self._pw_inicio_valida:
            self._mostrar_dialogo_password()

    def _mostrar_dialogo_password(self):
        pw_config = self._cfg_obtener("seguridad", "password", "")
        dlg = ctk.CTkToplevel(self)
        dlg.title("Acceso Restringido")
        dlg.geometry("400x215")
        dlg.configure(fg_color=Palette.BG_CARD)
        dlg.transient(self)
        dlg.lift()
        dlg.grab_set()
        dlg.minsize(400, 215)
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 215) // 2
        dlg.geometry(f"400x215+{px}+{py}")

        ctk.CTkLabel(
            dlg, text="Ingresá la contraseña para abrir la aplicación:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 6))

        entry_pw = ctk.CTkEntry(
            dlg, width=280, height=36, show="*",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6,
        )
        entry_pw.pack(pady=(0, 4))
        entry_pw.focus_set()
        dlg.after(50, lambda: entry_pw.focus_force())

        def verificar():
            if entry_pw.get() == pw_config:
                self._master_pw_cache = entry_pw.get()
                dlg.destroy()
                self._pw_inicio_valida = True
            else:
                lbl_status.configure(text="Contraseña incorrecta. Intentá de nuevo.", text_color=Palette.ERROR)
                entry_pw.delete(0, "end")
                entry_pw.focus_set()

        entry_pw.bind("<Return>", lambda e: verificar())

        # Botón de recuperación (entre la entry y los botones)
        def recuperar():
            self._recuperar_password(dlg, lbl_status)

        ctk.CTkButton(
            dlg, text="🔑  Recuperar Clave",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.WARNING, hover_color=Palette.WARNING_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=30, width=180,
            command=recuperar,
        ).pack(pady=(4, 2))

        # Un solo label para errores y recuperación
        lbl_status = ctk.CTkLabel(
            dlg, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_SECONDARY,
        )
        lbl_status.pack(pady=(0, 2))

        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(pady=(0, 4))
        ctk.CTkButton(
            btn_frame, text="Ingresar", width=110, height=36,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=verificar,
        ).pack(side="left", padx=5)
        ctk.CTkButton(
            btn_frame, text="Salir", width=110, height=36,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=lambda: (dlg.destroy(), self.destroy()),
        ).pack(side="left", padx=5)

    # ── Sidebar ──────────────────────────────────────────────────────────
    def _crear_sidebar(self):
        self.sidebar = ctk.CTkFrame(
            self, width=230, corner_radius=0, fg_color=Palette.BG_SIDEBAR
        )
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        # ── Logo / Título ────────────────────────────────────────────
        header = ctk.CTkFrame(
            self.sidebar, fg_color="transparent", height=80
        )
        header.pack(fill="x", padx=14, pady=(4, 0))
        header.pack_propagate(False)

        # Cargar icono de la aplicación
        try:
            icon_dir = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
            icon_path = os.path.join(icon_dir, "icono.ico")
            if os.path.exists(icon_path):
                pil_image = Image.open(icon_path)
                pil_image = pil_image.resize((72, 72), Image.LANCZOS)
                self._logo_icon = ctk.CTkImage(pil_image, size=(72, 72))
                ctk.CTkLabel(header, image=self._logo_icon, text="").pack(expand=True)
            else:
                ctk.CTkLabel(
                    header, text="MULTIAGENTE",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=20, weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).pack(anchor="w")
        except Exception:
            ctk.CTkLabel(
                header, text="MULTIAGENTE",
                font=ctk.CTkFont(family=FONT_FAMILY, size=20, weight="bold"),
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w")

        ctk.CTkFrame(
            self.sidebar, fg_color=Palette.DIVIDER, height=1
        ).pack(fill="x", padx=14, pady=(4, 4))

        # ── Zona de navegación con scroll ─────────────────────────────
        self._nav_scroll = ctk.CTkScrollableFrame(
            self.sidebar, fg_color="transparent",
            scrollbar_button_color=Palette.TEXT_MUTED,
            scrollbar_button_hover_color=Palette.TEXT_SECONDARY,
        )
        self._nav_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        nav_label = ctk.CTkLabel(
            self._nav_scroll,
            text="MÓDULOS",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
            text_color=Palette.TEXT_MUTED,
        )
        nav_label.pack(anchor="w", padx=14, pady=(6, 4))

        # ── Iconos Lucide (compartidos por todos los paneles) ─────────
        _icon_dir = os.path.join(os.path.dirname(__file__), "assets", "icons")
        _icon_size = (20, 20)
        self._icons = {}
        for _name in ["inbox", "file-text", "table", "clipboard-list", "send", "save",
                       "settings", "mail-search", "search", "mail", "cloud-download",
                       "refresh-cw", "bookmark", "printer", "play", "shield", "pencil",
                       "check", "x", "ticket", "zap", "folder-open", "hard-drive",
                       "folder", "banknote", "lock", "bot", "palette", "eye"]:
            self._icons[_name] = ctk.CTkImage(
                Image.open(os.path.join(_icon_dir, f"{_name}.png")), size=_icon_size
            )

        self.btn_descargar = self._crear_btn_nav(
            "Descargar Mails", "inbox", 0, lambda: self._cambiar_panel("descargar"),
            image=self._icons["inbox"],
        )
        self.btn_impresion = self._crear_btn_nav(
            "Impresión Documental", "file-text", 1, lambda: self._cambiar_panel("impresion"),
            image=self._icons["file-text"],
        )
        self.btn_planillas = self._crear_btn_nav(
            "Completar Planillas", "table", 2, lambda: self._cambiar_panel("planillas"),
            image=self._icons["table"],
        )
        self.btn_cargar_datos = self._crear_btn_nav(
            "Controlar Datos", "clipboard-list", 3, lambda: self._cambiar_panel("cargar-datos"),
            image=self._icons["clipboard-list"],
        )
        self.btn_correos = self._crear_btn_nav(
            "Enviar Correos", "send", 4, lambda: self._cambiar_panel("correos"),
            image=self._icons["send"],
        )
        self.btn_backup = self._crear_btn_nav(
            "Backup", "save", 5, lambda: self._cambiar_panel("backup"),
            image=self._icons["save"],
        )

        self.btn_ajustes = self._crear_btn_nav(
            "Ajustes", "settings", 6, lambda: self._cambiar_panel("ajustes"),
            image=self._icons["settings"],
        )

        # ── Separador ────────────────────────────────────────────────
        ctk.CTkFrame(
            self._nav_scroll, fg_color=Palette.DIVIDER, height=1
        ).pack(fill="x", padx=14, pady=(16, 10))

        # ── Súper Auto Toggle ────────────────────────────────────────

        ctk.CTkLabel(
            self._nav_scroll, text="AUTOMATIZACIÓN",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 4))

        self._super_switch = ctk.CTkSwitch(
            self._nav_scroll, text="⚡ Súper Auto",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            progress_color=Palette.ACCENT,
            button_color=Palette.TEXT_SECONDARY,
            button_hover_color=Palette.TEXT_PRIMARY,
            command=self._super_toggle,
        )
        self._super_switch.pack(anchor="w", padx=12, pady=(2, 2))

        self._super_lbl_guarda = ctk.CTkLabel(
            self._nav_scroll, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=10),
            text_color=Palette.TEXT_MUTED,
        )
        self._super_lbl_guarda.pack(anchor="w", padx=30, pady=(0, 6))

        # ── Botón Salir ──────────────────────────────────────────────
        self.btn_salir = ctk.CTkButton(
            self.sidebar,
            text="Salir del Sistema",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color="transparent",
            hover_color=Palette.ERROR_BG,
            text_color=Palette.TEXT_PRIMARY,
            border_color=Palette.BORDER,
            border_width=1,
            corner_radius=6,
            height=36,
            command=self._confirmar_salida,
        )
        self.btn_salir.pack(side="bottom", fill="x", padx=10, pady=10)

    def _crear_btn_nav(self, texto, icono, idx, comando, image=None):
        btn = ctk.CTkButton(
            self._nav_scroll,
            text=f"  {texto}",
            image=image, compound="left" if image else "none",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            fg_color="transparent",
            hover_color=Palette.BG_HOVER,
            text_color=Palette.WHITE,
            anchor="w",
            corner_radius=8,
            height=46,
            command=comando,
        )
        btn.pack(fill="x", padx=8, pady=2)
        btn._nav_idx = idx
        return btn

    def _marcar_nav_activo(self, idx):
        botones = [self.btn_descargar, self.btn_impresion, self.btn_planillas, self.btn_cargar_datos, self.btn_correos, self.btn_backup, self.btn_ajustes]
        for i, b in enumerate(botones):
            if i == idx:
                b.configure(
                    fg_color=Palette.ACCENT,
                    hover_color=Palette.ACCENT_HOVER,
                    text_color=Palette.WHITE,
                )
            else:
                b.configure(
                    fg_color="transparent",
                    hover_color=Palette.BG_HOVER,
                    text_color=Palette.WHITE,
                )

    # ── Área Principal ───────────────────────────────────────────────────
    def _crear_area_principal(self):
        self.main_area = ctk.CTkFrame(
            self, corner_radius=0, fg_color=Palette.BG_MAIN
        )
        self.main_area.pack(side="left", fill="both", expand=True)

        # ── Barra superior ───────────────────────────────────────────
        self.top_bar = ctk.CTkFrame(
            self.main_area, fg_color="transparent", height=56
        )
        self.top_bar.pack(fill="x", padx=16, pady=(16, 0))
        self.top_bar.pack_propagate(False)

        self.lbl_titulo_panel = ctk.CTkLabel(
            self.top_bar,
            text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=20, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        )
        self.lbl_titulo_panel.pack(side="left")

        # ── Controles OCR (mostrar solo en panel Control de Datos) ────
        self._ocr_topbar_frame = ctk.CTkFrame(
            self.top_bar, fg_color="transparent"
        )
        # Se packea/empaca dinámicamente en _cambiar_panel

        ctk.CTkLabel(
            self._ocr_topbar_frame,
            text="Modelo OCR:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(side="left")

        self._ocr_method_menu = ctk.CTkOptionMenu(
            self._ocr_topbar_frame,
            values=["Local", "API Visión"],
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_INPUT,
            button_color=Palette.ACCENT,
            button_hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            dropdown_fg_color=Palette.BG_CARD,
            dropdown_hover_color=Palette.BG_HOVER,
            dropdown_text_color=Palette.TEXT_PRIMARY,
            corner_radius=4, width=130, height=28,
            command=self._on_ocr_method_change,
        )
        self._ocr_method_menu.pack(side="left", padx=(4, 16))

        ctk.CTkLabel(
            self._ocr_topbar_frame,
            text="Modelo:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(side="left")

        self._modelo_vision_menu = ctk.CTkOptionMenu(
            self._ocr_topbar_frame,
            values=[],
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_INPUT,
            button_color=Palette.ACCENT,
            button_hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            dropdown_fg_color=Palette.BG_CARD,
            dropdown_hover_color=Palette.BG_HOVER,
            dropdown_text_color=Palette.TEXT_PRIMARY,
            corner_radius=4, width=200, height=28,
            state="disabled",
            command=self._on_modelo_vision_change,
        )
        self._modelo_vision_menu.pack(side="left")

        # Oculto por defecto — se muestra en _cambiar_panel para cargar-datos
        self._ocr_topbar_frame.pack_forget()

        self.lbl_badge = ctk.CTkLabel(
            self.top_bar,
            text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=10, weight="bold"),
            text_color=Palette.WHITE,
            fg_color=Palette.ACCENT,
            corner_radius=4,
            padx=8,
            pady=2,
        )
        # badge shown conditionally

        # ── Contenedor de paneles (Resizable PanedWindow) ────────────
        self.paned_window = tk.PanedWindow(
            self.main_area, orient="vertical",
            bg=Palette.BORDER, bd=0, sashwidth=4,
            relief="flat"
        )
        self.paned_window.pack(fill="both", expand=True, padx=16, pady=(8, 16))

        self.panel_container = ctk.CTkFrame(
            self.paned_window, fg_color="transparent"
        )
        self.paned_window.add(self.panel_container, minsize=200, stretch="always")

        # ── Área de log (inferior, compartida) ────────────────────────
        self.log_container = ctk.CTkFrame(self.paned_window, fg_color="transparent")
        self.paned_window.add(self.log_container, minsize=100, stretch="never")
        self._crear_panel_log(self.log_container)

        # Restaurar sash del panel inicial
        if "sash_planillas" in self.config:
            self.after(200, lambda: self.paned_window.sash_place(0, 0, self.config["sash_planillas"]))

    def _crear_panel_log(self, parent):
        log_frame = ctk.CTkFrame(
            parent, fg_color=Palette.BG_LOG, corner_radius=8,
            border_width=1, border_color=Palette.BORDER
        )
        log_frame.pack(fill="both", expand=True, padx=0, pady=(8, 0))

        # Header del log
        log_header = ctk.CTkFrame(log_frame, fg_color="transparent", height=28)
        log_header.pack(fill="x", padx=12, pady=(6, 0))
        log_header.pack_propagate(False)

        ctk.CTkLabel(
            log_header,
            text="CONSOLA DE SALIDA",
            font=ctk.CTkFont(family=FONT_FAMILY, size=15, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(side="left")

        self.lbl_log_count = ctk.CTkLabel(
            log_header,
            text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        )
        self.lbl_log_count.pack(side="right")

        self.btn_clear_log = ctk.CTkButton(
            log_header,
            text="Limpiar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color="transparent",
            hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            width=60,
            height=28,
            command=self._limpiar_log,
        )
        self.btn_clear_log.pack(side="right", padx=(0, 6))

        # Text widget con scroll
        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(family=FONT_MONO, size=13),
            fg_color=Palette.BG_LOG,
            text_color=Palette.TEXT_SECONDARY,
            border_width=0,
            corner_radius=0,
            wrap="word",
        )
        self.log_text.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        self.log_text.configure(state="disabled")

        # Configurar tags de color
        self.log_text.tag_config("error", foreground=Palette.ERROR)
        self.log_text.tag_config("success", foreground=Palette.SUCCESS)
        # warning eliminado — se usa color por defecto
        self.log_text.tag_config("info", foreground=Palette.INFO)

        self._log_line_count = 0
        self._log_tags = {}

    # ── Navegación de paneles ───────────────────────────────────────────
    def _cambiar_panel(self, nombre):
        """Navega al panel indicado sin verificar tareas activas."""
        # Verificar contraseña maestra para Ajustes (se hace en _forzado)
        self._cambiar_panel_forzado(nombre)

    def _cambiar_panel_forzado(self, nombre):
        # Verificar contraseña maestra para Ajustes
        if nombre == "ajustes" and not self._verificar_password_maestra():
            return

        # Guardar log del panel actual ANTES de ocultarlo
        if self.panel_actual and self.panel_actual in self.logs_por_panel:
            self.logs_por_panel[self.panel_actual] = self._capturar_lineas_log()

        # Ocultar el frame del panel actual (si existe)
        if self.panel_actual and self.panel_actual in self._panel_frames:
            self._panel_frames[self.panel_actual].pack_forget()

        # Construir el frame del panel destino si es primera vez
        if nombre not in self._panel_frames:
            if nombre == "impresion":
                self._panel_impresion()
            elif nombre == "planillas":
                self._panel_planillas()
            elif nombre == "descargar":
                self._panel_descargar()
            elif nombre == "correos":
                self._panel_correos()
            elif nombre == "backup":
                self._panel_backup()
            elif nombre == "ajustes":
                self._panel_ajustes()

        # "cargar-datos" se reintenta siempre — el guard interno detecta
        # si ya está completo (frame + tree_coordinacion) y sale rápido.
        if nombre == "cargar-datos":
            self._panel_cargar_datos()

        # Mostrar el frame del panel destino
        frame = self._panel_frames.get(nombre)
        if frame:
            frame.pack(in_=self.panel_container, fill="both", expand=True)

        # Reescanear carpetas al volver al panel de impresión
        if nombre == "impresion":
            self.after(50, self._imp_escanear_carpetas)

        # Actualizar título y navegación
        idx_map = {"descargar": 0, "impresion": 1, "planillas": 2, "cargar-datos": 3, "correos": 4, "backup": 5, "ajustes": 6}
        idx = idx_map.get(nombre, 1)
        self._marcar_nav_activo(idx)

        titles = {
            "impresion": "Impresión Documental",
            "planillas": "Completar Planillas",
            "descargar": "Descargar Mails",
            "correos": "Enviar Correos",
            "backup": "Backup",
            "cargar-datos": "Control de Datos",
            "ajustes": "Configuración del Sistema",
        }
        self.lbl_titulo_panel.configure(text=titles.get(nombre, ""))

        # Mostrar/ocultar controles OCR según panel
        if nombre == "cargar-datos":
            if not self._ocr_topbar_frame.winfo_ismapped():
                self._ocr_topbar_frame.pack(side="right")
                # Sincronizar método OCR desde config
                metodo_ocr = self.config.get("ocr_method", "api_vision")
                self._ocr_method_menu.set("API Visión" if metodo_ocr == "api_vision" else "Local")
                self._actualizar_modelos_vision()
                # Sincronizar state del modelo según método actual
                es_api = metodo_ocr == "api_vision"
                self._modelo_vision_menu.configure(state="normal" if es_api else "disabled")
        else:
            self._ocr_topbar_frame.pack_forget()

        if nombre == "planillas":
            self._actualizar_titulo_precintos()

        # Mostrar/ocultar consola según panel
        if nombre == "ajustes":
            self.log_container.pack_forget()
            # Sincronizar dropdown de modelo desde config actual
            if hasattr(self, '_ent_vision_model') and self._ent_vision_model.winfo_exists():
                import procesar_tickets
                modelo_default = self.config.get("api_vision", {}).get("model",
                                    procesar_tickets.MODELO_VISION_DEFAULT)
                modelos_guardados = self.config.get("api_vision", {}).get("custom_models", [])
                modelos = modelos_guardados if modelos_guardados else list(procesar_tickets.MODELOS_VISION)
                if modelo_default in modelos:
                    self._ent_vision_model.set(modelo_default)
                elif modelos:
                    self._ent_vision_model.set(modelos[0])
        else:
            if not self.log_container.winfo_ismapped():
                self.paned_window.add(self.log_container, minsize=100, stretch="never")

        # Guardar sash del panel anterior, restaurar sash del nuevo
        panel_anterior = self.panel_actual
        self.panel_actual = nombre
        if hasattr(self, 'paned_window') and self.paned_window.winfo_exists():
            try:
                sash_y = self.paned_window.sash_coord(0)[1]
                if panel_anterior and sash_y > 0:
                    self.config[f"sash_{panel_anterior}"] = sash_y
            except Exception:
                pass
            if f"sash_{nombre}" in self.config:
                self.after(150, lambda n=nombre: self._restaurar_sash(n))

        # Restaurar la consola guardada para este panel
        self._restaurar_log(nombre)

    def _restaurar_sash(self, nombre):
        try:
            self.paned_window.sash_place(0, 0, self.config.get(f"sash_{nombre}", 500))
        except Exception:
            pass

    # ═══════════════════════════════════════════════════════════════════
    # PANEL 1: IMPRESIÓN DOCUMENTAL
    # ═══════════════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════════════


    # ═══════════════════════════════════════════════════════════════════
    # SÚPER AUTO
    # ═══════════════════════════════════════════════════════════════════
    def _super_toggle(self):
        """Activa/desactiva el modo Súper Automatización."""
        if self._super_switch.get():
            guardas = self._cfg_obtener("valores", "guardas", ["Gonzalez", "Rodriguez", "Martinez", "Perez"])
            popup = ctk.CTkToplevel(self)
            popup.title("Súper Auto — Elegir Guarda")
            popup.geometry("380x180")
            popup.configure(fg_color=Palette.BG_CARD)
            popup.transient(self)
            popup.lift()
            popup.grab_set()
            popup.update_idletasks()
            px = self.winfo_x() + (self.winfo_width() - 380) // 2
            py = self.winfo_y() + (self.winfo_height() - 180) // 2
            popup.geometry(f"380x180+{px}+{py}")

            ctk.CTkLabel(popup, text="Seleccioná el guarda para Súper Auto:",
                         font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                         text_color=Palette.TEXT_PRIMARY).pack(pady=(20, 10))
            guarda_var = ctk.StringVar(value=guardas[0] if guardas else "")
            ctk.CTkOptionMenu(popup, variable=guarda_var, values=guardas,
                              font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                              fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                              button_hover_color=Palette.ACCENT_HOVER,
                              text_color=Palette.TEXT_PRIMARY, corner_radius=6,
                              width=220, height=34).pack(pady=(0, 12))

            def _confirmar():
                self._super_guarda = guarda_var.get()
                self._super_auto = True
                self._super_lbl_guarda.configure(text=f"Guarda: {self._super_guarda}")
                popup.destroy()

            def _cancelar():
                self._super_switch.deselect()
                self._super_auto = False
                self._super_guarda = ""
                self._super_lbl_guarda.configure(text="")
                popup.destroy()

            ctk.CTkButton(popup, text="Activar", width=100, height=32,
                          font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                          fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                          text_color=Palette.WHITE, corner_radius=6,
                          command=_confirmar).pack(side="left", padx=(100, 8))
            ctk.CTkButton(popup, text="Cancelar", width=100, height=32,
                          font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                          fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
                          text_color=Palette.TEXT_SECONDARY, corner_radius=6,
                          command=_cancelar).pack(side="left")
        else:
            self._super_auto = False
            self._super_guarda = ""
            self._super_lbl_guarda.configure(text="")

    def _super_ejecutar_cadena(self, resultados):
        """Ejecuta la cadena: imprimir → guarda → planillas."""
        carpetas = []
        for _, _, nombre_carpeta in resultados:
            escritorio = self._resolver_ruta("descarga_mails", "Desktop")
            ruta = os.path.join(escritorio, nombre_carpeta)
            if os.path.isdir(ruta):
                carpetas.append(ruta)

        if not carpetas:
            self._log("⚡ Súper Auto: sin carpetas para procesar.")
            return

        self._log(f"⚡ Súper Auto: procesando {len(carpetas)} carpetas...")
        self._ultimas_carpetas = carpetas

        def _cadena():
            cfg = self._cfg_obtener("super_auto", "pasos", {})
            res_imp = self._super_imprimir(carpetas, cfg) if cfg.get("sobre", True) or cfg.get("permiso", True) or cfg.get("hoja_ruta", True) or cfg.get("recibo_ata", True) else 0
            res_gua = self._super_aplicar_guarda(carpetas) if cfg.get("aplicar_guarda", True) else 0
            res_pla = self._super_completar_planillas() if cfg.get("completar_planillas", True) else False
            self.after(0, lambda: self._super_popup_final(res_imp, res_gua, res_pla))

        t = threading.Thread(target=_cadena, daemon=True)
        t.start()

    def _super_imprimir(self, carpetas, cfg=None):
        """Imprime usando la misma lógica que el panel de Impresión Documental."""
        if cfg is None:
            cfg = {}
        hacer_sobre    = cfg.get("sobre", True)
        hacer_permiso  = cfg.get("permiso", True)
        hacer_hoja_ruta = cfg.get("hoja_ruta", True)
        hacer_recibo   = cfg.get("recibo_ata", True)
        anio = datetime.now().strftime("%y")
        prefijo_permiso = f"{anio}069EC"
        impresora = self._detectar_impresoras()[0] if self._detectar_impresoras() else "Default"
        total_ok = 0

        for ruta in carpetas:
            nombre = os.path.basename(ruta)
            self.log_queue.put(f"[...] 📁 {nombre}")
            try:
                archivos = sorted(os.listdir(ruta))
            except Exception:
                continue

            # Detectar archivos (misma lógica que _imp_worker)
            sobres = [a for a in archivos if "CONTENEDORES" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]
            permisos = [a for a in archivos if a.upper().startswith(prefijo_permiso) and a.upper().endswith(".PDF")]
            hojas_ruta = [a for a in archivos if "HOJA" in a.upper() and "RUTA" in a.upper() and (a.upper().endswith(".XLSX") or a.upper().endswith(".XLS"))]

            # 1. SOBRE: imprimir hoja "SOBRE" de Contenedores (1 copia)
            if hacer_sobre and sobres:
                for sobre in sobres:
                    ruta_excel = os.path.join(ruta, sobre)
                    hojas = self._imp_hojas_sobre(ruta_excel)
                    if hojas:
                        self.log_queue.put(f"[...]   📄 Sobre: {sobre} → {hojas}")
                        ok = self._imp_enviar(ruta_excel, impresora, f"  Sobre ({sobre[:40]})", hojas=hojas, copias=1)
                        if ok:
                            total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin archivo Contenedores")

            # 2. Permisos (2 copias)
            if hacer_permiso and permisos:
                for a in permisos:
                    self.log_queue.put(f"[...]   📄 Permiso: {a} (2 copias)")
                    ok = self._imp_enviar(os.path.join(ruta, a), impresora, f"  Permiso ({a[:40]})", copias=2)
                    if ok:
                        total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Permiso (busca: {prefijo_permiso}*.PDF)")

            # 3. Hojas de Ruta (2 copias)
            if hacer_hoja_ruta and hojas_ruta:
                for a in hojas_ruta:
                    self.log_queue.put(f"[...]   📄 Hoja Ruta: {a} (2 copias)")
                    ok = self._imp_enviar(os.path.join(ruta, a), impresora, f"  Hoja Ruta ({a[:40]})", copias=2)
                    if ok:
                        total_ok += 1
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Hoja de Ruta")

            # 4. Servicio ATA / Recibo ATA: hojas del Excel Contenedores
            if hacer_recibo and sobres:
                tipo = self._detectar_tipo_carpeta(nombre)
                if tipo in ("ISO", "FLEXI"):
                    self.log_queue.put(f"[...]   ⏭ Saltando Recibo ATA: carpeta marítima ({tipo})")
                else:
                    nombres_ata = ["RECIBO ATA"] + [f"RECIBO ATA {i}" for i in range(2, 9)]
                    hojas_ata = []
                    for sobre in sobres:
                        ruta_excel = os.path.join(ruta, sobre)
                        try:
                            if sobre.lower().endswith(".xlsx"):
                                wb_tmp = self._abrir_excel_seguro(ruta_excel)
                                for sn in wb_tmp.sheetnames:
                                    if sn.upper() in [n.upper() for n in nombres_ata]:
                                        hojas_ata.append((sobre, sn))
                                wb_tmp.close()
                            else:
                                book = xlrd.open_workbook(ruta_excel)
                                for sn in book.sheet_names():
                                    if sn.upper() in [n.upper() for n in nombres_ata]:
                                        hojas_ata.append((sobre, sn))
                        except Exception:
                            pass
                    if hojas_ata:
                        for archivo, hoja in hojas_ata:
                            self.log_queue.put(f"[...]   📄 Recibo ATA: {archivo} → {hoja}")
                            ok = self._imp_enviar(os.path.join(ruta, archivo), impresora, f"  Recibo ATA ({archivo[:30]})", hojas=[hoja])
                            if ok:
                                total_ok += 1
                    else:
                        self.log_queue.put(f"[...]   ⚠ Sin hoja Recibo ATA en Contenedores")
            else:
                self.log_queue.put(f"[...]   ⚠ Sin Contenedores para Recibo ATA")

        self.log_queue.put(f"[...] ✓ Impresión: {total_ok} documentos enviados de {len(carpetas)} carpetas")
        return total_ok

    def _super_aplicar_guarda(self, carpetas):
        """Escribe el guarda en los archivos Contenedores de las carpetas."""
        aplicados = 0
        for carpeta in carpetas:
            try:
                archivos = os.listdir(carpeta)
            except Exception:
                continue
            for archivo in archivos:
                up = archivo.upper()
                if "CONTENEDORES" not in up or not (up.endswith(".XLSX") or up.endswith(".XLS")):
                    continue
                ruta = os.path.join(carpeta, archivo)
                es_xls = ruta.lower().endswith(".xls") and not ruta.lower().endswith(".xlsx")
                # Reintentos automáticos: el archivo puede estar bloqueado brevemente
                # si COM lo usó para imprimir justo antes.
                max_reintentos = 6
                ultimo_error = None
                for _reintento in range(max_reintentos):
                    try:
                        _abrir_ok = True
                        if es_xls:
                            import xlrd as xlrd_local
                            xlrd_local.open_workbook(ruta, formatting_info=True).release_resources()
                        else:
                            openpyxl.load_workbook(ruta, read_only=True).close()
                        break
                    except PermissionError as _pe:
                        _abrir_ok = False
                        ultimo_error = _pe
                        if _reintento < max_reintentos - 1:
                            self.log_queue.put(f"[...]   ⏳ {archivo[:40]} en uso, esperando... ({_reintento+1}/{max_reintentos-1})")
                            time.sleep(1.5)
                else:
                    self.log_queue.put(f"[...]   ⚠ Error guarda en {archivo}: {ultimo_error}")
                    continue
                try:
                    if not self._escribir_guarda_en_archivo(ruta, self._super_guarda, self.log_queue.put):
                        self.log_queue.put(f"[...]   ⚠ 'Guarda' no hallada en {archivo[:50]}")
                    else:
                        aplicados += 1
                except Exception as e:
                    self.log_queue.put(f"[...]   ⚠ Error guarda en {archivo}: {e}")
                    import traceback
                    self.log_queue.put(f"[...]     {traceback.format_exc()}")
        self.log_queue.put(f"[...] ✓ Guarda '{self._super_guarda}': {aplicados} planillas actualizadas")
        return aplicados

    def _escribir_guarda_en_archivo(
        self, ruta: str, guarda_nombre: str, log_func=None
    ) -> bool:
        """Write guarda_nombre to column H in the CHOFER sheet of an Excel file.

        Handles both .xls (Excel COM via win32com) and .xlsx (openpyxl) formats.
        Returns True if written, False if no 'GUARDA' found or no CHOFER sheet.
        """
        es_xls = ruta.lower().endswith(".xls") and not ruta.lower().endswith(".xlsx")
        escrito = False

        if es_xls:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    wb = excel.Workbooks.Open(ruta)
                    ws_chofer = None
                    for s in wb.Sheets:
                        if "CHOFER" in s.Name.upper():
                            ws_chofer = s
                            break
                    if not ws_chofer:
                        if log_func:
                            log_func("[...]   Hoja 'Choferes' no hallada")
                        return False
                    for row in range(2, 16):
                        val = ws_chofer.Cells(row, 7).Value  # col 7 = G
                        if val is not None and "GUARDA" in str(val).strip().upper():
                            ws_chofer.Cells(row, 8).Value = guarda_nombre  # col 8 = H
                            escrito = True
                            break
                    wb.Save()
                    wb.Close()
                finally:
                    excel.Quit()
            finally:
                pythoncom.CoUninitialize()
        else:
            wb = self._abrir_excel_seguro(ruta)
            try:
                ws_chofer = None
                for s in wb.sheetnames:
                    if "CHOFER" in s.upper():
                        ws_chofer = wb[s]
                        break
                if not ws_chofer:
                    if log_func:
                        log_func("[...]   Hoja 'Choferes' no hallada")
                    return False
                for row in range(2, 16):
                    cell = ws_chofer.cell(row=row, column=7)  # col 7 = G (1-indexed)
                    val = cell.value
                    if val is None:
                        for mr in ws_chofer.merged_cells.ranges:
                            if (mr.min_col <= 7 <= mr.max_col and mr.min_row <= row <= mr.max_row):
                                val = ws_chofer.cell(row=mr.min_row, column=mr.min_col).value
                                break
                    if val is not None and "GUARDA" in str(val).strip().upper():
                        ws_chofer.cell(row=row, column=8).value = guarda_nombre  # col 8 = H
                        escrito = True
                        break
                self._guardar_excel_seguro(wb, ruta)
            finally:
                wb.close()

        return escrito

    def _super_completar_planillas(self):
        """Ejecuta Completar Planillas con los datos extraídos."""
        try:
            self._planillas_core()
            self.log_queue.put(f"[...] ✓ Contenedores completadas (SOBRES, COBRO, PC)")
            return True
        except Exception as e:
            self.log_queue.put(f"[...] ⚠ Error completando planillas: {e}")
            return False

    def _super_popup_final(self, impresos, guardas, planillas_ok):
        """Popup resumen del Súper Auto."""
        popup = ctk.CTkToplevel(self)
        popup.title("Súper Auto Completado")
        popup.geometry("400x300")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 400) // 2
        py = self.winfo_y() + (self.winfo_height() - 300) // 2
        popup.geometry(f"400x300+{px}+{py}")

        ctk.CTkLabel(popup, text="⚡ SÚPER AUTO COMPLETADO",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=18, weight="bold"),
                     text_color=Palette.TEXT_PRIMARY).pack(pady=(24, 20))

        ctk.CTkLabel(popup, text=f"🖨  {impresos} archivos impresos",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS).pack(pady=3)
        txt_guarda = f"🛡  Guarda '{self._super_guarda}' aplicado a {guardas} planillas" if guardas > 0 else f"⚠  Guarda '{self._super_guarda}' — no se encontró etiqueta GUARDA"
        ctk.CTkLabel(popup, text=txt_guarda,
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS if guardas > 0 else Palette.WARNING).pack(pady=3)
        ctk.CTkLabel(popup, text=f"{'✓' if planillas_ok else '✗'} Planillas completadas (SOBRES, COBRO, PC)",
                     font=ctk.CTkFont(family=FONT_FAMILY, size=13),
                     text_color=Palette.SUCCESS if planillas_ok else Palette.ERROR).pack(pady=3)

        ctk.CTkButton(popup, text="Cerrar", width=120, height=34,
                      font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                      fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                      text_color=Palette.WHITE, corner_radius=6,
                      command=popup.destroy).pack(pady=(20, 16))

    # ═══════════════════════════════════════════════════════════════════
    # PANEL: BACKUP
    # ═══════════════════════════════════════════════════════════════════
    def _panel_backup(self):
        if "backup" in self._panel_frames:
            return
        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        self._panel_frames["backup"] = frame
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── Toolbar ──────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=44
        )
        toolbar.pack(fill="x", pady=(0, 6))
        toolbar.pack_propagate(False)

        self.btn_backup_pendrive = ctk.CTkButton(
            toolbar,
            text="Back Up",
            image=self._icons["hard-drive"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.SECONDARY, hover_color=Palette.SECONDARY_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=170,
            command=self._backup_pendrive_iniciar,
        )
        self.btn_backup_pendrive.pack(side="left", padx=4, pady=4)

        self.lbl_estado_backup = ctk.CTkLabel(
            toolbar,
            text="Listo para respaldar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_PRIMARY,
        )
        self.lbl_estado_backup.pack(side="left", padx=(8, 0))

        self.progress_backup = ctk.CTkProgressBar(
            toolbar, width=140, height=8, corner_radius=4,
            fg_color=Palette.BG_INPUT, progress_color=Palette.ACCENT,
        )
        self.progress_backup.pack(side="right", padx=12)
        self.progress_backup.set(0)

        # ── Lista de carpetas con checkboxes ─────────────────────────
        self._backup_scroll = ctk.CTkScrollableFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=180)
        self._backup_scroll.pack(fill="both", expand=True, pady=(0, 6))
        self._backup_checks = {}
        self._refrescar_lista_backup()

    def _refrescar_lista_backup(self):
        """Actualiza la lista de carpetas del escritorio con checkboxes."""
        for w in self._backup_scroll.winfo_children():
            w.destroy()
        self._backup_checks.clear()
        escritorio = self._resolver_ruta("planillas_carga", "Desktop")
        carpetas = []
        try:
            for item in sorted(os.listdir(escritorio)):
                ruta = os.path.join(escritorio, item)
                if not os.path.isdir(ruta) or item.startswith("."):
                    continue
                for archivo in os.listdir(ruta):
                    up = archivo.upper()
                    if "CONTENEDORES" in up or (up.startswith("PLANILLA DE CARGA") and (up.endswith(".XLSX") or up.endswith(".XLS"))):
                        carpetas.append((item, ruta)); break
        except Exception:
            pass
        if not carpetas:
            ctk.CTkLabel(self._backup_scroll,
                         text=f"No se encontraron carpetas de carga en:\n{escritorio}",
                         font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                         text_color=Palette.TEXT_MUTED, justify="center",
                         wraplength=400).pack(expand=True, pady=30)
        for nombre, _ in carpetas:
            var = ctk.BooleanVar(value=True)
            self._backup_checks[nombre] = var
            ctk.CTkCheckBox(self._backup_scroll, text=nombre, variable=var,
                            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                            text_color=Palette.TEXT_PRIMARY).pack(anchor="w", padx=10, pady=3)

    # ── Backup Pendrive ──────────────────────────────────────────────
    def _backup_pendrive_iniciar(self):
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        # Obtener carpetas chequeadas
        seleccionadas = []
        escritorio = self._resolver_ruta("planillas_carga", "Desktop")
        for nombre, var in self._backup_checks.items():
            if var.get():
                ruta = os.path.join(escritorio, nombre)
                if os.path.isdir(ruta):
                    seleccionadas.append(ruta)
        if not seleccionadas:
            messagebox.showwarning("Nada seleccionado", "Seleccioná al menos una carpeta.")
            return
        self._ejecutar_backup(seleccionadas)

    def _ejecutar_backup(self, seleccionadas):
        """Ejecuta el backup con las carpetas elegidas."""
        self.tarea_activa = True
        self._cancelar_tarea.clear()
        self.btn_backup_pendrive.configure(text="⏳  Moviendo...", state="disabled")
        self.progress_backup.configure(mode="indeterminate")
        self.progress_backup.start()
        self.lbl_estado_backup.configure(text="Detectando pendrive...", text_color=Palette.INFO)
        self._limpiar_log()
        self._log("💿 Iniciando Back Up...")
        t = threading.Thread(target=lambda: self._backup_pendrive_worker(seleccionadas), daemon=True)
        t.start()

    def _backup_pendrive_worker(self, carpetas=None):
        self._set_log_panel("backup")
        try:
            from ctypes import windll
            meses_es = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
                        7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
            ruta_base = self._cfg_obtener_rutas("backup_pendrive", "TRABAJO\\CARGAS")

            pendrive_encontrado = None
            # Primero buscar un pendrive que YA TENGA la carpeta base (evita Ventoy)
            for letra_ascii in range(ord('D'), ord('Z') + 1):
                unidad = f"{chr(letra_ascii)}:\\"
                try:
                    tipo = windll.kernel32.GetDriveTypeW(unidad)
                    if tipo == 2:  # DRIVE_REMOVABLE
                        test_path = os.path.join(unidad, ruta_base)
                        if os.path.exists(test_path):
                            pendrive_encontrado = unidad
                            self.log_queue.put(f"[...] ✓ Pendrive detectado con {ruta_base}: {pendrive_encontrado}")
                            break
                except Exception:
                    pass
            # Si ninguno tiene la carpeta, usar el primer removible
            if not pendrive_encontrado:
                for letra_ascii in range(ord('D'), ord('Z') + 1):
                    unidad = f"{chr(letra_ascii)}:\\"
                    try:
                        tipo = windll.kernel32.GetDriveTypeW(unidad)
                        if tipo == 2:
                            pendrive_encontrado = unidad
                            self.log_queue.put(f"[...] ✓ Pendrive (nuevo): {pendrive_encontrado}")
                            break
                    except Exception:
                        pass

            if not pendrive_encontrado:
                self.log_queue.put("[...] ⚠ No se detectó pendrive removible (D-Z)")
                self.after(0, self._backup_done)
                return

            self.log_queue.put(f"[...] ✓ Pendrive: {pendrive_encontrado}")

            self.after(0, lambda: self.progress_backup.configure(mode="determinate"))
            self.after(0, lambda: self.progress_backup.set(0))
            total = len(carpetas)
            nombres_movidos = []
            destinos_usados = set()
            for i, carpeta in enumerate(carpetas):
                if self._cancelar_tarea.is_set():
                    self.log_queue.put("[...] ⚠ Tarea cancelada."); break
                nombre = os.path.basename(carpeta)
                # Extraer año y mes de cada carpeta individualmente
                año = str(datetime.now().year)
                mes = datetime.now().month
                m = re.match(r"(\d{1,2})_(\d{1,2})_(\d{4})", nombre)
                if m:
                    mes = int(m.group(2))
                    año = m.group(3)
                mes_str = f"{mes:02d}_{meses_es.get(mes, meses_es[datetime.now().month])}"
                carpeta_destino = os.path.join(pendrive_encontrado, ruta_base, f"Cargas_{año}", mes_str)
                os.makedirs(carpeta_destino, exist_ok=True)
                destino = os.path.join(carpeta_destino, nombre)
                if os.path.exists(destino):
                    shutil.rmtree(destino)
                shutil.move(carpeta, destino)
                self.log_queue.put(f"[...]   ✓ {nombre} → {ruta_base}\\Cargas_{año}\\{mes_str}\\")
                nombres_movidos.append(nombre)
                destinos_usados.add(f"{ruta_base}\\Cargas_{año}\\{mes_str}")
                self.after(0, lambda p=i+1, t=total: self.progress_backup.set(p/t if t else 1))

            destinos_str = ", ".join(sorted(destinos_usados))
            self.log_queue.put(f"[...] ✓ COMPLETADO: Back Up — {len(carpetas)} carpetas movidas → {pendrive_encontrado}{destinos_str}")
            # Guardar para el popup
            self._backup_resultados = {"carpetas": nombres_movidos, "destino": f"{pendrive_encontrado}{destinos_str}"}
        except Exception as e:
            self.log_queue.put(f"[...] ⚠ Error: {e}")
            self._backup_resultados = None
        finally:
            self.after(0, self._backup_done)

    def _backup_done(self):
        self.tarea_activa = False
        try:
            self.btn_backup_pendrive.configure(text="Back Up", state="normal", fg_color=Palette.SECONDARY)
            self.progress_backup.stop()
            self.progress_backup.set(1)
            self.lbl_estado_backup.configure(text="Backup completado", text_color=Palette.SUCCESS)
            self._refrescar_lista_backup()
        except (AttributeError, Exception):
            pass

        # Popup de resultados
        res = getattr(self, '_backup_resultados', None)
        if res and res.get("carpetas"):
            carpetas = res["carpetas"]
            destino = res["destino"]
            detalle = ""
            for c in carpetas:
                detalle += f"  ▸ {c}\n"
            self.after(300, lambda d=detalle, t=len(carpetas), dest=destino: messagebox.showinfo(
                "💿 Back Up Completado",
                f"{t} carpeta(s) movida(s) al pendrive:\n\n"
                f"{d}\n"
                f"📁 Destino: {dest}"))

    # ═══════════════════════════════════════════════════════════════════
    # ═══════════════════════════════════════════════════════════════════

    def _mostrar_resumen(self, resultados):
        """Popup lindo con resumen de lo procesado (SOBRES, COBRO, PC)."""
        popup = ctk.CTkToplevel(self)
        popup.title("Resultado del Análisis")
        popup.geometry("440x340")
        popup.configure(fg_color=Palette.BG_CARD)
        popup.transient(self)
        popup.lift()

        # Centrar sobre la ventana principal
        popup.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 440) // 2
        py = self.winfo_y() + (self.winfo_height() - 340) // 2
        popup.geometry(f"440x340+{px}+{py}")

        # Título
        ctk.CTkLabel(
            popup, text="ANÁLISIS COMPLETADO",
            font=ctk.CTkFont(family=FONT_FAMILY, size=18, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(24, 20))

        # Items
        items = [
            ("SOBRES", resultados.get("sobres", {})),
            ("COBRO", resultados.get("cobro", {})),
            ("PC (Precintos)", resultados.get("pc", {})),
        ]

        for nombre, res in items:
            ok = res.get("ok", False)
            detalle = res.get("detalle", "No procesado")
            icono = "✓" if ok else "✗"
            color = Palette.SUCCESS if ok else Palette.ERROR

            row_frame = ctk.CTkFrame(popup, fg_color="transparent")
            row_frame.pack(fill="x", padx=32, pady=4)

            ctk.CTkLabel(
                row_frame, text=icono,
                font=ctk.CTkFont(family=FONT_FAMILY, size=20, weight="bold"),
                text_color=color, width=30,
            ).pack(side="left")

            ctk.CTkLabel(
                row_frame, text=nombre,
                font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
                text_color=Palette.TEXT_PRIMARY, width=120, anchor="w",
            ).pack(side="left")

            ctk.CTkLabel(
                row_frame, text=detalle,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.TEXT_SECONDARY,
            ).pack(side="left")

        # Botón cerrar
        ctk.CTkButton(
            popup, text="Cerrar", width=120, height=36,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=popup.destroy,
        ).pack(pady=(24, 16))

    def _abrir_excel_seguro(self, ruta, solo_lectura=False):
        """Abre un archivo Excel. Si está abierto, pide cerrarlo y reintenta."""
        nombre = os.path.basename(ruta)
        while True:
            try:
                return openpyxl.load_workbook(ruta, data_only=False, read_only=solo_lectura)
            except PermissionError:
                if not preguntar_reintentar(nombre, parent=self):
                    raise

    def _guardar_excel_seguro(self, wb, ruta):
        """Guarda un archivo Excel. Si está abierto, pide cerrarlo y reintenta."""
        nombre = os.path.basename(ruta)
        while True:
            try:
                wb.save(ruta)
                return
            except PermissionError:
                if not preguntar_reintentar(nombre, parent=self):
                    raise

    # ═══════════════════════════════════════════════════════════════════
    # SISTEMA DE LOG
    # ═══════════════════════════════════════════════════════════════════
    def _log(self, mensaje):
        """Agrega un mensaje al log desde cualquier hilo, etiquetado con el panel de origen."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        panel = getattr(self._log_ctx, "panel", self.panel_actual) or "general"
        self.log_queue.put(("_LOG_", panel, f"[{timestamp}] {mensaje}"))

    def _set_log_panel(self, panel):
        """Set the panel context for the current thread. All _log() calls from this thread
        will be tagged with this panel name."""
        self._log_ctx.panel = panel

    def _log_warning(self, mensaje):
        """Advertencia suprimida — no se muestra en el log."""
        pass

    def _log_error(self, mensaje):
        """Agrega un mensaje de error al log (hilo principal)."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.emit_log(f"[{timestamp}] ERROR: {mensaje}")

    def emit_log(self, texto):
        """Inserta texto coloreado en el widget de log (hilo principal)."""
        self.log_text.configure(state="normal")
        if self._log_line_count > 0:
            self.log_text.insert("end", "\n")
            
        tag = None
        txt_up = texto.upper()
        if "ERROR" in txt_up or "CRÍTICO" in txt_up or "NO SE PUDO" in txt_up:
            tag = "error"
        elif "ÉXITO" in txt_up or "FINALIZADO" in txt_up or "CORRECTAMENTE" in txt_up or "COMPLETADO" in txt_up:
            tag = "success"
        # ADVERTENCIA, FALTA, NO SE ENCONTRÓ → color por defecto (eliminado el amarillo)
        elif "ESCANEANDO" in txt_up or "PREPARANDO" in txt_up or "CONECTANDO" in txt_up:
            tag = "info"

        if tag:
            self.log_text.insert("end", texto, tag)
        else:
            self.log_text.insert("end", texto)
            
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self._log_line_count += 1

        # Actualizar contador
        self.lbl_log_count.configure(text=f"{self._log_line_count} líneas")

        # Limitar líneas a 500
        if self._log_line_count > 500:
            self._trim_log()

    def _poll_log_queue(self):
        """Procesa la cola de mensajes de log periódicamente.
        
        Soporta strings (emit_log directo) y tuplas:
        - ("_LOG_", panel, text) → log etiquetado por panel
        - (callable, mensaje) → llama callable(mensaje) en hilo principal
        - ("_OCR_RESULT_", data) → resultado de OCR — pobla TreeView (T7)
        - ("_OCR_DONE_", None) → señal de fin de OCR (T7+)
        
        Procesa max 20 msg/tick para no congelar el main thread.
        """
        try:
            _processed = 0
            while _processed < 20:
                msg = self.log_queue.get_nowait()
                if isinstance(msg, tuple):
                    if len(msg) == 3 and msg[0] == "_LOG_":
                        # Tagged log message: ("_LOG_", panel, text)
                        panel, text = msg[1], msg[2]
                        # Always store in the panel's persistent buffer
                        self.logs_por_panel.setdefault(panel, []).append(text)
                        # Only write to visible textbox if this panel is active
                        if panel == self.panel_actual:
                            self.emit_log(text)
                    elif len(msg) == 2:
                        if callable(msg[0]):
                            msg[0](msg[1])
                        elif msg[0] == "_OCR_RESULT_":
                            self._procesar_resultado_ocr(msg[1])
                        elif msg[0] == "_OCR_DONE_":
                            self._cargar_datos_done()
                        elif msg[0] == "_CONTROL_FINAL_RESULT_":
                            self._procesar_resultado_control_final(msg[1])
                        elif msg[0] == "_TAREA_COMPLETA_":
                            self._finalizar_tarea()
                else:
                    # Plain string fallback — tag with current panel
                    self.logs_por_panel.setdefault(self.panel_actual, []).append(msg)
                    self.emit_log(msg)
                _processed += 1
        except queue.Empty:
            pass
        self.after(100, self._poll_log_queue)

    @staticmethod
    def _fmt_kg(v):
        """Format kg value: no .0, with thousands separator. E.g. 15690 → '15.690'"""
        try:
            n = float(str(v).replace(',', '.'))
            return f"{int(n):,}".replace(",", ".")
        except (ValueError, TypeError):
            return str(v) if v else "0"

    @staticmethod
    def _fmt_dni(v):
        """Format DNI with thousands separator. E.g. 27981220 → '27.981.220'"""
        try:
            digits = ''.join(c for c in str(v) if c.isdigit())
            n = int(digits)
            return f"{n:,}".replace(",", ".")
        except (ValueError, TypeError):
            return str(v) if v else ""

    def _procesar_resultado_ocr(self, data):
        """Inserta una fila por ticket con color verde/rojo según match.
        Almacena datos de comparación para abrir detalle al hacer click."""
        ticket = data.get("ticket", {})
        cont_data = data.get("contenedor")
        ruta_match = data.get("match")
        shared_excels = data.get("shared_excels", [])
        shared_neto_sum = data.get("shared_neto_sum", 0)

        # Valores OCR
        archivo   = ticket.get("archivo", "")
        patente   = ticket.get("patente", "")
        semi      = ticket.get("semi", "")
        conductor = ticket.get("conductor", "")
        dni       = ticket.get("dni", "")
        neto_ocr  = ticket.get("neto", 0)
        tara_ocr  = ticket.get("tara", 0)
        contenedor_str = ticket.get("contenedor", "")
        contenedor_display = contenedor_str if contenedor_str else "-"
        permiso   = ticket.get("permiso", "")

        import re as _re

        def _normalizar_pat(p):
            """Normalización permisiva para MATCHING (buscar fila de Excel).
            L→I, 0→O, 1→I, 8→B, 5→S para tolerar errores comunes de OCR."""
            p = p.upper().strip()
            p = p.replace('0', 'O').replace('1', 'I').replace('L', 'I').replace('8', 'B').replace('5', 'S')
            p = _re.sub(r'[^A-Z0-9]', '', p)
            return p

        def _normalizar_txt(t):
            return str(t).upper().strip()

        def _comparar_num(a, b):
            try:
                return abs(float(a) - float(b)) < 1
            except (ValueError, TypeError):
                return False

        # Buscar camión matching en cont_data
        camion_match = None
        k_idx = None
        if cont_data and cont_data.get("camiones"):
            patente_norm = _normalizar_pat(patente)
            semi_norm    = _normalizar_pat(semi) if semi else ""
            dni_norm     = _re.sub(r'\D', '', str(dni).strip()) if dni else ""
            cond_norm    = _normalizar_txt(conductor) if conductor else ""
            cont_norm    = _normalizar_txt(contenedor_str) if contenedor_str else ""

            self._log(f"[Control Datos] Ticket: patente='{patente}' semi='{semi}' dni='{dni}' cond='{conductor}' cont='{contenedor_str}'")
            for camion in cont_data["camiones"]:
                cam_pat  = _normalizar_pat(camion.get("patente_camion", ""))
                cam_semi = _normalizar_pat(camion.get("patente_semi", ""))
                cam_dni  = _re.sub(r'\D', '', str(camion.get("dni", "")).strip())
                cam_cond = _normalizar_txt(camion.get("conductor", ""))
                cam_cont = _normalizar_txt(str(camion.get("contenedor", "")))

                def _nombres_equivalentes(a, b):
                    """True si al menos 2 palabras coinciden entre ambos nombres
                    (sin importar acentos, mayúsculas, puntuación, ni palabras extra).
                    
                    Ej: "RETA JORGE SEBASTIAN" y "RETA JORGE SEBASTIÁN" → 3/3 ✅
                        "CLAUDIO PAEZ" y "CLAUDIO CESAR PAEZ" → 2/3 ≥ 2 ✅
                        "JUAN PEREZ" y "CLAUDIO PAEZ" → 0 ❌
                    """
                    if not a or not b:
                        return False
                    import unicodedata
                    def _limpiar(s):
                        # Quitar acentos, puntuación, mayúsculas
                        s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
                        return re.sub(r'[^\w\s]', ' ', s).strip().upper()
                    a_set = set(_limpiar(a).split())
                    b_set = set(_limpiar(b).split())
                    if not a_set or not b_set:
                        return False
                    return len(a_set & b_set) >= 2

                # Intentar match por cualquier campo
                razon = ""
                if patente_norm and cam_pat and patente_norm == cam_pat:
                    razon = f"patente_camion ({patente})"
                elif semi_norm and cam_semi and semi_norm == cam_semi:
                    razon = f"patente_semi ({semi})"
                elif dni_norm and cam_dni and dni_norm == cam_dni:
                    razon = "DNI"
                elif cond_norm and cam_cond and _nombres_equivalentes(cond_norm, cam_cond):
                    razon = f"conductor ({conductor})"
                elif cont_norm and cam_cont and cont_norm == cam_cont:
                    razon = f"contenedor ({contenedor_str})"

                if razon:
                    camion_match = camion
                    k_idx = camion.get("k")
                    self._log(f"[Control Datos] ✅ Match por {razon}")
                    break

            if not camion_match:
                # Log de diagnóstico: qué campos no matchearon
                for camion in cont_data["camiones"]:
                    self._log(f"[Control Datos]   vs camión: '{camion.get('patente_camion','')}' "
                              f"semi:'{camion.get('patente_semi','')}' "
                              f"dni:'{camion.get('dni','')}' "
                              f"cond:'{camion.get('conductor','')}'")

        tiene_match = (camion_match is not None and ruta_match is not None)

        if not tiene_match:
            iid = f"ocr_{self._cargar_datos_idx}"
            self._cargar_datos_idx += 1
            estado = "⚠ No encontrado"
            valores = (data.get("cliente", archivo), patente, semi, conductor,
                       self._fmt_dni(dni), self._fmt_kg(neto_ocr), self._fmt_kg(tara_ocr),
                       contenedor_display, permiso, estado)
            try:
                self.tree_carga.insert("", "end", values=valores, iid=iid,
                                       tags=("tag_no_match",))
            except Exception:
                pass
            self._log(f"[Control Datos] {archivo}: ⚠ Sin match")
            return

        # Comparación campo-por-campo (normalización de formato, NO de caracteres)
        def _raw_txt(v):
            # Saca espacios/puntuación PERO preserva I≠L, 0≠O, 1≠I, etc.
            return _re.sub(r'[^A-Z0-9]', '', str(v).upper().strip())

        def _nombres_equivalentes(a, b):
            """True si al menos 2 palabras coinciden (sin acentos, puntuación, mayúsculas)."""
            if not a or not b:
                return False
            import unicodedata
            def _limpiar(s):
                s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
                return _re.sub(r"[^\w\s]", " ", s).strip().upper()
            a_set = set(_limpiar(a).split())
            b_set = set(_limpiar(b).split())
            if not a_set or not b_set:
                return False
            return len(a_set & b_set) >= 2

        cam_pat    = camion_match.get("patente_camion", "")
        cam_semi   = camion_match.get("patente_semi", "")
        cam_cond   = camion_match.get("conductor", "")
        cam_dni    = camion_match.get("dni", "")
        peso_carga = camion_match.get("peso_carga", 0)
        tara_cont  = camion_match.get("tara_cont", 0)
        pe_val     = data.get("pe", "") or ""

        ok_patente   = _raw_txt(patente) == _raw_txt(cam_pat)
        ok_semi      = _raw_txt(semi) == _raw_txt(cam_semi)
        ok_conductor = _nombres_equivalentes(conductor, cam_cond)
        ok_dni       = _raw_txt(dni) == _raw_txt(cam_dni)
        # Shared trips: compare against SUM of Neto from all Excels
        if shared_excels:
            ok_neto = _comparar_num(neto_ocr, shared_neto_sum)
        else:
            ok_neto = _comparar_num(neto_ocr, peso_carga)
        ok_tara      = _comparar_num(tara_ocr, tara_cont)
        ok_contenedor      = _raw_txt(contenedor_str) == _raw_txt(camion_match.get("contenedor", ""))
        ok_permiso   = _raw_txt(permiso) == _raw_txt(pe_val) if pe_val else True

        # Estado general: TODO verde si todos los campos coinciden
        all_ok = (ok_patente and ok_semi and ok_conductor and ok_dni
                  and ok_neto and ok_tara and ok_contenedor and ok_permiso)
        if all_ok:
            estado = "✅ Ok"
            tag = "tag_ok"
        else:
            estado = "❌ Diferencia"
            tag = "tag_mismatch"

        iid = f"ocr_{self._cargar_datos_idx}"
        self._cargar_datos_idx += 1
        nombre_contenedor = os.path.basename(ruta_match)
        valores = (
            data.get("cliente", archivo),
            patente, semi, conductor, self._fmt_dni(dni),
            self._fmt_kg(neto_ocr), self._fmt_kg(tara_ocr),
            contenedor_display,
            permiso, estado,
        )
        try:
            self.tree_carga.insert("", "end", values=valores, iid=iid, tags=(tag,))
            self._cargar_datos_rutas[iid] = (ruta_match, k_idx)
        except Exception:
            pass

        # Guardar datos para popup de comparación al hacer doble-click
        self._cargar_datos_comparacion[iid] = {
            "archivo": archivo,
            "ticket": {
                "Patente": patente,
                "Semirremolque": semi,
                "Conductor": conductor,
                "DNI": dni,
                "Neto (kg)": f"{neto_ocr:.0f}",
                "Tara (kg)": f"{tara_ocr:.0f}",
                "Contenedor": contenedor_display,
                "Permiso": permiso,
            },
            "contenedor": {
                "Patente": cam_pat,
                "Semirremolque": cam_semi,
                "Conductor": cam_cond,
                "DNI": cam_dni,
                "Neto (kg)": f"{peso_carga:.0f}",
                "Tara (kg)": f"{tara_cont:.0f}",
                "Contenedor": camion_match.get("contenedor", "") or "-",
                "Permiso": pe_val,
            },
            "ok": {
                "Patente": ok_patente,
                "Semirremolque": ok_semi,
                "Conductor": ok_conductor,
                "DNI": ok_dni,
                "Neto (kg)": ok_neto,
                "Tara (kg)": ok_tara,
                "Contenedor": ok_contenedor,
                "Permiso": ok_permiso,
            },
            "shared_excels": shared_excels,
            "shared_neto_sum": shared_neto_sum,
        }

        # Log
        if not all_ok:
            diffs = []
            if not ok_neto:
                diffs.append(f"neto OCR={neto_ocr:.0f} vs CONT={peso_carga:.0f}")
            if not ok_tara:
                diffs.append(f"tara OCR={tara_ocr:.0f} vs CONT={tara_cont:.0f}")
            self._log(f"[Control Datos] {archivo}: Diferencias — {'; '.join(diffs)}")
        else:
            self._log(f"[Control Datos] {archivo}: ✅ Todo OK")

    def _abrir_comparacion(self, event=None):
        """Abre popup con tabla vertical: Ticket | Excel, verde/rojo por campo."""
        sel = self.tree_carga.selection()
        if not sel:
            return
        iid = sel[0]
        datos = self._cargar_datos_comparacion.get(iid)
        if not datos:
            return

        def _build_popup(level):
            fsizes = self._get_font_sizes(level)
            geom = self._get_popup_geometry(520, 420, level)
            ancho = int(geom.split("x")[0])

            # Check if shared trip
            shared_excels = datos.get("shared_excels", [])
            shared_neto_sum = datos.get("shared_neto_sum", 0)
            is_shared = len(shared_excels) > 1

            # ── Number formatters ──
            _fmt_kg = self._fmt_kg
            _fmt_dni = self._fmt_dni

            if is_shared:
                geom = self._get_popup_geometry(720, 420, level)
                ancho = int(geom.split("x")[0])

            dlg = ctk.CTkToplevel(self)
            dlg.title(f"Comparación — {datos['archivo']}" + (" (Compartido)" if is_shared else ""))
            dlg.transient(self)
            dlg.grab_set()
            dlg.resizable(False, False)

            # ── Top bar: header + font level override ──────────────────
            top_bar = ctk.CTkFrame(dlg, fg_color=Palette.BG_SIDEBAR, corner_radius=6, height=36)
            top_bar.pack(fill="x", padx=12, pady=(12, 0))
            top_bar.pack_propagate(False)

            # Grid header: align with body columns
            top_bar.grid_columnconfigure(0, minsize=120, weight=0)

            if is_shared:
                # 4 columns: Campo | Ticket | Excel 1 | Excel 2
                headers = ["Campo", "Ticket (OCR)"]
                for i, ex in enumerate(shared_excels):
                    headers.append(f"Excel {i+1}")
                for col_i in range(1, len(headers)):
                    top_bar.grid_columnconfigure(col_i, weight=1)
                for col_i, txt in enumerate(headers):
                    lbl = ctk.CTkLabel(
                        top_bar, text=txt, width=120 if col_i == 0 else 160,
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                        text_color=Palette.TEXT_SECONDARY,
                    )
                    lbl.grid(row=0, column=col_i, sticky="ew", padx=4, pady=4)
            else:
                # Normal 3 columns: Campo | Ticket | Contenedor
                for col_i in range(1, 3):
                    top_bar.grid_columnconfigure(col_i, weight=1)
                for col_i, txt in enumerate(["Campo", "Ticket (OCR)", "Contenedor (Excel)"]):
                    lbl = ctk.CTkLabel(
                        top_bar, text=txt, width=120 if col_i == 0 else 160,
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                        text_color=Palette.TEXT_SECONDARY,
                    )
                    lbl.grid(row=0, column=col_i, sticky="ew", padx=4, pady=4)

            # Font level override selector
            override_menu = ctk.CTkOptionMenu(
                top_bar, values=["1", "2", "3"],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                width=50, height=26,
                command=lambda v: (dlg.destroy(), _build_popup(int(v))),
            )
            override_menu.set(str(level))
            override_menu.pack(side="right", padx=(0, 8))

            # Cuerpo: grid para columnas alineadas
            body = ctk.CTkFrame(dlg, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=12, pady=8)
            body.grid_columnconfigure(0, minsize=120, weight=0)  # Campo — fijo
            for col_i in range(1, 4 if is_shared else 3):
                body.grid_columnconfigure(col_i, weight=1)

            campos = [("Camion", "Patente"), ("Semirremolque", "Semirremolque"), ("Conductor", "Conductor"),
                       ("DNI", "DNI"), ("Neto (kg)", "Neto (kg)"), ("Tara (kg)", "Tara (kg)"),
                       ("Contenedor", "Contenedor"), ("Permiso", "Permiso")]

            for row_i, (label, key) in enumerate(campos):
                val_ticket = datos["ticket"].get(key, "")
                val_cont   = datos["contenedor"].get(key, "")
                ok         = datos["ok"].get(key, False)

                bg = "#C8FFC8" if ok else "#FFC8C8"
                fg = "#006400" if ok else "#8B0000"

                ctk.CTkLabel(
                    body, text=label, width=120, anchor="w",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"], weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).grid(row=row_i, column=0, sticky="w", padx=(4, 0), pady=2)

                if is_shared:
                    # Shared: Ticket (col 1) + Excel 1 (col 2) + Excel 2 (col 3)
                    val_ticket_neto = datos["ticket"].get("Neto (kg)", 0)

                    # ── Column 1: Ticket value (formatted) ──
                    if key in ("Neto (kg)", "Tara (kg)"):
                        display_ticket = _fmt_kg(val_ticket)
                    elif key == "DNI":
                        display_ticket = _fmt_dni(val_ticket)
                    else:
                        display_ticket = val_ticket
                    ctk.CTkLabel(
                        body, text=display_ticket, anchor="center",
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                        fg_color=bg, text_color=fg, corner_radius=4,
                    ).grid(row=row_i, column=1, sticky="ew", padx=4, pady=2)

                    # ── Columns 2..N: each Excel ──
                    for col_i, ex in enumerate(shared_excels, start=2):
                        ex_camion = ex.get("camion", {})
                        ex_data = ex.get("data", {})
                        ex_pe = ex.get("pe", "")

                        if key == "Neto (kg)":
                            # Neto: show individual + total
                            neto_individual = ex_camion.get("peso_carga", 0)
                            val = f"{_fmt_kg(neto_individual)} ({_fmt_kg(shared_neto_sum)})"
                            try:
                                ticket_neto = float(str(val_ticket_neto).replace('.', '').replace(',', '.'))
                                sums_ok = abs(shared_neto_sum - ticket_neto) < 1
                            except (ValueError, AttributeError):
                                sums_ok = False
                            cell_bg = "#C8FFC8" if sums_ok else "#FFC8C8"
                            cell_fg = "#006400" if sums_ok else "#8B0000"
                        elif key == "Tara (kg)":
                            val = _fmt_kg(ex_camion.get("tara_cont", ""))
                            cell_bg, cell_fg = bg, fg
                        elif key == "Patente":
                            val = str(ex_camion.get("patente_camion", ""))
                            cell_bg, cell_fg = bg, fg
                        elif key == "Semirremolque":
                            val = str(ex_camion.get("patente_semi", ""))
                            cell_bg, cell_fg = bg, fg
                        elif key == "Conductor":
                            val = str(ex_camion.get("conductor", ""))
                            cell_bg, cell_fg = bg, fg
                        elif key == "DNI":
                            val = _fmt_dni(ex_camion.get("dni", ""))
                            cell_bg, cell_fg = bg, fg
                        elif key == "Contenedor":
                            val = str(ex_camion.get("contenedor", ex_data.get("contenedor", "")))
                            cell_bg, cell_fg = bg, fg
                        elif key == "Permiso":
                            val = str(ex_pe)
                            cell_bg, cell_fg = bg, fg
                        else:
                            val = ""
                            cell_bg, cell_fg = bg, fg

                        lbl = ctk.CTkLabel(
                            body, text=val, anchor="center",
                            font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                            fg_color=cell_bg, text_color=cell_fg, corner_radius=4,
                        )
                        lbl.grid(row=row_i, column=col_i, sticky="ew", padx=4, pady=2)
                else:
                    # Normal: Ticket + single Excel — format numbers
                    if key in ("Neto (kg)", "Tara (kg)"):
                        display_ticket = _fmt_kg(val_ticket)
                        display_cont = _fmt_kg(val_cont)
                    elif key == "DNI":
                        display_ticket = _fmt_dni(val_ticket)
                        display_cont = _fmt_dni(val_cont)
                    else:
                        display_ticket = val_ticket
                        display_cont = val_cont
                    for col_i, val in enumerate([display_ticket, display_cont], start=1):
                        lbl = ctk.CTkLabel(
                            body, text=val, anchor="center",
                            font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                            fg_color=bg, text_color=fg, corner_radius=4,
                        )
                        lbl.grid(row=row_i, column=col_i, sticky="ew", padx=4, pady=2)

            # Leyenda al pie
            leyenda = ctk.CTkFrame(dlg, fg_color="transparent")
            leyenda.pack(fill="x", pady=(0, 12))
            inner = ctk.CTkFrame(leyenda, fg_color="transparent")
            inner.pack(anchor="center")
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#C8FFC8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Coincide",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#FFC8C8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Difiere",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            # Ajustar ventana al contenido real
            dlg.update_idletasks()
            alto = dlg.winfo_reqheight()
            x = (dlg.winfo_screenwidth() - ancho) // 2
            y = (dlg.winfo_screenheight() - alto) // 2
            dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

        _build_popup(self.config.get("font_level", 1))

    def _limpiar_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self._log_line_count = 0
        self.lbl_log_count.configure(text="")
        # También limpiar el historial guardado del panel actual
        if self.panel_actual in self.logs_por_panel:
            self.logs_por_panel[self.panel_actual] = []

    def _capturar_lineas_log(self):
        """Devuelve una lista con todas las líneas de texto del log (sin tags)."""
        contenido = self.log_text.get("1.0", "end-1c")
        if not contenido.strip():
            return []
        return [l for l in contenido.split("\n") if l.strip()]

    def _restaurar_log(self, nombre):
        """Limpia el widget y restaura las líneas guardadas para el panel dado."""
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")

        lineas = self.logs_por_panel.get(nombre, [])
        if lineas:
            for i, linea in enumerate(lineas):
                if i > 0:
                    self.log_text.insert("end", "\n")
                # Detectar tag para colorear
                txt_up = linea.upper()
                tag = None
                if any(kw in txt_up for kw in ("ERROR", "CRÍTICO", "NO SE PUDO")):
                    tag = "error"
                elif any(kw in txt_up for kw in ("ÉXITO", "FINALIZADO", "CORRECTAMENTE", "COMPLETADO", "GUARDADA")):
                    tag = "success"
                elif any(kw in txt_up for kw in ("ADVERTENCIA", "FALTA", "NO SE ENCONTRÓ")):
                    tag = "warning"
                elif any(kw in txt_up for kw in ("ESCANEANDO", "PREPARANDO", "CONECTANDO", "SUBIENDO", "LEYENDO", "COPIADO")):
                    tag = "info"
                if tag:
                    self.log_text.insert("end", linea, tag)
                else:
                    self.log_text.insert("end", linea)
            self._log_line_count = len(lineas)
            self.lbl_log_count.configure(text=f"{self._log_line_count} líneas")
        else:
            self._log_line_count = 0
            self.lbl_log_count.configure(text="")

        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _trim_log(self):
        """Elimina las líneas más antiguas para mantener un máximo de 500."""
        self.log_text.configure(state="normal")
        # Obtener todo, descartar primeras 200 líneas
        contenido = self.log_text.get("1.0", "end-1c")
        lineas = contenido.split("\n")
        if len(lineas) > 500:
            keep = lineas[-300:]  # Mantener últimas 300
            self.log_text.delete("1.0", "end")
            self.log_text.insert("1.0", "\n".join(keep))
            self._log_line_count = len(keep)
            self.lbl_log_count.configure(text=f"{self._log_line_count} líneas")
        self.log_text.configure(state="disabled")

    # ═══════════════════════════════════════════════════════════════════
    # BARRA DE ESTADO (Desactivada)
    # ═══════════════════════════════════════════════════════════════════
    def _set_status(self, texto):
        pass

    # ═══════════════════════════════════════════════════════════════════
    # CONFIG HELPERS
    # ═══════════════════════════════════════════════════════════════════
    def _cfg_obtener(self, seccion, clave, default):
        val = self.config.get(seccion, {}).get(clave, default)
        if seccion == "seguridad" and clave == "password" and val != default:
            val = self._decrypt_val(val, os.environ["MULTIAGENTE_SECRET_KEY"])
            if val.startswith("enc::"):
                return default
        return val

    def _cfg_obtener_correo(self, clave, default):
        val = self._cfg_obtener("correo", clave, default)
        if clave == "password" and val != default:
            key = os.environ["MULTIAGENTE_SECRET_KEY"]
            val = self._decrypt_val(val, key)
            if val.startswith("enc::"):
                return default
        return val

    def _cfg_obtener_docs(self, clave, default):
        return self._cfg_obtener("documentos", clave, default)

    def _cfg_obtener_rutas(self, clave, default):
        return self._cfg_obtener("rutas", clave, default)

    def _resolver_ruta(self, clave, default):
        r"""Resuelve una ruta configurable: 'Desktop' → %USERPROFILE%\Desktop, sino se usa tal cual."""
        valor = self._cfg_obtener_rutas(clave, default)
        if valor.lower() == "desktop":
            nombre_esc = self._cfg_obtener_rutas("escritorio_nombre", "Desktop")
            return os.path.join(os.path.expanduser("~"), nombre_esc)
        return valor

    def _get_font_sizes(self, level=None):
        """Return scaled font sizes for the given level."""
        if level is None:
            level = self.config.get("font_level", 1)
        scale = FONT_LEVEL_SCALES.get(level, 1.0)
        return {
            "data": int(FONT_BASE_SIZES["data"] * scale),
            "header": int(FONT_BASE_SIZES["header"] * scale),
            "legend": int(FONT_BASE_SIZES["legend"] * scale),
        }

    def _get_popup_geometry(self, base_w, base_h, level=None):
        """Return scaled geometry string for popups."""
        if level is None:
            level = self.config.get("font_level", 1)
        scale = FONT_LEVEL_SCALES.get(level, 1.0)
        w = int(base_w * scale)
        h = int(base_h * scale)
        return f"{w}x{h}"

    def _calc_popup_height(self, num_rows, level=None):
        """Calculate popup height based on number of data rows."""
        if level is None:
            level = self.config.get("font_level", 1)
        scale = FONT_LEVEL_SCALES.get(level, 1.0)
        top_bar = 48          # header bar + padding
        row_h = int(32 * scale)  # each data row
        legend = 42           # legend at bottom
        scroll_pad = 16       # scroll frame padding
        window_pad = 24       # top + bottom window padding
        h = top_bar + (num_rows * row_h) + legend + scroll_pad + window_pad
        return h

    def _recuperar_password(self, parent_dlg, lbl_status=None):
        """Envía la contraseña por SMTP directamente al correo configurado."""
        import smtplib
        from email.mime.text import MIMEText

        usuario = self._cfg_obtener_correo("usuario", "")
        pw_email = self._cfg_obtener_correo("password", "")
        pw_app = self._cfg_obtener("seguridad", "password", "")
        imap_srv = self._cfg_obtener_correo("imap_server", "imap.gmail.com")

        if not usuario or not pw_email:
            if lbl_status:
                self.after(0, lambda: lbl_status.configure(
                    text="No hay correo configurado en Ajustes > Correo.",
                    text_color=Palette.ERROR))
            return

        if not pw_app:
            if lbl_status:
                self.after(0, lambda: lbl_status.configure(
                    text="No hay contraseña maestra configurada.",
                    text_color=Palette.WARNING))
            return

        # Derivar SMTP del IMAP: imap.dominio.com → smtp.dominio.com
        if "gmail" in imap_srv.lower():
            smtp_srv = "smtp.gmail.com"
        else:
            smtp_srv = "smtp." + imap_srv.removeprefix("imap.")

        if lbl_status:
            self.after(0, lambda: lbl_status.configure(
                text=f"Enviando a {usuario} vía {smtp_srv}...",
                text_color=Palette.INFO))

        def _enviar():
            ultimo_error = ""
            for puerto in (587, 25, 465):
                try:
                    msg = MIMEText(f"Tu contraseña maestra del Sistema de Automatización es:\n\n"
                                   f"  {pw_app}\n\n"
                                   f"Usala para ingresar a la aplicación.")
                    msg["Subject"] = "Clave de Recuperación"
                    msg["From"] = usuario
                    msg["To"] = usuario

                    if puerto == 465:
                        server = smtplib.SMTP_SSL(smtp_srv, puerto, timeout=10)
                    else:
                        server = smtplib.SMTP(smtp_srv, puerto, timeout=10)
                        server.ehlo()
                        try:
                            server.starttls()
                            server.ehlo()
                        except Exception:
                            pass  # algunos servidores no soportan STARTTLS
                    server.login(usuario, pw_email)
                    server.send_message(msg)
                    server.quit()

                    self.after(0, lambda: lbl_status.configure(
                        text=f"Enviado a {usuario}",
                        text_color=Palette.SUCCESS)) if lbl_status else \
                        self.after(0, lambda: messagebox.showinfo("Contraseña enviada",
                            f"La contraseña fue enviada a:\n{usuario}\n\n"
                            f"Revisá tu bandeja de entrada."))
                    return
                except Exception as e:
                    ultimo_error = str(e)[:100]
                    continue

            self.after(0, lambda: lbl_status.configure(
                text=f"Error: {ultimo_error}",
                text_color=Palette.ERROR)) if lbl_status else \
                self.after(0, lambda: messagebox.showerror("Error al enviar",
                    f"No se pudo enviar en ningún puerto:\n\n{ultimo_error}\n\n"
                    f"SMTP: {smtp_srv}\nPuertos: 587, 25, 465\n"
                    f"Usuario: {usuario}\n\n"
                    f"Verificá Ajustes > Correo."))

        t = threading.Thread(target=_enviar, daemon=True)
        t.start()

    def _verificar_password_maestra(self):
        """Si hay contraseña maestra configurada, la pide. Retorna True si pasa."""
        pw_config = self._cfg_obtener("seguridad", "password", "")
        if not pw_config:
            return True
        return self._mostrar_dialogo_password_ajustes(pw_config)

    def _mostrar_dialogo_password_ajustes(self, pw_config):
        """Diálogo CTk para pedir contraseña al entrar a Ajustes (bloqueante)."""
        resultado = {"ok": False}

        dlg = ctk.CTkToplevel(self)
        dlg.title("Acceso Restringido")
        dlg.geometry("380x190")
        dlg.configure(fg_color=Palette.BG_CARD)
        dlg.transient(self)
        dlg.lift()
        dlg.grab_set()
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 380) // 2
        py = self.winfo_y() + (self.winfo_height() - 190) // 2
        dlg.geometry(f"380x190+{px}+{py}")

        ctk.CTkLabel(
            dlg, text="Ingresá la contraseña para acceder a Ajustes:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(24, 12))

        entry_pw = ctk.CTkEntry(
            dlg, width=260, height=36, show="*",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=6,
        )
        entry_pw.pack(pady=(0, 12))
        entry_pw.focus_set()
        dlg.after(50, lambda: entry_pw.focus_force())

        lbl_error = ctk.CTkLabel(
            dlg, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.ERROR,
        )
        lbl_error.pack()

        def verificar():
            if entry_pw.get() == pw_config:
                resultado["ok"] = True
                self._master_pw_cache = entry_pw.get()
                dlg.destroy()
            else:
                lbl_error.configure(text="Contraseña incorrecta. Intentá de nuevo.")
                entry_pw.delete(0, "end")
                entry_pw.focus_set()

        entry_pw.bind("<Return>", lambda e: verificar())
        btn_frame = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_frame.pack(pady=(12, 0))
        ctk.CTkButton(
            btn_frame, text="Ingresar", width=100, height=32,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=verificar,
        ).pack(side="left", padx=4)
        ctk.CTkButton(
            btn_frame, text="Cancelar", width=100, height=32,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=dlg.destroy,
        ).pack(side="left", padx=4)

        dlg.wait_window()
        return resultado["ok"]

    # ═══════════════════════════════════════════════════════════════════
    # PANEL: CARGA DE DATOS (T3+)
    # ═══════════════════════════════════════════════════════════════════
    def _panel_cargar_datos(self):
        if "cargar-datos" in self._panel_frames and hasattr(self, 'tree_coordinacion'):
            return
        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        # NOTA: _panel_frames se setea al FINAL del método para evitar
        # que una inicialización incompleta deje el panel huérfano.
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── T11: Verificar dependencias (Tesseract + Poppler) ───────────
        dependencias_ok = True
        errores_dep = []

        # ── Configurar rutas de binarios ──────────────────────────
        _app_base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        _res_base = getattr(sys, '_MEIPASS', _app_base)
        _POPPLER_BIN = os.path.join(_res_base, "poppler", "Library", "bin")
        _TESSERACT_EXE = os.path.join(_res_base, "engines", "tesseract", "tesseract.exe")
        if not os.path.isfile(_TESSERACT_EXE):
            _TESSERACT_EXE = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

        # Verificar Tesseract (solo archivo, sin ejecutarlo — evita ventana DOS)
        import pytesseract
        tesseract_sidecar = os.path.join(
            getattr(sys, '_MEIPASS', os.path.dirname(__file__)),
            "engines", "tesseract", "tesseract.exe")
        tesseract_paths = [
            tesseract_sidecar,
            _TESSERACT_EXE,
        ]
        if not any(os.path.isfile(p) for p in tesseract_paths):
            dependencias_ok = False
            errores_dep.append(
                "⚠ Tesseract-OCR no está instalado. "
                "Descargalo de https://github.com/UB-Mannheim/tesseract/wiki"
            )
        else:
            pytesseract.pytesseract.tesseract_cmd = next(p for p in tesseract_paths if os.path.isfile(p))

        # Verificar Poppler (pdf2image)
        poppler_binarios = [
            os.path.join(_POPPLER_BIN, "pdftoppm.exe"),
            os.path.join(_POPPLER_BIN, "pdfinfo.exe"),
        ]
        if not any(os.path.isfile(b) for b in poppler_binarios):
            dependencias_ok = False
            errores_dep.append(
                "⚠ Poppler no está instalado. "
                "Descargalo de https://github.com/oschwartz10612/poppler-windows/releases/"
            )

        if not dependencias_ok:
            error_frame = ctk.CTkFrame(
                frame, fg_color="#FFE0E0", corner_radius=8,
                border_width=1, border_color="#FFB0B0",
            )
            error_frame.pack(fill="x", pady=(0, 6))

            for msg in errores_dep:
                ctk.CTkLabel(
                    error_frame, text=msg,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                    text_color="red", wraplength=600, justify="left",
                ).pack(padx=12, pady=6, anchor="w")

            # ── Toolbar con botones deshabilitados ──────────────────────
            toolbar = ctk.CTkFrame(
                frame, fg_color=Palette.BG_CARD, corner_radius=8,
                border_width=1, border_color=Palette.BORDER, height=44
            )
            toolbar.pack(fill="x", pady=(0, 6))
            toolbar.pack_propagate(False)

            self.btn_controlar_tickets = ctk.CTkButton(
                toolbar,
                text="Controlar Tickets",
                image=self._icons["ticket"], compound="left",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.WHITE, corner_radius=6, height=34, width=170,
                state="disabled",
            )
            self.btn_controlar_tickets.pack(side="left", padx=4, pady=4)

            # Espacio para mantener layout
            ctk.CTkLabel(
                toolbar,
                text="Dependencias faltantes — resolver antes de continuar",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.ERROR,
            ).pack(side="left", padx=(8, 0))

            # ── Resultados vacío ──────────────────────────────────────
            scroll = ctk.CTkScrollableFrame(
                frame, fg_color=Palette.BG_CARD, corner_radius=8,
                border_width=1, border_color=Palette.BORDER
            )
            scroll.pack(fill="both", expand=True, pady=(0, 6))

            self._panel_frames["cargar-datos"] = frame
            return  # No seguir con el panel normal

        # ── Toolbar ──────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER, height=44
        )
        toolbar.pack(fill="x", pady=(0, 6))
        toolbar.pack_propagate(False)

        # ── Grupo Control Tickets ──────────────────────────────────
        self.btn_controlar_tickets = ctk.CTkButton(
            toolbar,
            text="Controlar Tickets",
            image=self._icons["ticket"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            command=self._cargar_datos_seleccionar_pdfs,
        )
        self.btn_controlar_tickets.pack(side="left", padx=(10, 2), pady=5)

        self._ct_auto_switch = ctk.CTkSwitch(
            toolbar, text="Auto",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            variable=self._ct_auto_var,
            progress_color=Palette.ACCENT,
            button_color=Palette.TEXT_SECONDARY,
            button_hover_color=Palette.TEXT_PRIMARY,
            command=self._cargar_datos_switch_toggle,
        )
        self._ct_auto_switch.pack(side="left", padx=2, pady=5)
        # Restore persisted state
        if self.config.get("cargar_datos_auto", False):
            self._ct_auto_var.set(True)
            self.btn_controlar_tickets.configure(text="Controlar Tickets ⚡", image=self._icons["zap"])

        # ── Separador vertical ──────────────────────────────────────
        ctk.CTkFrame(
            toolbar, width=2, height=34, fg_color=Palette.BORDER
        ).pack(side="left", padx=6, pady=5)

        self.btn_controlar_coordinacion = ctk.CTkButton(
            toolbar,
            text="Controlar Coordinación",
            image=self._icons["clipboard-list"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            command=self._controlar_coordinacion,
        )
        self.btn_controlar_coordinacion.pack(side="left", padx=2, pady=5)

        # ── Separador vertical ──────────────────────────────────────
        ctk.CTkFrame(
            toolbar, width=2, height=34, fg_color=Palette.BORDER
        ).pack(side="left", padx=6, pady=5)

        self.btn_control_final = ctk.CTkButton(
            toolbar,
            text="Control Final",
            image=self._icons["zap"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.SECONDARY, hover_color=Palette.SECONDARY_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=150,
            state="normal",
            command=self._control_final_seleccionar,
        )
        self.btn_control_final.pack(side="left", padx=2, pady=5)

        self._cf_auto_switch = ctk.CTkSwitch(
            toolbar, text="Auto",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            variable=self._cf_auto_var,
            progress_color=Palette.ACCENT,
            button_color=Palette.TEXT_SECONDARY,
            button_hover_color=Palette.TEXT_PRIMARY,
            command=self._control_final_switch_toggle,
        )
        self._cf_auto_switch.pack(side="left", padx=(0, 6), pady=5)

        # Restore auto mode state from config
        if self.config.get("control_final_auto", False):
            self._cf_auto_var.set(True)
            self.btn_control_final.configure(text="Control Final ⚡", image=self._icons["zap"])

        self.progress_carga = ctk.CTkProgressBar(
            toolbar, width=140, height=8, corner_radius=4,
            mode="indeterminate",
            fg_color=Palette.BG_INPUT, progress_color=Palette.ACCENT,
        )
        # Oculto por defecto — se muestra al iniciar procesamiento

        self.btn_limpiar_carga = ctk.CTkButton(
            toolbar,
            text="Limpiar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color="transparent",
            hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            border_width=0,
            corner_radius=4,
            height=30,
            width=70,
            command=self._limpiar_carga,
        )
        self.btn_limpiar_carga.pack(side="right", padx=4, pady=4)

        # ── Tab bar compacta ────────────────────────────────
        tab_bar = ctk.CTkFrame(
            frame, fg_color=Palette.BG_SIDEBAR, corner_radius=8, height=32,
            border_width=1, border_color=Palette.BORDER,
        )
        tab_bar.pack(fill="x", pady=(6, 2))
        tab_bar.pack_propagate(False)

        tab_font = ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold")
        _TAB_STEEL = "#546e7a"  # Distinct from toolbar blue/teal

        self._tab_tickets = ctk.CTkLabel(
            tab_bar, text="Control de Tickets", font=tab_font,
            fg_color=_TAB_STEEL, text_color=Palette.WHITE,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_tickets.pack(side="left", padx=4, pady=2)
        self._tab_tickets.bind("<Button-1>", lambda e: self._switch_tab("tickets"))

        self._tab_coord = ctk.CTkLabel(
            tab_bar, text="Coordinación", font=tab_font,
            fg_color="transparent", text_color=_TAB_STEEL,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_coord.pack(side="left", padx=2, pady=2)
        self._tab_coord.bind("<Button-1>", lambda e: self._switch_tab("coord"))

        self._tab_final = ctk.CTkLabel(
            tab_bar, text="Control Final", font=tab_font,
            fg_color="transparent", text_color=_TAB_STEEL,
            corner_radius=6, padx=14, pady=4,
        )
        self._tab_final.pack(side="left", padx=2, pady=2)
        self._tab_final.bind("<Button-1>", lambda e: self._switch_tab("final"))

        # ── Resultados (scroll frame + sections) ────────────────
        scroll = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER
        )
        scroll.pack(fill="both", expand=True, pady=(0, 6))

        # Treeview style oscuro
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "CargaDatos.Treeview",
            background=Palette.BG_TABLE,
            foreground=Palette.TEXT_PRIMARY,
            fieldbackground=Palette.BG_TABLE,
            borderwidth=0,
            font=(FONT_FAMILY, 10),
            rowheight=28,
        )
        style.map(
            "CargaDatos.Treeview",
            background=[("selected", Palette.ACCENT_DIM)],
            foreground=[("selected", Palette.WHITE)],
        )
        style.configure(
            "CargaDatos.Treeview.Heading",
            background=Palette.BG_SIDEBAR,
            foreground=Palette.TEXT_SECONDARY,
            font=(FONT_FAMILY, 10, "bold"),
            borderwidth=0,
            padding=(8, 6),
        )

        # ── Sección: Tickets ──────────────────────────────────────────
        self._section_tickets = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        columns = ("cliente", "camion", "semi", "conductor", "dni",
                   "neto", "tara", "contenedor", "permiso", "estado")
        headers = ("Cliente", "Camion", "Semi", "Conductor", "DNI",
                   "Neto", "Tara", "Contenedor", "Permiso", "Estado")
        anchos_default = (120, 80, 80, 100, 80, 70, 70, 100, 100, 100)
        _saved_cols = self.config.get("tree_carga_columns", {})

        self.tree_carga = ttk.Treeview(
            self._section_tickets, columns=columns, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(columns, headers, anchos_default):
            self.tree_carga.heading(col, text=hdr)
            w = _saved_cols.get(col, w_default)
            self.tree_carga.column(col, width=w, anchor="center", minwidth=40)

        # Scrollbar vertical
        scroll_y = ctk.CTkScrollbar(
            self._section_tickets, orientation="vertical",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        scroll_y.pack(side="right", fill="y", padx=(0, 2), pady=2)

        self.tree_carga.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self.tree_carga.yview)

        # Scrollbar horizontal
        self.scroll_carga_x = ctk.CTkScrollbar(
            self._section_tickets, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_carga.configure(xscrollcommand=self.scroll_carga_x.set)
        self.scroll_carga_x.configure(command=self.tree_carga.xview)
        self.scroll_carga_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_carga.pack(fill="both", expand=True, padx=2, pady=(0, 0))

        # Tags
        self.tree_carga.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_carga.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_carga.tag_configure("tag_no_match", background="#FFF3E0", foreground="#E65100")
        self.tree_carga.bind("<Double-1>", self._abrir_comparacion)

        # ── Sección: Coordinación ISO/FLEXI ──────────────────────────
        self._section_coord = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        cols_c = ("carpeta", "giro", "cliente", "destino",
                  "buque", "viaje", "booking", "pto_descarga", "pto_final",
                  "fecha_of_pe", "fecha_carga", "peso_flexi", "estado")
        hdrs_c = ("Carpeta", "Giro", "Cliente", "Destino",
                  "Buque", "Viaje", "Booking", "Pto Descarga", "Pto Final",
                  "Fec Of PE", "Fec Carga", "Peso Flexi", "Estado")
        anchos_c_default = (100, 70, 220, 120, 180, 70, 160, 120, 120, 110, 110, 80, 80)
        _saved_cols_c = self.config.get("tree_coordinacion_columns", {})

        self.tree_coordinacion = ttk.Treeview(
            self._section_coord, columns=cols_c, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(cols_c, hdrs_c, anchos_c_default):
            self.tree_coordinacion.heading(col, text=hdr)
            w = _saved_cols_c.get(col, w_default)
            self.tree_coordinacion.column(col, width=w, anchor="center", minwidth=40)

        self.tree_coordinacion.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_coordinacion.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_coordinacion.tag_configure("tag_error", background="#FFE0B0", foreground="#8B4500")

        self.tree_coordinacion.bind("<Double-1>", self._abrir_comparacion_coordinacion)

        # Scrollbar horizontal
        self.scroll_coord_x = ctk.CTkScrollbar(
            self._section_coord, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_coordinacion.configure(xscrollcommand=self.scroll_coord_x.set)
        self.scroll_coord_x.configure(command=self.tree_coordinacion.xview)
        self.scroll_coord_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_coordinacion.pack(fill="both", expand=True, padx=2, pady=(0, 8))

        # ── Sección: Control Final ──────────────────────────────────
        self._section_final = ctk.CTkFrame(
            scroll, fg_color="transparent",
        )

        cols_f = ("cliente", "camion", "semi", "conductor", "dni",
                  "neto", "tara", "contenedor", "permiso", "estado", "salida_aduana")
        hdrs_f = ("Cliente", "Camion", "Semi", "Conductor", "DNI",
                  "Neto", "Tara", "Contenedor", "Permiso", "Estado", "Salida Aduana")
        anchos_f_default = (120, 80, 80, 100, 80, 70, 70, 100, 100, 100, 100)
        _saved_cols_f = self.config.get("tree_control_final_columns", {})

        self.tree_control_final = ttk.Treeview(
            self._section_final, columns=cols_f, show="headings",
            style="CargaDatos.Treeview", selectmode="browse",
        )
        for col, hdr, w_default in zip(cols_f, hdrs_f, anchos_f_default):
            self.tree_control_final.heading(col, text=hdr)
            w = _saved_cols_f.get(col, w_default)
            self.tree_control_final.column(col, width=w, anchor="center", minwidth=40)

        self.tree_control_final.tag_configure("tag_ok", background="#C8FFC8", foreground="#006400")
        self.tree_control_final.tag_configure("tag_mismatch", background="#FFC8C8", foreground="#8B0000")
        self.tree_control_final.tag_configure("tag_no_match", background="#FFF3E0", foreground="#E65100")
        self.tree_control_final.bind("<Double-1>", self._abrir_comparacion_final)

        # Scrollbar horizontal
        self.scroll_control_final_x = ctk.CTkScrollbar(
            self._section_final, orientation="horizontal",
            fg_color=Palette.BG_CARD, button_color=Palette.TEXT_MUTED,
            button_hover_color=Palette.TEXT_SECONDARY,
        )
        self.tree_control_final.configure(xscrollcommand=self.scroll_control_final_x.set)
        self.scroll_control_final_x.configure(command=self.tree_control_final.xview)
        self.scroll_control_final_x.pack(fill="x", padx=2, pady=(2, 0))

        self.tree_control_final.pack(fill="both", expand=True, padx=2, pady=(0, 8))

        for section in (self._section_tickets, self._section_coord, self._section_final):
            section.pack(fill="both", expand=True, pady=(0, 6))
        self._section_coord.pack_forget()
        self._section_final.pack_forget()

        # Todo listo — registrar el frame como completo en _panel_frames
        self._panel_frames["cargar-datos"] = frame

    def _switch_tab(self, tab):
        """Switch visible section and highlight active tab."""
        _TAB_STEEL = "#546e7a"
        tabs = {
            "tickets": (self._tab_tickets, self._section_tickets),
            "coord":   (self._tab_coord,   self._section_coord),
            "final":   (self._tab_final,   self._section_final),
        }
        for name, (label, section) in tabs.items():
            if name == tab:
                label.configure(fg_color=_TAB_STEEL, text_color=Palette.WHITE)
                section.pack(fill="both", expand=True, pady=(0, 6))
            else:
                label.configure(fg_color="transparent", text_color=_TAB_STEEL)
                section.pack_forget()

    # ── Limpiar vista — planillas ──────────────────────────────────────
    def _limpiar_planillas(self):
        """Elimina filas y resetea estado del panel Planillas."""
        for row in self.tree_planillas.get_children():
            self.tree_planillas.delete(row)
        self.lbl_resumen_planillas.configure(text="Sin datos analizados")
        self.progress_planillas.set(0)

    # ── Limpiar vista — descargar ──────────────────────────────────────
    def _limpiar_mail(self):
        """Elimina filas y resetea estado del panel Descargar Mails."""
        for row in self._mail_tree.get_children():
            self._mail_tree.delete(row)
        self._mail_data.clear()
        self._mail_lbl_estado.configure(text="")
        self._mail_progress.set(0)

    # ── Limpiar vista — correos ────────────────────────────────────────
    def _limpiar_correos(self):
        """Elimina filas y resetea estado del panel Enviar Correos."""
        for row in self.tree_correos.get_children():
            self.tree_correos.delete(row)
        self.lbl_estado_correos.configure(text="Listo para despachar borradores")
        self.progress_correos.set(0)

    # ── Limpiar vista — cargar datos ───────────────────────────────────
    def _limpiar_carga(self):
        """Elimina todas las filas y resetea a vista tickets."""
        self.tree_carga.delete(*self.tree_carga.get_children())
        if hasattr(self, 'tree_control_final'):
            self.tree_control_final.delete(*self.tree_control_final.get_children())
            self.tree_control_final.heading("salida_aduana", text="Salida Aduana")
        if hasattr(self, 'tree_coordinacion'):
            self.tree_coordinacion.delete(*self.tree_coordinacion.get_children())
        # Resetear a vista tickets
        self._switch_tab("tickets")

    # ── Coordinación ISO/FLEXI ────────────────────────────────────────
    def _escanear_carpetas_coordinacion(self):
        """Escanea Desktop + 1 nivel buscando carpetas con CONTENEDORES Excel y PDF coordinación."""
        base = self._resolver_ruta("planillas_carga", "Desktop")
        carpetas = []

        def _tiene_coordinacion(nombre):
            """Detecta 'coordinacion' en el nombre sin importar acentos."""
            return "coordinaci" in nombre.lower()

        # 1 nivel de subcarpetas en Desktop
        for item in os.listdir(base):
            ruta_dir = os.path.join(base, item)
            if not os.path.isdir(ruta_dir):
                continue

            # Buscar CONTENEDORES Excel, PDF Coordinación, y Permiso dentro
            xls_path = None
            pdf_path = None
            permiso_path = None
            for f in os.listdir(ruta_dir):
                f_lower = f.lower()
                if "contenedor" in f_lower and (f_lower.endswith(".xlsx") or f_lower.endswith(".xls")):
                    xls_path = os.path.join(ruta_dir, f)
                elif f_lower.endswith(".pdf"):
                    if _tiene_coordinacion(f):
                        pdf_path = os.path.join(ruta_dir, f)
                    elif permiso_path is None:
                        permiso_path = os.path.join(ruta_dir, f)

            if xls_path:  # Si tiene Excel, lo consideramos
                carpetas.append({
                    "nombre": item,
                    "ruta": ruta_dir,
                    "pdf_path": pdf_path,
                    "xls_path": xls_path,
                    "permiso_path": permiso_path,
                    "pdf_ok": pdf_path is not None,
                    "xls_ok": True,
                    "permiso_ok": permiso_path is not None,
                })

        carpetas.sort(key=lambda c: c["nombre"])
        return carpetas

    def _controlar_coordinacion(self):
        """Escanea Desktop, muestra popup con checkboxes, procesa seleccionadas."""
        from procesar_tickets import (
            extraer_coordinacion, leer_choferes_coordinacion, comparar_coordinacion,
            extraer_fecha_permiso,
        )
        import unicodedata
        import tkinter as tk

        # 1. Escanear carpetas
        carpetas = self._escanear_carpetas_coordinacion()
        if not carpetas:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con archivos CONTENEDORES en el Desktop."
            )
            return

        # 2. Popup de selección con checkboxes
        seleccion = []

        top = ctk.CTkToplevel(self)
        top.title("Seleccionar envíos ISO/FLEXI para controlar")
        ancho, alto = 960, 520
        top.geometry(f"{ancho}x{alto}")
        top.transient(self)
        top.grab_set()
        top.update_idletasks()
        x = (top.winfo_screenwidth() - ancho) // 2
        y = (top.winfo_screenheight() - alto) // 2
        top.geometry(f"{ancho}x{alto}+{x}+{y}")

        # Encabezado
        ctk.CTkLabel(
            top,
            text="Carpetas con CONTENEDORES encontradas en el Desktop",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(12, 2))

        ctk.CTkLabel(
            top,
            text="Solo se pueden seleccionar carpetas con PDF + Excel presentes",
            font=ctk.CTkFont(family=FONT_FAMILY, size=10),
            text_color=Palette.TEXT_MUTED,
        ).pack(pady=(0, 8))

        # Scroll frame para la lista
        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        # Columnas fijas — una sola grilla compartida en el scroll
        cols_spec = [
            ("", 24, (6, 2)),           # checkbox
            ("Carpeta / Envío", 480, (4, 4)),
            ("PDF", 50, (0, 0)),
            ("Excel", 50, (0, 0)),
            ("Estado", 200, (4, 4)),
        ]

        # Configurar columnas UNA SOLA VEZ en el scroll
        for ci, (txt, w, pad) in enumerate(cols_spec):
            scroll.grid_columnconfigure(ci, minsize=w, weight=0)

        # Header de columnas — widgets directo en scroll, row 0
        for ci, (txt, w, pad) in enumerate(cols_spec):
            ctk.CTkLabel(
                scroll, text=txt, width=w,
                font=ctk.CTkFont(family=FONT_FAMILY, size=10, weight="bold"),
                text_color=Palette.TEXT_SECONDARY, anchor="w",
            ).grid(row=0, column=ci, padx=pad, pady=(8, 4), sticky="w")

        # Línea separadora — row 1
        sep = ctk.CTkFrame(scroll, height=1, fg_color=Palette.BORDER)
        sep.grid(row=1, column=0, columnspan=5, sticky="ew", padx=4, pady=2)

        vars_check = []

        for i, carp in enumerate(carpetas):
            row = i + 2
            puede = carp["pdf_ok"] and carp["xls_ok"]
            bg_fila = Palette.BG_CARD

            var = tk.BooleanVar(value=puede)
            vars_check.append(var)

            chk = ctk.CTkCheckBox(
                scroll, text="", variable=var, width=24,
                checkbox_width=20, checkbox_height=20,
                state="normal" if puede else "disabled",
                fg_color=Palette.ACCENT if puede else Palette.TEXT_MUTED,
            )
            chk.grid(row=row, column=0, padx=cols_spec[0][2], pady=3)
            if not puede:
                chk.configure(text=" " * 2)

            # Carpeta — tk.Label trunca al ancho de columna (no empuja ✅/❌)
            nombre_mostrar = carp["nombre"] if len(carp["nombre"]) <= 55 else carp["nombre"][:52] + "..."
            tk.Label(
                scroll, text=nombre_mostrar, width=58, anchor="w",
                font=(FONT_FAMILY, 10), fg=Palette.TEXT_PRIMARY,
                bg=bg_fila, bd=0, highlightthickness=0,
            ).grid(row=row, column=1, padx=cols_spec[1][2], sticky="w", pady=3)

            # PDF
            pdf_txt = "✅" if carp["pdf_ok"] else "❌"
            pdf_color = Palette.SUCCESS if carp["pdf_ok"] else Palette.ERROR
            ctk.CTkLabel(
                scroll, text=pdf_txt, width=50,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=pdf_color, anchor="center",
            ).grid(row=row, column=2, padx=cols_spec[2][2], pady=3)

            # Excel
            ctk.CTkLabel(
                scroll, text="✅", width=50,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=Palette.SUCCESS, anchor="center",
            ).grid(row=row, column=3, padx=cols_spec[3][2], pady=3)

            # Estado
            estado_txt = "✅ Listo para controlar" if puede else "❌ No se encontró PDF de Coordinación"
            ctk.CTkLabel(
                scroll, text=estado_txt, width=cols_spec[4][1],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                text_color=Palette.TEXT_MUTED, anchor="w",
            ).grid(row=row, column=4, padx=cols_spec[4][2], sticky="w", pady=3)

        # Botones
        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))

        def confirmar():
            nonlocal seleccion
            for i, (var, carp) in enumerate(zip(vars_check, carpetas)):
                if var.get() and carp["pdf_ok"] and carp["xls_ok"]:
                    seleccion.append(carp)
            top.destroy()

        def cancelar():
            top.destroy()

        ctk.CTkButton(
            btn_frame, text="Procesar seleccionados",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6, height=34, width=190,
            command=confirmar,
        ).pack(side="right", padx=4)

        ctk.CTkButton(
            btn_frame, text="Cancelar",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, hover_color=Palette.BG_HOVER,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6, height=34, width=120,
            command=cancelar,
        ).pack(side="right", padx=4)

        top.wait_window()

        if not seleccion:
            return

        # 3. Procesar seleccionadas
        if not hasattr(self, 'tree_coordinacion'):
            self._panel_cargar_datos()
            if not hasattr(self, 'tree_coordinacion'):
                self.log("ERROR: Panel de carga no disponible. Reinicie la aplicación.")
                return
        self.tree_coordinacion.delete(*self.tree_coordinacion.get_children())
        self._coord_comparaciones = {}

        # Mostrar sección coordinación, ocultar otras
        self._switch_tab("coord")

        ok_count = 0
        diff_count = 0
        error_count = 0

        for carp in seleccion:
            try:
                datos_pdf = extraer_coordinacion(carp["pdf_path"])
                if "_error" in datos_pdf:
                    self._insertar_fila_coordinacion(
                        carp["nombre"], "⚠ Error en PDF: " + datos_pdf["_error"],
                        "tag_error", None
                    )
                    error_count += 1
                    continue

                datos_xls = leer_choferes_coordinacion(carp["xls_path"])
                if "_error" in datos_xls:
                    self._insertar_fila_coordinacion(
                        carp["nombre"], "⚠ " + datos_xls["_error"],
                        "tag_error", None
                    )
                    error_count += 1
                    continue

                # OCR del permiso de exportación para fecha de oficialización
                if carp.get("permiso_path"):
                    fecha_of = extraer_fecha_permiso(carp["permiso_path"])
                    datos_pdf["fecha_of_pe"] = fecha_of
                else:
                    datos_pdf["fecha_of_pe"] = ""

                comparaciones, estado = comparar_coordinacion(datos_pdf, datos_xls)

                tag = "tag_ok" if estado == "ok" else "tag_mismatch"
                estado_label = "✅ OK" if estado == "ok" else "❌ Diferencias"

                valores = (
                    datos_pdf.get("carpeta", ""),
                    datos_pdf.get("giro", ""),
                    datos_pdf.get("cliente", ""),
                    datos_pdf.get("destino", ""),
                    datos_pdf.get("buque", ""),
                    datos_pdf.get("viaje", ""),
                    datos_pdf.get("booking", ""),
                    datos_pdf.get("pto_descarga", ""),
                    datos_pdf.get("pto_final", ""),
                    comparaciones.get("fecha_of_pe", {}).get("pdf", ""),
                    datos_xls.get("fecha_carga", ""),
                    datos_pdf.get("peso_flexi", ""),
                    estado_label,
                )

                iid = self.tree_coordinacion.insert("", "end", values=valores, tags=(tag,))
                self._coord_comparaciones[iid] = comparaciones

                if estado == "ok":
                    ok_count += 1
                else:
                    diff_count += 1

            except Exception as e:
                self._insertar_fila_coordinacion(
                    carp["nombre"], "⚠ Error: " + str(e),
                    "tag_error", None
                )
                error_count += 1

        total = ok_count + diff_count + error_count

    def _insertar_fila_coordinacion(self, carpeta, estado_texto, tag, comparaciones):
        """Inserta una fila en el treeview de coordinación."""
        valores = (carpeta, "", "", "", "", "", "", "", "", "", "", "", estado_texto)
        iid = self.tree_coordinacion.insert("", "end", values=valores, tags=(tag,))
        if comparaciones is not None:
            self._coord_comparaciones[iid] = comparaciones

    def _abrir_comparacion_coordinacion(self, event):
        """Popup con tabla verde/rojo campo por campo — mismo estilo que el de tickets."""
        sel = self.tree_coordinacion.selection()
        if not sel or not hasattr(self, "_coord_comparaciones"):
            return

        comps = self._coord_comparaciones.get(sel[0])
        if not comps:
            return

        etiquetas = {
            "giro": "Puerto Salida",
            "carpeta": "Carpeta", "cliente": "Cliente",
            "destino": "Destino", "buque": "Buque", "viaje": "Viaje",
            "booking": "Booking", "pto_descarga": "Pto Descarga",
            "pto_final": "Pto Final",
            "fecha_of_pe": "Fecha Ofic.",
            "fecha_carga": "Fecha Carga",
            "peso_flexi": "Peso Flexi (kg)",
        }

        def _build_popup(level):
            fsizes = self._get_font_sizes(level)
            geom = self._get_popup_geometry(760, 480, level)
            ancho = int(geom.split("x")[0])

            dlg = ctk.CTkToplevel(self)
            dlg.title("Comparación — Coordinación vs Excel")
            dlg.transient(self)
            dlg.resizable(False, False)

            # ── Top bar: header + font level override ──────────────────
            top_bar = ctk.CTkFrame(dlg, fg_color=Palette.BG_SIDEBAR, corner_radius=6, height=36)
            top_bar.pack(fill="x", padx=12, pady=(12, 0))
            top_bar.pack_propagate(False)

            # Grid header: align with body columns
            top_bar.grid_columnconfigure(0, minsize=120, weight=0)
            top_bar.grid_columnconfigure(1, weight=1)
            top_bar.grid_columnconfigure(2, weight=1)

            for col_i, txt in enumerate(["Campo", "PDF (Coordinación)", "Excel (Choferes)"]):
                lbl = ctk.CTkLabel(
                    top_bar, text=txt, width=120 if col_i == 0 else 280,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                    text_color=Palette.TEXT_SECONDARY,
                )
                lbl.grid(row=0, column=col_i, sticky="ew", padx=4, pady=4)

            # Font level override selector
            override_menu = ctk.CTkOptionMenu(
                top_bar, values=["1", "2", "3"],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                width=50, height=26,
                command=lambda v: (dlg.destroy(), _build_popup(int(v))),
            )
            override_menu.set(str(level))
            override_menu.pack(side="right", padx=(0, 8))

            # ── Cuerpo: grid para columnas alineadas (igual que Control Final) ──
            body = ctk.CTkFrame(dlg, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=12, pady=8)
            body.grid_columnconfigure(0, minsize=120, weight=0)  # Campo — fijo
            body.grid_columnconfigure(1, weight=1)  # PDF — expand
            body.grid_columnconfigure(2, weight=1)  # Excel — expand

            for row_i, (campo, c) in enumerate(comps.items()):
                etq = etiquetas.get(campo, campo)
                v_pdf = c["pdf"] or "—"
                v_xls = c["excel"] or "—"
                ok = c["match"]

                bg = "#C8FFC8" if ok else "#FFC8C8"
                fg = "#006400" if ok else "#8B0000"

                ctk.CTkLabel(
                    body, text=etq, width=120, anchor="w",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"], weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).grid(row=row_i, column=0, sticky="w", padx=(4, 0), pady=2)

                for col_i, val in enumerate([v_pdf, v_xls], start=1):
                    lbl = ctk.CTkLabel(
                        body, text=val, anchor="center",
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                        fg_color=bg, text_color=fg, corner_radius=4,
                    )
                    lbl.grid(row=row_i, column=col_i, sticky="ew", padx=4, pady=2)

            # ── Leyenda al pie (centrada) ─────────────────────────────
            leyenda = ctk.CTkFrame(dlg, fg_color="transparent")
            leyenda.pack(fill="x", pady=(0, 12))
            inner = ctk.CTkFrame(leyenda, fg_color="transparent")
            inner.pack(anchor="center")
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#C8FFC8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Coincide",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))
            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#FFC8C8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Difiere",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            # Ajustar ventana al contenido real
            dlg.update_idletasks()
            alto = dlg.winfo_reqheight()
            x = (dlg.winfo_screenwidth() - ancho) // 2
            y = (dlg.winfo_screenheight() - alto) // 2
            dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

        _build_popup(self.config.get("font_level", 1))

    def _control_final_switch_toggle(self):
        """Toggle auto mode. Changes btn text hint and persists state."""
        if self._cf_auto_var.get():
            self.btn_control_final.configure(text="Control Final ⚡", image=self._icons["zap"])
        else:
            self.btn_control_final.configure(text="Control Final", image=self._icons["zap"])
        # Persist state
        self.config["control_final_auto"] = self._cf_auto_var.get()
        self._guardar_config()

    def _control_final_seleccionar(self):
        """Seleccionar PDFs (tickets + aduanas) + Excel para Control Final."""
        if self._cf_auto_var.get():
            self._control_final_auto_scan()
            return
        from tkinter import filedialog as tk_filedialog
        rutas = tk_filedialog.askopenfilenames(
            title="Seleccionar PDFs (tickets + salidas aduana) + Excel",
            filetypes=[("Archivos soportados", "*.pdf *.xls *.xlsx")],
        )
        if not rutas:
            return
        # Separar Excel de PDFs
        exceles = [r for r in rutas if r.lower().endswith(('.xls', '.xlsx'))]
        pdfs = [r for r in rutas if r.lower().endswith('.pdf')]
        if not exceles:
            self._log("ERROR: Debe seleccionar al menos un Excel de Contenedores")
            return
        if not pdfs:
            self._log("ERROR: Debe seleccionar al menos un PDF")
            return

        self._log(f"[Control Final] {len(pdfs)} PDFs, {len(exceles)} Excel(s)")

        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self.tarea_activa = True

        for btn_name in ('btn_control_final', 'btn_controlar_tickets', 'btn_controlar_coordinacion'):
            btn = getattr(self, btn_name, None)
            if btn and btn.winfo_exists():
                btn.configure(state="disabled")

        if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
            self.progress_carga.pack(side="right", padx=12)
            self.progress_carga.configure(mode="indeterminate")
            self.progress_carga.start()

        # Mostrar sección control final, ocultar otras
        self._switch_tab("final")

        # Limpiar tree viejo
        for item in self.tree_control_final.get_children():
            self.tree_control_final.delete(item)

        t = threading.Thread(
            target=self._control_final_worker,
            args=(list(pdfs), list(exceles)),
            daemon=True,
        )
        t.start()

    def _control_final_auto_scan(self):
        """Auto-scan Desktop for folders with CONTENEDORES Excel."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return

        self._scan_desktop_folders(callback=self._control_final_auto_scan_done)

    def _control_final_auto_scan_done(self, folders):
        """Callback after background scan completes for control final auto."""
        if not folders:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con CONTENEDORES en el Desktop"
            )
            return

        self._log(f"[Auto] {len(folders)} carpetas con CONTENEDORES encontradas")
        self._control_final_auto_popup(folders)

    def _control_final_auto_popup(self, folders):
        """Popup to select folders for auto-scan Control Final (planilla de carga style)."""
        top = ctk.CTkToplevel(self)
        top.title("Control Final Automático")
        top.geometry("600x400")
        top.configure(fg_color=Palette.BG_CARD)
        top.transient(self)
        top.lift()
        top.grab_set()
        top.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 600) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        top.geometry(f"600x400+{px}+{py}")

        ctk.CTkLabel(
            top, text="CONTROL FINAL AUTOMÁTICO",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        ctk.CTkLabel(
            top, text="Seleccioná las carpetas del Desktop a procesar:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_TABLE, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for carp in folders:
            var = ctk.BooleanVar(value=True)
            checks[carp["name"]] = var

        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        def _update_btn(*_args):
            any_selected = any(v.get() for v in checks.values())
            btn_procesar.configure(state="normal" if any_selected else "disabled")

        def _todas():
            for v in checks.values():
                v.set(True)

        def _ninguna():
            for v in checks.values():
                v.set(False)

        def _confirmar():
            selected = [c for c in folders if checks[c["name"]].get()]
            top.destroy()

            if not selected:
                return

            # Collect PDFs + Excel from selected folders
            # Only scan_* (API Vision tickets), get* (Salidas de Aduana),
            # and YYAR* (MIC/DTA terrestre, e.g. 26AR240914H)
            import re as _re
            _permiso_pat = _re.compile(r'^\d{2}069EC', _re.IGNORECASE)
            _mic_pat = _re.compile(r'^\d{2}AR', _re.IGNORECASE)
            all_pdfs = []
            all_excels = []
            for carp in selected:
                for f in os.scandir(carp["path"]):
                    if f.is_file() and f.name.lower().endswith('.pdf'):
                        if _permiso_pat.match(f.name):
                            continue  # skip permisos de exportación
                        name_lower = f.name.lower()
                        if not (name_lower.startswith("scan")
                                or name_lower.startswith("get")
                                or _mic_pat.match(f.name)):
                            continue  # only scan_*, get*, and YYAR* (MIC) PDFs
                        all_pdfs.append(f.path)
                if carp["excel_path"]:
                    all_excels.append(carp["excel_path"])

            if not all_pdfs:
                messagebox.showwarning(
                    "Sin PDFs",
                    "Las carpetas seleccionadas no contienen PDFs"
                )
                return
            if not all_excels:
                messagebox.showwarning(
                    "Sin Excel",
                    "Las carpetas seleccionadas no contienen Excel CONTENEDORES"
                )
                return

            self._log(f"[Auto] {len(all_pdfs)} PDFs, {len(all_excels)} Excel(s)")

            if self.tarea_activa:
                messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
                return
            self.tarea_activa = True

            for btn_name in ('btn_control_final', 'btn_controlar_tickets',
                             'btn_controlar_coordinacion'):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="disabled")

            if (hasattr(self, 'progress_carga')
                    and self.progress_carga.winfo_exists()):
                self.progress_carga.pack(side="right", padx=12)
                self.progress_carga.configure(mode="indeterminate")
                self.progress_carga.start()

            # Show control final section
            self._switch_tab("final")

            for item in self.tree_control_final.get_children():
                self.tree_control_final.delete(item)

            t = threading.Thread(
                target=self._control_final_worker,
                args=(list(all_pdfs), list(all_excels)),
                daemon=True,
            )
            t.start()

        # Todas / Ninguna (izquierda)
        ctk.CTkButton(
            btn_frame, text="Todas", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_todas,
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            btn_frame, text="Ninguna", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_ninguna,
        ).pack(side="left", padx=4)

        # Procesar (derecha)
        btn_procesar = ctk.CTkButton(
            btn_frame, text="Procesar", width=140, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=_confirmar,
        )
        btn_procesar.pack(side="right", padx=4)

        # Checkboxes
        for carp in folders:
            var = checks[carp["name"]]
            var.trace_add("write", _update_btn)
            ctk.CTkCheckBox(
                scroll, text=carp["name"], variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

    def _control_final_worker(self, pdf_paths, excel_paths):
        """Worker thread para Control Final."""
        import procesar_tickets
        from datetime import date
        import os

        try:
            result_count = 0
            # 1. Clasificar PDFs
            tickets_pdf = []   # scanned, need OCR
            aduanas_data = []  # PLT / Salida Aduana (texto)
            mic_data = []      # MIC/DTA (texto)

            for ruta in pdf_paths:
                import fitz
                doc = fitz.open(ruta)
                text = ""
                for p in doc:
                    text += p.get_text()
                doc.close()

                if not text.strip():
                    tickets_pdf.append(ruta)
                    continue

                if "MIC/DTA" in text.upper() or "MANIFIESTO INTERNACIONAL DE CARGA" in text.upper():
                    # MIC/DTA → terrestre
                    md = procesar_tickets.extraer_mic_dta(ruta)
                    if "_error" not in md:
                        md["_archivo"] = os.path.basename(ruta)
                        mic_data.append(md)
                elif "SALIDA DE ZONA PRIMARIA" in text.upper():
                    # PLT Aduana
                    data = procesar_tickets.extraer_salida_aduana(ruta, modo="flexi")
                    data["_archivo"] = os.path.basename(ruta)
                    aduanas_data.append(data)
                else:
                    tickets_pdf.append(ruta)

            self._log(f"[Control Final] {len(tickets_pdf)} tickets, "
                      f"{len(aduanas_data)} salidas aduana, {len(mic_data)} mic/dta")

            # 2. OCR tickets
            self._contenedores_cache = {}
            self._cargar_cache_contenedores(excel_paths)

            ocr_method = self.config.get("ocr_method", "api_vision")
            textos_por_pdf = {}
            api_datos_raw = {}

            # Sort by client folder name for grouped display (always, before any processing)
            tickets_pdf.sort(key=lambda r: os.path.basename(os.path.dirname(r)).split("_")[7].lower() if len(os.path.basename(os.path.dirname(r)).split("_")) > 7 else "")

            if ocr_method == "api_vision":
                api_vision_conf = self.config.get("api_vision", {})
                api_key_raw = api_vision_conf.get("api_key", "")
                api_key = self._decrypt_api_key(api_key_raw)
                model = api_vision_conf.get("model", procesar_tickets.MODELO_VISION_DEFAULT)
                parallel_enabled = api_vision_conf.get("parallel_enabled", False)

                if parallel_enabled and len(tickets_pdf) > 1:
                    # Parallel mode: distribute across enabled models
                    model_states = api_vision_conf.get("parallel_model_states", {})
                    enabled_models = [m for m, on in model_states.items() if on]
                    if not enabled_models:
                        enabled_models = [model]
                    # Ensure selected model is first
                    if model in enabled_models:
                        enabled_models.remove(model)
                    enabled_models.insert(0, model)
                    temperature = api_vision_conf.get("temperature", 0.1)
                    max_tokens = api_vision_conf.get("max_tokens", 4000)
                    timeout = api_vision_conf.get("timeout", 60)
                    result = procesar_tickets.api_vision_con_fallback(
                        tickets_pdf, api_key, enabled_models,
                        temperature=temperature, max_tokens=max_tokens,
                        timeout=timeout, log_callback=self._log,
                    )
                    api_datos_raw = result["datos"]
                    textos_por_pdf = result["textos"]
                else:
                    # Sequential fallback: try models one by one per PDF
                    all_models = api_vision_conf.get("custom_models", [])
                    if not all_models:
                        all_models = list(procesar_tickets.MODELOS_VISION)
                    # Ensure selected model is first
                    if model in all_models:
                        all_models.remove(model)
                    all_models.insert(0, model)

                    self._log(f"[Control Final] API Visión secuencial: {len(tickets_pdf)} PDF(s), {len(all_models)} modelo(s)")
                    for ruta in tickets_pdf:
                        stem = os.path.splitext(os.path.basename(ruta))[0]
                        self._log(f"  API Visión: {stem}")
                        success = False
                        for m in all_models:
                            datos = procesar_tickets.api_vision_extraer_datos(
                                ruta, api_key, model=m
                            )
                            if "error" in datos:
                                err = datos["error"]
                                if "is not a valid model ID" in err:
                                    err = "modelo no disponible"
                                elif "429" in err or "rate" in err.lower():
                                    err = "rate limit"
                                elif "timeout" in err.lower():
                                    err = "timeout"
                                self._log(f"  ERROR {m}: {err}")
                                continue
                            # Success
                            api_datos_raw[stem] = datos
                            self._log(f"  ✓ {stem} procesado con {m}")
                            success = True
                            break
                        if not success:
                            self._log(f"  ⚠ Todos los modelos fallaron para {stem}, usando PaddleOCR")
                            texto = procesar_tickets.pdf_a_texto(ruta, engine="paddleocr")
                            textos_por_pdf[stem] = texto
            else:
                textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(tickets_pdf, engine="paddleocr")

            # 3. Build ALL lookup structures (terrestre + flexi)
            self._log(f"[Control Final] Procesando tickets...")
            import re as _re

            def _norm_pat_match(s):
                """Normalizar patente: sacar espacios, uppercase."""
                if not s: return ""
                return _re.sub(r'\s+', '', s).upper()

            # ── Aduana terrestre lookups (MIC/DTA) ──
            aduana_por_patente = {}
            aduana_por_precinto_mic = {}
            aduana_por_dni = {}
            for md in mic_data:
                pat = _norm_pat_match(md.get("patente_camion", ""))
                if pat: aduana_por_patente[pat] = md
                if md.get("precinto"):
                    for p in md["precinto"].upper().split():
                        p = p.strip()
                        if p: aduana_por_precinto_mic[p] = md
                dni = _re.sub(r'\D', '', md.get("cuil", ""))
                if dni: aduana_por_dni[dni] = md

            # ── Aduana flexi lookups (Salida de Aduana) ──
            aduana_por_contenedor = {}
            aduana_por_precinto_aduana = {}
            for ad in aduanas_data:
                if ad.get("contenedor"):
                    aduana_por_contenedor[ad["contenedor"].upper()] = ad
                if ad.get("precinto"):
                    for p in ad["precinto"].upper().split():
                        p = p.strip()
                        if p: aduana_por_precinto_aduana[p] = ad

            # ── Excel lookups (ALL cache, list-based for shared trips) ──
            excel_por_patente = {}  # patente -> [(cont_data, cont_idx), ...]
            excel_por_precinto = {}
            excel_por_dni = {}
            excel_por_contenedor = {}
            for k, v in self._contenedores_cache.items():
                camiones = v.get("camiones", [])
                for ci, cam in enumerate(camiones):
                    pat = _norm_pat_match(cam.get("patente_camion", ""))
                    if pat:
                        excel_por_patente.setdefault(pat, []).append((v, ci))
                    prec = cam.get("precinto", "").upper().strip()
                    if prec:
                        # Normalizar: dividir por guión para crear key individual
                        for pp in prec.replace("-", " ").split():
                            if pp:
                                excel_por_precinto.setdefault(pp, []).append((v, ci))
                    dni = _re.sub(r'\D', '', str(cam.get("dni", "")))
                    if dni:
                        excel_por_dni.setdefault(dni, []).append((v, ci))
                    ctn = _re.sub(r'[\s\-]+', '', cam.get("contenedor", "").upper().strip())
                    if ctn:
                        excel_por_contenedor.setdefault(ctn, []).append((v, ci))

            # Merged precinto aduana lookup (try both MIC and Salida)
            aduana_por_precinto_all = {**aduana_por_precinto_mic, **aduana_por_precinto_aduana}

            # 4. Process each ticket — try terrestre first, then flexi
            for ruta in tickets_pdf:
                stem = os.path.splitext(os.path.basename(ruta))[0]
                carpeta = os.path.basename(os.path.dirname(ruta))
                cliente = carpeta.split("_")[7] if len(carpeta.split("_")) > 7 else carpeta
                try:
                    shared_excel_matches = []
                    modo_result = "flexi"  # default per-ticket

                    # Build ticket_data
                    if stem in api_datos_raw:
                        raw = api_datos_raw[stem]
                        ticket_data = {
                            "archivo":   stem,
                            "Patente":   raw.get("Patente", ""),
                            "Semirremolque": raw.get("Semirremolque", ""),
                            "Conductor": raw.get("Conductor", ""),
                            "DNI":       raw.get("DNI", ""),
                            "Neto":      raw.get("Neto", ""),
                            "Tara Contenedor": raw.get("Tara Contenedor", ""),
                            "Contenedor": raw.get("Contenedor", ""),
                            "Permiso":   raw.get("Permiso", ""),
                        }
                    else:
                        texto = textos_por_pdf.get(stem, "")
                        if not texto:
                            continue
                        extraido = procesar_tickets.extraer_datos(texto)
                        ticket_data = {
                            "archivo":   stem,
                            "Patente":   extraido.get("patente", ""),
                            "Semirremolque": extraido.get("semi", ""),
                            "Conductor": extraido.get("conductor", ""),
                            "DNI":       extraido.get("dni", ""),
                            "Neto":      extraido.get("neto", ""),
                            "Tara Contenedor": extraido.get("tara", ""),
                            "Contenedor": extraido.get("contenedor", ""),
                            "Permiso":   extraido.get("oferta", ""),
                        }

                    patente_ticket = _norm_pat_match(ticket_data.get("Patente", ""))
                    dni_ticket = _re.sub(r'\D', '', ticket_data.get("DNI", ""))
                    contenedor_ticket = _re.sub(r'[\s\-]+', '',
                                                ticket_data.get("Contenedor", "").upper().strip())

                    aduana = {}
                    cont_data = None
                    cont_idx = -1

                    # ── Step 1: Try terrestre (patente → MIC/DTA + patente → Excel) ──
                    if patente_ticket:
                        aduana = aduana_por_patente.get(patente_ticket, {})
                        all_excel = excel_por_patente.get(patente_ticket, [])
                        if all_excel:
                            cont_data, cont_idx = all_excel[0]
                            shared_excel_matches = all_excel[1:]
                        if aduana:
                            modo_result = "terrestre"

                        # Variants O/0 (Mercosur patentes)
                        if not cont_data and not aduana:
                            for i, ch in enumerate(patente_ticket):
                                variants = set()
                                if ch == 'O': variants.add(patente_ticket[:i] + '0' + patente_ticket[i+1:])
                                if ch == '0': variants.add(patente_ticket[:i] + 'O' + patente_ticket[i+1:])
                                if ch == 'I': variants.add(patente_ticket[:i] + '1' + patente_ticket[i+1:])
                                if ch == '1': variants.add(patente_ticket[:i] + 'I' + patente_ticket[i+1:])
                                if ch == 'S': variants.add(patente_ticket[:i] + '5' + patente_ticket[i+1:])
                                if ch == '5': variants.add(patente_ticket[:i] + 'S' + patente_ticket[i+1:])
                                for vr in variants:
                                    if vr in excel_por_patente:
                                        vm = excel_por_patente[vr]
                                        if not cont_data:
                                            cont_data, cont_idx = vm[0]
                                            shared_excel_matches = vm[1:] if len(vm) > 1 else []
                                            self._log(f"  ⚠ Patente OCR corregida: {patente_ticket}→{vr} (Excel)")
                                    if not aduana and vr in aduana_por_patente:
                                        aduana = aduana_por_patente[vr]
                                        self._log(f"  ⚠ Patente OCR corregida: {patente_ticket}→{vr} (Aduana)")
                                    if cont_data and aduana: break
                                if cont_data and aduana: break

                    # ── Step 2: If no terrestre match, try flexi (contenedor → Salida Aduana) ──
                    if not aduana and contenedor_ticket:
                        aduana = aduana_por_contenedor.get(contenedor_ticket, {})
                        if aduana:
                            modo_result = "flexi"
                            self._log(f"  ✓ {stem}: flexi match por contenedor='{contenedor_ticket}'")
                            # Match Excel by contenedor
                            all_excel = excel_por_contenedor.get(contenedor_ticket, [])
                            if all_excel:
                                cont_data, cont_idx = all_excel[0]
                                shared_excel_matches = all_excel[1:]

                    # ── Step 3: Fallback por precinto ──
                    if not cont_data or not aduana:
                        # 3a: Have Excel but need Aduana
                        if cont_data and cont_idx >= 0 and not aduana:
                            cam = cont_data["camiones"][cont_idx]
                            prec = cam.get("precinto", "").upper().strip()
                            if prec and prec in aduana_por_precinto_all:
                                aduana = aduana_por_precinto_all[prec]
                                modo_result = "terrestre" if prec in aduana_por_precinto_mic else "flexi"
                                self._log(f"  ⚠ Aduana x precinto Excel: {stem} — precinto={prec}")

                        # 3b: Have Aduana but need Excel
                        if aduana and (not cont_data or cont_idx < 0):
                            prec = aduana.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    self._log(f"  ⚠ Excel x precinto Aduana: {stem} — precinto={p}")
                                    break

                    # ── Step 3c: DNI fallback ──
                    if not cont_data and not aduana and dni_ticket:
                        if dni_ticket in excel_por_dni:
                            matches = excel_por_dni[dni_ticket]
                            cont_data, cont_idx = matches[0]
                            shared_excel_matches = matches[1:] if len(matches) > 1 else []
                            self._log(f"  ⚠ Excel x DNI: {stem} — dni={dni_ticket}")
                        if dni_ticket in aduana_por_dni:
                            aduana = aduana_por_dni[dni_ticket]
                            modo_result = "terrestre"
                            self._log(f"  ⚠ Aduana x DNI: {stem} — dni={dni_ticket}")

                    # ── Step 3d: Puente post-DNI ──
                    if not cont_data or not aduana:
                        if cont_data and cont_idx >= 0 and not aduana:
                            cam = cont_data["camiones"][cont_idx]
                            prec = cam.get("precinto", "").upper().strip()
                            if prec and prec in aduana_por_precinto_all:
                                aduana = aduana_por_precinto_all[prec]
                                self._log(f"  ⚠ Aduana x precinto post-DNI: {stem} — precinto={prec}")
                        if aduana and (not cont_data or cont_idx < 0):
                            prec = aduana.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    self._log(f"  ⚠ Excel x precinto post-DNI: {stem} — precinto={p}")
                                    break

                    # ── Step 3e: Brute force — scan all MIC by precinto ──
                    if not cont_data and not aduana:
                        for md in mic_data:
                            prec = md.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    aduana = md
                                    modo_result = "terrestre"
                                    self._log(f"  ⚠ Sin match x precinto (último): {stem} — precinto={p}")
                                    break
                            if cont_data: break

                    # ── Step 3f: Brute force — scan all Salida Aduana by precinto ──
                    if not cont_data and not aduana:
                        for ad in aduanas_data:
                            prec = ad.get("precinto", "").upper().strip()
                            for p in prec.split():
                                if p in excel_por_precinto:
                                    matches = excel_por_precinto[p]
                                    cont_data, cont_idx = matches[0]
                                    shared_excel_matches = matches[1:] if len(matches) > 1 else []
                                    aduana = ad
                                    modo_result = "flexi"
                                    self._log(f"  ⚠ Sin match x precinto (flexi): {stem} — precinto={p}")
                                    break
                            if cont_data: break

                    result = self._build_fila_control_final(
                        ticket_data, cont_data, cont_idx, aduana, stem, cliente=cliente, modo=modo_result,
                        shared_excel_matches=shared_excel_matches
                    )
                    if result:
                        self.log_queue.put(("_CONTROL_FINAL_RESULT_", result))
                        result_count += 1

                except Exception as e:
                    self._log(f"  ✗ Error {stem}: {e}")
                    import traceback
                    self._log(traceback.format_exc())

        except Exception as e:
            self._log(f"[Control Final] Error general: {e}")
            import traceback
            self._log(traceback.format_exc())
        finally:
            try: rc = result_count
            except NameError: rc = 0
            self._log(f"[Control Final] Finalizado — {rc} ticket(s) procesado(s)")
            self.log_queue.put(("_TAREA_COMPLETA_", None))

    def _build_fila_control_final(self, ticket_data, cont_data, cont_idx, aduana, stem, cliente="",
                                   modo="flexi", shared_excel_matches=None):
        """Build row values and comparison data for Control Final.

        Args:
            modo: "flexi" (ISO/Flexi containers) or "terrestre" (bulk/granel).
            shared_excel_matches: list of (cont_data, cont_idx) from additional
                Excels matching the same truck (shared trips). The primary match
                is NOT in this list — it's passed via cont_data/cont_idx.

        Returns dict with keys:
          valores    → tuple for tree row (11 cols)
          tag        → tag_ok / tag_mismatch
          ticket     → dict {Patente, Semirremolque, ..., Permiso} (normalized ticket)
          contenedor → dict {Patente, Semirremolque, ..., Permiso} (normalized Excel)
          aduana     → dict {Patente Camión, ..., CUIL, Contenedor, ...}
          ok         → dict per field: True if ticket matches reference
          modo       → the mode used
          shared_excels → list of {peso_carga, archivo} for shared trips (empty if not shared)
          shared_neto_sum → sum of all Excel peso_carga (0 if not shared)
        """
        import procesar_tickets
        import re as _re

        # ── Normalization helpers ──────────────────────────────
        def _norm_pat(s):
            return str(s).replace(" ", "") if s else s

        def _norm_num(s):
            """Remove trailing .0 / .00 + thousands-separator dots from numeric values."""
            if not s:
                return s
            v = str(s).strip()
            v = _re.sub(r'\.0+$', '', v)      # "23440.0" → "23440"
            v = v.replace(".", "")             # "23.440" → "23440"
            return v

        def _fmt_neto(s):
            """Normalize number then format with dot as thousands separator: '27890' → '27.890'."""
            if not s:
                return s
            v = str(s).strip()
            # Quitar comas y puntos existentes
            v = v.replace(",", "")
            v = _re.sub(r'\.0+$', '', v)
            v = v.replace(".", "")
            if v.isdigit():
                return f"{int(v):,}".replace(",", ".")
            return v

        def _norm_ctn(s):
            """Remove spaces and dashes from contenedor numbers."""
            if not s:
                return s
            return _re.sub(r'[\s\-]+', '', str(s).strip())

        def _dni_solo(v):
            """Extraer solo DNI, limpiando CUIL/CUIT y no-dígitos."""
            if not v:
                return ""
            d = _re.sub(r'\D', '', v)
            if len(d) == 11 and d[:2] in ("20", "23", "27", "30"):
                return d[2:-1]
            return d

        # ── Ticket data (ORIGINAL, never overwrite) ──
        t_patente    = ticket_data.get("Patente", "")
        t_semi       = ticket_data.get("Semirremolque", "")
        t_conductor  = ticket_data.get("Conductor", "")
        t_dni        = ticket_data.get("DNI", "")
        t_neto       = ticket_data.get("Neto", "")
        t_tara       = ticket_data.get("Tara Contenedor", "")
        # Normalize ticket OCR values: strip thousand-separator dots
        # API sometimes returns "24.000" (with dots) instead of "24000"
        _tmp = str(t_neto).replace(".", "").replace(",", ".")
        if _tmp.replace(".", "", 1).isdigit():
            t_neto = _tmp
        _tmp = str(t_tara).replace(".", "").replace(",", ".")
        if _tmp.replace(".", "", 1).isdigit():
            t_tara = _tmp
        t_contenedor = ticket_data.get("Contenedor", "")
        t_permiso    = ticket_data.get("Permiso", "")
        t_dni        = _dni_solo(t_dni)

        # ── Aduana data ──
        a_plt            = aduana.get("plt", "")
        a_peso_bruto     = aduana.get("peso_bruto", "")
        a_cuil           = aduana.get("cuil", "")
        a_patente_camion = aduana.get("patente_camion", "")
        a_patente_semi   = aduana.get("patente_semi", "")
        a_conductor      = aduana.get("conductor", "")
        a_id_destinacion = aduana.get("id_destinacion", "")
        a_id_destinacion_1 = aduana.get("id_destinacion_1", "")
        a_id_destinacion_2 = aduana.get("id_destinacion_2", "")
        a_exportador     = aduana.get("exportador", "")
        a_contenedor     = aduana.get("contenedor", "")
        a_precinto       = aduana.get("precinto", "")

        a_dni = _dni_solo(a_cuil)

        # ── Excel / Contenedor data ──
        e_patente = ""
        e_semi = ""
        e_conductor = ""
        e_dni = ""
        e_neto = ""
        e_tara = ""
        e_contenedor = ""
        e_precinto = ""
        e_permiso = ""
        e_peso_flexi = "0"
        if cont_data and cont_idx >= 0:
            camiones = cont_data.get("camiones", [])
            if cont_idx < len(camiones):
                cam = camiones[cont_idx]
                e_patente    = cam.get("patente_camion", "")
                e_semi       = cam.get("patente_semi", "")
                e_conductor  = cam.get("conductor", "")
                e_dni        = cam.get("dni", "")
                e_neto       = str(cam.get("peso_carga", ""))
                e_tara       = str(cam.get("tara_cont", ""))
                e_contenedor = cam.get("contenedor", "")
                e_precinto   = cam.get("precinto", "")
                e_permiso    = cont_data.get("pe", "")
                e_peso_flexi = str(cam.get("peso_flexi", "0"))
        e_dni = _dni_solo(e_dni)

        # ── Shared Excel data ──
        shared_excels = []
        shared_neto_sum = 0
        if shared_excel_matches:
            for s_data, s_idx in shared_excel_matches:
                s_camiones = s_data.get("camiones", [])
                if s_idx < len(s_camiones):
                    s_cam = s_camiones[s_idx]
                    s_neto = s_cam.get("peso_carga", 0)
                    shared_excels.append({
                        "peso_carga": s_neto,
                        "archivo": os.path.basename(s_data.get("_archivo", "")),
                    })
                    shared_neto_sum += s_neto

        # Also count primary match
        if cont_data and cont_idx >= 0:
            camiones = cont_data.get("camiones", [])
            if cont_idx < len(camiones):
                primary_neto = camiones[cont_idx].get("peso_carga", 0)
                shared_excels.insert(0, {
                    "peso_carga": primary_neto,
                    "archivo": os.path.basename(cont_data.get("_archivo", "")),
                })
                shared_neto_sum += primary_neto

        is_shared = len(shared_excels) > 1

        # ── Normalize all values ──
        t_patente    = _norm_pat(t_patente)
        t_semi       = _norm_pat(t_semi)
        t_neto       = _norm_num(t_neto)
        t_contenedor = _norm_ctn(t_contenedor)

        e_patente    = _norm_pat(e_patente)
        e_semi       = _norm_pat(e_semi)
        e_neto       = _norm_num(e_neto)
        e_contenedor = _norm_ctn(e_contenedor)

        a_patente_camion = _norm_pat(a_patente_camion)
        a_patente_semi   = _norm_pat(a_patente_semi)
        a_peso_bruto     = _norm_num(a_peso_bruto)
        a_contenedor     = _norm_ctn(a_contenedor)

        # ── Calculate aduana neto (solo flexi: peso_bruto - flexi) ──
        if modo == "terrestre":
            a_neto_calc = a_peso_bruto
        else:
            a_neto_calc = a_peso_bruto
            if a_peso_bruto:
                try:
                    pb = float(a_peso_bruto.replace(",", ""))
                    pf_val = (cont_data.get("peso_flexi_global") if cont_data else None) or e_peso_flexi
                    pf = float(pf_val) if pf_val and pf_val != "0" else 0
                    a_neto_calc = str(int(pb - pf)) if pf > 0 else str(int(pb))
                except (ValueError, TypeError):
                    pass

        # ── Apply terrestre field overrides ──
        if modo == "terrestre":
            t_contenedor = "—"
            e_contenedor = "—"
            if not a_contenedor:
                a_contenedor = "—"

        # ── Format neto/tara with dot thousands separator ──
        t_neto       = _fmt_neto(t_neto)
        e_neto       = _fmt_neto(e_neto)
        a_neto_calc  = _fmt_neto(a_neto_calc)
        t_tara       = _fmt_neto(t_tara)
        e_tara       = _fmt_neto(e_tara)
        # No format a_peso_bruto display — a_neto_calc is used for comparison

        # ── Shared trips: override Excel Neto with TOTAL (all fractions) ──
        if is_shared and shared_neto_sum:
            e_neto = _fmt_neto(str(int(shared_neto_sum)))

        # ── Format ticket PE: si tiene "/", mantener primera parte + últimos 5 de la segunda ──
        if "/" in t_permiso:
            parts = t_permiso.split("/")
            if len(parts) >= 2 and len(parts[-1].strip()) >= 5:
                t_permiso = f"{parts[0].strip()}/{parts[-1].strip()[-5:]}"

        # ── For shared trips: combine PEs (Ticket + Excel + MIC) ──
        if is_shared:
            # Ticket: primary PE / last 5 of secondary PE (to match Excel format)
            # Only if ticket PE doesn't already have "/" (already formatted above)
            if t_permiso and "/" not in t_permiso and shared_excel_matches:
                s_data, s_idx = shared_excel_matches[0]
                s_pe = s_data.get("pe", "")
                if s_pe and len(s_pe) >= 5:
                    t_permiso = f"{t_permiso}/{s_pe[-5:]}"
            # Excel: primary PE / last 5 of secondary PE
            if e_permiso and shared_excel_matches:
                s_data, s_idx = shared_excel_matches[0]
                s_pe = s_data.get("pe", "")
                if s_pe and len(s_pe) >= 5:
                    e_permiso = f"{e_permiso}/{s_pe[-5:]}"
            # MIC: hoja 1 PE / last 5 of hoja 2 PE
            if modo == "terrestre" and a_id_destinacion_1 and a_id_destinacion_2:
                a_id_destinacion = f"{a_id_destinacion_1}/{a_id_destinacion_2[-5:]}"

        # ── Build display dicts (all normalized) ──
        ticket_d = {
            "Patente": t_patente, "Semirremolque": t_semi,
            "Conductor": t_conductor, "DNI": t_dni,
            "Neto (kg)": t_neto, "Tara (kg)": t_tara,
            "Contenedor": t_contenedor, "Permiso": t_permiso,
            "Precinto": "-",
        }
        cont_d = {
            "Patente": e_patente, "Semirremolque": e_semi,
            "Conductor": e_conductor, "DNI": e_dni,
            "Neto (kg)": e_neto, "Tara (kg)": e_tara,
            "Contenedor": e_contenedor, "Permiso": e_permiso,
            "Precinto": e_precinto,
        }
        dni_key = "DNI" if modo == "terrestre" else "CUIL"
        # MIC/DTA campo 42/44/46 — shared trip breakdown (terrestre only)
        a_campo_42 = aduana.get("campo_42", "")
        a_campo_44 = aduana.get("campo_44", "")
        a_campo_46 = aduana.get("campo_46", "")
        aduana_d = {
            "PLT": a_plt,
            "Peso Bruto Aduana": a_neto_calc,
            "Id Destinación": a_id_destinacion,
            "Exportador": a_exportador,
            "Patente Camión": a_patente_camion,
            "Patente Semi": a_patente_semi,
            "Conductor Aduana": a_conductor,
            dni_key: a_dni,
            "Contenedor": a_contenedor,
            "Precinto": a_precinto,
            "campo_42": a_campo_42,
            "campo_44": a_campo_44,
            "campo_46": a_campo_46,
        }

        # ── 3-way comparison ──
        ok_map = {}
        all_ok = True
        compare_fields = [
            "Patente", "Semirremolque", "Conductor",
            "DNI", "Neto (kg)", "Tara (kg)",
            "Contenedor", "Permiso", "Precinto",
        ]

        for campo in compare_fields:
            v_ticket = ticket_d.get(campo, "").strip().upper()
            v_cont   = cont_d.get(campo, "").strip().upper()

            # Map field → aduana variable
            aduana_val = ""
            if   campo == "Patente":      aduana_val = a_patente_camion.strip().upper()
            elif campo == "Semirremolque": aduana_val = a_patente_semi.strip().upper()
            elif campo == "Conductor":    aduana_val = a_conductor.strip().upper()
            elif campo == "DNI":          aduana_val = a_dni.strip().upper()
            elif campo == "Neto (kg)":
                # Shared terrestre: use campo_46 (total) instead of a_peso_bruto (hoja 1 only)
                if is_shared and modo == "terrestre" and a_campo_46:
                    aduana_val = _fmt_neto(a_campo_46).strip().upper()
                else:
                    aduana_val = a_neto_calc.strip().upper()
            elif campo == "Contenedor":   aduana_val = a_contenedor.strip().upper()
            elif campo == "Permiso":      aduana_val = a_id_destinacion.strip().upper()
            elif campo == "Precinto":     aduana_val = a_precinto.strip().upper()
            # Tara → no aduana equivalent, stays ""

            # Normalize conductor accents before compare
            if campo == "Conductor":
                import unicodedata
                _norm = lambda s: unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().strip()
                if v_ticket:   v_ticket   = _norm(v_ticket)
                if v_cont:     v_cont     = _norm(v_cont)
                if aduana_val: aduana_val = _norm(aduana_val)

                # Word-level matching: al menos 2 palabras de cada fuente en Aduana
                t_words = set(v_ticket.split()) if v_ticket else set()
                c_words = set(v_cont.split()) if v_cont else set()
                a_words = set(aduana_val.split()) if aduana_val else set()

                ticket_ok = len(t_words & a_words) >= 2 if len(t_words) >= 2 else t_words.issubset(a_words) if t_words else False
                excel_ok  = len(c_words & a_words) >= 2 if len(c_words) >= 2 else c_words.issubset(a_words) if c_words else False
                ok_map[campo] = ticket_ok and excel_ok
                if not ok_map[campo]:
                    all_ok = False
                continue  # skip the generic comparison below

            # Normalize DNI: strip non-digits for comparison
            if campo == "DNI":
                v_ticket   = _re.sub(r'\D', '', v_ticket)
                v_cont     = _re.sub(r'\D', '', v_cont)
                aduana_val = _re.sub(r'\D', '', aduana_val)

            # Collect non-empty values from all 3 sources
            vals = set()
            # Precinto: ticket siempre "-", no comparar
            if campo == "Precinto":
                if v_cont:     vals.add(v_cont)
                if aduana_val: vals.add(aduana_val)
            elif campo == "Tara (kg)":
                # Tara: solo Ticket vs Excel (Aduana no tiene tara)
                if v_ticket:   vals.add(v_ticket)
                if v_cont:     vals.add(v_cont)
            else:
                if v_ticket:   vals.add(v_ticket)
                if v_cont:     vals.add(v_cont)
                if aduana_val: vals.add(aduana_val)

            # 0 or 1 unique value → match; 2+ unique → mismatch
            if len(vals) <= 1:
                ok_map[campo] = True
            else:
                # Para Precinto: si hay multi-precinto, verificar solapamiento
                if campo == "Precinto" and len(vals) > 1:
                    cont_codes = set(v_cont.replace("-", " ").split()) if v_cont else set()
                    adu_codes = set(aduana_val.replace("-", " ").split()) if aduana_val else set()
                    if cont_codes & adu_codes:
                        ok_map[campo] = True
                    else:
                        ok_map[campo] = False
                        all_ok = False
                # Para Permiso en compartidos: al menos 2 de 3 fuentes coinciden
                elif campo == "Permiso" and is_shared:
                    sources = [v for v in [v_ticket, v_cont, aduana_val] if v]
                    if sources and max(sources.count(v) for v in sources) >= 2:
                        ok_map[campo] = True
                    else:
                        ok_map[campo] = False
                        all_ok = False
                else:
                    ok_map[campo] = False
                    all_ok = False

        estado = "ok" if all_ok else "mismatch"
        estado_label = "✅ OK" if all_ok else "❌ Diferencia"
        tag = "tag_ok" if all_ok else "tag_mismatch"

        # Tree row: all values normalized
        valores = (
            cliente,
            t_patente,
            t_semi,
            t_conductor,
            self._fmt_dni(t_dni),
            t_neto,
            t_tara,
            t_contenedor,
            t_permiso,
            estado_label,
            a_plt,
        )

        return {
            "valores": valores,
            "tag": tag,
            "ticket": ticket_d,
            "contenedor": cont_d,
            "aduana": aduana_d,
            "ok": ok_map,
            "modo": modo,
            "shared_excels": shared_excels,
            "shared_neto_sum": shared_neto_sum,
        }

    def _procesar_resultado_control_final(self, data):
        """Inserta fila en tree_control_final."""
        try:
            stem = data.get("valores", ())[0] if data.get("valores") else "?"
            iid = self.tree_control_final.insert("", "end", values=data["valores"], tags=(data["tag"],))
            # Actualizar header según modo (terrestre → MIC/DTA, flexi → Salida Aduana)
            if len(self.tree_control_final.get_children()) == 1:
                modo = data.get("modo", "flexi")
                self.tree_control_final.heading(
                    "salida_aduana",
                    text="MIC/DTA" if modo == "terrestre" else "Salida Aduana"
                )
            self._log(f"  + Fila insertada: {stem}")
            # Store comparison data
            if not hasattr(self, '_control_final_comparacion'):
                self._control_final_comparacion = {}
            self._control_final_comparacion[iid] = {
                "ticket": data["ticket"],
                "contenedor": data["contenedor"],
                "ok": data["ok"],
                "aduana": data["aduana"],
                "modo": data.get("modo", "flexi"),
                "shared_excels": data.get("shared_excels", []),
                "shared_neto_sum": data.get("shared_neto_sum", 0),
            }
        except Exception as e:
            self._log(f"Error insertando fila: {e}")

    def _abrir_comparacion_final(self, event):
        """Popup con tabla comparativa: Ticket vs Aduana vs Excel."""
        sel = self.tree_control_final.selection()
        if not sel or not hasattr(self, '_control_final_comparacion'):
            return

        datos = self._control_final_comparacion.get(sel[0])
        if not datos:
            return

        modo = datos.get("modo", "flexi")
        aduana_header = "MIC/DTA" if modo == "terrestre" else "Salida Aduana"

        def _build_popup(level):
            fsizes = self._get_font_sizes(level)
            geom = self._get_popup_geometry(700, 420, level)
            ancho = int(geom.split("x")[0])

            dlg = ctk.CTkToplevel(self)
            dlg.title("Comparación — Control Final")
            dlg.transient(self)
            dlg.grab_set()
            dlg.resizable(False, False)

            # ── Top bar: header + font level override ──────────────────
            top_bar = ctk.CTkFrame(dlg, fg_color=Palette.BG_SIDEBAR, corner_radius=6, height=36)
            top_bar.pack(fill="x", padx=12, pady=(12, 0))
            top_bar.pack_propagate(False)

            for col_i, txt in enumerate(["Campo", "Ticket (OCR)", aduana_header, "Contenedor (Excel)"]):
                lbl = ctk.CTkLabel(
                    top_bar, text=txt, width=100 if col_i == 0 else 160,
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["header"], weight="bold"),
                    text_color=Palette.TEXT_SECONDARY,
                )
                kwargs = {"side": "left", "padx": 4}
                if col_i > 0:
                    kwargs["fill"] = "x"
                    kwargs["expand"] = True
                lbl.pack(**kwargs)

            # Font level override selector
            override_menu = ctk.CTkOptionMenu(
                top_bar, values=["1", "2", "3"],
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
                width=50, height=26,
                command=lambda v: (dlg.destroy(), _build_popup(int(v))),
            )
            override_menu.set(str(level))
            override_menu.pack(side="right", padx=(0, 8))

            # Cuerpo: grid para columnas alineadas
            body = ctk.CTkFrame(dlg, fg_color="transparent")
            body.pack(fill="both", expand=True, padx=12, pady=8)
            body.grid_columnconfigure(0, minsize=100, weight=0)  # Campo — fijo
            body.grid_columnconfigure(1, weight=1)  # Ticket — expand
            body.grid_columnconfigure(2, weight=1)  # Aduana — expand
            body.grid_columnconfigure(3, weight=1)  # Contenedor — expand

            if modo == "terrestre":
                campos = [
                    ("Camion", "Patente", "Patente Camión"),
                    ("Semi", "Semirremolque", "Patente Semi"),
                    ("Conductor", "Conductor", "Conductor Aduana"),
                    ("DNI", "DNI", "DNI"),
                    ("Neto (kg)", "Neto (kg)", "Peso Bruto Aduana"),
                    ("Tara (kg)", "Tara (kg)", None),
                    ("Contenedor", "Contenedor", "Contenedor"),
                    ("Permiso", "Permiso", "Id Destinación"),
                    ("Precinto", "Precinto", "Precinto"),
                ]
            else:
                campos = [
                    ("Camion", "Patente", "Patente Camión"),
                    ("Semi", "Semirremolque", "Patente Semi"),
                    ("Conductor", "Conductor", "Conductor Aduana"),
                    ("DNI", "DNI", "CUIL"),
                    ("Neto (kg)", "Neto (kg)", "Peso Bruto Aduana"),
                    ("Tara (kg)", "Tara (kg)", None),
                    ("Contenedor", "Contenedor", "Contenedor"),
                    ("Permiso", "Permiso", "Id Destinación"),
                    ("Precinto", "Precinto", "Precinto"),
                ]

            # Campos con coloreado individual por celda (mayoría decide)
            campos_mayoria = {"Camion", "Semi", "DNI",
                              "Neto (kg)", "Contenedor", "Permiso"}

            shared_excels = datos.get("shared_excels", [])
            is_shared = len(shared_excels) > 1

            for row_i, (campo, key_ticket, key_aduana) in enumerate(campos):
                val_ticket = datos["ticket"].get(key_ticket, "")
                val_cont = datos["contenedor"].get(key_ticket, "")
                val_aduana = datos["aduana"].get(key_aduana, "—") if key_aduana else "—"
                # Format DNI with thousand separators: 36987654 -> 36.987.654
                if campo == "DNI":
                    val_ticket = self._fmt_dni(val_ticket)
                    val_cont = self._fmt_dni(val_cont)
                    val_aduana = self._fmt_dni(val_aduana)

                # Override Neto Excel column for shared trips
                if is_shared and campo == "Neto (kg)":
                    excel_neto_parts = " + ".join(
                        self._fmt_kg(ex["peso_carga"]) for ex in shared_excels
                    )
                    excel_neto_total = self._fmt_kg(datos["shared_neto_sum"])
                    val_cont = f"{excel_neto_parts} = {excel_neto_total}"

                # Override Neto MIC/DTA column for shared terrestre trips
                if is_shared and modo == "terrestre" and campo == "Neto (kg)":
                    c42 = datos["aduana"].get("campo_42", "")
                    c44 = datos["aduana"].get("campo_44", "")
                    c46 = datos["aduana"].get("campo_46", "")
                    if c42 and c44 and c46:
                        val_aduana = (f"{self._fmt_kg(c42)} + "
                                      f"{self._fmt_kg(c44)} = "
                                      f"{self._fmt_kg(c46)}")
                ok = datos["ok"].get(key_ticket, False)

                vals_list = [str(val_ticket), str(val_aduana), str(val_cont)]

                if campo in campos_mayoria:
                    colors = []
                    # For shared trips Neto: compare totals (after "="), not full strings
                    if is_shared and campo == "Neto (kg)":
                        totals = []
                        for v in vals_list:
                            if "=" in v:
                                totals.append(v.split("=")[-1].strip())
                            else:
                                totals.append(v.strip())
                        all_match = len(set(totals)) == 1
                        for _ in vals_list:
                            colors.append(("#C8FFC8", "#006400") if all_match else ("#FFC8C8", "#8B0000"))
                    else:
                        for v in vals_list:
                            count = sum(1 for x in vals_list if x == v)
                            if count >= 2:
                                colors.append(("#C8FFC8", "#006400"))
                            else:
                                colors.append(("#FFC8C8", "#8B0000"))
                else:
                    bg = "#C8FFC8" if ok else "#FFC8C8"
                    fg = "#006400" if ok else "#8B0000"
                    colors = [(bg, fg)] * 3

                # Campo label (column 0)
                ctk.CTkLabel(
                    body, text=campo, anchor="w",
                    font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"], weight="bold"),
                    text_color=Palette.TEXT_PRIMARY,
                ).grid(row=row_i, column=0, padx=(4, 8), pady=2, sticky="w")

                # Data columns (1, 2, 3)
                for col_i, (val, (bg, fg)) in enumerate(zip(vals_list, colors)):
                    lbl = ctk.CTkLabel(
                        body, text=val, anchor="center",
                        font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["data"]),
                        fg_color=bg, text_color=fg, corner_radius=4,
                    )
                    lbl.grid(row=row_i, column=col_i + 1, padx=4, pady=2, sticky="ew")

            # Legend
            leyenda = ctk.CTkFrame(dlg, fg_color="transparent")
            leyenda.pack(fill="x", pady=(0, 12))
            inner = ctk.CTkFrame(leyenda, fg_color="transparent")
            inner.pack(anchor="center")

            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#C8FFC8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Coincide",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            ctk.CTkLabel(
                inner, text="", width=14, height=14,
                fg_color="#FFC8C8", corner_radius=7,
            ).pack(side="left", padx=(4, 2))
            ctk.CTkLabel(
                inner, text="Diferencia",
                font=ctk.CTkFont(family=FONT_FAMILY, size=fsizes["legend"]),
                text_color=Palette.TEXT_MUTED,
            ).pack(side="left", padx=(0, 14))

            # Ajustar ventana al contenido real
            dlg.update_idletasks()
            alto = dlg.winfo_reqheight()
            x = (dlg.winfo_screenwidth() - ancho) // 2
            y = (dlg.winfo_screenheight() - alto) // 2
            dlg.geometry(f"{ancho}x{alto}+{x}+{y}")

        _build_popup(self.config.get("font_level", 1))

    def _cargar_datos_switch_toggle(self):
        """Toggle auto mode for Control Tickets. Changes btn text hint and persists state."""
        if self._ct_auto_var.get():
            self.btn_controlar_tickets.configure(text="Controlar Tickets ⚡", image=self._icons["zap"])
        else:
            self.btn_controlar_tickets.configure(text="Controlar Tickets", image=self._icons["ticket"])
        self.config["cargar_datos_auto"] = self._ct_auto_var.get()
        self._guardar_config()

    def _cargar_datos_auto_scan(self):
        """Auto-scan Desktop for folders with CONTENEDORES Excel (Control Tickets)."""
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self._scan_desktop_folders(
            pattern=None,
            callback=self._cargar_datos_auto_popup,
            button=self.btn_controlar_tickets,
        )

    def _cargar_datos_auto_popup(self, folders):
        """Popup to select folders for auto-scan Control Tickets (planilla de carga style)."""
        if not folders:
            messagebox.showinfo(
                "Sin resultados",
                "No se encontraron carpetas con CONTENEDORES en el Desktop"
            )
            return

        self._log(f"[Auto Tickets] {len(folders)} carpetas con CONTENEDORES encontradas")

        top = ctk.CTkToplevel(self)
        top.title("Control de Tickets Automático")
        top.geometry("600x400")
        top.configure(fg_color=Palette.BG_CARD)
        top.transient(self)
        top.lift()
        top.grab_set()
        top.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - 600) // 2
        py = self.winfo_y() + (self.winfo_height() - 400) // 2
        top.geometry(f"600x400+{px}+{py}")

        ctk.CTkLabel(
            top, text="CONTROL DE TICKETS AUTOMÁTICO",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(pady=(16, 8))

        ctk.CTkLabel(
            top, text="Seleccioná las carpetas del Desktop a procesar:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w", padx=16, pady=(0, 6))

        scroll = ctk.CTkScrollableFrame(
            top, fg_color=Palette.BG_TABLE, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        checks = {}
        for carp in folders:
            var = ctk.BooleanVar(value=True)
            checks[carp["name"]] = var

        btn_frame = ctk.CTkFrame(top, fg_color="transparent")
        btn_frame.pack(fill="x", padx=16, pady=(0, 12))

        def _update_btn(*_args):
            any_selected = any(v.get() for v in checks.values())
            btn_procesar.configure(state="normal" if any_selected else "disabled")

        def _todas():
            for v in checks.values():
                v.set(True)

        def _ninguna():
            for v in checks.values():
                v.set(False)

        def _confirmar():
            selected = [c for c in folders if checks[c["name"]].get()]
            top.destroy()

            if not selected:
                return

            # Build per-folder groups, merging COMPARTIDO folders together
            import re as _re
            _permiso_pat = _re.compile(r'^\d{2}069EC', _re.IGNORECASE)

            # Separate shared vs normal folders
            shared_groups = {}  # key -> [carp1, carp2, ...]
            normal_groups = []

            for carp in selected:
                name_upper = carp["name"].upper()
                if "COMPARTIDO" in name_upper:
                    # Extract a grouping key: the permiso code from folder name
                    # e.g. "19_06_2026_1_TERRESTRE_465R_560284_EWOS_F7_COMPARTIDO_CERRADO"
                    # Use the date prefix as grouping key (same trip, same date)
                    parts = carp["name"].split("_")
                    # Group by date+number prefix: "19_06_2026_1"
                    if len(parts) >= 3:
                        key = "_".join(parts[:3])  # "19_06_2026_1"
                    else:
                        key = carp["name"]
                    shared_groups.setdefault(key, []).append(carp)
                else:
                    normal_groups.append(carp)

            folder_groups = []
            total_pdfs = 0

            # Process shared folders: merge scans + combine Excels
            for key, shared_carp in shared_groups.items():
                all_pdfs = []
                all_excels = []
                for carp in shared_carp:
                    for f in os.scandir(carp["path"]):
                        if f.is_file() and f.name.lower().endswith('.pdf'):
                            if _permiso_pat.match(f.name):
                                continue
                            if not (f.name.lower().startswith("scan") or f.name.lower().startswith("escan")):
                                continue
                            all_pdfs.append(f.path)
                    if carp["excel_path"]:
                        all_excels.append(carp["excel_path"])

                if all_pdfs and all_excels:
                    # Pass ALL Excels for shared trips — worker sums Neto
                    folder_groups.append((all_pdfs, all_excels))
                    total_pdfs += len(all_pdfs)
                    names = [c["name"] for c in shared_carp]
                    self._log(
                        f"[Auto Tickets] Compartido detectado: {len(names)} carpetas, "
                        f"{len(all_excels)} Excels — Neto se sumará"
                    )

            # Process normal folders: each folder's scans paired with ITS OWN Excel
            for carp in normal_groups:
                pdfs = []
                for f in os.scandir(carp["path"]):
                    if f.is_file() and f.name.lower().endswith('.pdf'):
                        if _permiso_pat.match(f.name):
                            continue
                        if not (f.name.lower().startswith("scan") or f.name.lower().startswith("escan")):
                            continue
                        pdfs.append(f.path)
                if pdfs and carp["excel_path"]:
                    folder_groups.append((pdfs, [carp["excel_path"]]))
                    total_pdfs += len(pdfs)

            if not folder_groups:
                messagebox.showwarning(
                    "Sin datos",
                    "No se encontraron PDFs scan_ con Excel CONTENEDORES"
                )
                return

            self._log(
                f"[Auto Tickets] {len(folder_groups)} grupos, {total_pdfs} PDFs — "
                f"(compartidos: carpetas COMPARTIDO se procesan juntas)"
            )

            if self.tarea_activa:
                messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
                return
            self.tarea_activa = True

            for btn_name in ('btn_controlar_tickets',):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="disabled")

            if (hasattr(self, 'progress_carga')
                    and self.progress_carga.winfo_exists()):
                self.progress_carga.pack(side="right", padx=12)
                self.progress_carga.configure(mode="indeterminate")
                self.progress_carga.start()

            # Show tickets section
            self._switch_tab("tickets")

            # Store groups for serial processing
            self._ct_folder_groups = folder_groups
            self._ct_group_idx = 0

            # Start first group
            self._procesar_siguiente_grupo_ct()

        # Todas / Ninguna (izquierda)
        ctk.CTkButton(
            btn_frame, text="Todas", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_todas,
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            btn_frame, text="Ninguna", width=80, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ERROR,
            text_color=Palette.TEXT_SECONDARY, corner_radius=6,
            command=_ninguna,
        ).pack(side="left", padx=4)

        # Procesar (derecha)
        btn_procesar = ctk.CTkButton(
            btn_frame, text="Procesar", width=140, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.WHITE, corner_radius=6,
            command=_confirmar,
        )
        btn_procesar.pack(side="right", padx=4)

        # Checkboxes
        for carp in folders:
            var = checks[carp["name"]]
            var.trace_add("write", _update_btn)
            ctk.CTkCheckBox(
                scroll, text=carp["name"], variable=var,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                text_color=Palette.TEXT_PRIMARY,
            ).pack(anchor="w", padx=12, pady=4)

    def _procesar_siguiente_grupo_ct(self):
        """Process the next folder group (scan PDFs + Excel(s)) serially.
        Shared groups pass multiple Excels — worker sums Neto."""
        if self._ct_group_idx >= len(self._ct_folder_groups):
            # All groups done
            self._cargar_datos_done()
            return

        pdfs, excels = self._ct_folder_groups[self._ct_group_idx]
        self._ct_group_idx += 1
        remaining = len(self._ct_folder_groups) - self._ct_group_idx
        excel_note = f"{len(excels)} Excels" if len(excels) > 1 else "1 Excel"
        self._log(
            f"[Auto Tickets] Grupo {self._ct_group_idx}/{len(self._ct_folder_groups)}: "
            f"{len(pdfs)} PDFs, {excel_note} — {remaining} grupos restantes"
        )

        t = threading.Thread(
            target=self._cargar_datos_worker,
            args=("local", list(pdfs), list(excels)),
            daemon=True,
        )
        t.start()

    def _cargar_datos_seleccionar_pdfs(self):
        if self._ct_auto_var.get():
            self._cargar_datos_auto_scan()
            return
        from tkinter import filedialog as tk_filedialog
        rutas = tk_filedialog.askopenfilenames(
            title="Seleccionar PDFs y Excel CONTENEDORES",
            filetypes=[("Archivos soportados", "*.pdf *.xlsx *.xls")],
        )
        if not rutas:
            return
        rutas_pdf = [r for r in rutas if r.lower().endswith('.pdf')]
        rutas_excel = [r for r in rutas if r.lower().endswith(('.xlsx', '.xls'))]
        if not rutas_pdf:
            self._log("[Control Datos] ⚠ No seleccionaste ningún PDF.")
            return
        if not rutas_excel:
            self._log("[Control Datos] ⚠ No seleccionaste ningún archivo Excel CONTENEDORES.")
            return
        # Mostrar sección tickets, ocultar otras
        self._switch_tab("tickets")
        self._log(f"[Control Datos] Procesando {len(rutas_pdf)} PDFs contra {len(rutas_excel)} Excel(s)...")
        if self.tarea_activa:
            messagebox.showwarning("Agente ocupado", "Hay una tarea en ejecución.")
            return
        self.tarea_activa = True
        for btn_name in ('btn_controlar_tickets',):
            btn = getattr(self, btn_name, None)
            if btn and btn.winfo_exists():
                btn.configure(state="disabled")
        if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
            self.progress_carga.pack(side="right", padx=12)
            self.progress_carga.configure(mode="indeterminate")
            self.progress_carga.start()
        t = threading.Thread(
            target=self._cargar_datos_worker,
            args=("local", list(rutas_pdf), list(rutas_excel)),
            daemon=True,
        )
        t.start()

    def _cargar_datos_escribir(self):
        """T8: Escribe valores OCR validados en las celdas DATOS del CONTENEDOR."""
        ok_count = 0
        fail_count = 0
        items = list(self.tree_carga.get_children())

        if not items:
            self._log("[Control Datos] No hay filas para escribir.")
            return

        for item in items:
            vals = self.tree_carga.item(item, "values")
            if not vals or len(vals) < 9:
                continue
            estado = vals[-1]
            if estado != "✅ Ok":
                continue

            iid = item
            if iid not in self._cargar_datos_rutas:
                self._log(f"[Control Datos] ⚠ Sin ruta almacenada para {vals[0]}")
                fail_count += 1
                continue

            ruta, k_idx = self._cargar_datos_rutas[iid]
            neto_val = vals[5]   # columna neto
            tara_val = vals[6]   # columna tara
            nombre_archivo = vals[0]

            try:
                wb = openpyxl.load_workbook(ruta)
                if "DATOS" not in wb.sheetnames:
                    self._log(f"[Control Datos] ⚠ {nombre_archivo}: no tiene hoja DATOS")
                    wb.close()
                    fail_count += 1
                    continue
                ws = wb["DATOS"]

                # Calcular coordenadas según bloque k
                row_base = 20 + (k_idx // 2) * 13
                label_col = 1 + (k_idx % 2) * 12
                value_col = label_col + 6

                # PESO CARGA = row_base + 8, value_col
                ws.cell(row=row_base + 8, column=value_col, value=float(neto_val))
                # TARA CONT = row_base + 9, value_col
                ws.cell(row=row_base + 9, column=value_col, value=float(tara_val))

                wb.save(ruta)
                wb.close()
                ok_count += 1
                self._log(f"[Control Datos] ✓ {nombre_archivo}: Neto={neto_val}, Tara={tara_val}")

            except PermissionError:
                self._log(
                    f"[Control Datos] ✗ El archivo {os.path.basename(ruta)} "
                    "está bloqueado. Cerralo e intentá de nuevo."
                )
                fail_count += 1
            except Exception as e:
                self._log(f"[Control Datos] ✗ Error escribiendo {nombre_archivo}: {e}")
                fail_count += 1

        resumen = f"Escritura completada: {ok_count} ok, {fail_count} errores"
        self._log(f"[Control Datos] {resumen}")

    # ── T5: Matching CONTENEDOR ──────────────────────────────────────
    def _cargar_cache_contenedores(self, rutas_excel=None):
        """Escanea Desktop UNA VEZ y llena self._contenedores_cache abriendo
        cada workbook UNA SOLA VEZ (PE + DATOS + Choferes).
        Si se pasan rutas_excel explícitas, las usa en lugar de escanear."""
        if rutas_excel:
            self._contenedores_cache = {}
            for ruta in rutas_excel:
                data = self._leer_wb_completo(ruta)
                if data:
                    self._contenedores_cache[ruta] = data
            return
        self._contenedores_cache = {}
        base = self._resolver_ruta("planillas_carga", "Desktop")
        if not os.path.isdir(base):
            self._log("[CACHE] Desktop no encontrado")
            return

        archivos = []
        try:
            # Archivos sueltos en Desktop
            for f in os.listdir(base):
                ruta = os.path.join(base, f)
                if os.path.isfile(ruta) and 'CONTENEDORES' in f.upper() and \
                   (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')):
                    archivos.append(ruta)
            # 1 nivel de subcarpetas
            for item in os.listdir(base):
                ruta_item = os.path.join(base, item)
                if os.path.isdir(ruta_item):
                    for f in os.listdir(ruta_item):
                        if 'CONTENEDORES' in f.upper() and \
                           (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')):
                            archivos.append(os.path.join(ruta_item, f))
        except Exception:
            pass

        self._log(f"[CACHE] Escaneando {len(archivos)} archivos CONTENEDORES...")
        for ruta in archivos:
            data = self._leer_wb_completo(ruta)
            if data:
                self._contenedores_cache[ruta] = data
        n = len(self._contenedores_cache)
        self._log(f"[CACHE] Caché completado: {n} archivos")

    def _match_contenedor(self, permiso_ticket):
        """Busca archivos CONTENEDORES asociados al permiso del ticket.

        Usa self._contenedores_cache (cargado por _cargar_cache_contenedores).
        NO abre workbooks — lee PE desde caché.

        Soporta viajes compartidos: si el permiso tiene '/' (ej: '26069EC01000372Y/1000579A'),
        prueba cada parte por separado y combina resultados.

        Retorna lista de rutas de archivos que coinciden (ordenadas por
        estrategia 1 primero, luego estrategia 2), o lista vacía."""
        import re as _re

        if not permiso_ticket:
            self.log_queue.put((self._log_warning, "[MATCH] permiso_ticket vacío, abortando"))
            return []

        # ── Viajes compartidos: probar cada permiso por separado ──────
        partes = [p.strip() for p in str(permiso_ticket).split('/') if p.strip()]
        if len(partes) > 1:
            self.log_queue.put((self._log_warning,
                f"[MATCH] Viaje compartido: {len(partes)} permisos -> {[p for p in partes]}"))
            todas = []
            for parte in partes:
                sub = self._match_contenedor(parte)
                todas.extend(sub)
            # Deduplicar manteniendo orden
            seen = set()
            unicas = []
            for r in todas:
                if r not in seen:
                    seen.add(r)
                    unicas.append(r)
            if unicas:
                self.log_queue.put((self._log_warning,
                    f"[MATCH] ✅ Total candidatos (compartido): {len(unicas)}"))
            return unicas

        # ── Preparar patrones desde el permiso del ticket ────────────
        ticket_alfanum = _re.sub(r'[^a-zA-Z0-9]', '', str(permiso_ticket))
        self.log_queue.put((self._log_warning, f"[MATCH] Permiso limpio: '{ticket_alfanum}'"))

        if not ticket_alfanum:
            return []

        sufijo_ticket = ticket_alfanum[-5:].upper()
        self.log_queue.put((self._log_warning, f"[MATCH] Sufijo (últimos 5): '{sufijo_ticket}'"))

        codigo_corto = ""
        m = _re.search(r'0+([1-9A-Z][A-Z0-9]+)$', ticket_alfanum)
        if m:
            codigo_corto = m.group(1).upper()
            self.log_queue.put((self._log_warning, f"[MATCH] Código corto (tras ceros): '{codigo_corto}'"))

        # ── Buscar en caché en vez de abrir workbooks ──────────────
        if not self._contenedores_cache:
            self.log_queue.put((self._log_warning,
                f"[MATCH] Caché vacío, sin archivos CONTENEDORES"))
            return []

        self.log_queue.put((self._log_warning,
            f"[MATCH] Evaluando {len(self._contenedores_cache)} archivos desde caché..."))

        # ── Evaluar TODOS los archivos desde caché ─────────────────
        estrategia1 = []  # match exacto por sufijo PE
        estrategia2 = []  # código corto en nombre de carpeta

        for ruta, data in self._contenedores_cache.items():
            carpeta = os.path.basename(os.path.dirname(ruta))
            nombre = os.path.basename(ruta)
            pe_val = data.get("pe")
            try:
                if pe_val is not None:
                    pe_alfanum = _re.sub(r'[^a-zA-Z0-9]', '', str(pe_val))
                    pe_sufijo = pe_alfanum[-5:].upper() if len(pe_alfanum) >= 5 else pe_alfanum.upper()
                    if pe_sufijo == sufijo_ticket:
                        estrategia1.append(ruta)
                        self.log_queue.put((self._log_warning,
                            f"[MATCH]   ✅ E1: '{pe_sufijo}' == '{sufijo_ticket}' -> {ruta}"))
                        continue  # no evaluar E2 si ya es E1
                # Estrategia 2: código corto en carpeta
                if codigo_corto and codigo_corto in carpeta.upper():
                    estrategia2.append(ruta)
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   📁 E2: '{codigo_corto}' en carpeta -> {ruta}"))
            except Exception:
                pass

        # Combinar: E1 primero (match exacto), luego E2
        todas = estrategia1 + [r for r in estrategia2 if r not in estrategia1]

        if todas:
            self.log_queue.put((self._log_warning,
                f"[MATCH] ✅ Total candidatos: {len(todas)}"))
            return todas

        self.log_queue.put((self._log_warning,
            f"[MATCH] ❌ Sin candidatos para permiso '{permiso_ticket}' "
            f"(sufijo='{sufijo_ticket}')"))
        return []

    # ── Helper: leer PE de hoja Choferes ──────────────────────────
    def _leer_pe_choferes(self, ruta):
        """Busca etiqueta 'PE' en hoja Choferes y retorna el valor
        de la celda de la derecha. Retorna None si no encuentra."""
        # Cache check
        if ruta in getattr(self, '_contenedores_cache', {}):
            return self._contenedores_cache[ruta].get("pe")
        try:
            if ruta.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
                if 'Choferes' not in wb.sheetnames:
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   {os.path.basename(ruta)}: NO tiene hoja Choferes"))
                    wb.close()
                    return None
                ws = wb['Choferes']
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(r, c).value
                        if v is not None and str(v).strip().upper() == 'PE':
                            pe = ws.cell(r, c + 1).value
                            wb.close()
                            if pe:
                                return str(pe).strip()
                            return None
                wb.close()
                self.log_queue.put((self._log_warning,
                    f"[MATCH]   {os.path.basename(ruta)}: No se encontró etiqueta 'PE' en Choferes"))
                return None
            else:
                book = xlrd.open_workbook(ruta)
                if 'Choferes' not in book.sheet_names():
                    self.log_queue.put((self._log_warning,
                        f"[MATCH]   {os.path.basename(ruta)}: NO tiene hoja Choferes"))
                    return None
                ws = book.sheet_by_name('Choferes')
                for r in range(ws.nrows):
                    for c in range(ws.ncols):
                        v = ws.cell_value(r, c)
                        if v and str(v).strip().upper() == 'PE':
                            pe = ws.cell_value(r, c + 1)
                            if pe:
                                return str(pe).strip()
                            return None
                self.log_queue.put((self._log_warning,
                    f"[MATCH]   {os.path.basename(ruta)}: No se encontró etiqueta 'PE' en Choferes"))
                return None
        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[MATCH]   Error al leer {os.path.basename(ruta)}: {ex}"))
            return None

    def _leer_contenedores_desde_choferes(self, ruta_wb):
        """Lee la columna 'NUMERO DE CONTENEDORES' (col D) de hoja Choferes.

        Retorna lista de strings (container numbers vacíos o no),
        en el mismo orden que aparecen en la hoja (skip header row).
        """
        # Cache check
        if ruta_wb in getattr(self, '_contenedores_cache', {}):
            return self._contenedores_cache[ruta_wb].get("contenedores", [])
        import re as _re
        contenedores = []
        try:
            if ruta_wb.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)
                if 'Choferes' not in wb.sheetnames:
                    wb.close()
                    return contenedores
                ws = wb['Choferes']
                # Buscar header "NUMERO DE CONTENEDORES" en fila 1, col D
                col_idx = None
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(1, c).value
                    if v and 'NUMERO' in str(v).upper() and 'CONTENEDOR' in str(v).upper():
                        col_idx = c
                        break
                if col_idx is None:
                    col_idx = 4  # fallback a col D
                # Leer desde fila 2 hasta el final
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(r, col_idx).value
                    contenedores.append(str(val).strip() if val else "")
                wb.close()
            else:
                book = xlrd.open_workbook(ruta_wb)
                if 'Choferes' not in book.sheet_names():
                    return contenedores
                ws = book.sheet_by_name('Choferes')
                col_idx = None
                for c in range(ws.ncols):
                    v = ws.cell_value(0, c)
                    if v and 'NUMERO' in str(v).upper() and 'CONTENEDOR' in str(v).upper():
                        col_idx = c
                        break
                if col_idx is None:
                    col_idx = 3  # fallback a col D (0-indexed)
                for r in range(1, ws.nrows):
                    val = ws.cell_value(r, col_idx)
                    contenedores.append(str(val).strip() if val else "")

            n_reales = sum(1 for c in contenedores if c)
            self.log_queue.put((self._log_warning,
                f"[Choferes] {n_reales} contenedores encontrados en col {col_idx+1}"))
            return contenedores
        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[Choferes] Error leyendo contenedores: {ex}"))
            return contenedores

    # ── T5b: Lector combinado (una sola apertura) ──────────────────────
    def _leer_wb_completo(self, ruta_wb):
        """Abre el workbook UNA VEZ y extrae datos desde Choferes + peso/tara desde DATOS.

        Choferes es la fuente primaria: patente, semi, chofer, DNI, contenedor, precintos, PE.
        DATOS solo provee peso_carga y tara_cont.
        Patentes se limpian al leer (sin guiones/espacios).

        Retorna dict {pe, camiones[], contenedores[], precintos[], peso_flexi_global} o None.
        """
        import re as _re

        def _clean_pat(p):
            """Elimina guiones, espacios y puntuación de patentes."""
            if not p:
                return ""
            return _re.sub(r'[^A-Z0-9]', '', str(p).upper().strip())

        pe = None
        contenedores = []
        precintos = []
        camiones = []
        peso_flexi_global = None

        try:
            if ruta_wb.lower().endswith('.xlsx'):
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)

                # ── Hoja Choferes: fuente primaria ──
                if 'Choferes' in wb.sheetnames:
                    ws = wb['Choferes']

                    # Detectar columnas por headers
                    h_cols = {}
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(1, c).value
                        if v:
                            h = str(v).upper()
                            if 'DOMINIO' in h and 'TRACTOR' in h:
                                h_cols['tractor'] = c
                            if 'DOMINIO' in h and 'SEMI' in h:
                                h_cols['semi'] = c
                            if 'NUMERO' in h and 'CONTENEDOR' in h:
                                h_cols['contenedor'] = c
                            if 'NOMBRE' in h and 'CHOFER' in h:
                                h_cols['chofer'] = c
                            if 'DNI' in h and 'CHOFER' in h:
                                h_cols['dni'] = c
                            if 'PRECINTO' in h and 'ADUANA' in h:
                                h_cols['precinto'] = c
                            if 'PRECINTO' in h and 'LINEA' in h:
                                h_cols['precinto_linea'] = c

                    # Buscar PE
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(r, c).value
                            if v is not None and str(v).strip().upper() == 'PE':
                                pe_v = ws.cell(r, c + 1).value
                                if pe_v:
                                    pe = str(pe_v).strip()
                                break
                        if pe is not None:
                            break

                    # Buscar PESO FLEXI
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(r, c).value
                            if v is not None and 'PESO FLEXI' in str(v).upper():
                                pfv = ws.cell(r, c + 1).value
                                if pfv is not None:
                                    try:
                                        peso_flexi_global = float(str(pfv).strip())
                                    except ValueError:
                                        pass
                                break
                        if peso_flexi_global is not None:
                            break

                    # Leer camiones desde filas de Choferes (empieza en fila 2)
                    for r in range(2, ws.max_row + 1):
                        tractor_val = ws.cell(r, h_cols.get('tractor', 1)).value if 'tractor' in h_cols else None
                        if not tractor_val:
                            break  # Fin de datos de camiones
                        camion = {"k": len(camiones), "peso_carga": 0, "tara_cont": 0}
                        if 'tractor' in h_cols:
                            v = ws.cell(r, h_cols['tractor']).value
                            camion["patente_camion"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_camion"] = ""
                        if 'semi' in h_cols:
                            v = ws.cell(r, h_cols['semi']).value
                            camion["patente_semi"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_semi"] = ""
                        if 'contenedor' in h_cols:
                            v = ws.cell(r, h_cols['contenedor']).value
                            camion["contenedor"] = str(v).strip() if v else ""
                        else:
                            camion["contenedor"] = ""
                        if 'chofer' in h_cols:
                            v = ws.cell(r, h_cols['chofer']).value
                            camion["conductor"] = str(v).strip() if v else ""
                        else:
                            camion["conductor"] = ""
                        if 'dni' in h_cols:
                            v = ws.cell(r, h_cols['dni']).value
                            if v is not None:
                                camion["dni"] = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
                            else:
                                camion["dni"] = ""
                        else:
                            camion["dni"] = ""
                        # Precinto: unir aduana + linea con guion
                        prec_parts = []
                        if 'precinto' in h_cols:
                            pv = ws.cell(r, h_cols['precinto']).value
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        if 'precinto_linea' in h_cols:
                            pv = ws.cell(r, h_cols['precinto_linea']).value
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        camion["precinto"] = "-".join(prec_parts) if prec_parts else ""
                        contenedores.append(camion["contenedor"])
                        precintos.append(camion["precinto"])
                        camiones.append(camion)

                # ── Hoja DATOS: SOLO peso_carga y tara_cont ──
                if 'DATOS' in wb.sheetnames and camiones:
                    ws_d = wb['DATOS']
                    max_row = ws_d.max_row or 0
                    max_col = ws_d.max_column or 0

                    def _celda(row, col):
                        if row > max_row or col > max_col:
                            return None
                        return ws_d.cell(row=row, column=col).value

                    def _val_num(raw_val):
                        if raw_val is None:
                            return 0
                        if isinstance(raw_val, (int, float)):
                            return float(raw_val)
                        return 0

                    # Buscar posiciones de CAMION / PATENTE CAMION en DATOS
                    camion_positions = []
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            raw = _celda(r, c)
                            if raw and isinstance(raw, str):
                                lbl = raw.strip().upper()
                                if lbl in ('CAMION', 'PATENTE CAMION'):
                                    camion_positions.append((r, c))

                    # Para cada CAMION, buscar PESO CARGA y TARA CONT en filas cercanas
                    peso_tara_list = []
                    for cam_row, cam_col in camion_positions:
                        peso = 0
                        tara = 0
                        for offset in range(1, 12):
                            check_row = cam_row + offset
                            if check_row > max_row:
                                break
                            raw = _celda(check_row, cam_col)
                            if raw and isinstance(raw, str):
                                lbl = raw.strip().upper()
                                if lbl in ('PESO CARGA', 'PESO CARGA (KG)'):
                                    peso = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 <= max_col else 0
                                elif lbl == 'TARA CONT':
                                    tara = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 <= max_col else 0
                        peso_tara_list.append((peso, tara))

                    # Asignar peso/tara a camiones por orden
                    for i, camion in enumerate(camiones):
                        if i < len(peso_tara_list):
                            camion["peso_carga"] = peso_tara_list[i][0]
                            camion["tara_cont"] = peso_tara_list[i][1]

                wb.close()

            else:
                # .xls con xlrd
                book = xlrd.open_workbook(ruta_wb)

                # ── Hoja Choferes: fuente primaria ──
                if 'Choferes' in book.sheet_names():
                    ws = book.sheet_by_name('Choferes')

                    # Detectar columnas por headers (row 0 en xlrd)
                    h_cols = {}
                    for c in range(ws.ncols):
                        v = ws.cell_value(0, c)
                        if v:
                            h = str(v).upper()
                            if 'DOMINIO' in h and 'TRACTOR' in h:
                                h_cols['tractor'] = c
                            if 'DOMINIO' in h and 'SEMI' in h:
                                h_cols['semi'] = c
                            if 'NUMERO' in h and 'CONTENEDOR' in h:
                                h_cols['contenedor'] = c
                            if 'NOMBRE' in h and 'CHOFER' in h:
                                h_cols['chofer'] = c
                            if 'DNI' in h and 'CHOFER' in h:
                                h_cols['dni'] = c
                            if 'PRECINTO' in h and 'ADUANA' in h:
                                h_cols['precinto'] = c
                            if 'PRECINTO' in h and 'LINEA' in h:
                                h_cols['precinto_linea'] = c

                    # Buscar PE
                    for r in range(ws.nrows):
                        for c in range(ws.ncols):
                            v = ws.cell_value(r, c)
                            if v and str(v).strip().upper() == 'PE':
                                pe_v = ws.cell_value(r, c + 1)
                                if pe_v:
                                    pe = str(pe_v).strip()
                                break
                        if pe is not None:
                            break

                    # Buscar PESO FLEXI
                    for r in range(ws.nrows):
                        for c in range(ws.ncols):
                            v = ws.cell_value(r, c)
                            if v and 'PESO FLEXI' in str(v).upper():
                                pfv = ws.cell_value(r, c + 1)
                                if pfv:
                                    try:
                                        peso_flexi_global = float(str(pfv).strip())
                                    except ValueError:
                                        pass
                                break
                        if peso_flexi_global is not None:
                            break

                    # Leer camiones desde filas de Choferes (empieza en row 1 en xlrd)
                    for r in range(1, ws.nrows):
                        tractor_val = ws.cell_value(r, h_cols.get('tractor', 0)) if 'tractor' in h_cols else None
                        if not tractor_val:
                            break  # Fin de datos de camiones
                        camion = {"k": len(camiones), "peso_carga": 0, "tara_cont": 0}
                        if 'tractor' in h_cols:
                            v = ws.cell_value(r, h_cols['tractor'])
                            camion["patente_camion"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_camion"] = ""
                        if 'semi' in h_cols:
                            v = ws.cell_value(r, h_cols['semi'])
                            camion["patente_semi"] = _clean_pat(v) if v else ""
                        else:
                            camion["patente_semi"] = ""
                        if 'contenedor' in h_cols:
                            v = ws.cell_value(r, h_cols['contenedor'])
                            camion["contenedor"] = str(v).strip() if v else ""
                        else:
                            camion["contenedor"] = ""
                        if 'chofer' in h_cols:
                            v = ws.cell_value(r, h_cols['chofer'])
                            camion["conductor"] = str(v).strip() if v else ""
                        else:
                            camion["conductor"] = ""
                        if 'dni' in h_cols:
                            v = ws.cell_value(r, h_cols['dni'])
                            if v is not None:
                                camion["dni"] = str(int(v)) if isinstance(v, float) and v == int(v) else str(v).strip()
                            else:
                                camion["dni"] = ""
                        else:
                            camion["dni"] = ""
                        # Precinto
                        prec_parts = []
                        if 'precinto' in h_cols:
                            pv = ws.cell_value(r, h_cols['precinto'])
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        if 'precinto_linea' in h_cols:
                            pv = ws.cell_value(r, h_cols['precinto_linea'])
                            if pv and str(pv).strip() and str(pv).strip() != '-':
                                prec_parts.extend(str(pv).strip().replace("-", " ").split())
                        camion["precinto"] = "-".join(prec_parts) if prec_parts else ""
                        contenedores.append(camion["contenedor"])
                        precintos.append(camion["precinto"])
                        camiones.append(camion)

                # ── Hoja DATOS: SOLO peso_carga y tara_cont ──
                if 'DATOS' in book.sheet_names() and camiones:
                    ws_d = book.sheet_by_name('DATOS')
                    max_row = ws_d.nrows
                    max_col = ws_d.ncols

                    def _celda(row, col):
                        if row > max_row or col > max_col:
                            return None
                        return ws_d.cell_value(row - 1, col - 1)

                    def _val_num(raw_val):
                        if raw_val is None:
                            return 0
                        if isinstance(raw_val, (int, float)):
                            return float(raw_val)
                        return 0

                    camion_positions = []
                    for r in range(1, max_row + 1):
                        for c in range(0, max_col):
                            raw = _celda(r, c)
                            if raw is not None:
                                lbl = str(raw).strip().upper()
                                if lbl in ('CAMION', 'PATENTE CAMION'):
                                    camion_positions.append((r, c))

                    peso_tara_list = []
                    for cam_row, cam_col in camion_positions:
                        peso = 0
                        tara = 0
                        for offset in range(1, 12):
                            check_row = cam_row + offset
                            if check_row > max_row:
                                break
                            raw = _celda(check_row, cam_col)
                            if raw is not None:
                                lbl = str(raw).strip().upper()
                                if lbl in ('PESO CARGA', 'PESO CARGA (KG)'):
                                    peso = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 < max_col else 0
                                elif lbl == 'TARA CONT':
                                    tara = _val_num(_celda(check_row, cam_col + 6)) if cam_col + 6 < max_col else 0
                        peso_tara_list.append((peso, tara))

                    for i, camion in enumerate(camiones):
                        if i < len(peso_tara_list):
                            camion["peso_carga"] = peso_tara_list[i][0]
                            camion["tara_cont"] = peso_tara_list[i][1]

        except Exception as ex:
            self.log_queue.put((self._log_warning,
                f"[CACHE] Error al leer {os.path.basename(ruta_wb)}: {ex}"))
            return None

        n_ctn = sum(1 for c in contenedores if c)
        self.log_queue.put((self._log_warning,
            f"[CACHE] Cacheado: {os.path.basename(ruta_wb)} — "
            f"PE={pe} camiones={len(camiones)} contenedores={n_ctn}"))
        return {"pe": pe, "camiones": camiones, "contenedores": contenedores,
                "peso_flexi_global": peso_flexi_global,
                "precintos": precintos}

    # ── T6: Lectura de DATOS y comparación ────────────────────────────
    def _leer_datos_contenedor(self, ruta_wb):
        """Lee la hoja DATOS del CONTENEDOR y devuelve dict con camiones.

        Retorna:
            {"camiones": [{patente_camion, patente_semi, conductor,
                           dni, peso_carga, tara_cont}, ...]}
            o None si no hay hoja DATOS.
        """
        # Cache check: si ya fue cargado por _leer_wb_completo, evitar otro open
        if ruta_wb in getattr(self, '_contenedores_cache', {}):
            camiones = self._contenedores_cache[ruta_wb].get("camiones", [])
            return {"camiones": camiones} if camiones else None

        if ruta_wb.lower().endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(ruta_wb, read_only=True, data_only=True)
            except Exception as e:
                self.log_queue.put((self._log_warning, f"No se pudo abrir {os.path.basename(ruta_wb)}: {e}"))
                return None
            if 'DATOS' not in wb.sheetnames:
                wb.close()
                self.log_queue.put((self._log_warning, f"El archivo {os.path.basename(ruta_wb)} no tiene hoja DATOS"))
                return None
            ws = wb['DATOS']
            usar_xlrd = False
        else:
            # .xls con xlrd
            try:
                book = xlrd.open_workbook(ruta_wb)
            except Exception as e:
                self.log_queue.put((self._log_warning, f"No se pudo abrir {os.path.basename(ruta_wb)}: {e}"))
                return None
            if 'DATOS' not in book.sheet_names():
                self.log_queue.put((self._log_warning, f"El archivo {os.path.basename(ruta_wb)} no tiene hoja DATOS"))
                return None
            ws = book.sheet_by_name('DATOS')
            usar_xlrd = True

        # Obtener límites del sheet para evitar IndexError
        if usar_xlrd:
            max_row = ws.nrows
            max_col = ws.ncols
        else:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

        def _celda(row, col):
            """Lee valor de celda (1-indexed row/col)."""
            if row > max_row or col > max_col:
                return None
            if usar_xlrd:
                return ws.cell_value(row - 1, col - 1)
            else:
                return ws.cell(row=row, column=col).value

        def _val_num(raw_val):
            """Retorna float si raw_val es numérico, 0 en otro caso."""
            if raw_val is None:
                return 0
            if isinstance(raw_val, (int, float)):
                return float(raw_val)
            return 0

        camiones = []

        # ── Scan dinámico: busca todas las etiquetas PATENTE CAMION en DATOS ──
        # Soporta layout horizontal (C1..C3 en cols 1,13,25 misma fila)
        # y el layout antiguo de 2 por bloque de filas.
        bloques = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                raw = _celda(r, c)
                if raw is not None:
                    etiqueta = str(raw).strip().upper()
                    if 'PATENTE' in etiqueta and 'CAMION' in etiqueta:
                        if c + 6 <= max_col:
                            bloques.append((r, c))

        # Ordenar por fila, luego columna (izquierda→derecha, arriba→abajo)
        bloques.sort(key=lambda b: (b[0], b[1]))

        for k, (patente_row, label_col) in enumerate(bloques):
            row_base = patente_row - 1  # PATENTE CAMION está en row_base + 1
            value_col = label_col + 6

            patente_camion = str(_celda(patente_row, value_col) or "").strip()
            patente_semi   = str(_celda(patente_row + 1, value_col) or "").strip()
            conductor      = str(_celda(patente_row + 2, value_col) or "").strip()
            dni_raw        = _celda(patente_row + 3, value_col)
            if isinstance(dni_raw, (int, float)):
                dni_celda = str(int(dni_raw))
            else:
                dni_celda = str(dni_raw or "").strip()
            peso_carga_num = _val_num(_celda(patente_row + 7, value_col))
            tara_cont_num  = _val_num(_celda(patente_row + 8, value_col))

            camion = {
                "k": k,
                "patente_camion": patente_camion,
                "patente_semi": patente_semi,
                "conductor": conductor,
                "dni": dni_celda,
                "contenedor": "",  # se mergea desde Choferes después
                "peso_carga": peso_carga_num,
                "tara_cont": tara_cont_num,
            }
            camiones.append(camion)

        if not usar_xlrd:
            wb.close()

        # ── Merge container numbers from Choferes sheet (col D: NUMERO DE CONTENEDORES) ──
        try:
            contenedores = self._leer_contenedores_desde_choferes(ruta_wb)
            if contenedores and camiones:
                for i, camion in enumerate(camiones):
                    if i < len(contenedores) and contenedores[i]:
                        camion["contenedor"] = contenedores[i]
                        self.log_queue.put((self._log_warning,
                            f"[DATOS]   Contenedor[{i}]: '{contenedores[i]}' -> Patente: {camion['patente_camion']}"))
        except Exception as e:
            self.log_queue.put((self._log_warning, f"[DATOS] Error merge contenedores: {e}"))

        resultado = {"camiones": camiones} if camiones else None
        self.log_queue.put((self._log_warning,
            f"[DATOS] _leer_datos_contenedor: {len(camiones)} camiones encontrados "
            f"en {os.path.basename(ruta_wb)}"
            + (f" -> {[c['patente_camion'] for c in camiones]}" if camiones else "")))
        return resultado

    # ── T4: Worker OCR ────────────────────────────────────────────────
    def _cargar_datos_worker(self, fuente, rutas, rutas_excel=None):
        """Worker que corre en threading.Thread daemon.
        
        Procesa cada PDF: OCR → extracción → matching CONTENEDOR → lectura DATOS.
        Pushea resultados a log_queue como ("_OCR_RESULT_", data).
        Al finalizar pushea ("_OCR_DONE_", None).
        
        Args:
            rutas_excel: lista de rutas a archivos CONTENEDORES seleccionados
                         por el usuario. Si es None, escanea Desktop (viejo).
        """
        self._set_log_panel("cargar-datos")
        import procesar_tickets
        import re as _re
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] [Control Datos] Fuente: {fuente} — {len(rutas)} PDF(s)")

        total = len(rutas)
        # ── 0. Cargar caché de CONTENEDORES (una vez, zero opens repetidos) ──
        self._contenedores_cache = {}
        if rutas_excel:
            self._log(f"[Control Datos] Usando {len(rutas_excel)} Excel(s) seleccionados por el usuario")
            self._cargar_cache_contenedores(rutas_excel)
        else:
            self._cargar_cache_contenedores()  # escanea Desktop (compatibilidad)
        # ── 1. OCR: según método seleccionado ──
        ocr_method = self.config.get("ocr_method", "api_vision")
        textos_por_pdf = {}
        api_datos_raw = {}  # dict crudo de API Visión para mapeo directo

        if ocr_method == "api_vision":
            api_vision_conf = self.config.get("api_vision", {})
            api_key_raw = api_vision_conf.get("api_key", "")
            api_key = self._decrypt_api_key(api_key_raw)

            if not api_key:
                self.log_queue.put(f"[{timestamp}] [Control Datos] ⚠ No hay API Key configurada, usando OCR local")
                textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(rutas, engine="paddleocr")
            else:
                model = api_vision_conf.get("model", procesar_tickets.MODELO_VISION_DEFAULT)
                temperature = api_vision_conf.get("temperature", 0.1)
                max_tokens = api_vision_conf.get("max_tokens", 4000)
                timeout = api_vision_conf.get("timeout", 60)
                parallel_enabled = api_vision_conf.get("parallel_enabled", False)

                if parallel_enabled and total > 1:
                    # Parallel mode: distribute across enabled models
                    model_states = api_vision_conf.get("parallel_model_states", {})
                    enabled_models = [m for m, on in model_states.items() if on]
                    if not enabled_models:
                        enabled_models = [model]
                    # Ensure selected model is first
                    if model in enabled_models:
                        enabled_models.remove(model)
                    enabled_models.insert(0, model)

                    def _log_fn(msg):
                        self.log_queue.put(f"[{timestamp}] {msg}")

                    self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión paralelo: {total} PDF(s), {len(enabled_models)} modelo(s)...")
                    result = procesar_tickets.api_vision_con_fallback(
                        rutas, api_key, enabled_models,
                        temperature=temperature, max_tokens=max_tokens,
                        timeout=timeout, log_callback=_log_fn,
                    )
                    api_datos_raw = result["datos"]
                    textos_por_pdf = result["textos"]
                else:
                    # Sequential fallback: try models one by one per PDF
                    all_models = api_vision_conf.get("custom_models", [])
                    if not all_models:
                        all_models = list(procesar_tickets.MODELOS_VISION)
                    # Ensure selected model is first
                    if model in all_models:
                        all_models.remove(model)
                    all_models.insert(0, model)

                    self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión secuencial: {total} PDF(s), {len(all_models)} modelo(s) disponible(s)")
                    for i, ruta in enumerate(rutas):
                        stem = os.path.splitext(os.path.basename(ruta))[0]
                        carpeta = os.path.basename(os.path.dirname(ruta))
                        cliente = carpeta.split("_")[7] if len(carpeta.split("_")) > 7 else carpeta
                        self.log_queue.put(f"[{timestamp}] [Control Datos]   [{i+1}/{total}] {stem}")
                        success = False
                        for m in all_models:
                            try:
                                datos = procesar_tickets.api_vision_extraer_datos(
                                    ruta, api_key, model=m,
                                    temperature=temperature, max_tokens=max_tokens,
                                    timeout=timeout,
                                )
                                if "error" in datos:
                                    # Clean error message
                                    err = datos["error"]
                                    if "is not a valid model ID" in err:
                                        err = "modelo no disponible"
                                    elif "429" in err or "rate" in err.lower():
                                        err = "rate limit"
                                    elif "timeout" in err.lower():
                                        err = "timeout"
                                    self.log_queue.put(f"[{timestamp}] [Control Datos]   ERROR {m}: {err}")
                                    continue
                                # Success
                                api_datos_raw[stem] = datos
                                texto_simulado = "\n".join(f"{k}: {v}" for k, v in datos.items() if v)
                                textos_por_pdf[stem] = texto_simulado
                                self.log_queue.put(f"[{timestamp}] [Control Datos]   ✓ {stem} procesado con {m}")
                                success = True
                                break
                            except Exception as e:
                                self.log_queue.put(f"[{timestamp}] [Control Datos]   ERROR {m}: {e}")
                                continue
                        if not success:
                            self.log_queue.put(f"[{timestamp}] [Control Datos]   ⚠ Todos los modelos fallaron para {stem}, usando PaddleOCR")
                            textos_local = procesar_tickets.pdfs_a_texto_batch([ruta], engine="paddleocr")
                            textos_por_pdf[stem] = textos_local.get(stem, "")

                vision_ok = len(api_datos_raw)
                self.log_queue.put(f"[{timestamp}] [Control Datos] API Visión completado: {vision_ok}/{total} PDFs OK")
        else:
            self.log_queue.put(f"[{timestamp}] [Control Datos] OCR batch: {total} PDF(s)...")
            textos_por_pdf = procesar_tickets.pdfs_a_texto_batch(rutas, engine="paddleocr")
            self.log_queue.put(f"[{timestamp}] [Control Datos] OCR batch completado ({len(textos_por_pdf)} PDFs)")

        # ── 2. Parsear cada resultado + match CONTENEDOR ──
        def _normalizar_simple(s):
            return _re.sub(r'[^A-Z0-9]', '', str(s).upper().strip())

        def _procesar_texto(stem, texto):
            """Parsea el texto OCR y busca match CONTENEDOR.
            Si hay datos crudos de API Visión, mapea directamente."""
            
            if stem in api_datos_raw:
                # Mapeo directo desde JSON de API Visión
                raw = api_datos_raw[stem]
                patente  = raw.get("Patente", "")
                semi     = raw.get("Semirremolque", "")
                conductor = raw.get("Conductor", "")
                dni      = raw.get("DNI", "")
                permiso  = raw.get("Permiso", "")
                neto_str = raw.get("Neto", "")
                tara_str = raw.get("Tara Contenedor", "")
                contenedor_str = raw.get("Contenedor", "")
            else:
                # Parseo tradicional desde texto OCR
                datos = procesar_tickets.extraer_datos(texto)
                patente  = datos.get("Patente Camion", "")
                semi     = datos.get("Patente Acoplado", "")
                conductor_full = datos.get("Conductor", "")
                permiso  = datos.get("Merc./Permiso", "")
                neto_str = datos.get("Peso Neto (kg)", "")
                tara_str = datos.get("Tara Contenedor", "")
                contenedor_str = datos.get("Contenedor", "")
                dni = datos.get("DNI Conductor", "") or ""
                conductor = datos.get("Conductor", conductor_full) or conductor_full

            try:
                neto = float(neto_str.replace('.', '').replace(',', '.')) if neto_str else 0
            except (ValueError, AttributeError):
                neto = 0
            try:
                tara = float(tara_str.replace('.', '').replace(',', '.')) if tara_str else 0
            except (ValueError, AttributeError):
                tara = 0

            ticket_data = {
                "archivo": stem,
                "patente": patente, "semi": semi,
                "conductor": conductor, "dni": dni,
                "neto": neto, "tara": tara,
                "contenedor": contenedor_str,
                "permiso": permiso,
            }

            # Match CONTENEDOR — buscar por patente/semi/DNI en cache
            # Patentes en cache ya están limpias (sin guiones/espacios)
            ruta_match = None
            cont_data = None
            pe_val = None
            pn = _normalizar_simple(patente)
            sn = _normalizar_simple(semi)
            dn = _re.sub(r'\D', '', str(dni).strip())

            # ── Buscar por patente/semi/DNI en cache ──
            all_matches = []  # [(ruta, data, camion)]
            shared_excels = []
            shared_neto_sum = 0
            if self._contenedores_cache:
                for ruta, data in self._contenedores_cache.items():
                    if not data or not data.get("camiones"):
                        continue
                    for camion in data["camiones"]:
                        cp = camion.get("patente_camion", "")  # ya limpio en cache
                        cs = camion.get("patente_semi", "")    # ya limpio en cache
                        cdn = _re.sub(r'\D', '', str(camion.get("dni", "")).strip())
                        if (pn and cp and pn == cp) or \
                           (sn and cs and sn == cs) or \
                           (dn and cdn and dn == cdn):
                            all_matches.append((ruta, data, camion))
                            break

            if all_matches:
                ruta_match = all_matches[0][0]
                cont_data = all_matches[0][1]
                pe_val = data.get("pe")  # ya leído en cache

                if len(all_matches) > 1:
                    for ruta, data, camion in all_matches:
                        carpeta = os.path.basename(os.path.dirname(ruta))
                        shared_excels.append({
                            "ruta": ruta,
                            "nombre": carpeta,
                            "data": data,
                            "camion": camion,
                            "pe": data.get("pe"),
                        })
                        shared_neto_sum += camion.get("peso_carga", 0)
                    self.log_queue.put(
                        f"[MATCH] ✅ Fast compartido: patente '{patente}' "
                        f"-> {len(all_matches)} Excels, Neto total: {shared_neto_sum} kg")
                else:
                    self.log_queue.put(
                        f"[MATCH] ✅ Fast: patente '{patente}' -> {os.path.basename(ruta_match)}")

            return ticket_data, cont_data, ruta_match, pe_val, permiso, shared_excels, shared_neto_sum

        # ── 3. Procesar resultados y enviarlos a la UI ──
        # Sort by parent folder for grouped display
        rutas.sort(key=lambda r: os.path.basename(os.path.dirname(r)).split("_")[7].lower() if len(os.path.basename(os.path.dirname(r)).split("_")) > 7 else "")
        for hechos, ruta in enumerate(rutas, start=1):
            stem = os.path.splitext(os.path.basename(ruta))[0]
            has_data = stem in api_datos_raw or stem in textos_por_pdf

            if has_data:
                texto = textos_por_pdf.get(stem, "")
                ticket_data, cont_data, ruta_match, pe_val, permiso, \
                    shared_excels, shared_neto_sum = _procesar_texto(stem, texto)
                nombre = stem

                self.log_queue.put(
                    f"[{timestamp}] [{hechos}/{total}] ✓ {nombre} — permiso {permiso}")
                if shared_excels:
                    self.log_queue.put(
                        f"     ✓ COMPARTIDO: {len(shared_excels)} Excels, "
                        f"Neto total: {shared_neto_sum} kg")
                elif ruta_match:
                    self.log_queue.put(
                        f"     ✓ CONTENEDOR: {os.path.basename(ruta_match)}")
                else:
                    self.log_queue.put(
                        f"     ⚠ Sin CONTENEDOR match para permiso '{permiso}'")
                self.log_queue.put(("_OCR_RESULT_", {
                    "ticket": ticket_data,
                    "contenedor": cont_data,
                    "match": ruta_match,
                    "pe": pe_val,
                    "carpeta": os.path.basename(os.path.dirname(ruta)),
                    "cliente": os.path.basename(os.path.dirname(ruta)).split("_")[7] if len(os.path.basename(os.path.dirname(ruta)).split("_")) > 7 else "",
                    "shared_excels": shared_excels,
                    "shared_neto_sum": shared_neto_sum,
                }))
            else:
                nombre = os.path.basename(ruta)
                self.log_queue.put(
                    f"[{timestamp}] [{hechos}/{total}] ✗ {nombre}: OCR falló")

        # Finalizar
        self.log_queue.put(("_OCR_DONE_", None))
        self._contenedores_cache = {}  # liberar memoria

    def _cargar_datos_done(self):
        """Callback al recibir _OCR_DONE_. Desbloquea UI y habilita escritura.
        Si hay grupos pendientes (auto mode), procesa el siguiente."""
        # Chain: process next folder group if pending
        if (hasattr(self, '_ct_folder_groups')
                and self._ct_group_idx < len(self._ct_folder_groups)):
            self._procesar_siguiente_grupo_ct()
            return

        # No more groups — final cleanup
        self._ct_folder_groups = []
        self._ct_group_idx = 0
        self.tarea_activa = False
        try:
            for btn_name in ('btn_controlar_tickets',):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="normal")
            if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
                self.progress_carga.stop()
                self.progress_carga.pack_forget()
        except Exception:
            pass
        self._log("[Control Datos] OCR finalizado. Revisar resultados en la tabla.")

    def _finalizar_tarea(self):
        """Callback al recibir _TAREA_COMPLETA_. Rehabilita botones."""
        self.tarea_activa = False
        try:
            for btn_name in ('btn_control_final', 'btn_controlar_tickets', 'btn_controlar_coordinacion'):
                btn = getattr(self, btn_name, None)
                if btn and btn.winfo_exists():
                    btn.configure(state="normal")
            if hasattr(self, 'progress_carga') and self.progress_carga.winfo_exists():
                self.progress_carga.stop()
                self.progress_carga.pack_forget()
        except Exception:
            pass
        self._log("[Control] Tarea completada.")

    def _on_ocr_method_change(self, choice):
        """Persiste el método OCR elegido."""
        val = "api_vision" if choice == "API Visión" else "local"
        self.config["ocr_method"] = val
        # Habilitar/deshabilitar selector de modelo según método
        es_api = val == "api_vision"
        self._modelo_vision_menu.configure(state="normal" if es_api else "disabled")

    def _on_modelo_vision_change(self, choice):
        """Persiste el modelo de API Visión elegido."""
        self.config.setdefault("api_vision", {})["model"] = choice
        # Sincronizar también el dropdown de Ajustes si existe
        if hasattr(self, '_ent_vision_model') and self._ent_vision_model.winfo_exists():
            self._ent_vision_model.set(choice)

    def _actualizar_modelos_vision(self):
        """Sincroniza el dropdown de modelo con la config actual."""
        import procesar_tickets
        modelos_guardados = self.config.get("api_vision", {}).get("custom_models", [])
        modelos = modelos_guardados if modelos_guardados else list(procesar_tickets.MODELOS_VISION)
        modelo_default = self.config.get("api_vision", {}).get("model", procesar_tickets.MODELO_VISION_DEFAULT)
        self._modelo_vision_menu.configure(values=modelos)
        if modelo_default in modelos:
            self._modelo_vision_menu.set(modelo_default)
        elif modelos:
            self._modelo_vision_menu.set(modelos[0])

    # ═══════════════════════════════════════════════════════════════════
    # PANEL: AJUSTES
    # ═══════════════════════════════════════════════════════════════════
    def _panel_ajustes(self):
        if "ajustes" in self._panel_frames:
            return
        frame = ctk.CTkFrame(self.panel_container, fg_color="transparent")
        self._panel_frames["ajustes"] = frame
        # frame.pack(fill="both", expand=True) — packed by _cambiar_panel_forzado

        # ── Toolbar de tabs ──────────────────────────────────────────
        tabs_frame = ctk.CTkFrame(
            frame, fg_color=Palette.BG_CARD, corner_radius=8,
            border_width=1, border_color=Palette.BORDER,
        )
        tabs_frame.pack(fill="x", pady=(0, 4))

        self._ajustes_tabs = {}
        self._ajustes_frames = {}
        tab_names = [
            ("correo", "Correo", "settings"),
            ("documentos", "Documentos", "file-text"),
            ("descarga", "Descarga Mails", "inbox"),
            ("rutas", "Rutas", "folder"),
            ("valores", "Valores", "banknote"),
            ("seguridad", "Seguridad", "lock"),
            ("ocr", "OCR", "bot"),
            ("apariencia", "Apariencia", "palette"),
        ]

        def cambiar_tab(nombre):
            for key, fr in self._ajustes_frames.items():
                fr.pack_forget() if fr.winfo_ismapped() else None
            for key, btn in self._ajustes_tabs.items():
                if key == nombre:
                    btn.configure(fg_color=Palette.ACCENT, text_color=Palette.WHITE)
                else:
                    btn.configure(fg_color="transparent", text_color=Palette.TEXT_PRIMARY)
            # Construir contenido del tab si es primera vez que se clickea
            fr = self._ajustes_frames[nombre]
            if not fr.winfo_children() and nombre in self._ajustes_builders:
                self._ajustes_builders[nombre](fr)
            fr.pack(fill="both", expand=True)

        tab_keys = [t[0] for t in tab_names]

        def _reflow_tabs(event=None):
            fw = tabs_frame.winfo_width()
            if fw < 50:
                return
            pitch = 114  # 110 ancho + 4 padx
            cols = max(1, fw // pitch)
            for i, key in enumerate(tab_keys):
                self._ajustes_tabs[key].grid(row=i // cols, column=i % cols, padx=2, pady=2, sticky="w")

        for key, label, icon_key in tab_names:
            btn = ctk.CTkButton(
                tabs_frame, text=label,
                image=self._icons[icon_key] if icon_key else None,
                compound="left" if icon_key else "none",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                fg_color="transparent", hover_color=Palette.BG_HOVER,
                text_color=Palette.TEXT_SECONDARY,
                border_width=1, border_color=Palette.BORDER,
                corner_radius=6, height=32, width=110,
                command=lambda k=key: cambiar_tab(k),
            )
            btn.grid(row=0, column=0)  # placeholder, _reflow_tabs reposiciona
            self._ajustes_tabs[key] = btn

        tabs_frame.bind("<Configure>", _reflow_tabs)
        tabs_frame.after_idle(_reflow_tabs)

        self._ajustes_tabs["correo"].configure(fg_color=Palette.ACCENT, text_color=Palette.WHITE)

        # ── Contenedor de contenido ───────────────────────────────────
        content_frame = ctk.CTkFrame(frame, fg_color="transparent")
        content_frame.pack(fill="both", expand=True)

        for key, _, _ in tab_names:
            sub = ctk.CTkScrollableFrame(
                content_frame, fg_color=Palette.BG_CARD, corner_radius=8,
                border_width=1, border_color=Palette.BORDER,
            )
            self._ajustes_frames[key] = sub

        # Builders diferidos — el contenido de cada tab se crea al primer clic
        self._ajustes_builders = {
            "correo": self._ajustes_tab_correo,
            "documentos": self._ajustes_tab_documentos,
            "descarga": self._ajustes_tab_descarga,
            "rutas": self._ajustes_tab_rutas,
            "valores": self._ajustes_tab_valores,
            "seguridad": self._ajustes_tab_seguridad,
            "ocr": self._ajustes_tab_ocr,
            "apariencia": self._ajustes_tab_apariencia,
        }
        # Solo construir el tab Correo (visible por defecto)
        self._ajustes_tab_correo(self._ajustes_frames["correo"])
        self._ajustes_frames["correo"].pack(fill="both", expand=True)

        # ── Botón Guardar ────────────────────────────────────────────
        save_frame = ctk.CTkFrame(frame, fg_color=Palette.BG_CARD, corner_radius=8,
                                  border_width=1, border_color=Palette.BORDER, height=44)
        save_frame.pack(fill="x", pady=(4, 0))
        save_frame.pack_propagate(False)

        ctk.CTkButton(
            save_frame,
            text="Guardar",
            image=self._icons["save"], compound="left",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=Palette.SUCCESS, hover_color="#00c853",
            text_color=Palette.WHITE, corner_radius=6, height=34, width=190,
            command=self._guardar_ajustes,
        ).pack(side="left", padx=10, pady=4)

        self._ajustes_lbl_status = ctk.CTkLabel(
            save_frame, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.SUCCESS,
        )
        self._ajustes_lbl_status.pack(side="left", padx=8)

    def _ajustes_tab_apariencia(self, parent):
        """Tab Apariencia en Ajustes: configuración visual (tamaño de fuentes en popups)."""
        self._ajustes_seccion(parent, "Tamaño de Tablas Comparativas")

        ctk.CTkLabel(
            parent,
            text="Ajusta el tamaño de fuente de las ventanas de comparación (Tickets, Coordinación, Final). "
                 "El nivel se aplica al reiniciar o abrir un popup.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
            wraplength=480,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 8))

        # Font level selector
        font_level_row = ctk.CTkFrame(parent, fg_color="transparent")
        font_level_row.pack(fill="x", padx=14, pady=3)
        ctk.CTkLabel(
            font_level_row, text="Tamaño de tablas (1-3)",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(anchor="w")
        self._ent_font_level = ctk.CTkOptionMenu(
            font_level_row,
            values=["1", "2", "3"],
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
            button_hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            dropdown_fg_color=Palette.BG_CARD,
            dropdown_hover_color=Palette.BG_HOVER,
            dropdown_text_color=Palette.TEXT_PRIMARY,
            width=80, height=30,
        )
        self._ent_font_level.set(str(self.config.get("font_level", 1)))
        self._ent_font_level.pack(anchor="w", pady=(2, 0))

    # ── Helpers de layout ────────────────────────────────────────────
    def _ajustes_seccion(self, parent, titulo):
        ctk.CTkLabel(
            parent, text=titulo,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=Palette.TEXT_PRIMARY,
        ).pack(anchor="w", padx=14, pady=(12, 4))

    def _ajustes_row(self, parent, label, default="", show="", extra=None, width=260, toggle_pw=False):
        """Crea fila compacta: label arriba, entry abajo (ancho fijo). Retorna el CTkEntry."""
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=14, pady=3)
        ctk.CTkLabel(
            row, text=label,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY, anchor="w",
        ).pack(anchor="w")
        if toggle_pw:
            # Entry + botón lado a lado
            erow = ctk.CTkFrame(row, fg_color="transparent")
            erow.pack(anchor="w", pady=(2, 0))
            e = ctk.CTkEntry(
                erow, width=width - 34, height=30,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                text_color=Palette.TEXT_PRIMARY, corner_radius=4,
                show=show,
            )
            e.pack(side="left")
            self._pw_visible = False
            pw_toggle_btn = ctk.CTkButton(
                erow, text="",
                image=self._icons["eye"],
                width=30, height=30,
                font=ctk.CTkFont(size=14),
                fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
                text_color=Palette.TEXT_MUTED, corner_radius=4,
                command=None,
            )
            pw_toggle_btn.pack(side="left", padx=(4, 0))
            def _toggle():
                self._pw_visible = not self._pw_visible
                e.configure(show="" if self._pw_visible else "*")
            pw_toggle_btn.configure(command=_toggle)
        else:
            e = ctk.CTkEntry(
                row, width=width, height=30,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
                text_color=Palette.TEXT_PRIMARY, corner_radius=4,
                show=show,
            )
            e.pack(anchor="w", pady=(2, 0))
        e.insert(0, str(default) if default else "")
        if extra:
            ctk.CTkLabel(
                row, text=extra,
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                text_color=Palette.TEXT_MUTED, anchor="w",
            ).pack(anchor="w", pady=(1, 0))
        return e

    # ── TAB: CORREO ──────────────────────────────────────────────────
    def _ajustes_tab_correo(self, parent):
        self._ajustes_seccion(parent, "Credenciales IMAP")
        self._ent_correo_usuario = self._ajustes_row(
            parent, "Usuario (email):", self._cfg_obtener_correo("usuario", ""), width=300)
        self._ent_correo_password = self._ajustes_row(
            parent, "Contraseña:", self._cfg_obtener_correo("password", ""),
            show="*", width=240, toggle_pw=True)
        self._ent_correo_imap = self._ajustes_row(
            parent, "Servidor IMAP:", self._cfg_obtener_correo("imap_server", IMAP_SERVER), width=280)
        self._ent_correo_puerto = self._ajustes_row(
            parent, "Puerto IMAP:", str(self._cfg_obtener_correo("imap_puerto", PUERTO_IMAP)), width=80)
        self._sent_correo_remitente_papeles = self._ajustes_row(
            parent, "Remitente Papeles (filtro):", self._cfg_obtener_correo("remitente_papeles", ""), width=280)

        self._ajustes_seccion(parent, "Destinatarios — Planillas de Carga")
        default_grupal = "\n".join(self._cfg_obtener_correo("destinatarios_grupal", DESTINATARIOS_GRUPAL))
        self._ajustes_texto_grupal = ctk.CTkTextbox(
            parent, height=100, width=400,
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=4, border_width=1,
        )
        self._ajustes_texto_grupal.insert("1.0", default_grupal)
        self._ajustes_texto_grupal.pack(anchor="w", padx=14, pady=(2, 8))

        self._ajustes_seccion(parent, "Destinatarios — Correo Individual")
        default_ind = "\n".join(self._cfg_obtener_correo("destinatarios_individual", DESTINATARIOS_INDIVIDUAL))
        self._ajustes_texto_ind = ctk.CTkTextbox(
            parent, height=70, width=400,
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=4, border_width=1,
        )
        self._ajustes_texto_ind.insert("1.0", default_ind)
        self._ajustes_texto_ind.pack(anchor="w", padx=14, pady=(2, 8))

        self._ajustes_seccion(parent, "Remitente Balanza")
        self._ent_remitente_balanza = self._ajustes_row(
            parent, "Remitente Balanza:", self.config.get("correo", {}).get("remitente_balanza", ""), width=300)

    # ── TAB: DOCUMENTOS ──────────────────────────────────────────────
    def _ajustes_tab_documentos(self, parent):
        # Contenedor horizontal: izquierda (dorsos) | derecha (copias)
        split = ctk.CTkFrame(parent, fg_color="transparent")
        split.pack(fill="x", padx=10, pady=(4, 0))

        # ── Columna Izquierda: Dorsos ────────────────────────────────
        col_izq = ctk.CTkFrame(split, fg_color="transparent")
        col_izq.pack(side="left", fill="y", padx=(4, 8))

        self._ajustes_seccion(col_izq, "Hojas — Dorsos")
        self._ent_dorso_mic = self._ajustes_row(
            col_izq, "Dorso MIC:", str(self._cfg_obtener_docs("dorso_mic", 15)), width=70)
        self._ent_dorso_crt = self._ajustes_row(
            col_izq, "Dorso CRT:", str(self._cfg_obtener_docs("dorso_crt", 4)), width=70)
        self._ent_dorso_pe = self._ajustes_row(
            col_izq, "Dorso PE:", str(self._cfg_obtener_docs("dorso_pe", 2)), width=70)

        # ── Columna Derecha: Copias ──────────────────────────────────
        col_der = ctk.CTkFrame(split, fg_color="transparent")
        col_der.pack(side="left", fill="y", padx=(8, 4))

        self._ajustes_seccion(col_der, "Copias — Impresión")
        self._ent_permiso_exp = self._ajustes_row(
            col_der, "Permiso de Exportación:", str(self._cfg_obtener_docs("permiso_exportacion", 2)), width=70)
        self._ent_hoja_ruta = self._ajustes_row(
            col_der, "Hoja de Ruta:", str(self._cfg_obtener_docs("hoja_ruta", 2)), width=70)
        self._ent_sobre = self._ajustes_row(
            col_der, "Sobre (Planilla de Carga):", str(self._cfg_obtener_docs("sobre", 1)), width=70)

    def _ajustes_row_browse(self, parent, label, default="", extra=None, width=290):
        """Como _ajustes_row pero con botón 📂 Examinar.
           Detecta automáticamente si es pendrive (ruta relativa) o disco fijo (absoluta)."""
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=14, pady=3)
        ctk.CTkLabel(
            row, text=label,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY, anchor="w",
        ).pack(anchor="w")
        entry_row = ctk.CTkFrame(row, fg_color="transparent")
        entry_row.pack(fill="x", pady=(2, 0))
        e = ctk.CTkEntry(
            entry_row, width=width, height=30,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=4,
        )
        e.insert(0, str(default) if default else "")
        e.pack(side="left")
        btn = ctk.CTkButton(
            entry_row, text="📂 Examinar", width=100, height=30,
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_HOVER, hover_color=Palette.ACCENT_DIM,
            text_color=Palette.TEXT_SECONDARY, corner_radius=4,
            command=lambda: self._ajustes_examinar_carpeta(e),
        )
        btn.pack(side="left", padx=(6, 0))
        if extra:
            ctk.CTkLabel(
                row, text=extra,
                font=ctk.CTkFont(family=FONT_FAMILY, size=10),
                text_color=Palette.TEXT_MUTED, anchor="w",
            ).pack(anchor="w", pady=(1, 0))
        return e

    def _ajustes_examinar_carpeta(self, entry_widget):
        r"""Abre diálogo para seleccionar carpeta.
           Si la unidad es removible (pendrive), recorta la letra (F:\TRABAJO → TRABAJO).
           Si es disco fijo/red, deja la ruta absoluta completa."""
        from tkinter import filedialog as tk_filedialog
        ruta = tk_filedialog.askdirectory(title="Seleccionar carpeta")
        if ruta:
            ruta = ruta.replace("/", "\\")
            match = re.match(r"^([A-Za-z]):\\(.*)", ruta)
            if match:
                letra = match.group(1)
                resto = match.group(2)
                # Detectar si es pendrive (DRIVE_REMOVABLE = 2) o disco fijo
                try:
                    from ctypes import windll
                    tipo = windll.kernel32.GetDriveTypeW(f"{letra}:\\")
                    if tipo == 2:  # DRIVE_REMOVABLE → pendrive
                        ruta = resto
                except Exception:
                    pass  # si falla, deja la ruta como está
            entry_widget.delete(0, "end")
            entry_widget.delete(0, "end")
            entry_widget.insert(0, ruta)

    # ── TAB: RUTAS ───────────────────────────────────────────────────
    def _ajustes_tab_rutas(self, parent):
        self._ajustes_seccion(parent, "Archivos en Pendrive / PC — Carpetas de búsqueda")
        ctk.CTkLabel(
            parent,
            text="Usá Examinar para elegir la carpeta. Si es un pendrive, la letra\n"
                 "se recorta sola. Si es un disco fijo, se guarda la ruta absoluta.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=10),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 4))

        self._ent_ruta_sobres = self._ajustes_row_browse(
            parent, "SOBRES (carpeta):",
            self._cfg_obtener_rutas("sobres", "TRABAJO\\01_PLANILLAS"),
            extra="Busca SOBRES_2026.xlsx")
        self._ent_ruta_cobro = self._ajustes_row_browse(
            parent, "COBRO (carpeta):",
            self._cfg_obtener_rutas("cobro", "TRABAJO\\01_PLANILLAS"),
            extra="Busca COBRO_2026.xlsx")
        self._ent_ruta_pc = self._ajustes_row_browse(
            parent, "PC (carpeta):",
            self._cfg_obtener_rutas("pc", "TRABAJO\\01_PLANILLAS"),
            extra="Busca PC.xlsx / PC_2026.xlsx")

        # Separador
        ctk.CTkFrame(parent, fg_color=Palette.DIVIDER, height=1).pack(fill="x", padx=14, pady=(10, 8))

        self._ajustes_seccion(parent, "Planilla Maestra — CARGA TERRESTRE")
        self._ent_ruta_ct_carpeta = self._ajustes_row_browse(
            parent, "Carpeta:",
            self._cfg_obtener_rutas("carga_terrestre_carpeta", "TRABAJO\\01_PLANILLAS"))
        self._ent_ruta_ct_nombre = self._ajustes_row(
            parent, "Nombre del archivo:",
            self._cfg_obtener_rutas("carga_terrestre_nombre", "CARGA TERRESTRE.xlsx"), width=280)

        # Separador
        ctk.CTkFrame(parent, fg_color=Palette.DIVIDER, height=1).pack(fill="x", padx=14, pady=(10, 8))

        self._ajustes_seccion(parent, "Contenedores de Carga y Descargas")
        self._ent_ruta_planillas = self._ajustes_row_browse(
            parent, "Carpeta de Contenedores de Carga:",
            self._cfg_obtener_rutas("planillas_carga", "Desktop"),
            extra="Donde se buscan las carpetas con 'PLANILLA DE CARGA'")
        self._ent_ruta_descarga = self._ajustes_row_browse(
            parent, "Carpeta de descarga de mails:",
            self._cfg_obtener_rutas("descarga_mails", "Desktop"),
            extra="Donde se guardan los adjuntos descargados de los mails")

        # Escritorio
        ctk.CTkFrame(parent, fg_color=Palette.DIVIDER, height=1).pack(fill="x", padx=14, pady=(10, 8))
        self._ajustes_seccion(parent, "Nombre del Escritorio")
        self._ent_ruta_escritorio = self._ajustes_row(
            parent, "Nombre de la carpeta Escritorio:",
            self._cfg_obtener_rutas("escritorio_nombre", "Desktop"),
            extra="Solo cambiar si tu SO usa otro nombre (ej: 'Desktop' en inglés)", width=200)

        # Backup y Sellos
        ctk.CTkFrame(parent, fg_color=Palette.DIVIDER, height=1).pack(fill="x", padx=14, pady=(10, 8))
        self._ajustes_seccion(parent, "Backup Pendrive y Sellos")
        self._ent_ruta_backup = self._ajustes_row_browse(
            parent, "Backup Pendrive (carpeta):",
            self._cfg_obtener_rutas("backup_pendrive", "TRABAJO\\CARGAS"))
        self._ent_ruta_mic_sellos = self._ajustes_row_browse(
            parent, "FECHA MIC Y SELLOS (carpeta):",
            self._cfg_obtener_rutas("mic_sellos", "TRABAJO\\01_PLANILLAS"),
            extra="Busca FECHA MIC Y SELLOS.xlsx")
        self._ent_ruta_crt_original = self._ajustes_row_browse(
            parent, "FECHA CRT Y ORIGINAL (carpeta):",
            self._cfg_obtener_rutas("crt_original", "TRABAJO\\01_PLANILLAS"),
            extra="Busca FECHA CRT Y ORIGINAL.xlsx")

    # ── TAB: VALORES ──────────────────────────────────────────────────
    def _ajustes_tab_descarga(self, parent):
        self._ajustes_seccion(parent, "Cantidad de Mails a Buscar")
        ctk.CTkLabel(
            parent,
            text="Estos valores se usan en el panel Correos para limitar la búsqueda.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 4))

        dm = self.config.get("descarga_mails", {})
        papeles_default = dm.get("papeles", "2")
        reglas_default = dm.get("reglas", "4")
        sin_filtro_default = dm.get("sin_filtro", "20")

        self._ent_descarga_papeles = self._ajustes_row(
            parent, "📥  Buscar y Descargar (papeles):", papeles_default,
            extra="Mails más nuevos con 'papeles' en el asunto.", width=80)
        self._ent_descarga_reglas = self._ajustes_row(
            parent, "🔍  Buscar con reglas:", reglas_default,
            extra="Mails que coinciden con las reglas configuradas.", width=80)
        self._ent_descarga_sin_filtro = self._ajustes_row(
            parent, "📋  Mail sin filtros:", sin_filtro_default,
            extra="Últimos mails sin ningún filtro.", width=80)

    def _ajustes_tab_valores(self, parent):
        self._ajustes_seccion(parent, "Tarifas de Planilla COBRO")
        self._ent_precio_carpeta = self._ajustes_row(
            parent, "Valor Carpeta ($):",
            str(self._cfg_obtener("valores", "precio_carpeta", 49000)),
            extra="Primer ítem de cada fecha. Los siguientes llevan la mitad.",
            width=120,
        )
        self._ent_ata_tares = self._ajustes_row(
            parent, "ATA y TARES ($):",
            str(self._cfg_obtener("valores", "ata_tares", 65000)),
            extra="Valor del servicio ATA por contenedor.",
            width=120,
        )

        self._ajustes_seccion(parent, "Planilla de Carga")
        self._ent_clave_pdf = self._ajustes_row(
            parent, "Clave de protección:",
            self._cfg_obtener("valores", "clave_pdf", "123"),
            extra="Contraseña que se aplica a la hoja 'Planilla de Carga' al extraerla.",
            width=120,
        )

        self._ajustes_seccion(parent, "Guardas Disponibles")
        ctk.CTkLabel(
            parent,
            text="Un guarda por línea. Se usan en 'Agregar Guarda' del panel Contenedores.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 4))

        guardas_actuales = self._cfg_obtener("valores", "guardas", ["Gonzalez", "Rodriguez", "Martinez", "Perez"])
        self._ajustes_texto_guardas = ctk.CTkTextbox(
            parent, height=100, width=320,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=4,
        )
        self._ajustes_texto_guardas.pack(anchor="w", padx=14, pady=(0, 8))
        self._ajustes_texto_guardas.insert("1.0", "\n".join(guardas_actuales))

        # ── Config Súper Auto ──────────────────────────────────────
        self._ajustes_seccion(parent, "⚡ Súper Auto — Funciones Automáticas")
        sa = self._cfg_obtener("super_auto", "pasos", {})
        self._super_check_sobre    = ctk.BooleanVar(value=sa.get("sobre", True))
        self._super_check_permiso  = ctk.BooleanVar(value=sa.get("permiso", True))
        self._super_check_hoja_ruta = ctk.BooleanVar(value=sa.get("hoja_ruta", True))
        self._super_check_recibo   = ctk.BooleanVar(value=sa.get("recibo_ata", True))
        self._super_check_guarda   = ctk.BooleanVar(value=sa.get("aplicar_guarda", True))
        self._super_check_planillas = ctk.BooleanVar(value=sa.get("completar_planillas", True))

        for var, txt in [(self._super_check_sobre, "☑ Imprimir Sobre (1 copia)"),
                          (self._super_check_permiso, "☑ Imprimir Permiso (2 copias)"),
                          (self._super_check_hoja_ruta, "☑ Imprimir Hoja de Ruta (2 copias)"),
                          (self._super_check_recibo, "☑ Imprimir Recibo ATA (1 copia)"),
                          (self._super_check_guarda, "☑ Aplicar Guarda"),
                          (self._super_check_planillas, "☑ Completar Planillas")]:
            ctk.CTkCheckBox(parent, text=txt, variable=var,
                            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                            text_color=Palette.TEXT_PRIMARY,
                            border_color=Palette.BORDER).pack(anchor="w", padx=20, pady=3)

    # ── TAB: SEGURIDAD ───────────────────────────────────────────────
    def _ajustes_tab_seguridad(self, parent):
        self._ajustes_seccion(parent, "Contraseña Maestra")
        ctk.CTkLabel(
            parent,
            text="Protege el acceso a la aplicación y los Ajustes.\n"
                 "Dejá ambos campos vacíos para desactivar la protección.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 8))

        pw_actual = self._cfg_obtener("seguridad", "password", "")
        self._ent_master_pw = self._ajustes_row(
            parent, "Contraseña nueva:", pw_actual, show="*", width=220)
        self._ent_master_pw_confirm = self._ajustes_row(
            parent, "Confirmar contraseña:", pw_actual, show="*", width=220,
            extra="Debe coincidir con la anterior. Vacío = sin protección.")

        self._ajustes_seccion(parent, "Recuperación")
        ctk.CTkLabel(
            parent,
            text="Si olvidás la contraseña, se enviará al correo configurado en la pestaña Correo.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
        ).pack(anchor="w", padx=14, pady=(0, 4))

    # ── TAB: OCR ─────────────────────────────────────────────────────
    def _verificar_motores_ocr(self):
        """Verifica Tesseract y PaddleOCR después de que el panel se renderice (evita demora al abrir)."""
        import procesar_tickets
        # Tesseract — priorizar sidecar portátil sobre instalación del sistema
        import pytesseract
        tess_sidecar = os.path.join(procesar_tickets.TESSERACT_SIDECAR, "tesseract.exe")
        tess_ok = False
        tess_path = ""
        if os.path.isfile(tess_sidecar):
            try:
                pytesseract.pytesseract.tesseract_cmd = tess_sidecar
                pytesseract.get_tesseract_version()
                tess_ok = True
                tess_path = tess_sidecar
            except Exception:
                pass
        if not tess_ok:
            try:
                pytesseract.pytesseract.tesseract_cmd = procesar_tickets.TESSERACT_CMD
                pytesseract.get_tesseract_version()
                tess_ok = True
                tess_path = procesar_tickets.TESSERACT_CMD
            except Exception:
                tess_ok = False
                tess_path = ""
        if hasattr(self, '_ocr_lbl_tesseract') and self._ocr_lbl_tesseract.winfo_exists():
            self._ocr_lbl_tesseract.configure(
                text="✓ Disponible" if tess_ok else "✗ No encontrado",
                text_color=Palette.SUCCESS if tess_ok else Palette.ERROR,
            )
        if hasattr(self, '_ocr_lbl_tess_path') and self._ocr_lbl_tess_path.winfo_exists():
            self._ocr_lbl_tess_path.configure(text=f"({tess_path})" if tess_path else "")

        # PaddleOCR — background thread (no congelar la UI)
        if hasattr(self, '_ocr_lbl_paddle') and self._ocr_lbl_paddle.winfo_exists():
            self._ocr_lbl_paddle.configure(
                text="⏳ Verificando...",
                text_color=Palette.TEXT_MUTED,
            )
        threading.Thread(target=self._verificar_paddle_bg, daemon=True).start()

    def _verificar_paddle_bg(self):
        """Corre _detectar_paddle() en background y actualiza UI via after()."""
        import procesar_tickets
        paddle_ok = procesar_tickets._detectar_paddle() is not None
        paddle_path = procesar_tickets.PADDLE_PORTABLE_PYTHON if paddle_ok else procesar_tickets.PADDLE_SIDECAR
        self.after(0, lambda: self._actualizar_estado_paddle(paddle_ok, paddle_path))

    def _actualizar_estado_paddle(self, ok: bool, path: str):
        """Callback desde el thread: actualiza labels de PaddleOCR en la UI."""
        if hasattr(self, '_ocr_lbl_paddle') and self._ocr_lbl_paddle.winfo_exists():
            self._ocr_lbl_paddle.configure(
                text="✓ Disponible" if ok else "✗ No encontrado",
                text_color=Palette.SUCCESS if ok else Palette.ERROR,
            )
        if hasattr(self, '_ocr_lbl_paddle_path') and self._ocr_lbl_paddle_path.winfo_exists():
            self._ocr_lbl_paddle_path.configure(text=f"({path})" if path else "")

    def _ajustes_tab_ocr(self, parent):
        """Tab OCR en Ajustes: muestra estado de disponibilidad de motores."""
        import procesar_tickets

        self._ajustes_seccion(parent, "Motores OCR")

        ctk.CTkLabel(
            parent,
            text="Tesseract → permiso de exportación (rápido).  PaddleOCR → tickets escaneados (preciso).",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED,
            wraplength=480,
            justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 8))

        self._ajustes_seccion(parent, "Estado de Motores")

        # Crear labels (placeholder), verificar motores después
        row1 = ctk.CTkFrame(parent, fg_color="transparent")
        row1.pack(fill="x", padx=14, pady=2)
        ctk.CTkLabel(
            row1, text="Tesseract:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(side="left")
        self._ocr_lbl_tesseract = ctk.CTkLabel(
            row1, text="⌛ Verificando...",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_MUTED,
        )
        self._ocr_lbl_tesseract.pack(side="left", padx=(8, 0))
        self._ocr_lbl_tess_path = ctk.CTkLabel(
            row1, text="",
            font=ctk.CTkFont(family=FONT_MONO, size=10),
            text_color=Palette.TEXT_MUTED,
        )
        self._ocr_lbl_tess_path.pack(side="left", padx=(4, 0))

        row2 = ctk.CTkFrame(parent, fg_color="transparent")
        row2.pack(fill="x", padx=14, pady=2)
        ctk.CTkLabel(
            row2, text="PaddleOCR:",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY,
        ).pack(side="left")
        self._ocr_lbl_paddle = ctk.CTkLabel(
            row2, text="⌛ Verificando...",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
            text_color=Palette.TEXT_MUTED,
        )
        self._ocr_lbl_paddle.pack(side="left", padx=(8, 0))
        self._ocr_lbl_paddle_path = ctk.CTkLabel(
            row2, text="",
            font=ctk.CTkFont(family=FONT_MONO, size=10),
            text_color=Palette.TEXT_MUTED,
        )
        self._ocr_lbl_paddle_path.pack(side="left", padx=(4, 0))

        # Verificar motores después de que el panel se haya renderizado
        parent.after(100, self._verificar_motores_ocr)

        # ── API Visión (OpenRouter) ───────────────────────────────────
        self._ajustes_seccion(parent, "API Visión (OpenRouter)")

        ctk.CTkLabel(
            parent,
            text="API para extraer datos de tickets escaneados mediante modelos con visión. Usa OpenRouter como gateway.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED, wraplength=480, justify="left",
        ).pack(anchor="w", padx=14, pady=(0, 8))

        # API Key (encriptada)
        self._ent_vision_api_key = self._ajustes_row(
            parent, "API Key",
            default=self.config.get("api_vision", {}).get("api_key", ""),
            show="*", width=300, toggle_pw=True,
        )

        # Modelo (dropdown) — mismo layout que _ajustes_row: label arriba, control abajo
        modelos_guardados = self.config.get("api_vision", {}).get("custom_models", [])
        modelos = modelos_guardados if modelos_guardados else list(procesar_tickets.MODELOS_VISION)
        modelo_default = self.config.get("api_vision", {}).get("model", procesar_tickets.MODELO_VISION_DEFAULT)
        vis_row = ctk.CTkFrame(parent, fg_color="transparent")
        vis_row.pack(fill="x", padx=14, pady=3)
        ctk.CTkLabel(
            vis_row, text="Modelo",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=Palette.TEXT_SECONDARY, anchor="w",
        ).pack(anchor="w")
        self._ent_vision_model = ctk.CTkOptionMenu(
            vis_row,
            values=modelos,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.BG_INPUT, button_color=Palette.ACCENT,
            button_hover_color=Palette.ACCENT_HOVER,
            text_color=Palette.TEXT_PRIMARY,
            dropdown_fg_color=Palette.BG_CARD,
            dropdown_hover_color=Palette.BG_HOVER,
            dropdown_text_color=Palette.TEXT_PRIMARY,
            corner_radius=4, width=320, height=30,
        )
        self._ent_vision_model.set(modelo_default if modelo_default in modelos else modelos[0] if modelos else "")
        self._ent_vision_model.pack(anchor="w", pady=(2, 0))

        # Modelos disponibles (text box, editar/quitar modelos)
        ctk.CTkLabel(
            parent, text="Modelos disponibles (uno por línea)",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED, anchor="w",
        ).pack(anchor="w", padx=14, pady=(6, 0))
        self._ent_vision_custom_models = ctk.CTkTextbox(
            parent, height=70, width=400,
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=Palette.BG_INPUT, border_color=Palette.BORDER,
            text_color=Palette.TEXT_PRIMARY, corner_radius=4, border_width=1,
        )
        self._ent_vision_custom_models.insert("1.0", "\n".join(modelos))
        self._ent_vision_custom_models.pack(anchor="w", padx=14, pady=(2, 8))
        self._ent_vision_custom_models.bind("<KeyRelease>", lambda e: self._sync_modelos_desde_textbox())
        self._sync_modelos_desde_textbox()

        # ── API Vision en Paralelo ───────────────────────────────────
        parallel_cfg = self.config.get("api_vision", {})
        self._chk_parallel_enabled = ctk.CTkCheckBox(
            parent, text="Habilitar API Vision en Paralelo",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
            border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
            text_color=Palette.TEXT_PRIMARY,
            command=self._toggle_parallel_panel,
        )
        self._chk_parallel_enabled.pack(anchor="w", padx=14, pady=(4, 2))
        if parallel_cfg.get("parallel_enabled", False):
            self._chk_parallel_enabled.select()

        # Model panel frame (hidden initially)
        self._parallel_panel = ctk.CTkFrame(parent, fg_color="transparent")
        ctk.CTkLabel(
            self._parallel_panel, text="Modelos para procesamiento paralelo",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=Palette.TEXT_MUTED, anchor="w",
        ).pack(anchor="w", padx=14, pady=(2, 0))
        self._parallel_checks_frame = ctk.CTkFrame(self._parallel_panel, fg_color="transparent")
        self._parallel_checks_frame.pack(anchor="w", padx=14, pady=(0, 4))
        self._parallel_model_checks = {}

        # Build checkboxes for existing models
        saved_states = parallel_cfg.get("parallel_model_states", {})
        for m in modelos:
            is_enabled = saved_states.get(m, True)
            chk = ctk.CTkCheckBox(
                self._parallel_checks_frame, text=m,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            chk.pack(anchor="w", padx=2, pady=1)
            if is_enabled:
                chk.select()
            self._parallel_model_checks[m] = chk

        self._toggle_parallel_panel()

        # Temperature
        self._ent_vision_temperature = self._ajustes_row(
            parent, "Temperature (0-1)",
            default=str(self.config.get("api_vision", {}).get("temperature", 0.1)),
            width=100,
        )

        # Max Tokens
        self._ent_vision_max_tokens = self._ajustes_row(
            parent, "Max Tokens",
            default=str(self.config.get("api_vision", {}).get("max_tokens", 4000)),
            width=100,
        )

        # Timeout
        self._ent_vision_timeout = self._ajustes_row(
            parent, "Timeout (segundos)",
            default=str(self.config.get("api_vision", {}).get("timeout", 60)),
            width=100,
        )

    # ── Sincronizar dropdown de modelo desde el textbox ─────────────
    def _sync_modelos_desde_textbox(self):
        raw = self._ent_vision_custom_models.get("1.0", "end-1c")
        lines = [l.strip() for l in raw.split("\n") if l.strip()]
        if lines:
            current = self._ent_vision_model.get()
            self._ent_vision_model.configure(values=lines)
            if current in lines:
                self._ent_vision_model.set(current)
            else:
                self._ent_vision_model.set(lines[0])
            # También sincronizar el dropdown de la top_bar si existe
            if hasattr(self, '_modelo_vision_menu') and self._modelo_vision_menu.winfo_exists():
                top_current = self._modelo_vision_menu.get()
                self._modelo_vision_menu.configure(values=lines)
                if top_current in lines:
                    self._modelo_vision_menu.set(top_current)
                elif lines:
                    self._modelo_vision_menu.set(lines[0])
        # Rebuild parallel model checkboxes when model list changes
        if hasattr(self, '_parallel_model_checks'):
            self._rebuild_parallel_model_checks()

    # ── Toggle panel de modelos paralelos ───────────────────────────
    def _toggle_parallel_panel(self):
        if self._chk_parallel_enabled.get() == 1:
            self._parallel_panel.pack(anchor="w", padx=0, pady=(0, 4), after=self._chk_parallel_enabled)
        else:
            self._parallel_panel.pack_forget()

    # ── Reconstruir checkboxes de modelos paralelos ─────────────────
    def _rebuild_parallel_model_checks(self):
        """Rebuild parallel model checkboxes from the textbox content."""
        raw = self._ent_vision_custom_models.get("1.0", "end-1c")
        modelos = [l.strip() for l in raw.split("\n") if l.strip()]
        # Preserve existing states
        old_states = {}
        for m, chk in self._parallel_model_checks.items():
            try:
                old_states[m] = chk.get() == 1
            except Exception:
                old_states[m] = True
        # Clear frame
        for w in self._parallel_checks_frame.winfo_children():
            w.destroy()
        self._parallel_model_checks = {}
        for m in modelos:
            is_enabled = old_states.get(m, True)
            chk = ctk.CTkCheckBox(
                self._parallel_checks_frame, text=m,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                fg_color=Palette.ACCENT, hover_color=Palette.ACCENT_HOVER,
                border_color=Palette.BORDER, checkmark_color=Palette.WHITE,
                text_color=Palette.TEXT_PRIMARY,
            )
            chk.pack(anchor="w", padx=2, pady=1)
            if is_enabled:
                chk.select()
            self._parallel_model_checks[m] = chk

    # ── GUARDAR AJUSTES ──────────────────────────────────────────────
    def _guardar_ajustes(self):
        try:
            _g = lambda a: getattr(self, a, None)

            # Seguridad — validar que coincidan y actualizar cache
            ent_pw1 = _g('_ent_master_pw')
            ent_pw2 = _g('_ent_master_pw_confirm')
            pw1 = ent_pw1.get().strip() if ent_pw1 else self.config.get("seguridad", {}).get("password", "")
            pw2 = ent_pw2.get().strip() if ent_pw2 else ""
            if ent_pw1 and ent_pw2 and pw1 != pw2:
                self._ajustes_lbl_status.configure(
                    text="✗ Las contraseñas no coinciden. Corregí y volvé a guardar.",
                    text_color=Palette.ERROR)
                return

            self._master_pw_cache = pw1

            # Correo (siempre encriptado con la key del .env, no con la master)
            ent = _g('_ent_correo_password')
            mail_pw = ent.get().strip() if ent else ""
            key = os.environ["MULTIAGENTE_SECRET_KEY"]
            correo_cfg = {}
            for attr, cfg_key in [('_ent_correo_usuario', 'usuario'), ('_ent_correo_password', 'password'),
                                  ('_ent_correo_imap', 'imap_server'), ('_ent_correo_puerto', 'imap_puerto'),
                                  ('_sent_correo_remitente_papeles', 'remitente_papeles'),
                                  ('_ent_remitente_balanza', 'remitente_balanza')]:
                w = _g(attr)
                if w is not None:
                    raw = w.get().strip()
                    correo_cfg[cfg_key] = int(raw) if cfg_key == 'imap_puerto' else raw
            if mail_pw:
                correo_cfg["password"] = self._encrypt_val(mail_pw, key)
            w = _g('_ajustes_texto_grupal')
            if w is not None:
                correo_cfg["destinatarios_grupal"] = [
                    l.strip() for l in w.get("1.0", "end-1c").split("\n") if l.strip()
                ]
            w = _g('_ajustes_texto_ind')
            if w is not None:
                correo_cfg["destinatarios_individual"] = [
                    l.strip() for l in w.get("1.0", "end-1c").split("\n") if l.strip()
                ]
            if correo_cfg:
                self.config["correo"] = {**self.config.get("correo", {}), **correo_cfg}

            # Documentos
            docs_cfg = {}
            for key_doc, attr in [("dorso_mic", "_ent_dorso_mic"), ("dorso_crt", "_ent_dorso_crt"),
                                  ("dorso_pe", "_ent_dorso_pe"), ("permiso_exportacion", "_ent_permiso_exp"),
                                  ("hoja_ruta", "_ent_hoja_ruta"), ("sobre", "_ent_sobre")]:
                w = _g(attr)
                if w is not None:
                    try:
                        docs_cfg[key_doc] = int(w.get().strip())
                    except ValueError:
                        pass
            if docs_cfg:
                self.config["documentos"] = {**self.config.get("documentos", {}), **docs_cfg}

            # Rutas
            rutas_cfg = {}
            for attr, cfg_key in [("_ent_ruta_sobres", "sobres"), ("_ent_ruta_cobro", "cobro"),
                                  ("_ent_ruta_pc", "pc"), ("_ent_ruta_ct_carpeta", "carga_terrestre_carpeta"),
                                  ("_ent_ruta_ct_nombre", "carga_terrestre_nombre"),
                                  ("_ent_ruta_planillas", "planillas_carga"),
                                  ("_ent_ruta_descarga", "descarga_mails"),
                                  ("_ent_ruta_escritorio", "escritorio_nombre"),
                                  ("_ent_ruta_backup", "backup_pendrive"),
                                  ("_ent_ruta_mic_sellos", "mic_sellos"),
                                  ("_ent_ruta_crt_original", "crt_original")]:
                w = _g(attr)
                if w is not None:
                    rutas_cfg[cfg_key] = w.get().strip()
            if rutas_cfg:
                self.config["rutas"] = {**self.config.get("rutas", {}), **rutas_cfg}

            # Descarga Mails — guardar y sincronizar entries del panel Correos
            descarga_cfg = {}
            for attr, cfg_key in [("_ent_descarga_papeles", "papeles"),
                                  ("_ent_descarga_reglas", "reglas"),
                                  ("_ent_descarga_sin_filtro", "sin_filtro")]:
                w = _g(attr)
                if w is not None:
                    descarga_cfg[cfg_key] = w.get().strip()
            if descarga_cfg:
                self.config["descarga_mails"] = {**self.config.get("descarga_mails", {}), **descarga_cfg}
            # Sincronizar los entries del panel Correos (si aún existen)
            for attr, src in [('_mail_entry_cantidad', '_ent_descarga_papeles'),
                              ('_mail_entry_cantidad_reglas', '_ent_descarga_reglas'),
                              ('_mail_entry_sin_filtro', '_ent_descarga_sin_filtro'),
                              ('_mail_entry_sin_filtro_r1', '_ent_descarga_sin_filtro')]:
                w = _g(attr)
                if w is not None and w.winfo_exists():
                    src_w = _g(src)
                    if src_w is not None:
                        w.delete(0, "end")
                        w.insert(0, src_w.get().strip())

            # Valores
            valores_cfg = {}
            w = _g('_ent_precio_carpeta')
            if w is not None:
                try:
                    valores_cfg["precio_carpeta"] = int(w.get().strip())
                except ValueError:
                    valores_cfg["precio_carpeta"] = 49000
            w = _g('_ent_ata_tares')
            if w is not None:
                try:
                    valores_cfg["ata_tares"] = int(w.get().strip())
                except ValueError:
                    valores_cfg["ata_tares"] = 65000
            w = _g('_ent_clave_pdf')
            if w is not None:
                valores_cfg["clave_pdf"] = w.get().strip() or "123"
            w = _g('_ajustes_texto_guardas')
            if w is not None:
                guardas = [l.strip() for l in w.get("1.0", "end-1c").split("\n") if l.strip()]
                if guardas:
                    valores_cfg["guardas"] = guardas
            if valores_cfg:
                self.config["valores"] = {**self.config.get("valores", {}), **valores_cfg}

            # Súper Auto configuración
            super_cfg = {}
            for attr, cfg_key in [("_super_check_sobre", "sobre"), ("_super_check_permiso", "permiso"),
                                  ("_super_check_hoja_ruta", "hoja_ruta"), ("_super_check_recibo", "recibo_ata"),
                                  ("_super_check_guarda", "aplicar_guarda"),
                                  ("_super_check_planillas", "completar_planillas")]:
                w = _g(attr)
                if w is not None:
                    super_cfg[cfg_key] = w.get()
            if super_cfg:
                pasos = dict(self.config.get("super_auto", {}).get("pasos", {}))
                pasos.update(super_cfg)
                self.config["super_auto"] = {"pasos": pasos}

            # Seguridad — guardar contraseña encriptada
            if ent_pw1 is not None:
                self.config["seguridad"] = {
                    "password": self._encrypt_val(pw1, os.environ["MULTIAGENTE_SECRET_KEY"]),
                }

            # API Visión
            api_cfg = {}
            w = _g('_ent_vision_custom_models')
            if w is not None:
                raw = w.get("1.0", "end-1c")
                api_cfg["custom_models"] = [l.strip() for l in raw.split("\n") if l.strip()]
            for attr, cfg_key, conv in [
                ("_ent_vision_api_key", "api_key", lambda v: self._encrypt_val(v.strip(), self._clave_encriptacion())),
                ("_ent_vision_model", "model", lambda v: v.strip() or procesar_tickets.MODELO_VISION_DEFAULT),
                ("_ent_vision_temperature", "temperature", lambda v: float(v.strip() or "0.1")),
                ("_ent_vision_max_tokens", "max_tokens", lambda v: int(v.strip() or "4000")),
                ("_ent_vision_timeout", "timeout", lambda v: int(v.strip() or "60")),
            ]:
                w = _g(attr)
                if w is not None:
                    api_cfg[cfg_key] = conv(w.get())
            if api_cfg:
                self.config["api_vision"] = {**self.config.get("api_vision", {}), **api_cfg}

            # Parallel settings
            w = _g('_chk_parallel_enabled')
            if w is not None:
                self.config["api_vision"]["parallel_enabled"] = w.get() == 1
            model_states = {}
            for attr_name, widget in getattr(self, '_parallel_model_checks', {}).items():
                model_states[attr_name] = widget.get() == 1
            if model_states:
                self.config["api_vision"]["parallel_model_states"] = model_states

            # Apariencia — font level
            w = _g('_ent_font_level')
            if w is not None:
                try:
                    self.config["font_level"] = int(w.get())
                except (ValueError, TypeError):
                    self.config["font_level"] = 1

            self._guardar_config()
            self._ajustes_lbl_status.configure(text="✓ Configuración guardada correctamente.")
            self.after(3000, lambda: self._ajustes_lbl_status.configure(text="")
                       if self._ajustes_lbl_status.winfo_exists() else None)
        except Exception as e:
            self._ajustes_lbl_status.configure(text=f"✗ Error al guardar: {e}", text_color=Palette.ERROR)

    def _cargar_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    self.config = json.load(f)
            except Exception as e:
                self._log(f"ERROR al cargar configuración: {e}")

    def _guardar_config(self):
        try:
            self.config["window_geo"] = self.geometry()
            if hasattr(self, 'paned_window') and self.paned_window.winfo_exists():
                sash = self.paned_window.sash_coord(0)
                if self.panel_actual and sash[1] > 0:
                    self.config[f"sash_{self.panel_actual}"] = sash[1]
            # Guardar anchos de columna del árbol de mails
            if hasattr(self, '_mail_tree') and self._mail_tree.winfo_exists():
                self.config["mail_tree_columns"] = {
                    col: self._mail_tree.column(col, "width")
                    for col in ("sel", "asunto", "fecha", "adjuntos", "carpeta")
                }
            # Guardar anchos de columna de Control de Tickets
            if hasattr(self, 'tree_carga') and self.tree_carga.winfo_exists():
                self.config["tree_carga_columns"] = {
                    col: self.tree_carga.column(col, "width")
                    for col in ("cliente", "camion", "semi", "conductor", "dni",
                                "neto", "tara", "contenedor", "permiso", "estado")
                }
            # Guardar anchos de columna de Coordinacion
            if hasattr(self, 'tree_coordinacion') and self.tree_coordinacion.winfo_exists():
                self.config["tree_coordinacion_columns"] = {
                    col: self.tree_coordinacion.column(col, "width")
                    for col in ("carpeta", "giro", "cliente", "destino",
                                "buque", "viaje", "booking", "pto_descarga", "pto_final",
                                "fecha_of_pe", "fecha_carga", "peso_flexi", "estado")
                }
            # Guardar anchos de columna de Control Final
            if hasattr(self, 'tree_control_final') and self.tree_control_final.winfo_exists():
                self.config["tree_control_final_columns"] = {
                    col: self.tree_control_final.column(col, "width")
                    for col in ("cliente", "camion", "semi", "conductor", "dni",
                                "neto", "tara", "contenedor", "permiso", "estado", "salida_aduana")
                }
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2)
        except Exception as e:
            self._log(f"ERROR al guardar configuración: {e}")

    def _encrypt_val(self, value, key):
        if not value:
            return ""
        if value.startswith("enc::"):
            return value
        import hashlib
        import base64
        try:
            salt = os.urandom(16)
            derived = hashlib.pbkdf2_hmac('sha256', key.encode('utf-8'), salt, 10000, dklen=len(value))
            encrypted = bytes(a ^ b for a, b in zip(value.encode('utf-8'), derived))
            payload = base64.b64encode(salt + encrypted).decode('utf-8')
            return f"enc::{payload}"
        except Exception as e:
            self._log(f"ERROR al encriptar: {e}")
            return value

    def _decrypt_val(self, value, key):
        if not value or not value.startswith("enc::"):
            return value
        import hashlib
        import base64
        try:
            payload = value[5:]
            data = base64.b64decode(payload.encode('utf-8'))
            if len(data) < 16:
                return value
            salt = data[:16]
            encrypted = data[16:]
            derived = hashlib.pbkdf2_hmac('sha256', key.encode('utf-8'), salt, 10000, dklen=len(encrypted))
            decrypted = bytes(a ^ b for a, b in zip(encrypted, derived))
            return decrypted.decode('utf-8')
        except Exception as e:
            try:
                self._log(f"ERROR al desencriptar: {e}")
            except AttributeError:
                print(f"[WARN] Decrypt failed (log_queue not ready): {e}")
            return value

    def _clave_encriptacion(self):
        """Clave de encriptación desde variable de entorno, con fallback."""
        return os.environ.get("MULTIAGENTE_SECRET_KEY", "api_vision")

    def _decrypt_api_key(self, raw):
        """Desencripta API key probando clave actual, con fallback a la vieja."""
        result = self._decrypt_val(raw, self._clave_encriptacion())
        if result and result.startswith("enc::"):
            result = self._decrypt_val(raw, "api_vision")
        return result

    # ═══════════════════════════════════════════════════════════════════
    # CONFIRMACIÓN DE SALIDA
    # ═══════════════════════════════════════════════════════════════════
    def _confirmar_salida(self):
        """Cierra la aplicación con confirmación genérica."""
        ok = messagebox.askyesno(
            "Salir",
            "¿Está seguro de que desea salir del sistema?"
        )
        if not ok:
            return
        self._guardar_config()
        self.destroy()

# ── Punto de entrada ─────────────────────────────────────────────────────
def main():
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    ctk.set_widget_scaling(1.0)
    ctk.set_window_scaling(1.0)

    # Forzar escala DPI para pantallas HiDPI
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = App()
    app.mainloop()

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    main()
