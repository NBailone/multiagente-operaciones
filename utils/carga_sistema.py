# -*- coding: utf-8 -*-
"""CARGA SISTEMA — genera la planilla de carga para el sistema aduanero.

Puerto a Python de ``Creacion Excel/CARGA_SISTEMA.ps1``: a partir del Excel
con "Contenedores" en el nombre (hojas Choferes y Hoja Continuacion), genera
``CARGA SISTEMA <sufijo>.xlsx`` en la misma carpeta con el formato esperado
por SIMWEB.

Reglas heredadas del script original:
  - KG NETOS: si Peso Flexi tiene un numero -> 1; si no -> K. NETOS.
  - KG BRUTOS siempre sale de K. BRUTOS por contenedor.
  - CUIT calculado desde el DNI (modulo 11, prefijo 20, fallback 23).
  - Si la salida ya existia, se conserva el CUIT TRANSPORTE cargado (B3)
    y se sobrescribe sin backup.
"""

import os
import re
from copy import copy
from datetime import datetime

import xlrd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _normalizado(s):
    """Sin espacios ni guiones, en mayusculas. Ej: 'OCO 564' -> 'OCO564'."""
    if s is None:
        return ""
    return re.sub(r"[\s\-]", "", str(s)).upper()


def _cuit_from_dni(dni):
    """CUIT modulo 11 con pesos oficiales. Prefijo 20 (hombre), fallback 23."""
    dni8 = str(dni).zfill(8)
    pesos = (5, 4, 3, 2, 7, 6, 5, 4, 3, 2)

    def digito(pref):
        num = pref + dni8
        suma = sum(int(num[i]) * pesos[i] for i in range(10))
        resto = suma % 11
        if resto == 0:
            return "0"
        if resto == 1:
            return None  # señal de excepcion -> cambiar prefijo
        return str(11 - resto)

    pref = "20"
    dv = digito(pref)
    if dv is None:
        pref = "23"
        dv = digito(pref)
    if dv is None:
        dv = "0"
    return f"{pref}-{dni8}-{dv}"


def _celda_str(v):
    """Normaliza el valor crudo de una celda a string sin ruido de floats."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# Lectura de hojas (.xls via xlrd, .xlsx/.xlsm via openpyxl)
# Representacion uniforme: lista de (nro_fila_1based, {letra_col: valor_raw})
# ---------------------------------------------------------------------------

def _abrir_libro(ruta):
    """Devuelve (hojas_nombres, lector_hoja) donde lector_hoja(nombre) da filas.

    Las celdas fecha van bajo la clave '<col>__fecha' como datetime; el resto
    como valor crudo bajo '<col>'.
    """
    if ruta.lower().endswith(".xls"):
        book = xlrd.open_workbook(ruta)

        def leer(nombre):
            ws = book.sheet_by_name(nombre)
            filas = []
            for r in range(ws.nrows):
                celdas = {}
                for c in range(ws.ncols):
                    v = ws.cell_value(r, c)
                    if v == "" or v is None:
                        continue
                    letra = openpyxl.utils.get_column_letter(c + 1)
                    # ctype 3 = XL_CELL_DATE (fecha/hora real, no numero)
                    if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                        try:
                            celdas[letra + "__fecha"] = xlrd.xldate_as_datetime(v, book.datemode)
                            continue
                        except Exception:
                            pass
                    celdas[letra] = v
                if celdas:
                    filas.append((r + 1, celdas))
            return filas

        return book.sheet_names(), leer

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)

    def leer(nombre):
        ws = wb[nombre]
        filas = []
        for fila in ws.iter_rows():
            celdas = {}
            for cell in fila:
                if cell.value is None:
                    continue
                letra = openpyxl.utils.get_column_letter(cell.column)
                if isinstance(cell.value, datetime):
                    celdas[letra + "__fecha"] = cell.value
                else:
                    celdas[letra] = cell.value
            if celdas:
                filas.append((cell.row if fila else 0, celdas))
        wb.close()
        return filas

    return wb.sheetnames, leer


def _siguiente_no_vacio(celdas, letra):
    """Valor de la primera celda no vacia a la derecha de `letra` (max 40)."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    idx = column_index_from_string(letra)
    for k in range(1, 41):
        cand = get_column_letter(idx + k)
        if cand in celdas:
            v = _celda_str(celdas[cand])
            if v != "":
                return v
    return ""


def _mapear_columnas(filas, definiciones):
    """Busca la fila encabezado y mapea nombre_logico -> letra de columna.

    definiciones: lista de (nombre_logico, regex_header). La fila debe matchear
    el primer par (ancla); el resto se mapea si aparece.
    """
    ancla = definiciones[0][1]
    resto = definiciones[1:]
    for nro, celdas in filas:
        unidos = "|".join(
            _celda_str(v) for k, v in celdas.items() if not k.endswith("__fecha")
        ).upper()
        if not re.search(ancla, unidos):
            continue
        mapa = {}
        for letra, v in celdas.items():
            if letra.endswith("__fecha"):
                continue
            h = _celda_str(v).upper()
            for nombre, patron in resto:
                if nombre not in mapa and re.search(patron, h):
                    mapa[nombre] = letra
        return nro, mapa
    return None, {}


# ---------------------------------------------------------------------------
# Generador principal
# ---------------------------------------------------------------------------

def encontrar_excel_contenedores(carpeta):
    """Excel con 'Contenedor' en el nombre dentro de la carpeta, o None."""
    candidatos = []
    for nombre in sorted(os.listdir(carpeta)):
        if not re.search(r"Contenedor", nombre, re.IGNORECASE):
            continue
        if not re.search(r"\.xls[xmb]?$", nombre, re.IGNORECASE):
            continue
        if nombre.startswith("~$") or ".backup-" in nombre:
            continue
        if nombre.upper().startswith("CARGA SISTEMA"):
            continue
        candidatos.append(os.path.join(carpeta, nombre))
    return candidatos


def generar_carga_sistema(ruta_carpeta, log=None):
    """Genera CARGA SISTEMA <sufijo>.xlsx en `ruta_carpeta`.

    Args:
        ruta_carpeta: carpeta que contiene el Excel de contenedores.
        log: callable opcional (str) para mensajes de progreso.

    Returns:
        dict con claves: ok, mensaje, salida, filas, avisos, origen.
    """
    res = {"ok": False, "mensaje": "", "salida": "", "filas": 0, "avisos": [], "origen": ""}
    log = log or (lambda s: None)

    candidatos = encontrar_excel_contenedores(ruta_carpeta)
    if len(candidatos) == 0:
        res["mensaje"] = "No encontré Excel con 'Contenedores' en la carpeta."
        return res
    if len(candidatos) > 1:
        res["mensaje"] = (
            f"Hay {len(candidatos)} archivos con 'Contenedores'; dejá uno solo "
            f"por carpeta: {[os.path.basename(c) for c in candidatos]}"
        )
        return res
    origen = candidatos[0]
    res["origen"] = os.path.basename(origen)
    log(f"Archivo: {res['origen']}")

    hojas, leer = _abrir_libro(origen)

    # --- Hoja Choferes -------------------------------------------------------
    nombre_ch = next((h for h in hojas if h.lower().startswith("chof")), None)
    if not nombre_ch:
        res["mensaje"] = "El Excel no tiene hoja 'Choferes'."
        return res
    filas_ch = leer(nombre_ch)

    _, map_ch = _mapear_columnas(filas_ch, [
        ("ANCLA", r"PRECINTO"),
        ("PRECINTO", r"^PRECINTO"),
        ("TRACTOR", r"DOMINIO TRACTOR"),
        ("SEMI", r"DOMINIO SEMI"),
        ("CONTENEDOR", r"CONTENEDOR"),
        ("NOMBRE", r"NOMBRE"),
        ("DNI", r"DNI"),
    ])
    faltantes = [k for k in ("PRECINTO", "TRACTOR", "SEMI", "CONTENEDOR", "NOMBRE", "DNI") if k not in map_ch]
    if faltantes:
        res["mensaje"] = f"No encontré columnas {faltantes} en hoja Choferes."
        return res

    choferes = []
    for nro, celdas in filas_ch:
        dni = _celda_str(celdas.get(map_ch["DNI"], ""))
        dni_limpio = dni.split(".")[0] if re.fullmatch(r"\d+\.0", dni) else dni
        cuit = ""
        if re.fullmatch(r"\d{6,8}", dni_limpio):
            cuit = _cuit_from_dni(int(dni_limpio))
        elif re.fullmatch(r"\d{2}-\d{7,8}-\d", dni):
            cuit = dni
        else:
            continue
        choferes.append({
            "Precinto": _celda_str(celdas.get(map_ch["PRECINTO"], "")),
            "Tractor": _normalizado(celdas.get(map_ch["TRACTOR"], "")),
            "Semi": _normalizado(celdas.get(map_ch["SEMI"], "")),
            "Contenedor": _normalizado(celdas.get(map_ch["CONTENEDOR"], "")),
            "Nombre": _celda_str(celdas.get(map_ch["NOMBRE"], "")),
            "Cuit": cuit,
        })

    # --- Etiquetas de cabecera: PE, Fecha Carga, Peso Flexi ------------------
    pe = fecha_carga = peso_flexi = ""
    fecha_carga_dt = None
    for nro, celdas in filas_ch:
        for letra, v in list(celdas.items()):
            if letra.endswith("__fecha"):
                continue
            txt = _celda_str(v)
            if txt == "PE" and not pe:
                pe = _siguiente_no_vacio(celdas, letra)
            elif re.match(r"^Fecha Carga", txt, re.IGNORECASE) and not fecha_carga and not fecha_carga_dt:
                # la celda vecina puede ser fecha real (clave '<col>__fecha')
                from openpyxl.utils import column_index_from_string, get_column_letter
                vecina = get_column_letter(column_index_from_string(letra) + 1)
                if (vecina + "__fecha") in celdas:
                    fecha_carga_dt = celdas[vecina + "__fecha"]
                else:
                    fecha_carga = _siguiente_no_vacio(celdas, letra)
            elif re.match(r"^Peso Flexi", txt, re.IGNORECASE) and not peso_flexi:
                peso_flexi = _siguiente_no_vacio(celdas, letra)

    # --- Hoja Continuación ---------------------------------------------------
    nombre_ct = next((h for h in hojas if re.match(r"hoja continu", h.lower())), None)
    if not nombre_ct:
        res["mensaje"] = "El Excel no tiene hoja 'Hoja Continuacion'."
        return res
    filas_ct = leer(nombre_ct)

    _, map_ct = _mapear_columnas(filas_ct, [
        ("ANCLA", r"CONTENEDOR"),
        ("CONTENEDOR", r"^CONTENEDOR"),
        ("PRECINTO", r"^PRECINTO"),
        ("NETO", r"NETOS"),
        ("BRUTO", r"BRUTOS"),
    ])
    faltantes = [k for k in ("CONTENEDOR", "NETO", "BRUTO") if k not in map_ct]
    if faltantes:
        res["mensaje"] = f"No encontré columnas {faltantes} en Hoja Continuacion."
        return res

    neto_por_cont, bruto_por_cont = {}, {}
    neto_por_prec, bruto_por_prec = {}, {}
    for nro, celdas in filas_ct:
        cont = _normalizado(celdas.get(map_ct["CONTENEDOR"], ""))
        if not cont:
            continue
        unidos = "|".join(_celda_str(v) for k, v in celdas.items() if not k.endswith("__fecha"))
        if re.search(r"total", unidos, re.IGNORECASE):
            continue
        neto = _celda_str(celdas.get(map_ct["NETO"], ""))
        bruto = _celda_str(celdas.get(map_ct["BRUTO"], ""))
        neto_por_cont.setdefault(cont, neto)
        bruto_por_cont.setdefault(cont, bruto)
        prec = _celda_str(celdas.get(map_ct.get("PRECINTO", ""), ""))
        if prec:
            neto_por_prec.setdefault(prec, neto)
            bruto_por_prec.setdefault(prec, bruto)

    # --- Armar filas de salida ----------------------------------------------
    peso_flexi_num = 0.0
    m = re.match(r"^(\d+(?:[.,]\d+)?)$", peso_flexi)
    if m:
        peso_flexi_num = float(m.group(1).replace(",", "."))

    salida_filas, avisos = [], []
    for ch in choferes:
        neto = bruto = ""
        if ch["Contenedor"] in neto_por_cont:
            neto = neto_por_cont[ch["Contenedor"]]
            bruto = bruto_por_cont[ch["Contenedor"]]
        elif ch["Precinto"] in neto_por_prec:
            neto = neto_por_prec[ch["Precinto"]]
            bruto = bruto_por_prec[ch["Precinto"]]
        else:
            avisos.append(
                f"Contenedor {ch['Contenedor'] or '(sin dato)'} / precinto "
                f"{ch['Precinto'] or '(sin dato)'} ({ch['Nombre']}): sin KG en Hoja Continuacion"
            )
        kg_neto = "1" if peso_flexi_num != 0 else neto
        salida_filas.append([ch["Precinto"], ch["Tractor"], ch["Semi"],
                             ch["Contenedor"], ch["Nombre"], ch["Cuit"], kg_neto, bruto])

    # --- Nombre de salida ----------------------------------------------------
    base = os.path.splitext(os.path.basename(origen))[0]
    m = re.search(r"(\d+)\s*T", os.path.basename(origen))
    sufijo = m.group(1) + "T" if m else base
    out_name = f"CARGA SISTEMA {sufijo}.xlsx"
    out_path = os.path.join(ruta_carpeta, out_name)

    # --- Conservar CUIT TRANSPORTE ya cargado -------------------------------
    cuit_transporte = ""
    if os.path.isfile(out_path):
        try:
            wb_prev = openpyxl.load_workbook(out_path)
            ws_prev = wb_prev.worksheets[0]
            if "CUIT" in str(ws_prev["A3"].value or "").upper():
                cuit_transporte = _celda_str(ws_prev["B3"].value)
            wb_prev.close()
        except Exception:
            pass

    log(f"Datos: PE={pe} | Fecha Carga={fecha_carga_dt or fecha_carga} | "
        f"Peso Flexi={peso_flexi or '-'} | Choferes={len(salida_filas)}")
    for a in avisos:
        log(f"AVISO: {a}")

    if not salida_filas:
        res["ok"] = True
        res["filas"] = 0
        res["mensaje"] = "No se encontraron choferes con DNI en la hoja Choferes; no se generó nada."
        res["salida"] = ""
        return res

    # --- Escribir xlsx -------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    anchos = {"A": 22.28, "B": 22.14, "C": 16.14, "D": 21.14,
              "E": 27.57, "F": 16.43, "G": 15.29, "H": 13.86}
    for letra, w in anchos.items():
        ws.column_dimensions[letra].width = w

    fino = Side(style="thin", color="FF404040")
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    fuente_base = Font(name="Arial Narrow", size=11)
    fuente_bold = Font(name="Arial Narrow", size=11, bold=True)
    fill_header = PatternFill("solid", fgColor="FFD9E1F2")
    centro = Alignment(horizontal="center", vertical="center")
    izq = Alignment(horizontal="left", vertical="center")
    der = Alignment(horizontal="right", vertical="center")

    def celda_texto(fila, letra, val, bold=False, borde_on=True, centro_on=True):
        c = ws[f"{letra}{fila}"]
        c.value = val
        c.font = fuente_bold if bold else fuente_base
        if borde_on:
            c.border = borde
        if centro_on:
            c.alignment = centro
        return c

    def celda_cabecera(fila, letra, val, alineacion):
        """Etiqueta/valor de cabecera (filas 1-3): negrita + alineacion fija."""
        c = ws[f"{letra}{fila}"]
        c.value = val
        c.font = fuente_bold
        c.alignment = alineacion
        return c

    # Fila 1: Fecha Carga | Fila 2: PE | Fila 3: CUIT TRANSPORTE
    # Etiqueta en A (negrita, izquierda) — valor en B (negrita, derecha)
    celda_cabecera(1, "A", "Fecha Carga", izq)
    c_b1 = ws["B1"]
    c_b1.font = fuente_bold
    c_b1.alignment = der
    if fecha_carga_dt is not None:
        c_b1.value = fecha_carga_dt
        c_b1.number_format = "DD/MM/YYYY"
    else:
        c_b1.value = fecha_carga

    celda_cabecera(2, "A", "PE", izq)
    celda_cabecera(2, "B", pe, der)

    celda_cabecera(3, "A", "CUIT TRANSPORTE ", izq)
    if cuit_transporte:
        celda_cabecera(3, "B", cuit_transporte, der)

    # Fila 5: cabecera de tabla
    headers = ["PRECINTO ADUANA", "DOMINIO TRACTOR", "DOMINIO SEMI",
               "NUMERO  DEL CONTENEDOR", "NOMBRE DEL CHOFER", "CUIL  DEL CHOFER",
               "KG NETOS", "KG BRUOS "]
    ws.row_dimensions[5].height = 33
    for i, h in enumerate(headers):
        letra = chr(ord("A") + i)
        celda_texto(5, letra, h, bold=True)

    # Filas de datos
    fila = 6
    for vals in salida_filas:
        ws.row_dimensions[fila].height = 25
        for i, val in enumerate(vals):
            letra = chr(ord("A") + i)
            if i >= 6 and re.match(r"^\d+(\.\d+)?$", str(val)):
                c = ws[f"{letra}{fila}"]
                c.value = float(val) if "." in str(val) else int(float(val))
            else:
                celda_texto(fila, letra, val)
            if i >= 6:
                ws[f"{letra}{fila}"].border = borde
                ws[f"{letra}{fila}"].alignment = centro
        fila += 1

    try:
        wb.save(out_path)
    except PermissionError:
        res["mensaje"] = f"'{out_name}' está abierto en Excel. Cerralo y reintentá."
        res["avisos"] = avisos
        return res

    res["ok"] = True
    res["salida"] = out_name
    res["filas"] = len(salida_filas)
    res["avisos"] = avisos
    res["mensaje"] = f"{out_name}: {len(salida_filas)} chofer(es)"
    return res


def generar_lote(carpetas, log=None):
    """Genera en varias carpetas. Devuelve lista de resultados."""
    log = log or (lambda s: None)
    resultados = []
    for carpeta in carpetas:
        nombre = os.path.basename(carpeta.rstrip("\\/"))
        log(f"── {nombre}")
        res = generar_carga_sistema(carpeta, log=log)
        res["carpeta"] = nombre
        resultados.append(res)
        log(("✓ " if res["ok"] else "✗ ") + res["mensaje"])
    return resultados
